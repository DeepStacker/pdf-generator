# ============================================================
# FINAL PRODUCTION VERSION
# GOLD LOAN AUDIT PDF + SMART AUDIT EXCEL GENERATOR
#
# INSTALL:
# pip install pandas openpyxl reportlab
# ============================================================

import os
import re
import uuid
from pathlib import Path

import pandas as pd

# ============================================================
# OPENPYXL
# ============================================================

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)

from openpyxl.utils import get_column_letter

from openpyxl.formatting.rule import FormulaRule

from openpyxl.worksheet.page import PageMargins


# ============================================================
# REPORTLAB
# ============================================================

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    PageBreak,
)

from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# ============================================================
# CONFIG
# ============================================================

PDF_OUTPUT_DIR = "output_pdfs"

EXCEL_OUTPUT_DIR = "output_excels"

ROWS_PER_PAGE = 13


# ============================================================
# NORMAL SHEET COLUMNS
# ============================================================

NORMAL_COLUMNS = {
    "loan_no": "SVS_LOAN_NO",
    "sole_id": "SOLE_ID",
    "branch": "BRANCH_NAME",
}


# ============================================================
# JSR SHEET COLUMNS
# ============================================================

JSR_COLUMNS = {

    "loan_no": "LOAN NO",

    "loan_date": "DATE OF LOAN",

    "customer": "CUSTOMER NAME",

    "gross_weight": "GROSSWT",

    "net_weight": "NETWT",

    "sanctioned_amount": "SANCTIONED AMOUNT",

    "ornament": "ORNAMENT DETAILS",

    "quantity": "QUANTITY",

    "carat_master": "CARAT MASTER",

    "old_packet_no": "NEW GOLD PACKET NO"
}


# ============================================================
# PDF STYLES
# ============================================================

header_style = ParagraphStyle(
    "header_style",
    fontName="Helvetica-Bold",
    fontSize=7.2,
    alignment=TA_CENTER,
    leading=8,
    wordWrap='CJK'
)

body_style = ParagraphStyle(
    "body_style",
    fontName="Helvetica",
    fontSize=6.2,
    alignment=TA_CENTER,
    leading=7,
    wordWrap='CJK'
)

left_body_style = ParagraphStyle(
    "left_body_style",
    fontName="Helvetica",
    fontSize=6.2,
    alignment=TA_LEFT,
    leading=7,
    wordWrap='CJK'
)


# ============================================================
# HELPERS
# ============================================================

def normalize_columns(df):

    df.columns = [
        str(col).strip().upper()
        for col in df.columns
    ]

    return df


def clean_text(value):

    if pd.isna(value):
        return ""

    value = str(value)

    value = re.sub(
        r'([a-z])([A-Z])',
        r'\1 \2',
        value
    )

    value = re.sub(
        r'\s+',
        ' ',
        value
    )

    return value.strip()


def safe_value(value):

    if pd.isna(value):
        return ""

    return str(value).strip()


def safe_float(value):

    if pd.isna(value):
        return ""

    try:

        value = float(value)

        if value.is_integer():
            return str(int(value))

        return f"{value:.3f}"

    except:
        return str(value)


def format_date(value):

    if pd.isna(value):
        return ""

    try:

        parsed = pd.to_datetime(
            value,
            errors="coerce"
        )

        if pd.isna(parsed):
            return str(value)

        return parsed.strftime("%d-%m-%Y")

    except:
        return str(value)


def base_sheet_name(sheet_name):

    name = sheet_name.strip()

    if name.upper().endswith("_JSR"):
        return name[:-4].strip().lower()

    return name.lower()


def generate_verification_id():

    return str(uuid.uuid4().int)[:10]


def sanitize_filename(name):

    return re.sub(
        r'[\\/*?:"<>|]',
        "_",
        str(name)
    )


# ============================================================
# VALIDATION
# ============================================================

def validate_required_columns(
    df,
    required_columns,
    sheet_name
):

    missing = [
        col
        for col in required_columns
        if col not in df.columns
    ]

    if missing:

        raise Exception(
            f"Missing columns in {sheet_name}: "
            f"{missing}"
        )


# ============================================================
# BUILD MASTER DATAFRAME
# ============================================================

def build_master_dataframe(file_path):

    excel = pd.ExcelFile(file_path)

    sheet_map = {}

    for sheet in excel.sheet_names:

        base = base_sheet_name(sheet)

        if base not in sheet_map:

            sheet_map[base] = {
                "normal": None,
                "jsr": None
            }

        if sheet.upper().endswith("_JSR"):
            sheet_map[base]["jsr"] = sheet
        else:
            sheet_map[base]["normal"] = sheet

    final_frames = []

    for base, sheets in sheet_map.items():

        normal_sheet = sheets["normal"]
        jsr_sheet = sheets["jsr"]

        if not normal_sheet or not jsr_sheet:

            print(
                f"Skipping incomplete pair: {base}"
            )

            continue

        try:

            print(
                f"Processing: "
                f"{normal_sheet} <--> {jsr_sheet}"
            )

            normal_df = pd.read_excel(
                file_path,
                sheet_name=normal_sheet
            )

            jsr_df = pd.read_excel(
                file_path,
                sheet_name=jsr_sheet
            )

            normal_df = normalize_columns(normal_df)
            jsr_df = normalize_columns(jsr_df)

            required_normal = [
                NORMAL_COLUMNS["loan_no"],
                NORMAL_COLUMNS["sole_id"],
                NORMAL_COLUMNS["branch"],
            ]

            required_jsr = [

                JSR_COLUMNS["loan_no"],

                JSR_COLUMNS["loan_date"],

                JSR_COLUMNS["customer"],

                JSR_COLUMNS["gross_weight"],

                JSR_COLUMNS["net_weight"],

                JSR_COLUMNS["sanctioned_amount"],

                JSR_COLUMNS["ornament"],

                JSR_COLUMNS["quantity"],

                JSR_COLUMNS["carat_master"],

                JSR_COLUMNS["old_packet_no"]
            ]

            validate_required_columns(
                normal_df,
                required_normal,
                normal_sheet
            )

            validate_required_columns(
                jsr_df,
                required_jsr,
                jsr_sheet
            )

            normal_selected = normal_df[
                required_normal
            ].copy()

            jsr_selected = jsr_df[
                required_jsr
            ].copy()

            merged = pd.merge(

                normal_selected,
                jsr_selected,

                left_on=NORMAL_COLUMNS["loan_no"],
                right_on=JSR_COLUMNS["loan_no"],

                how="left"
            )

            merged["_GROUP_KEY"] = (

                merged[NORMAL_COLUMNS["loan_no"]]
                .astype(str)
                .str.strip()
            )

            merged = merged.sort_values(
                by=["_GROUP_KEY"]
            ).reset_index(drop=True)

            unique_keys = (
                merged["_GROUP_KEY"]
                .drop_duplicates()
                .tolist()
            )

            serial_map = {

                key: idx + 1
                for idx, key in enumerate(unique_keys)
            }

            merged["SR_NO"] = merged[
                "_GROUP_KEY"
            ].map(serial_map)

            merged.drop(
                columns=["_GROUP_KEY"],
                inplace=True
            )

            final_frames.append(merged)

        except Exception as e:

            print(
                f"Error processing {base}: {e}"
            )

    if not final_frames:
        raise Exception("No valid data found")

    final_df = pd.concat(
        final_frames,
        ignore_index=True
    )

    return final_df


# ============================================================
# TABLE DATA
# ============================================================

def create_table_data(
    rows,
    branch_name,
    verification_id,
    branch_sole_id,
    page_no,
    total_pages,
    total_accounts
):

    data = []

    data.append([

        Paragraph("<b>BRANCH NAME</b>", header_style),
        "",

        Paragraph(
            safe_value(branch_name),
            header_style
        ),

        Paragraph("<b>BRANCH SOL ID</b>", header_style),
        "",

        Paragraph(
            safe_value(branch_sole_id),
            header_style
        ),

        Paragraph("<b>BRANCH In Time</b>", header_style),
        "",
        "",

        Paragraph(
            f"<b>PAGE</b><br/>{page_no}/{total_pages}",
            header_style
        ),

        Paragraph(
            f"<b>TOTAL</b><br/>{total_accounts}",
            header_style
        ),

        Paragraph("<b>Audit Start Time</b>", header_style),
        "",

        "",
        "",

        Paragraph("<b>Audit Completed Time</b>", header_style),
        "",

        ""
    ])

    data.append([

        Paragraph("Sr No", header_style),

        Paragraph("ACCOUNT NUMBER", header_style),

        Paragraph("APPLICANT NAME", header_style),

        Paragraph(
            "<b>Branch Valuation Details</b>",
            header_style
        ),

        "",
        "",
        "",
        "",
        "",

        Paragraph(
            f"<b>Verification Details</b><br/>{verification_id}",
            header_style
        ),

        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ])

    data.append([

        "",
        "",
        "",

        Paragraph("ORNAMENT DETAILS", header_style),

        Paragraph("SANCTIONED AMOUNT", header_style),

        Paragraph("OLD PACKET NO", header_style),

        Paragraph("DATE OF LOAN", header_style),

        Paragraph("QTY", header_style),

        Paragraph("GROSS WEIGHT", header_style),

        Paragraph("TARE WEIGHT", header_style),

        Paragraph("NEW PACKET", header_style),

        Paragraph("CARRAT", header_style),

        Paragraph("ACTUAL QTY", header_style),

        Paragraph("ACTUAL GROSS", header_style),

        Paragraph("STONE WT", header_style),

        Paragraph("IMPURITY", header_style),

        Paragraph("ACTUAL NET WT", header_style),

        Paragraph("REMARK", header_style)
    ])

    for _, row in rows.iterrows():

        customer_name = clean_text(
            row.get(
                JSR_COLUMNS["customer"],
                ""
            )
        )

        ornament = clean_text(
            row.get(
                JSR_COLUMNS["ornament"],
                ""
            )
        )

        data.append([

            Paragraph(
                safe_value(row["SR_NO"]),
                body_style
            ),

            Paragraph(
                safe_value(
                    row.get(
                        NORMAL_COLUMNS["loan_no"],
                        ""
                    )
                ),
                body_style
            ),

            Paragraph(
                customer_name,
                left_body_style
            ),

            Paragraph(
                ornament,
                left_body_style
            ),

            Paragraph(
                safe_float(
                    row.get(
                        JSR_COLUMNS["sanctioned_amount"],
                        ""
                    )
                ),
                body_style
            ),

            Paragraph(
                safe_value(
                    row.get(
                        JSR_COLUMNS["old_packet_no"],
                        ""
                    )
                ),
                body_style
            ),

            Paragraph(
                format_date(
                    row.get(
                        JSR_COLUMNS["loan_date"],
                        ""
                    )
                ),
                body_style
            ),

            Paragraph(
                safe_float(
                    row.get(
                        JSR_COLUMNS["quantity"],
                        ""
                    )
                ),
                body_style
            ),

            Paragraph(
                safe_float(
                    row.get(
                        JSR_COLUMNS["gross_weight"],
                        ""
                    )
                ),
                body_style
            ),

            "", "", "", "", "", "", "", ""
        ])

    return data


# ============================================================
# SUMMARY PAGE
# ============================================================

def generate_summary_page(
    branch_name,
    total_accounts,
    branch_sole_id,
    verification_id
):

    data = [

        [
            "SOL ID",
            branch_sole_id,
            "BRANCH NAME",
            branch_name
        ],

        [
            "BANK NAME",
            "Equitas Small Finance Bank",
            "AUDIT DATE",
            ""
        ],

        [
            "BRANCH IN TIME",
            "",
            "AUDIT START TIME",
            ""
        ],

        [
            "AUDIT COMPLETED TIME",
            "",
            "VERIFICATION ID",
            verification_id
        ],

        [
            "TOTAL ACCOUNTS RECEIVED",
            str(total_accounts),

            "TOTAL ACCOUNTS AUDITED",
            str(total_accounts)
        ],

        [
            "AUDITOR NAME",
            "",
            "AUDITOR SIGNATURE",
            ""
        ],

        [
            "BRANCH MANAGER NAME",
            "",
            "BRANCH MANAGER SIGNATURE",
            ""
        ],

        [
            "REMARKS / OBSERVATIONS",
            "",
            "",
            ""
        ]
    ]

    table = Table(

        data,

        colWidths=[190, 220, 190, 220],

        rowHeights=[
            40,
            40,
            55,
            55,
            45,
            75,
            75,
            140
        ]
    )

    table.setStyle(TableStyle([

        ("BOX", (0, 0), (-1, -1),
         1.5, colors.black),

        ("GRID", (0, 0), (-1, -1),
         1.2, colors.black),

        ("FONTNAME", (0, 0), (-1, -1),
         "Helvetica"),

        ("FONTNAME", (0, 0), (0, -1),
         "Helvetica-Bold"),

        ("FONTNAME", (2, 0), (2, -1),
         "Helvetica-Bold"),

        ("FONTSIZE", (0, 0), (-1, -1),
         12),

        ("VALIGN", (0, 0), (-1, -1),
         "MIDDLE"),

        ("ALIGN", (0, 0), (-1, -1),
         "LEFT"),

        ("SPAN", (1, 7), (3, 7)),

        ("TOPPADDING", (0, 0), (-1, -1), 10),

        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),

        ("LEFTPADDING", (0, 0), (-1, -1), 12),

        ("RIGHTPADDING", (0, 0), (-1, -1), 12),

    ]))

    return table


# ============================================================
# PDF GENERATION
# ============================================================

def generate_branch_pdf(
    branch_name,
    branch_df
):

    os.makedirs(
        PDF_OUTPUT_DIR,
        exist_ok=True
    )

    safe_branch = sanitize_filename(
        branch_name
    )

    output_path = os.path.join(
        PDF_OUTPUT_DIR,
        f"{safe_branch}_Audit_Report.pdf"
    )

    doc = SimpleDocTemplate(

        output_path,

        pagesize=landscape(A4),

        leftMargin=4,
        rightMargin=4,
        topMargin=4,
        bottomMargin=4,

        title=f"{branch_name} Audit Report",
        author="Audit System",
        subject="Gold Loan Audit Verification",
    )

    elements = []

    branch_df = branch_df.sort_values(
        by=["SR_NO"]
    )

    usable_width = landscape(A4)[0] - 20

    weights = [

        2, 7, 11,
        9, 5, 6, 5, 3, 4,
        5, 7, 3, 3, 5, 5, 4,
        5, 5
    ]

    total_weight = sum(weights)

    col_widths = [
        usable_width * (w / total_weight)
        for w in weights
    ]

    total_rows = len(branch_df)

    total_pages = (
        total_rows // ROWS_PER_PAGE
    ) + (
        1 if total_rows % ROWS_PER_PAGE else 0
    )

    verification_id = generate_verification_id()

    branch_sole_id = safe_value(
        branch_df.iloc[0].get(
            NORMAL_COLUMNS["sole_id"],
            ""
        )
    )

    page_no = 1

    for start in range(
        0,
        total_rows,
        ROWS_PER_PAGE
    ):

        end = start + ROWS_PER_PAGE

        page_rows = branch_df.iloc[start:end]

        table_data = create_table_data(

            page_rows,
            branch_name,
            verification_id,
            branch_sole_id,
            page_no,
            total_pages,
            total_rows
        )

        row_heights = [
            25,
            16,
            35
        ] + [36] * len(page_rows)

        table = Table(

            table_data,

            colWidths=col_widths,

            rowHeights=row_heights,

            repeatRows=3
        )

        style = TableStyle([

            ("BOX", (0, 0), (-1, -1),
             1.2, colors.black),

            ("GRID", (0, 0), (-1, -1),
             0.8, colors.black),

            ("SPAN", (0, 0), (1, 0)),
            ("SPAN", (3, 0), (4, 0)),
            ("SPAN", (6, 0), (8, 0)),
            ("SPAN", (11, 0), (12, 0)),
            ("SPAN", (15, 0), (16, 0)),

            ("SPAN", (0, 1), (0, 2)),
            ("SPAN", (1, 1), (1, 2)),
            ("SPAN", (2, 1), (2, 2)),

            ("SPAN", (3, 1), (8, 1)),
            ("SPAN", (9, 1), (16, 1)),

            ("BACKGROUND", (0, 0), (-1, 0),
             colors.HexColor("#d9d9d9")),

            ("BACKGROUND", (3, 1), (8, 1),
             colors.HexColor("#ffe699")),

            ("BACKGROUND", (3, 2), (8, 2),
             colors.HexColor("#fff2cc")),

            ("BACKGROUND", (9, 1), (16, 1),
             colors.HexColor("#9dc3e6")),

            ("BACKGROUND", (9, 2), (16, 2),
             colors.HexColor("#ddebf7")),

            ("BACKGROUND", (0, 1), (2, 2),
             colors.HexColor("#d9d9d9")),

            ("LINEBELOW", (0, 2), (-1, 2),
             1.2, colors.black),

            ("FONTNAME", (0, 0), (-1, 2),
             "Helvetica-Bold"),

            ("ALIGN", (0, 0), (-1, -1),
             "CENTER"),

            ("VALIGN", (0, 0), (-1, -1),
             "MIDDLE"),

            ("TOPPADDING", (0, 0), (-1, -1), 2),

            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),

            ("LEFTPADDING", (0, 0), (-1, -1), 1),

            ("RIGHTPADDING", (0, 0), (-1, -1), 1),

        ])

        table.setStyle(style)

        elements.append(table)

        if end < total_rows:
            elements.append(PageBreak())

        page_no += 1

    elements.append(PageBreak())

    summary_table = generate_summary_page(

        branch_name,
        total_rows,
        branch_sole_id,
        verification_id
    )

    elements.append(summary_table)

    doc.build(elements)

    print(f"Generated PDF: {output_path}")


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_branch_excel(
    branch_name,
    branch_df
):

    os.makedirs(
        EXCEL_OUTPUT_DIR,
        exist_ok=True
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Audit"

    thin = Side(
        border_style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left_align = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    bold = Font(
        bold=True,
        size=11
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFE699"
    )

    blue_fill = PatternFill(
        "solid",
        fgColor="BDD7EE"
    )

    green_fill = PatternFill(
        "solid",
        fgColor="C6E0B4"
    )

    grey_fill = PatternFill(
        "solid",
        fgColor="BFBFBF"
    )

    red_fill = PatternFill(
        "solid",
        fgColor="F4CCCC"
    )

    # ========================================================
    # TOP HEADERS
    # ========================================================

    top_headers = [

        "BRANCH SOL ID",
        "BRANCH NAME",
        "STATE",
        "VERIFIER NAME",
        "VERIFICATION DATE",
        "BRANCH IN TIME",
        "AUDIT START TIME",
        "AUDIT END TIME",
        "TOTAL ACCOUNTS RECEIVED",
        "TOTAL ACCOUNTS AUDITED",
        "REMAINING ACCOUNTS",
        "REMARKS"
    ]

    for idx, value in enumerate(top_headers, start=1):

        c = ws.cell(
            row=1,
            column=idx,
            value=value
        )

        c.font = bold
        c.border = border
        c.alignment = center
        c.fill = grey_fill

    branch_sole_id = safe_value(
        branch_df.iloc[0].get(
            NORMAL_COLUMNS["sole_id"],
            ""
        )
    )

    second_row = [

        branch_sole_id,
        branch_name,
        "",
        "",
        "",
        "",
        "",
        "",
        branch_df["SR_NO"].nunique(),
        branch_df["SR_NO"].nunique(),
        "",
        ""
    ]

    for idx, value in enumerate(second_row, start=1):

        c = ws.cell(
            row=2,
            column=idx,
            value=value
        )

        c.border = border
        c.alignment = center

    # ========================================================
    # MAIN TABLE HEADERS
    # ========================================================

    headers = [

        "SR",
        "ACCOUNT NO",
        "APPLICANT NAME",

        "SANCTION AMT",
        "LOAN DATE",

        "OLD PKT NO",
        "NEW PKT NO",

        "BANK QTY",
        "ACTUAL QTY",
        "QTY DIFF",

        "BANK GROSS",
        "ACTUAL GROSS",
        "GROSS DIFF",

        "BANK NET",
        "ACTUAL NET",
        "NET DIFF",

        "MASTER CARAT",
        "ACTUAL CARAT",
        "CARAT DIFF",

        "REMARKS",
        "ADDITIONAL DETAILS",

        "ORNAMENT DETAILS"
    ]

    header_row = 4

    for idx, value in enumerate(headers, start=1):

        c = ws.cell(
            row=header_row,
            column=idx,
            value=value
        )

        c.font = bold
        c.border = border
        c.alignment = center

        # BANK DATA
        if idx in [
            1,2,3,4,5,6,8,11,14,17,22
        ]:
            c.fill = yellow_fill

        # MANUAL ENTRY
        elif idx in [
            7,9,12,15,18
        ]:
            c.fill = blue_fill

        # SYSTEM GENERATED
        else:
            c.fill = green_fill

    # ========================================================
    # DATA ROWS
    # ========================================================

    row_num = 5

    LEFT_ALIGN_COLS = [3,21,22]

    for _, row in branch_df.iterrows():

        values = [

            # A
            row["SR_NO"],

            # B
            safe_value(
                row.get(
                    NORMAL_COLUMNS["loan_no"],
                    ""
                )
            ),

            # C
            clean_text(
                row.get(
                    JSR_COLUMNS["customer"],
                    ""
                )
            ),

            # D
            safe_float(
                row.get(
                    JSR_COLUMNS["sanctioned_amount"],
                    ""
                )
            ),

            # E
            format_date(
                row.get(
                    JSR_COLUMNS["loan_date"],
                    ""
                )
            ),

            # F
            safe_value(
                row.get(
                    JSR_COLUMNS["old_packet_no"],
                    ""
                )
            ),

            # G
            "",

            # H
            safe_float(
                row.get(
                    JSR_COLUMNS["quantity"],
                    ""
                )
            ),

            # I
            "",

            # J
            f'=IF(I{row_num}="","",ROUND(I{row_num}-H{row_num},3))',

            # K
            safe_float(
                row.get(
                    JSR_COLUMNS["gross_weight"],
                    ""
                )
            ),

            # L
            "",

            # M
            f'=IF(L{row_num}="","",ROUND(L{row_num}-K{row_num},3))',

            # N
            safe_float(
                row.get(
                    JSR_COLUMNS["net_weight"],
                    ""
                )
            ),

            # O
            "",

            # P
            f'=IF(O{row_num}="","",ROUND(O{row_num}-N{row_num},3))',

            # Q
            safe_value(
                row.get(
                    JSR_COLUMNS["carat_master"],
                    ""
                )
            ),

            # R
            "",

            # S
            f'=IF(R{row_num}="","",IF(R{row_num}=Q{row_num},"DIFF: 0","DIFF: "&ROUND((R{row_num}-Q{row_num}),3)))',

            # T
            f'=IF(S{row_num}<>"DIFF: 0","Carat mismatch","OK - no discrepancy")',

            # U
            f'=IF(S{row_num}<>"DIFF: 0",'
            f'I{row_num}&" "&V{row_num}&" ("&TEXT(L{row_num},"0.000")&") - ACTUAL "&R{row_num}&" CARAT","")',

            # V
            clean_text(
                row.get(
                    JSR_COLUMNS["ornament"],
                    ""
                )
            )
        ]

        for col_idx, value in enumerate(values, start=1):

            c = ws.cell(
                row=row_num,
                column=col_idx,
                value=value
            )

            c.border = border

            if col_idx in LEFT_ALIGN_COLS:

                c.alignment = left_align

            else:

                c.alignment = center

        ws.row_dimensions[row_num].height = 32

        row_num += 1

    # ========================================================
    # COLUMN WIDTHS
    # ========================================================

    widths = {

        1: 6,
        2: 16,
        3: 24,

        4: 14,
        5: 14,

        6: 14,
        7: 14,

        8: 11,
        9: 11,
        10: 11,

        11: 13,
        12: 13,
        13: 13,

        14: 13,
        15: 13,
        16: 13,

        17: 12,
        18: 12,
        19: 14,

        20: 24,

        21: 45,

        22: 40
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # ========================================================
    # ROW HEIGHTS
    # ========================================================

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[4].height = 42

    # ========================================================
    # CONDITIONAL FORMATTING
    # ========================================================

    for col in ["J", "M", "P"]:

        ws.conditional_formatting.add(

            f"{col}5:{col}{row_num}",

            FormulaRule(
                formula=[f'{col}5<>0'],
                fill=red_fill
            )
        )

    ws.conditional_formatting.add(

        f"S5:S{row_num}",

        FormulaRule(
            formula=['S5<>"DIFF: 0"'],
            fill=red_fill
        )
    )

    # ========================================================
    # FILTER
    # ========================================================

    ws.auto_filter.ref = (
        f"A4:V{row_num}"
    )

    # ========================================================
    # FREEZE
    # ========================================================

    ws.freeze_panes = "D5"

    # ========================================================
    # PAGE SETUP
    # ========================================================

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.3,
        bottom=0.3
    )

    safe_branch = sanitize_filename(
        branch_name
    )

    output_path = os.path.join(
        EXCEL_OUTPUT_DIR,
        f"{safe_branch}_Audit.xlsx"
    )

    wb.save(output_path)

    print(f"Generated Excel: {output_path}")

# ============================================================
# MAIN
# ============================================================

def main(file_path):

    master_df = build_master_dataframe(
        file_path
    )

    branch_col = NORMAL_COLUMNS["branch"]

    grouped = master_df.groupby(
        branch_col
    )

    for branch_name, branch_df in grouped:

        try:

            generate_branch_pdf(
                branch_name,
                branch_df
            )

            generate_branch_excel(
                branch_name,
                branch_df
            )

        except Exception as e:

            print(
                f"Error processing "
                f"{branch_name}: {e}"
            )


# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":

    import argparse

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "file",
        help="Excel file path"
    )

    args = parser.parse_args()

    file_path = Path(args.file)

    if not file_path.exists():

        print(f"File not found: {file_path}")

    else:

        main(file_path)
