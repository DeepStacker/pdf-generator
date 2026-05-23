# ============================================================
# STAGE 2
# FINAL ENTERPRISE VERSION
# CONSOLIDATED ACCOUNT LEVEL AUDIT REPORT
# ============================================================

import os
import math
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)

from openpyxl.utils import get_column_letter

from openpyxl.worksheet.page import PageMargins

from openpyxl.formatting.rule import FormulaRule


# ============================================================
# OUTPUT
# ============================================================

OUTPUT_DIR = "stage2_output"


# ============================================================
# HELPERS
# ============================================================

def safe_float(v):

    if pd.isna(v):
        return 0.0

    try:

        if isinstance(v, str):

            v = v.strip()

            if not v:
                return 0.0

            if "DIFF:" in v.upper():

                v = (
                    v.upper()
                    .replace("DIFF:", "")
                    .strip()
                )

        value = float(v)

        if math.isnan(value):
            return 0.0

        if math.isinf(value):
            return 0.0

        return round(value, 3)

    except:
        return 0.0


def safe_str(v):

    if pd.isna(v):
        return ""

    return str(v).strip()


def avg_int(series):

    vals = []

    for x in series:

        try:

            val = safe_float(x)

            if val == 0:
                continue

            vals.append(val)

        except:
            continue

    if not vals:
        return ""

    avg = sum(vals) / len(vals)

    if math.isnan(avg):
        return ""

    return int(round(avg))


def merge_issue_only(series):

    vals = []

    for x in series:

        x = safe_str(x)

        if not x:
            continue

        if x.lower() == "ok - no discrepancy":
            continue

        if x not in vals:
            vals.append(x)

    if vals:
        return ", ".join(vals)

    return "OK - no discrepancy"


def merge_non_empty(series):

    vals = []

    for x in series:

        x = safe_str(x)

        if not x:
            continue

        if x not in vals:
            vals.append(x)

    return ", ".join(vals)


def first_non_empty(series):

    for x in series:

        x = safe_str(x)

        if x:
            return x

    return ""


def format_3(v):

    try:
        return f"{float(v):.3f}"
    except:
        return "0.000"


# ============================================================
# LOAD EXCEL
# ============================================================

def load_stage1_excel(file_path):

    df = pd.read_excel(
        file_path,
        header=3
    )

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    df = df.dropna(
        how="all"
    )

    df = df[
        df["ACCOUNT NO"]
        .notna()
    ]

    return df


# ============================================================
# CONSOLIDATION
# ============================================================

def consolidate_accounts(df):

    grouped = df.groupby(
        "ACCOUNT NO",
        dropna=False
    )

    rows = []

    sr_no = 1

    for acc_no, grp in grouped:

        try:

            # ================================================
            # SANCTION AMOUNT
            # DO NOT SUM
            # ================================================

            sanction_amt = safe_float(
                first_non_empty(
                    grp["SANCTION AMT"]
                )
            )

            # ================================================
            # QUANTITY
            # ================================================

            bank_qty = round(
                grp["BANK QTY"]
                .apply(safe_float)
                .sum(),
                3
            )

            actual_qty = round(
                grp["ACTUAL QTY"]
                .apply(safe_float)
                .sum(),
                3
            )

            # ================================================
            # GROSS
            # ================================================

            bank_gross = round(
                grp["BANK GROSS"]
                .apply(safe_float)
                .sum(),
                3
            )

            actual_gross = round(
                grp["ACTUAL GROSS"]
                .apply(safe_float)
                .sum(),
                3
            )

            # ================================================
            # NET
            # ================================================

            bank_net = round(
                grp["BANK NET"]
                .apply(safe_float)
                .sum(),
                3
            )

            actual_net = round(
                grp["ACTUAL NET"]
                .apply(safe_float)
                .sum(),
                3
            )

            # ================================================
            # REMARKS
            # ================================================

            remarks = merge_issue_only(
                grp["REMARKS"]
            )

            # ================================================
            # ADDITIONAL DETAILS
            # ================================================

            additional_details = merge_non_empty(
                grp["ADDITIONAL DETAILS"]
            )

            # ================================================
            # AVG CARAT
            # ================================================

            avg_carat = avg_int(
                grp["ACTUAL CARAT"]
            )

            row = {

                "S.NO":
                    sr_no,

                "ACCOUNT NUMBER":
                    safe_str(acc_no),

                "APPLICANT NAME":
                    first_non_empty(
                        grp["APPLICANT NAME"]
                    ),

                "SANCTIONED_AMOUNT":
                    format_3(
                        sanction_amt
                    ),

                "DATE_OF_LOAN":
                    first_non_empty(
                        grp["LOAN DATE"]
                    ),

                "OLD PACKET NO.":
                    first_non_empty(
                        grp["OLD PKT NO"]
                    ),

                "NEW PACKET NO.":
                    first_non_empty(
                        grp["NEW PKT NO"]
                    ),

                # ============================================
                # INTEGER COUNTS
                # ============================================

                "Bank Valuation No. Of Ornaments":
                    int(round(bank_qty)),

                "Actual No. of Ornaments":
                    int(round(actual_qty)),

                # ============================================
                # WEIGHTS
                # ============================================

                "Bank Valuation Gross Wt.":
                    format_3(
                        bank_gross
                    ),

                "Actual Gross Wt.":
                    format_3(
                        actual_gross
                    ),

                "Bank Valuation Net Wt.":
                    format_3(
                        bank_net
                    ),

                "Actual Net Wt.":
                    format_3(
                        actual_net
                    ),

                "REMARKS":
                    remarks,

                "ADDITIONAL DETAILS":
                    additional_details,

                "AVG CARAT":
                    avg_carat
            }

            rows.append(row)

            sr_no += 1

        except Exception as e:

            print(
                f"Error processing "
                f"account {acc_no}: {e}"
            )

    return pd.DataFrame(rows)


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_stage2_excel(
    original_file,
    consolidated_df
):

    os.makedirs(
        OUTPUT_DIR,
        exist_ok=True
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Consolidated Audit"

    # ========================================================
    # STYLES
    # ========================================================

    thin = Side(
        border_style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left_align = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    bold = Font(
        bold=True,
        size=11
    )

    grey_fill = PatternFill(
        "solid",
        fgColor="BFBFBF"
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFE699"
    )

    green_fill = PatternFill(
        "solid",
        fgColor="C6E0B4"
    )

    red_fill = PatternFill(
        "solid",
        fgColor="F4CCCC"
    )

    # ========================================================
    # COPY TOP HEADER
    # ========================================================

    old_wb = load_workbook(
        original_file,
        data_only=True
    )

    old_ws = old_wb.active

    for r in range(1, 3):

        for c in range(1, 13):

            val = old_ws.cell(
                row=r,
                column=c
            ).value

            cell = ws.cell(
                row=r,
                column=c,
                value=val
            )

            cell.border = border
            cell.alignment = center

            if r == 1:

                cell.fill = grey_fill
                cell.font = bold

    # ========================================================
    # HEADERS
    # ========================================================

    headers = [

        "S.NO",
        "ACCOUNT NUMBER",
        "APPLICANT NAME",

        "SANCTIONED_AMOUNT",
        "DATE_OF_LOAN",

        "OLD PACKET NO.",
        "NEW PACKET NO.",

        "Bank Valuation No. Of Ornaments",
        "Actual No. of Ornaments",
        "Difference in No. Of Ornaments",

        "Bank Valuation Gross Wt.",
        "Actual Gross Wt.",
        "Gross Weight Difference",

        "Bank Valuation Net Wt.",
        "Actual Net Wt.",
        "Net Weight Difference",

        "REMARKS",
        "ADDITIONAL DETAILS"
    ]

    header_row = 4

    for idx, h in enumerate(headers, start=1):

        c = ws.cell(
            row=header_row,
            column=idx,
            value=h
        )

        c.font = bold
        c.border = border
        c.alignment = center

        if idx <= 7:
            c.fill = yellow_fill
        else:
            c.fill = green_fill

    # ========================================================
    # DATA
    # ========================================================

    row_num = 5

    LEFT_ALIGN_COLS = [3,17,18]

    for _, row in consolidated_df.iterrows():

        vals = [

            # A
            row["S.NO"],

            # B
            row["ACCOUNT NUMBER"],

            # C
            row["APPLICANT NAME"],

            # D
            row["SANCTIONED_AMOUNT"],

            # E
            row["DATE_OF_LOAN"],

            # F
            row["OLD PACKET NO."],

            # G
            row["NEW PACKET NO."],

            # H
            row["Bank Valuation No. Of Ornaments"],

            # I
            row["Actual No. of Ornaments"],

            # J
            f'=ROUND(I{row_num}-H{row_num},0)',

            # K
            row["Bank Valuation Gross Wt."],

            # L
            row["Actual Gross Wt."],

            # M
            f'=ROUND(L{row_num}-K{row_num},3)',

            # N
            row["Bank Valuation Net Wt."],

            # O
            row["Actual Net Wt."],

            # P
            f'=ROUND(O{row_num}-N{row_num},3)',

            # Q
            row["REMARKS"],

            # R
            row["ADDITIONAL DETAILS"]
        ]

        for col_idx, v in enumerate(vals, start=1):

            cell = ws.cell(
                row=row_num,
                column=col_idx,
                value=v
            )

            cell.border = border

            if col_idx in LEFT_ALIGN_COLS:
                cell.alignment = left_align
            else:
                cell.alignment = center

        ws.row_dimensions[row_num].height = 38

        row_num += 1

    # ========================================================
    # WIDTHS
    # ========================================================

    widths = {

        1: 8,
        2: 18,
        3: 30,

        4: 18,
        5: 16,

        6: 18,
        7: 18,

        8: 18,
        9: 18,
        10: 18,

        11: 18,
        12: 18,
        13: 18,

        14: 18,
        15: 18,
        16: 18,

        17: 26,
        18: 65
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # ========================================================
    # ROW HEIGHTS
    # ========================================================

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[4].height = 42

    # ========================================================
    # CONDITIONAL FORMATTING
    # ========================================================

    ws.conditional_formatting.add(

        f"Q5:Q{row_num}",

        FormulaRule(
            formula=['Q5<>"OK - no discrepancy"'],
            fill=red_fill
        )
    )

    # ========================================================
    # FILTER
    # ========================================================

    ws.auto_filter.ref = (
        f"A4:R{row_num}"
    )

    # ========================================================
    # FREEZE
    # ========================================================

    ws.freeze_panes = "C5"

    # ========================================================
    # PAGE SETUP
    # ========================================================

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.3,
        bottom=0.3
    )

    # ========================================================
    # SAVE
    # ========================================================

    input_name = Path(
        original_file
    ).stem

    output_path = os.path.join(

        OUTPUT_DIR,

        f"{input_name}_Consolidated.xlsx"
    )

    wb.save(output_path)

    print(
        f"\nGenerated Stage-2 Report:\n"
        f"{output_path}"
    )


# ============================================================
# MAIN
# ============================================================

def main(file_path):

    print("\nLoading stage-1 audit excel...")

    df = load_stage1_excel(
        file_path
    )

    print(
        f"Loaded rows: {len(df)}"
    )

    print(
        "\nConsolidating accounts..."
    )

    consolidated_df = consolidate_accounts(
        df
    )

    print(
        f"Consolidated accounts: "
        f"{len(consolidated_df)}"
    )

    generate_stage2_excel(
        file_path,
        consolidated_df
    )


# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":

    import argparse

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "file",
        help="Stage-1 Audit Excel"
    )

    args = parser.parse_args()

    file_path = Path(args.file)

    if not file_path.exists():

        print(
            f"File not found: {file_path}"
        )

    else:

        main(file_path)
