#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Audit Engine Elite v5 — Headless Server & WSGI Portal

Professional-grade PDF report generator for IDFC FIRST Bank & Equitas Small Finance Bank gold-loan audits.
Replaces retro retro Tkinter visuals completely with a beautiful, high-fidelity responsiveSingle-Page browser dashboard,
resolving all Parallels DLL crash issues.
"""

import os
import sys
import time
import socket
import threading
import subprocess
import zipfile
import sqlite3
import logging
import json
import webbrowser
from datetime import datetime
import openpyxl

# Import local WSGI micro-framework
from bottle import route, run, request, response, static_file

# Import core business logic and web dashboard skeleton
import pdf_logic
import equitas_logic
import web_assets

# =========================================================
# VERSION & CONSTANTS
# =========================================================
VERSION = "5.2.167"
APP_TITLE = "Audit Engine v5.0"

# File logging setup
LOG_FILE = os.path.join(os.path.expanduser("~"), ".idfc_audit_engine.log")
file_logger = logging.getLogger("audit_engine")
file_logger.setLevel(logging.INFO)
try:
    _fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    _fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    file_logger.addHandler(_fh)
except OSError:
    pass

# =========================================================
# PLATFORM UTILITIES
# =========================================================
def open_path(path):
    """Open a file or folder in the system's default handler."""
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except OSError as e:
        file_logger.warning(f"Failed to open path: {path} — {e}")

def trigger_desktop_notification(title, message):
    """Trigger a native desktop notification balloon or bubble."""
    try:
        if sys.platform == "darwin":
            safe_msg = message.replace('"', '\\"').replace("'", "\\'")
            safe_title = title.replace('"', '\\"').replace("'", "\\'")
            cmd = ["osascript", "-e", f'display notification "{safe_msg}" with title "{safe_title}"']
            subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        elif sys.platform == "win32":
            ps_script = f"""
            [void] [System.Reflection.Assembly]::LoadWithPartialName("System.Windows.Forms");
            $notification = New-Object System.Windows.Forms.NotifyIcon;
            $notification.Icon = [System.Drawing.SystemIcons]::Information;
            $notification.BalloonTipIcon = "Info";
            $notification.BalloonTipTitle = "{title}";
            $notification.BalloonTipText = "{message}";
            $notification.Visible = $True;
            $notification.ShowBalloonTip(5000);
            """
            kwargs = {}
            kwargs["creationflags"] = 0x08000000
            subprocess.run(
                ["powershell", "-Command", ps_script],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                **kwargs
            )
    except Exception as ne:
        file_logger.warning(f"Failed to trigger desktop notification: {ne}")

def preprocess_mapped_excel(filepath, column_mappings, bank):
    """Creates a copy of the Excel file with columns renamed according to user mappings."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        target_sheet = None
        max_matches = -1
        
        mapped_headers_set = set(column_mappings.values())
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
                matches = sum(1 for h in headers if h in mapped_headers_set)
                if matches > max_matches:
                    max_matches = matches
                    target_sheet = sname
            except StopIteration:
                continue
                
        if not target_sheet:
            target_sheet = wb.sheetnames[0]
            
        ws = wb[target_sheet]
        
        first_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
        
        header_indices = {h: idx for idx, h in enumerate(headers, 1)}
        
        if bank == "IDFC First Bank":
            prospect_col = column_mappings.get('prospect')
            cuid_col = column_mappings.get('cuid')
            tare_col = column_mappings.get('tare')
            branch_col = column_mappings.get('branch')
            
            rename_map = {}
            if prospect_col: rename_map["Prospectno"] = prospect_col
            if cuid_col: rename_map["CUID"] = cuid_col
            if tare_col: rename_map["Tare Weight"] = tare_col
            if branch_col: rename_map["CurrentBranchName"] = branch_col
            
            for target_name, orig_name in rename_map.items():
                if orig_name in header_indices:
                    col_idx = header_indices[orig_name]
                    ws.cell(row=1, column=col_idx).value = target_name
            
            updated_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            if "CurrentBranch" not in updated_headers:
                branch_name_idx = updated_headers.index("CurrentBranchName") + 1 if "CurrentBranchName" in updated_headers else None
                if branch_name_idx:
                    new_col_idx = len(updated_headers) + 1
                    ws.cell(row=1, column=new_col_idx).value = "CurrentBranch"
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=new_col_idx).value = ws.cell(row=r, column=branch_name_idx).value
            
            updated_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "State" not in updated_headers:
                new_col_idx = len(updated_headers) + 1
                ws.cell(row=1, column=new_col_idx).value = "State"
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=new_col_idx).value = "UNKNOWN"
        else:
            svs_col = column_mappings.get('svs')
            sole_col = column_mappings.get('sole')
            branch_col = column_mappings.get('branch')
            loan_col = column_mappings.get('loan')
            
            rename_map = {}
            if svs_col: rename_map["SVS_LOAN_NO"] = svs_col
            if sole_col: rename_map["SOLE_ID"] = sole_col
            if branch_col: rename_map["BRANCH_NAME"] = branch_col
            if loan_col: rename_map["LOAN_NO"] = loan_col
            
            for target_name, orig_name in rename_map.items():
                if orig_name in header_indices:
                    col_idx = header_indices[orig_name]
                    ws.cell(row=1, column=col_idx).value = target_name
                    
        temp_dir = os.path.join(os.path.expanduser("~"), ".temp_audit_engine")
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, f"mapped_{os.path.basename(filepath)}")
        wb.save(temp_path)
        wb.close()
        return temp_path
    except Exception as e:
        file_logger.warning(f"Column mapping preprocessing failed: {e}")
        return filepath

# =========================================================
# DATABASE HANDLING
# =========================================================
DB_PATH = os.path.join(os.path.expanduser("~"), ".idfc_pdf_generator_v3.db")

def _connect_db():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS history
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      timestamp TEXT,
                      excel_name TEXT,
                      pdf_count INTEGER,
                      output_path TEXT,
                      audit_type TEXT)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS config
                     (key TEXT PRIMARY KEY, value TEXT)''')
    try:
        cursor.execute("ALTER TABLE history ADD COLUMN full_path TEXT")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()

def set_config(key, value):
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()
    conn.close()

def get_config(key, default=None):
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM config WHERE key = ?", (key,))
    res = cursor.fetchone()
    conn.close()
    return res[0] if res else default

def log_generation(excel_name, pdf_count, output_path, audit_type, full_path=None):
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO history (timestamp, excel_name, pdf_count, output_path, audit_type, full_path) VALUES (?, ?, ?, ?, ?, ?)",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), excel_name, pdf_count, output_path, audit_type, full_path)
    )
    conn.commit()
    conn.close()

def get_stats():
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*), SUM(pdf_count) FROM history")
    res = cursor.fetchone()
    conn.close()
    return (res[0] or 0, res[1] or 0)

def get_analytics():
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT audit_type, COUNT(*) FROM history GROUP BY audit_type")
    types = dict(cursor.fetchall())
    cursor.execute("SELECT strftime('%Y-%m-%d', timestamp), COUNT(*) FROM history GROUP BY 1 ORDER BY 1 DESC LIMIT 7")
    trend = cursor.fetchall()
    conn.close()
    return types, trend

def get_recent_history(search="", limit=100):
    conn = _connect_db()
    cursor = conn.cursor()
    if search:
        cursor.execute(
            "SELECT id, timestamp, excel_name, pdf_count, output_path, audit_type, full_path FROM history WHERE excel_name LIKE ? ORDER BY id DESC LIMIT ?",
            (f"%{search}%", limit)
        )
    else:
        cursor.execute(
            "SELECT id, timestamp, excel_name, pdf_count, output_path, audit_type, full_path FROM history ORDER BY id DESC LIMIT ?",
            (limit,)
        )
    res = cursor.fetchall()
    conn.close()
    return res

MAX_RECENT_FILES = 8

def _get_recent_files():
    files = []
    for i in range(MAX_RECENT_FILES):
        f = get_config(f"recent_file_{i}", "")
        if f and os.path.exists(f) and f not in files:
            files.append(f)
    return files

def _add_recent_file(filepath):
    files = _get_recent_files()
    filepath = os.path.abspath(os.path.normpath(filepath))
    files = [f for f in files if f != filepath]
    files.insert(0, filepath)
    files = files[:MAX_RECENT_FILES]
    for i, f in enumerate(files):
        set_config(f"recent_file_{i}", f)

# Auto-detect Bank Profile fingerprint
IDFC_FINGERPRINT_COLS = {"prospectno", "cuid", "tare weight", "currentbranch"}
EQUITAS_FINGERPRINT_COLS = {"svs_loan_no", "sole_id", "branch_name", "loan no"}

def _detect_bank_from_file(filepath):
    """Peek at Excel headers to determine which bank the file is for."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                try:
                    header_row = [
                        str(cell.value).strip().lower().replace("\n", "") if cell.value else ""
                        for cell in next(ws.iter_rows(min_row=1, max_row=1))
                    ]
                    headers_set = set(header_row)
                    idfc_score = len(IDFC_FINGERPRINT_COLS & headers_set)
                    equitas_score = len(EQUITAS_FINGERPRINT_COLS & headers_set)
                    if idfc_score >= 3:
                        return "IDFC First Bank"
                    if equitas_score >= 3:
                        return "Equitas Small Finance Bank"
                except StopIteration:
                    continue
        finally:
            wb.close()
    except Exception:
        pass
    return None

# Initialize Database
init_db()

# =========================================================
# AUTO-UPDATER LOGIC
# =========================================================
UPDATE_REPO = "DeepStacker/pdf-generator"
GITHUB_API = f"https://api.github.com/repos/{UPDATE_REPO}/releases/latest"

import urllib.request as _urllib
import urllib.error as _urlerror
import tempfile as _tempfile
import shutil as _shutil
import ssl as _ssl

def _make_ssl_context():
    try:
        import certifi
        cafile = certifi.where()
        if os.path.exists(cafile):
            return _ssl.create_default_context(cafile=cafile)
    except Exception:
        pass
    return _ssl.create_default_context()

def _urlopen_with_fallback(req, timeout=10):
    try:
        ctx = _make_ssl_context()
        return _urllib.urlopen(req, context=ctx, timeout=timeout)
    except _urlerror.URLError as e:
        if "CERTIFICATE_VERIFY_FAILED" in str(e):
            ctx = _ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = _ssl.CERT_NONE
            return _urllib.urlopen(req, context=ctx, timeout=timeout)
        raise

def _get_platform_suffix():
    if sys.platform == "darwin":
        return "macos"
    elif sys.platform == "win32":
        return "windows"
    return "linux"

def _check_latest_release():
    req = _urllib.Request(GITHUB_API, headers={"User-Agent": f"AuditEngine/{VERSION}"})
    with _urlopen_with_fallback(req, timeout=10) as resp:
        data = json.loads(resp.read().decode())
    tag = data["tag_name"]
    source_url = data["zipball_url"]
    body = data.get("body", "")
    suffix = f"binary_{_get_platform_suffix()}.zip"
    binary_url = ""
    for asset in data.get("assets", []):
        if suffix in asset["name"] and asset["name"].endswith(".zip"):
            binary_url = asset["browser_download_url"]
            break
    return tag, source_url, body, binary_url

def _parse_version(tag):
    return tuple(int(x) for x in tag.lstrip("vV").split("."))

def _download_update(url, dest_path, progress_callback=None):
    req = _urllib.Request(url, headers={"User-Agent": f"AuditEngine/{VERSION}"})
    with _urlopen_with_fallback(req, timeout=60) as resp:
        total = int(resp.headers.get("Content-Length", 0))
        downloaded = 0
        with open(dest_path, "wb") as f:
            while True:
                chunk = resp.read(65536)
                if not chunk:
                    break
                f.write(chunk)
                downloaded += len(chunk)
                if progress_callback and total:
                    progress_callback(downloaded / total * 100)
    return dest_path

def _install_binary_update(zip_path, install_dir, log_callback=print):
    import shutil
    import stat
    extract_to = _tempfile.mkdtemp(prefix="audit_bin_")
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_to)
    for root, dirs, files in os.walk(extract_to):
        for f in files:
            fp = os.path.join(root, f)
            st = os.stat(fp)
            os.chmod(fp, st.st_mode | stat.S_IXUSR | stat.S_IXGRP | stat.S_IXOTH)
    src = None
    for root, dirs, files in os.walk(extract_to):
        for f in files:
            fp = os.path.join(root, f)
            if os.access(fp, os.X_OK):
                src = fp
                break
        if src:
            break
    if not src:
        raise RuntimeError("No executable found in update ZIP")
    old_exe = sys.executable
    log_callback(f"Replacing {old_exe} with {src}")
    if sys.platform == "win32":
        new_exe = os.path.join(extract_to, os.path.basename(old_exe))
        if src != new_exe:
            shutil.copy2(src, new_exe)
        bat_path = os.path.join(extract_to, "_update.bat")
        bat_contents = (
            "@echo off\r\n"
            "REM === Audit Engine Auto-Update Script ===\r\n"
            "REM Wait for the parent process to exit cleanly\r\n"
            ":wait_exit\r\n"
            'tasklist /fi "PID eq %PARENT_PID%" 2>nul | find "%PARENT_PID%" >nul\r\n'
            "if not errorlevel 1 (\r\n"
            "  timeout /t 1 /nobreak >nul\r\n"
            "  goto wait_exit\r\n"
            ")\r\n"
            "REM Extra wait for PyInstaller _MEI temp dir cleanup\r\n"
            "timeout /t 3 /nobreak >nul\r\n"
            "REM Copy new exe over old with retry (antivirus may briefly lock)\r\n"
            "set RETRIES=0\r\n"
            ":retry_copy\r\n"
            'copy /y "' + new_exe.replace('/', '\\') + '" "' + old_exe.replace('/', '\\') + '" >nul 2>&1\r\n'
            "if errorlevel 1 (\r\n"
            "  set /a RETRIES+=1\r\n"
            "  if %RETRIES% lss 10 (\r\n"
            "    timeout /t 2 /nobreak >nul\r\n"
            "    goto retry_copy\r\n"
            "  )\r\n"
            "  REM Copy failed after retries — launch old exe anyway\r\n"
            ")\r\n"
            "REM Clean up stale _MEI temp directories from previous runs\r\n"
            'for /d %%D in ("%TEMP%\\_MEI*") do rd /s /q "%%D" 2>nul\r\n'
            "REM Launch the updated application\r\n"
            'start "" "' + old_exe.replace('/', '\\') + '"\r\n'
            "REM Clean up staging directory and self-delete\r\n"
            'rd /s /q "' + extract_to.replace('/', '\\') + '" 2>nul\r\n'
        )
        with open(bat_path, "w") as f:
            f.write(bat_contents)
        startup = subprocess.STARTUPINFO()
        startup.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startup.wShowWindow = 0
        subprocess.Popen(
            ["cmd.exe", "/c", bat_path],
            env={**os.environ, "PARENT_PID": str(os.getpid())},
            close_fds=True,
            startupinfo=startup,
        )
        log_callback("Update staged; batch script launched. Waiting for clean exit.")
        return bat_path
    backup = old_exe + ".bak"
    if os.path.exists(old_exe):
        os.rename(old_exe, backup)
    os.makedirs(os.path.dirname(old_exe), exist_ok=True)
    shutil.copy2(src, old_exe)
    os.chmod(old_exe, 0o755)
    if sys.platform == "darwin":
        subprocess.run(["xattr", "-dr", "com.apple.quarantine", old_exe], check=False)
    if os.path.exists(backup):
        os.remove(backup)
    shutil.rmtree(extract_to, ignore_errors=True)
    log_callback(f"Binary updated at {old_exe}")
    return None

def _get_install_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _cleanup_stale_mei():
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return
    try:
        temp_dir = _tempfile.gettempdir()
        current_mei = getattr(sys, '_MEIPASS', '')
        current_mei_name = os.path.basename(current_mei).lower() if current_mei else ""
        for entry in os.listdir(temp_dir):
            if not entry.startswith('_MEI'):
                continue
            if entry.lower() == current_mei_name:
                continue
            mei_path = os.path.join(temp_dir, entry)
            if not os.path.isdir(mei_path):
                continue
            try:
                _shutil.rmtree(mei_path)
            except (PermissionError, OSError):
                pass
    except Exception:
        pass

def _restart_app():
    if sys.platform == "win32":
        subprocess.Popen([sys.executable] + (sys.argv if not getattr(sys, "frozen", False) else []))
        sys.exit(0)
    if getattr(sys, "frozen", False):
        os.execl(sys.executable, sys.executable)
    else:
        os.execl(sys.executable, sys.executable, *sys.argv)

# Clean up stale temp folders in background
if getattr(sys, "frozen", False):
    threading.Thread(target=_cleanup_stale_mei, daemon=True).start()

# =========================================================
# STATE TRACKERS
# =========================================================
class ProgressTracker:
    def __init__(self):
        self.pct = 0
        self.active_branch = ""
        self.logs = []
        self.is_running = False
        self.summary = None
        self.cancel_requested = False

    def log(self, level, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.logs.append({
            "timestamp": timestamp,
            "level": level,
            "message": str(message)
        })
        file_logger.info(f"[{level}] {message}")
        print(f"[{timestamp}] [{level}] {message}")

    def update_pct(self, pct, active_branch=""):
        self.pct = pct
        if active_branch:
            self.active_branch = active_branch

    def reset(self):
        self.pct = 0
        self.active_branch = ""
        self.logs = []
        self.is_running = True
        self.summary = None
        self.cancel_requested = False

global_tracker = ProgressTracker()

class UpdateState:
    def __init__(self):
        self.update_ready = False
        self.latest_version = ""
        self.binary_url = ""
        self.dest_zip_path = ""
        self.progress_pct = 0
        self.is_downloading = False
        self.success = False
        self.error = ""
        self.staged_bat = None

update_state = UpdateState()

cancel_event = threading.Event()

# =========================================================
# HEADLESS DIALOG POPUPS
# =========================================================
def ask_file_dialog():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    file_path = filedialog.askopenfilename(
        parent=root,
        title="Select Master Excel File",
        filetypes=[("Excel Files", "*.xlsx;*.xls")]
    )
    root.destroy()
    return file_path

def ask_directory_dialog():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    dir_path = filedialog.askdirectory(
        parent=root,
        title="Select Output Directory"
    )
    root.destroy()
    return dir_path

# =========================================================
# THREAD WORKERS
# =========================================================
def worker_idfc_thread(inp, out_base, typ, output_mode, auto_open, naming_pattern):
    try:
        global_tracker.log("INFO", f"Initializing Build: {os.path.basename(inp)}")
        
        inp = os.path.abspath(os.path.normpath(inp))
        out_base = os.path.abspath(os.path.normpath(out_base))

        excel_name = os.path.splitext(os.path.basename(inp))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        out = os.path.join(out_base, f"{excel_name}_{timestamp}")
        out = os.path.abspath(os.path.normpath(out))
        os.makedirs(out, exist_ok=True)

        s, h, rows = pdf_logic.read_excel(inp, lambda x: global_tracker.log("INFO", x))
        groups = pdf_logic.group_by_branch(rows)

        total = len(groups)
        count = 0
        needs_zip = output_mode in ("ZIP ONLY", "BOTH")
        pdf_pct_max = 90.0 if needs_zip else 100.0

        import time as _time
        _start_time = _time.time()
        _per_item_times = []

        for c, br in sorted(groups.items()):
            if cancel_event.is_set():
                global_tracker.log("WARN", f"CANCELLED by user after {count}/{total} branches.")
                break

            name = str(br[0].get("CurrentBranchName", "Branch")).strip()
            st = str(br[0].get("State", "")).strip()
            safe_name = "".join(x for x in name if x.isalnum() or x in " -_").strip()
            if not safe_name:
                safe_name = str(c)
                
            branch_part = safe_name
            type_part = typ
            filename = naming_pattern.replace("{branch}", branch_part).replace("{type}", type_part)
            filename = "".join(x for x in filename if x.isalnum() or x in " -_.").strip()
            if not filename.endswith(".pdf"):
                filename += ".pdf"
            path = os.path.join(out, filename)
            path = os.path.abspath(os.path.normpath(path))

            global_tracker.update_pct((count / total) * pdf_pct_max, active_branch=f"Building: {safe_name}")
            global_tracker.log("INFO", f"Building: {safe_name}")
            pdf_logic.generate_pdf(typ, c, name, st, br, path)
            count += 1

            # ETA calculation
            elapsed = _time.time() - _start_time
            _per_item_times.append(elapsed / count)
            avg_time = sum(_per_item_times) / len(_per_item_times)
            remaining = avg_time * (total - count)

        was_cancelled = cancel_event.is_set()

        if not was_cancelled:
            log_generation(excel_name, count, out, typ, full_path=inp)
            _elapsed = _time.time() - _start_time
            global_tracker.log("OK", f"SUCCESS: {count} Reports Created in {_elapsed:.1f}s.")

            zip_path = f"{out}.zip"

            if output_mode in ("ZIP ONLY", "BOTH"):
                try:
                    global_tracker.log("INFO", "Compressing files...")
                    global_tracker.update_pct(92)
                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root_dir, _, files in os.walk(out):
                            for file in files:
                                zipf.write(os.path.join(root_dir, file), file)
                    global_tracker.update_pct(97)
                    global_tracker.log("OK", f"ZIP READY: {os.path.basename(zip_path)}")
                except OSError as ze:
                    global_tracker.log("ERROR", f"Zip Error: {ze}")

            if output_mode == "ZIP ONLY":
                try:
                    global_tracker.log("INFO", "Cleaning up raw PDFs (Space Saved)...")
                    import shutil
                    shutil.rmtree(out)
                except OSError as re:
                    global_tracker.log("ERROR", f"Cleanup Error: {re}")

            global_tracker.update_pct(100)

            # Calculate total output size
            final_target = zip_path if (output_mode == "ZIP ONLY" and os.path.exists(zip_path)) else out
            total_size = 0
            if os.path.isfile(final_target):
                total_size = os.path.getsize(final_target)
            elif os.path.isdir(final_target):
                for root_dir, _, files in os.walk(final_target):
                    for f in files:
                        total_size += os.path.getsize(os.path.join(root_dir, f))

            size_str = f"{total_size / 1024:.1f} KB" if total_size < 1024*1024 else f"{total_size / 1024 / 1024:.1f} MB"
            global_tracker.log("INFO", f"Total output size: {size_str}")

            if auto_open:
                if os.path.exists(final_target):
                    open_path(final_target)

            global_tracker.summary = {
                "title": "Generation Complete",
                "items": [
                    {"label": "Status", "value": "✓ Success"},
                    {"label": "Audit Type", "value": typ},
                    {"label": "Branches Completed", "value": str(count)},
                    {"label": "Time Elapsed", "value": f"{_elapsed:.1f}s"},
                    {"label": "Total Output Size", "value": size_str},
                    {"label": "Output Staged At", "value": final_target}
                ]
            }
            trigger_desktop_notification("Audit Engine Elite", f"✓ Generation complete! Created {count} branch reports.")
        else:
            log_generation(excel_name, count, out, f"{typ} (CANCELLED)", full_path=inp)
            global_tracker.log("WARN", f"Wound down gracefully after {count}/{total} branches.")
            global_tracker.summary = {
                "title": "Generation Cancelled",
                "message": f"Processing was stopped. Successfully saved reports for {count} branches."
            }

    except Exception as e:
        global_tracker.log("ERROR", f"FAILURE: {e}")
        global_tracker.summary = {
            "title": "Generation Failed",
            "message": str(e)
        }
        trigger_desktop_notification("Generation Failed", f"✗ Batch compilation failed: {e}")
    finally:
        global_tracker.is_running = False
        if "mapped_" in os.path.basename(inp) and ".temp_audit_engine" in inp:
            try:
                if os.path.exists(inp):
                    os.remove(inp)
            except OSError:
                pass

def worker_equitas_thread(inp, out_base, stage, equitas_format, equitas_pack):
    try:
        global_tracker.log("INFO", f"Initializing Equitas {stage}: {os.path.basename(inp)}")

        inp = os.path.abspath(os.path.normpath(inp))
        out_base = os.path.abspath(os.path.normpath(out_base))

        excel_name = os.path.splitext(os.path.basename(inp))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        out = os.path.join(out_base, f"{excel_name}_EQ_{stage.replace(' ', '')}_{timestamp}")
        out = os.path.abspath(os.path.normpath(out))
        os.makedirs(out, exist_ok=True)

        if stage == "STAGE 1":
            import time as _time
            _eqs1_start = _time.time()
            pdf_c, exc_c = equitas_logic.run_equitas_stage1(
                inp, out, 
                lambda x, level="INFO": global_tracker.log(level, x), 
                cancel_event, 
                lambda pct: global_tracker.update_pct(pct),
                output_format=equitas_format, output_mode=equitas_pack
            )
            _eqs1_elapsed = _time.time() - _eqs1_start
            was_cancelled = cancel_event.is_set()

            if not was_cancelled:
                log_generation(excel_name, pdf_c + exc_c, out, "Equitas-S1", full_path=inp)
                global_tracker.log("OK", f"SUCCESS: {pdf_c} PDFs, {exc_c} Excels created. ({_eqs1_elapsed:.1f}s)")
                global_tracker.update_pct(100)

                # Calculate size
                total_size = 0
                for root_dir, _, files in os.walk(out):
                    for f in files:
                        total_size += os.path.getsize(os.path.join(root_dir, f))
                size_str = f"{total_size / 1024:.1f} KB" if total_size < 1024*1024 else f"{total_size / 1024 / 1024:.1f} MB"

                if get_config("auto_open", "True") == "True":
                    open_path(out)

                global_tracker.summary = {
                    "title": "Stage 1 Complete",
                    "items": [
                        {"label": "Status", "value": "✓ Success"},
                        {"label": "PDF Worksheets", "value": str(pdf_c)},
                        {"label": "Excel Templates", "value": str(exc_c)},
                        {"label": "Time Taken", "value": f"{_eqs1_elapsed:.1f}s"},
                        {"label": "Total Size", "value": size_str},
                        {"label": "Output Directory", "value": out}
                    ]
                }
                trigger_desktop_notification("Audit Engine Elite", f"✓ Stage 1 Complete! {pdf_c} PDFs and {exc_c} Excels built.")
            else:
                log_generation(excel_name, pdf_c + exc_c, out, "Equitas-S1 (CANCELLED)", full_path=inp)
                global_tracker.log("WARN", "Stage 1 Generation Cancelled.")
                global_tracker.summary = {
                    "title": "Stage 1 Cancelled",
                    "message": f"Stage 1 stopped by user. Generated {pdf_c} PDFs and {exc_c} Excels."
                }

        else:
            # Stage 2 Consolidation
            out_path = equitas_logic.run_equitas_stage2(
                inp, out, 
                lambda x, level="INFO": global_tracker.log(level, x), 
                cancel_event, 
                lambda pct: global_tracker.update_pct(pct)
            )
            was_cancelled = cancel_event.is_set()

            if not was_cancelled and out_path:
                log_generation(excel_name, 1, out_path, "Equitas-S2", full_path=inp)
                total_size = os.path.getsize(out_path) if os.path.isfile(out_path) else 0
                size_str = f"{total_size / 1024:.1f} KB" if total_size < 1024*1024 else f"{total_size / 1024 / 1024:.1f} MB"
                global_tracker.log("OK", f"SUCCESS: Consolidated worksheet created. ({size_str})")
                global_tracker.update_pct(100)

                if get_config("auto_open", "True") == "True":
                    open_path(out)

                global_tracker.summary = {
                    "title": "Stage 2 Complete",
                    "items": [
                        {"label": "Status", "value": "✓ Success"},
                        {"label": "Consolidated Excel", "value": os.path.basename(out_path)},
                        {"label": "Total File Size", "value": size_str},
                        {"label": "Output Excel File", "value": out_path}
                    ]
                }
                trigger_desktop_notification("Audit Engine Elite", "✓ Stage 2 Complete! Consolidated worksheet created.")
            else:
                log_generation(excel_name, 0, out, "Equitas-S2 (CANCELLED)", full_path=inp)
                global_tracker.log("WARN", "Stage 2 Consolidation Cancelled.")
                global_tracker.summary = {
                    "title": "Stage 2 Cancelled",
                    "message": "Consolidation workbook creation stopped."
                }

    except Exception as e:
        global_tracker.log("ERROR", f"FAILURE: {e}")
        global_tracker.summary = {
            "title": "Equitas Generation Failed",
            "message": str(e)
        }
        trigger_desktop_notification("Equitas Generation Failed", f"✗ Batch compilation failed: {e}")
    finally:
        global_tracker.is_running = False
        if "mapped_" in os.path.basename(inp) and ".temp_audit_engine" in inp:
            try:
                if os.path.exists(inp):
                    os.remove(inp)
            except OSError:
                pass

# =========================================================
# WSGI BOTTLE WEB SERVICE ROUTING
# =========================================================
@route('/')
def serve_index():
    return web_assets.HTML_CONTENT

@route('/api/dashboard')
def api_dashboard():
    total_sessions, total_pdfs = get_stats()
    
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT timestamp FROM history ORDER BY id DESC LIMIT 1")
    last_run_row = cursor.fetchone()
    last_run = last_run_row[0] if last_run_row else "No activity yet"
    conn.close()

    recent_files = _get_recent_files()

    return {
        "bank": get_config("bank", "IDFC First Bank"),
        "last_file": get_config("last_file", ""),
        "out_path": get_config("out_path", os.path.join(os.path.expanduser("~"), "Desktop")),
        "audit_type": get_config("audit_type", "POA"),
        "output_mode": get_config("pkg_mode", "BOTH"),
        "equitas_format": get_config("equitas_format", "BOTH"),
        "equitas_pack": get_config("equitas_pack", "FOLDER"),
        "auto_open": get_config("auto_open", "True") == "True",
        "eq_auto_open": get_config("auto_open", "True") == "True",
        "db_path": DB_PATH,
        "log_path": LOG_FILE,
        "naming_pattern": get_config("naming_pattern", "{branch}_{type}"),
        "recent_files": recent_files,
        "total_sessions": total_sessions,
        "total_pdfs": total_pdfs,
        "last_run": last_run
    }

@route('/api/config/save', method='POST')
def api_config_save():
    data = request.json
    set_config(data['key'], data['value'])
    return {"success": True}

def peek_excel_data(filepath):
    """Peek at the first sheet, returning headers and first 5 rows of raw data."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheet = wb.active
        headers = []
        rows = []
        for i, row_cells in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c is not None else f"Column {idx}" for idx, c in enumerate(row_cells)]
            elif i <= 5:
                row_data = [str(c) if c is not None else "" for c in row_cells]
                # Pad row_data to match headers length just in case
                if len(row_data) < len(headers):
                    row_data.extend([""] * (len(headers) - len(row_data)))
                rows.append(row_data[:len(headers)])
            else:
                break
        wb.close()
        return headers, rows
    except Exception as e:
        file_logger.warning(f"Failed to peek Excel data: {e}")
        return [], []

@route('/api/validate', method='POST')
def api_validate():
    data = request.json
    filepath = data.get('filepath', '')
    if not filepath or not os.path.exists(filepath):
        return {"success": False, "error": "Spreadsheet file path does not exist."}
    
    detected_bank = _detect_bank_from_file(filepath)
    try:
        headers, preview_rows = peek_excel_data(filepath)
        
        if detected_bank == "IDFC First Bank":
            valid, err = pdf_logic.validate_excel(filepath)
            if valid:
                s, h, rows = pdf_logic.read_excel(filepath, lambda x: None)
                groups = pdf_logic.group_by_branch(rows)
                return {
                    "success": True, 
                    "detected_bank": detected_bank, 
                    "rows": len(rows), 
                    "branches": len(groups),
                    "headers": headers,
                    "preview": preview_rows
                }
            return {"success": False, "error": err, "detected_bank": detected_bank, "headers": headers, "preview": preview_rows}
        
        elif detected_bank == "Equitas Small Finance Bank":
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if "JSR" in sheets or "Normal" in sheets:
                valid, err = equitas_logic.validate_equitas_stage1_file(filepath)
                return {
                    "success": valid, 
                    "error": err, 
                    "detected_bank": detected_bank, 
                    "rows": "stage 1 peek", 
                    "branches": "N/A",
                    "headers": headers,
                    "preview": preview_rows
                }
            else:
                valid, err = equitas_logic.validate_equitas_stage2_file(filepath)
                return {
                    "success": valid, 
                    "error": err, 
                    "detected_bank": detected_bank, 
                    "rows": "stage 2 peek", 
                    "branches": "N/A",
                    "headers": headers,
                    "preview": preview_rows
                }
        
        return {
            "success": False, 
            "error": "Invalid Gold Loan Audit Excel format structure.", 
            "detected_bank": None,
            "headers": headers,
            "preview": preview_rows
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

@route('/api/browse/file')
def api_browse_file():
    path = ask_file_dialog()
    return {"path": path}

@route('/api/browse/folder')
def api_browse_folder():
    path = ask_directory_dialog()
    return {"path": path}

@route('/api/run', method='POST')
def api_run():
    if global_tracker.is_running:
        return {"success": False, "error": "A background generation thread is currently active."}

    data = request.json
    bank = data.get('bank')
    filepath = data.get('filepath')
    out_path = data.get('out_path')
    auto_open = data.get('auto_open', True)
    naming_pattern = data.get('naming_pattern', '{branch}_{type}')

    if not filepath or not os.path.exists(filepath):
        return {"success": False, "error": "Excel spreadsheet file is missing."}
    if not out_path or not os.path.exists(out_path):
        return {"success": False, "error": "Output staging directory path is invalid."}

    # Save configs and recent links
    set_config("bank", bank)
    set_config("last_file", filepath)
    set_config("out_path", out_path)
    set_config("auto_open", str(auto_open))
    set_config("naming_pattern", naming_pattern)
    _add_recent_file(filepath)

    cancel_event.clear()
    global_tracker.reset()

    # Column mappings preprocessing
    column_mappings = data.get('column_mappings')
    if column_mappings:
        filepath = preprocess_mapped_excel(filepath, column_mappings, bank)

    if bank == "IDFC First Bank":
        audit_type = data.get('audit_type', 'POA')
        output_mode = data.get('output_mode', 'BOTH')
        
        set_config("audit_type", audit_type)
        set_config("pkg_mode", output_mode)
        
        t = threading.Thread(
            target=worker_idfc_thread,
            args=(filepath, out_path, audit_type, output_mode, auto_open, naming_pattern),
            daemon=True
        )
        t.start()
    else:
        equitas_stage = data.get('equitas_stage', 'STAGE 1')
        equitas_format = data.get('equitas_format', 'BOTH')
        equitas_pack = data.get('equitas_pack', 'FOLDER')
        
        set_config("equitas_format", equitas_format)
        set_config("equitas_pack", equitas_pack)
        
        t = threading.Thread(
            target=worker_equitas_thread,
            args=(filepath, out_path, equitas_stage, equitas_format, equitas_pack),
            daemon=True
        )
        t.start()

    return {"success": True}

@route('/api/progress')
def api_progress():
    return {
        "pct": global_tracker.pct,
        "active_branch": global_tracker.active_branch,
        "logs": global_tracker.logs,
        "is_running": global_tracker.is_running,
        "summary": global_tracker.summary
    }

@route('/api/cancel')
def api_cancel():
    cancel_event.set()
    global_tracker.cancel_requested = True
    return {"success": True}

@route('/api/history')
def api_history():
    search = request.query.get('search', '').strip()
    history = get_recent_history(search)
    results = []
    for row in history:
        results.append({
            "id": row[0],
            "timestamp": row[1],
            "excel_name": row[2],
            "pdf_count": row[3],
            "output_path": row[4],
            "audit_type": row[5],
            "full_path": row[6]
        })
    return json.dumps(results)

@route('/api/history/clear', method='POST')
def api_history_clear():
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM history")
    conn.commit()
    conn.close()
    return {"success": True}

@route('/api/recent/clear', method='POST')
def api_recent_clear():
    conn = _connect_db()
    cursor = conn.cursor()
    for i in range(MAX_RECENT_FILES):
        cursor.execute("DELETE FROM config WHERE key = ?", (f"recent_file_{i}",))
    conn.commit()
    conn.close()
    return {"success": True}

@route('/api/stats')
def api_stats():
    types, trend = get_analytics()
    total_sessions, total_pdfs = get_stats()
    
    conn = _connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(DISTINCT excel_name) FROM history")
    total_excels = cursor.fetchone()[0] or 0
    conn.close()
    
    return {
        "distribution": types,
        "trend": trend,
        "total_sessions": total_sessions,
        "total_pdfs": total_pdfs,
        "total_excels": total_excels
    }

@route('/api/open', method='POST')
def api_open():
    data = request.json
    path = data.get('path', '')
    if path and os.path.exists(path):
        open_path(path)
    return {"success": True}

# Heartbeat self-termination logic
last_heartbeat = time.time()

@route('/api/heartbeat')
def api_heartbeat():
    global last_heartbeat
    last_heartbeat = time.time()
    return {"status": "ok"}

def heartbeat_monitor():
    global last_heartbeat
    # Generous initial launch wait buffer time for tab opening
    time.sleep(30)
    while True:
        # Browsers throttle background tab timers to 1 minute max.
        # Use a 120 second timeout to avoid killing the server while the tab is inactive.
        if time.time() - last_heartbeat > 120:
            print("No heartbeat received for 120 seconds. Terminating WSGI background portal cleanly.")
            file_logger.info("Heartbeat lost. Cleaner self-terminating backend process now.")
            os._exit(0)
        time.sleep(5)

# =========================================================
# AUTO-UPDATER WSGI BRIDGES
# =========================================================
@route('/api/update/check')
def update_check():
    try:
        tag, source_url, body, binary_url = _check_latest_release()
        current_v = _parse_version(VERSION)
        latest_v = _parse_version(tag)
        if latest_v > current_v:
            if binary_url:
                update_state.update_ready = True
                update_state.latest_version = tag
                update_state.binary_url = binary_url
                return {
                    "update_ready": True,
                    "current": VERSION,
                    "latest": tag,
                    "body": body
                }
            else:
                file_logger.info(f"Update tag {tag} exists, but binary for OS is missing/building. Deferring.")
    except Exception as e:
        file_logger.warning(f"Background updates repo search query failed: {e}")
    return {"update_ready": False, "current": VERSION}

def download_update_worker():
    try:
        update_state.is_downloading = True
        update_state.progress_pct = 0
        update_state.success = False
        update_state.error = ""

        dest_dir = _tempfile.mkdtemp(prefix="audit_update_")
        dest_zip = os.path.join(dest_dir, "update.zip")
        update_state.dest_zip_path = dest_zip

        def progress_cb(pct):
            update_state.progress_pct = pct

        _download_update(update_state.binary_url, dest_zip, progress_cb)
        
        install_dir = _get_install_dir()
        bat = _install_binary_update(dest_zip, install_dir, log_callback=lambda x: file_logger.info(x))
        if bat:
            update_state.staged_bat = bat

        update_state.success = True
        update_state.progress_pct = 100
    except Exception as e:
        update_state.error = str(e)
        update_state.success = False
    finally:
        update_state.is_downloading = False

@route('/api/update/install', method='POST')
def update_install():
    if not update_state.binary_url:
        return {"success": False, "error": "No new staged update release staged."}
    if update_state.is_downloading:
        return {"success": True}
    threading.Thread(target=download_update_worker, daemon=True).start()
    return {"success": True}

@route('/api/update/progress')
def update_progress():
    return {
        "pct": update_state.progress_pct,
        "is_downloading": update_state.is_downloading,
        "success": update_state.success,
        "error": update_state.error
    }

def _delayed_apply():
    import time
    time.sleep(1)
    if update_state.staged_bat:
        # Exit forcefully so Windows batch script can overwrite the binary
        os._exit(0)
    else:
        _restart_app()

@route('/api/update/apply', method='POST')
def update_apply():
    threading.Thread(target=_delayed_apply, daemon=True).start()
    return {"success": True}

# =========================================================
# APPLICATION STARTUP ENTRY
# =========================================================
def find_free_port(start_port=52140):
    for port in range(start_port, start_port + 20):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(('127.0.0.1', port))
                return port
            except OSError:
                continue
    return start_port

def open_browser(port):
    time.sleep(0.6)
    webbrowser.open(f"http://127.0.0.1:{port}")

if __name__ == "__main__":
    # Required for PyInstaller frozen executables on Windows
    import multiprocessing
    multiprocessing.freeze_support()

    # Find free local port
    port = find_free_port()

    # Spawn auto-browser launch thread
    threading.Thread(target=open_browser, args=(port,), daemon=True).start()

    # Spawn background client tab heartbeat auto-closer daemon thread
    threading.Thread(target=heartbeat_monitor, daemon=True).start()

    # Run the WSGI Micro webserver
    print(f"Audit Engine Headless Server listening on http://127.0.0.1:{port}")
    run(host='127.0.0.1', port=port, quiet=True)
