import re
import sys
from copy import copy
from datetime import datetime, date
from pathlib import Path


import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ─── Highlight Colors ─────────────────────────────────────────────────
DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def map_columns(ws):
    """Auto-detect column mapping from header row (row 2) and section row (row 1).
    Uses ws._cells to avoid iterating empty columns (performance).
    First occurrence wins (main display columns take priority over extra columns)."""
    mapping = {}
    row2 = {}
    row1 = {}
    for (_r, _c), _cell in ws._cells.items():
        if _r == 1 and _cell.value is not None:
            row1[_c] = str(_cell.value).strip()
        elif _r == 2 and _cell.value is not None:
            row2[_c] = str(_cell.value).strip()
    for c in sorted(set(row2.keys()) | set(row1.keys())):
        h = row2.get(c)
        s = row1.get(c)
        if h:
            key = h.upper()
            if key not in mapping:
                mapping[key] = c
        if s and not h:
            key = f"__SECTION__{s.upper()}"
            if key not in mapping:
                mapping[key] = c
    return mapping


def get_col(mapping, *names):
    """Get column number for any of the given names."""
    for name in names:
        key = name.upper()
        if key in mapping:
            return mapping[key]
    return None


def build_col_map(col):
    """Build the canonical column map from a map_columns() header mapping."""
    return {
        "packet": get_col(col, "PACKET NO."),
        "account": get_col(col, "ACCOUNT NUMBER"),
        "applicant": get_col(col, "APPLICANT NAME"),
        "status": get_col(col, "FRESH/RENEWAL/CLOSED/ALREADY VERIFIED"),
        "taf": get_col(col, "TAF/POA"),
        "remarks": get_col(col, "AGENCY REMARKS"),
        "magnet": get_col(col, "MAGNET TEST RESULT"),
        "tampered": get_col(col, "PACKET TAMPERED YES/NO"),
        "gdr_gross": get_col(col, "GROSS WEIGHT AS PER GDR/PACKET"),
        "actual_gross": get_col(col, "ACTUAL GROSS WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "tare": get_col(col, "ACTUAL TARE WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "gross_diff": get_col(col, "DIFFERENCE IN GROSS WEIGHT"),
        "gdr_net": get_col(col, "NET WEIGHT AS PER GDR/PACKET"),
        "actual_net": get_col(col, "ACTUAL NET WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "net_diff": get_col(col, "DIFFERENCE IN NET WEIGHT"),
        "spur_count": get_col(col, "TOTAL NO.OF SPURIOUS ORNAMENTS"),
        "spur_weight": get_col(col, "SPURIOUS ORNAMENTS GROSS WEIGHT"),
        "spur_pct": get_col(col, "% OF SPURIOUS ORNAMENTS GROSS WEIGHT"),
        "carat_count": get_col(col, "TOTAL NO.OF ORNAMENTS WITH CARAT MISMATCH"),
        "uncommon_count": get_col(col, "TOTAL NO.OF UNCOMMON ORNAMENTS"),
        "sanction_date": get_col(col, "SANCTION DATE"),
        "verification_date": get_col(col, "AGENCY VERIFICATION DATE"),
        "sanction_limit": get_col(col, "SANCTION LIMIT"),
        "gdr_no": get_col(col, "GDR NUMBER"),
        "ornaments_gdr": get_col(col, "TOTAL NO.OF ORNAMENTS AS PER THE GDR/PACKET"),
        "ornaments_actual": get_col(col, "ACTUAL AVAILABLE ORNAMENTS AT THE TIME OF FCU VERIFICATION"),
        "ornaments_diff": get_col(col, "DIFFRENCE IN ACTUAL ORNAMENTS"),
        "renewal_date": get_col(col, "RENEWAL/CLOSED DATE"),
    }


def is_topup_remark(remarks):
    """True if the remarks text marks the row as a top-up ("Top Up", "Top-up",
    "Topup of account ..." and similar spellings).

    Matches on a word boundary so unrelated remarks that merely contain "top"
    and "up" as parts of other words (e.g. "Desktop Update pending") are not
    mistaken for a top-up row.
    """
    return bool(re.search(r"\bTOP[\s\-]*UP\b", str(remarks or "").upper()))


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def find_data_end(ws, data_start=5):
    """Return the last row index that contains actual cell *values*.

    Real-world sheets often carry stray formatting / empty styled cells far
    below the real data (ws.max_row can be 200k+ while only a few hundred rows
    hold values). Iterating all_rows up to ws.max_row then becomes quadratic
    and can take minutes. We bound the scan to the highest row that holds a
    non-empty cell value, ignoring style-only cells."""
    max_row = ws.max_row
    if max_row <= data_start:
        return max_row
    end = data_start
    for (r, c), cell in ws._cells.items():
        if r < data_start or r > max_row:
            continue
        v = cell.value
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        if r > end:
            end = r
    return end


def is_empty(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.upper() == "NAN" or s == "None"


def parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.date()
        return v
    s = str(v).strip()
    if not s or s.upper() in ("NAN", "NONE", "NAT", "-"):
        return None
    
    months_map = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
    }
    
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
            
    for sep in ["-", "/", ".", " "]:
        parts = [p.strip() for p in s.split(sep) if p.strip()]
        if len(parts) == 3:
            p0, p1, p2 = parts[0], parts[1], parts[2]
            try:
                if p2.isdigit() and len(p2) in (2, 4):
                    yr = int(p2)
                    if yr < 100:
                        yr += 2000
                    day = int(p0)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return date(yr, m, day)
                elif p0.isdigit() and len(p0) == 4:
                    yr = int(p0)
                    day = int(p2)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return date(yr, m, day)
            except Exception:
                pass
    return None


def get_custom_output_filename(ws, col, col_map, all_rows, _cell_cache, default_filename):
    col_branch = col.get("BRANCH NAME")
    branch_name = "UNKNOWN"
    if col_branch:
        for r in all_rows:
            cell_obj = _cell_cache.get((r, col_branch))
            v = cell_obj.value if cell_obj else None
            if v and str(v).strip():
                branch_name = str(v).strip().upper()
                break
                
    import re
    branch_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', branch_name).strip()
    branch_name = re.sub(r'_+', '_', branch_name)
    
    col_ver_date = col_map.get("verification_date")
    valid_dates = []
    if col_ver_date:
        for r in all_rows:
            cell_obj = _cell_cache.get((r, col_ver_date))
            v = cell_obj.value if cell_obj else None
            if v:
                parsed_d = parse_any_date(v)
                if parsed_d:
                    valid_dates.append(parsed_d)
                    
    if valid_dates:
        unique_dates = sorted(list(set(valid_dates)))
        min_date = unique_dates[0]
        max_date = unique_dates[-1]
        
        months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        def fmt_d(d):
            return f"{d.day:02d}-{months[d.month-1]}-{d.year}"
            
        if min_date == max_date:
            date_str = fmt_d(min_date)
        else:
            date_str = f"{fmt_d(min_date)}_to_{fmt_d(max_date)}"
    else:
        date_str = "UNKNOWN_DATE"
        
    return f"{branch_name}_Audit-MIS_{date_str}.xlsx"


def extract_accounts_from_pdf(pdf_path):
    from pypdf import PdfReader
    import re
    accounts = []
    try:
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                found = re.findall(r'\b\d{15}\b', text)
                accounts.extend(found)
    except Exception as e:
        print(f"Error parsing PDF: {e}")
    return accounts


def rearrange_rows_by_pdf(ws, col_map, all_rows, pdf_accounts):
    col_acct = col_map.get("account")
    if not col_acct:
        return
        
    acct_to_rows = defaultdict(list)
    for r in all_rows:
        acct = safe_str(ws.cell(row=r, column=col_acct).value)
        if acct:
            acct_to_rows[acct].append(r)
            
    new_row_order = []
    used_rows = set()
    
    for acct in pdf_accounts:
        if acct in acct_to_rows:
            for r in acct_to_rows[acct]:
                if r not in used_rows:
                    new_row_order.append(r)
                    used_rows.add(r)
                    
    for r in all_rows:
        if r not in used_rows:
            new_row_order.append(r)
            
    # Bound the column range to the true data width: the rightmost column that
    # holds an actual value in the header rows or the data rows. This keeps
    # extra columns beyond the mapped ones moving together with their rows
    # (a fixed cap would silently leave far-right data behind, misaligning
    # rows), while still ignoring style-only stray cells that can push
    # ws.max_column into the thousands.
    row_set = set(all_rows)
    max_val_col = 0
    for (r, c), cl in ws._cells.items():
        if c <= max_val_col or (r not in row_set and r not in (1, 2)):
            continue
        v = cl.value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        max_val_col = c
    known_cols = [c for c in col_map.values() if c]
    if known_cols:
        max_val_col = max(max_val_col, max(known_cols))
    max_col = min(ws.max_column, max_val_col + 2) if max_val_col else min(ws.max_column, 150)
    row_data_cache = {}
    row_heights = {}
    
    for r in all_rows:
        row_heights[r] = ws.row_dimensions[r].height
        cells_dict = {}
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.has_style:
                cells_dict[c] = {
                    "value": cell.value,
                    "font": copy(cell.font) if cell.font else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "border": copy(cell.border) if cell.border else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                    "number_format": cell.number_format,
                }
            else:
                cells_dict[c] = {
                    "value": cell.value,
                    "font": None,
                    "fill": None,
                    "border": None,
                    "alignment": None,
                    "number_format": cell.number_format,
                }
        row_data_cache[r] = cells_dict
        
    for new_r, old_r in zip(all_rows, new_row_order):
        ws.row_dimensions[new_r].height = row_heights[old_r]
        cells_dict = row_data_cache[old_r]
        for c in range(1, max_col + 1):
            cell = ws.cell(row=new_r, column=c)
            data = cells_dict[c]
            # If the source cell carried no explicit style, the target must be
            # reset to defaults — otherwise the moved row keeps the previous
            # occupant's formatting.
            target_had_style = cell.has_style
            cell.value = data["value"]
            if data["font"]:
                cell.font = data["font"]
            elif target_had_style:
                cell.font = Font()
            if data["fill"]:
                cell.fill = data["fill"]
            elif target_had_style:
                cell.fill = PatternFill()
            if data["border"]:
                cell.border = data["border"]
            elif target_had_style:
                cell.border = Border()
            if data["alignment"]:
                cell.alignment = data["alignment"]
            elif target_had_style:
                cell.alignment = Alignment()
            cell.number_format = data["number_format"]


def process_file(filepath: str, output_suffix: str = "_VALIDATED", pdf_path: str = None):
    """Process a single report file."""
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        return

    print(f"\n{'='*80}")
    print(f"PROCESSING: {filepath.name}")
    if pdf_path:
        print(f"USING PDF SEQUENCE: {Path(pdf_path).name}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, keep_vba=False, keep_links=False)
    wb_data = openpyxl.load_workbook(filepath, data_only=True, keep_vba=False, keep_links=False)
    sheet_name = "Purity Verification Format"
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        return
    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]

    col = map_columns(ws)
    data_start = 5
    data_end = find_data_end(ws, data_start)

    col_map = build_col_map(col)

    all_rows = list(range(data_start, data_end + 1))

    if pdf_path:
        pdf_accounts = extract_accounts_from_pdf(pdf_path)
        if pdf_accounts:
            rearrange_rows_by_pdf(ws, col_map, all_rows, pdf_accounts)
            rearrange_rows_by_pdf(ws_data, col_map, all_rows, pdf_accounts)

    _cell_cache = {}
    for (_r, _c), _cell in list(ws._cells.items()):
        _cell_cache[(_r, _c)] = _cell

    rows_data = {}
    for r in all_rows:
        rows_data[r] = {
            "packet": safe_str(_cell_cache.get((r, col_map["packet"])).value if col_map["packet"] and (r, col_map["packet"]) in _cell_cache else None),
            "account": safe_str(_cell_cache.get((r, col_map["account"])).value if col_map["account"] and (r, col_map["account"]) in _cell_cache else None),
            "name": safe_str(_cell_cache.get((r, col_map["applicant"])).value if col_map["applicant"] and (r, col_map["applicant"]) in _cell_cache else None),
            "status": safe_str(_cell_cache.get((r, col_map["status"])).value if col_map["status"] and (r, col_map["status"]) in _cell_cache else None),
            "taf": safe_str(_cell_cache.get((r, col_map["taf"])).value if col_map["taf"] and (r, col_map["taf"]) in _cell_cache else None),
            "remarks": safe_str(_cell_cache.get((r, col_map["remarks"])).value if col_map["remarks"] and (r, col_map["remarks"]) in _cell_cache else None),
            "magnet": safe_str(_cell_cache.get((r, col_map["magnet"])).value if col_map["magnet"] and (r, col_map["magnet"]) in _cell_cache else None),
            "tampered": safe_str(_cell_cache.get((r, col_map["tampered"])).value if col_map["tampered"] and (r, col_map["tampered"]) in _cell_cache else None),
        }

    _data_cache = {}
    for (_r, _c), _cell in ws_data._cells.items():
        _data_cache[(_r, _c)] = _cell

    summary = defaultdict(list)
    run_validation(ws, ws_data, col_map, all_rows, rows_data, summary, _cell_cache=_cell_cache, _data_cache=_data_cache)

    # ══════════════════════════════════════════════════════════════════
    # SAVE & REPORT
    # ══════════════════════════════════════════════════════════════════
    custom_filename = get_custom_output_filename(ws, col, col_map, all_rows, _cell_cache, filepath.name)
    output_path = filepath.parent / custom_filename
    wb.save(output_path)

    print(f"\n{'─'*60}")
    print(f"PROCESSING COMPLETE: {output_path.name}")
    print(f"{'─'*60}")

    for key, items in sorted(summary.items()):
        if items:
            color = ""
            reset = ""
            if any(
                kw in key
                for kw in ["duplicate", "outlier", "mismatch", "error", "negative"]
            ):
                color = "\033[91m"  # Red
            elif any(kw in key for kw in ["found", "updated", "fixed", "copied"]):
                color = "\033[92m"  # Green
            elif any(kw in key for kw in ["warning", "multiple", "unexpected", "no_match"]):
                color = "\033[93m"  # Yellow

            label = key.replace("_", " ").title()
            print(f"\n  {color}{label}: {len(items)}{reset}")
            for item in items[:10]:  # Limit display
                print(f"    • {item}")
            if len(items) > 10:
                print(f"    ... and {len(items)-10} more")

    total_issues = sum(len(v) for v in summary.values())
    print(f"\n  TOTAL ISSUES FOUND: {total_issues}")
    print(f"  Saved to: {output_path}")

    wb.close()
    return summary


def run_validation(ws, ws_data, col_map, all_rows, rows_data, summary, _cell_cache=None, _data_cache=None):
    """Run all validation passes against a worksheet.
    
    Args:
        ws: openpyxl Worksheet (for writing)
        ws_data: data_only Worksheet (for formula values)
        col_map: dict of column name → column index
        all_rows: list of row numbers to process
        rows_data: dict of row → {packet, account, name, status, taf, remarks, magnet, tampered}
        summary: defaultdict(list) to collect issues
        _cell_cache: optional dict {(row, col): cell} for fast access (app.py), or None
        
    Returns:
        summary (same dict, mutated in place)
    """
    resolved_packet_rows = set()

    # Computed once (not per-highlight-call): columns excluded from highlighting
    # on rows already resolved by the Pass 2 top-up packet match.
    _diff_cols_excluded = frozenset(
        c for c in (
            col_map.get("gdr_gross"), col_map.get("actual_gross"), col_map.get("tare"), col_map.get("gross_diff"),
            col_map.get("gdr_net"), col_map.get("actual_net"), col_map.get("net_diff"),
            col_map.get("ornaments_gdr"), col_map.get("ornaments_actual"), col_map.get("ornaments_diff"),
        )
        if c is not None
    )

    # Computed once (not re-derived in every pass): uppercased status keyed by row.
    row_status_upper = {}
    for r in all_rows:
        rd = rows_data[r]
        row_status_upper[r] = rd["status"].upper() if rd["status"] else ""

    # Rows reset to default values by the Closed/Top-Up cleaning pass; later
    # passes skip these so cleaned defaults aren't re-flagged as issues.
    cleaned_rows = set()

    # ── Cell helpers (use cache if available, else direct cell access) ──
    def cell(row, c):
        if c is None:
            return None
        if _cell_cache is not None:
            cl = _cell_cache.get((row, c))
            if cl is not None:
                return cl
        return ws.cell(row=row, column=c)

    def val(row, c):
        cl = cell(row, c)
        return cl.value if cl is not None else None

    def set_val(row, c, new_val):
        cl = cell(row, c)
        if cl is not None:
            cl.value = new_val

    def highlight(row, c, fill, font=None, force=False, skip_excluded=True):
        if skip_excluded and row in resolved_packet_rows and c in _diff_cols_excluded:
            return
        cl = cell(row, c)
        if cl is not None:
            if not force and cl.fill and cl.fill.start_color and cl.fill.start_color.rgb:
                rgb = str(cl.fill.start_color.rgb)
                if not rgb.endswith("000000") and rgb != "00000000":
                    return
            cl.fill = copy(fill)
            if font:
                # Preserve the cell's original font (family, size, italics…)
                # and only apply the requested emphasis attributes on top.
                new_font = copy(cl.font) if cl.font else Font()
                if font.color is not None:
                    new_font.color = copy(font.color)
                if font.bold:
                    new_font.bold = True
                cl.font = new_font

    def num_or_none(v):
        try:
            s = str(v).strip()
            if not s or s.upper() in ("NA", "N/A", "N A", "NAN", "NONE", "-"):
                return None
            return float(s)
        except (ValueError, TypeError):
            return None

    def val_data(row, c):
        v = val(row, c)
        if v is not None and isinstance(v, str) and v.startswith("="):
            if _data_cache is not None:
                cv_cell = _data_cache.get((row, c))
                cv = cv_cell.value if cv_cell is not None else None
            else:
                cv = ws_data.cell(row=row, column=c).value if ws_data else None
            return cv if cv is not None else v
        return v

    # ══════════════════════════════════════════════════════════════════
    # PASS 1: Duplicate Packet Numbers
    # ══════════════════════════════════════════════════════════════════
    col_packet = col_map.get("packet")
    if col_packet:
        packet_map = defaultdict(list)
        for r in all_rows:
            pkt = rows_data[r]["packet"]
            if pkt and pkt.upper() not in ("", "NAN", "NONE"):
                packet_map[pkt].append(r)
        for pkt, dup_rows in packet_map.items():
            if len(dup_rows) > 1:
                for r in dup_rows:
                    is_topup = is_topup_remark(rows_data[r]["remarks"])
                    if is_topup:
                        highlight(r, col_packet, LIGHT_BLUE_FILL)
                        summary["duplicate_packet_topup"].append(f"Row {r}: Packet '{pkt}' (Top-Up)")
                    else:
                        highlight(r, col_packet, DARK_RED_FILL, WHITE_FONT)
                        summary["duplicate_packet"].append(f"Row {r}: Packet '{pkt}' (appears {len(dup_rows)}x)")

    # ══════════════════════════════════════════════════════════════════
    # PASS 2: Empty Packet → Resolve by Applicant Name
    # ══════════════════════════════════════════════════════════════════
    col_remarks = col_map.get("remarks")
    if col_packet and col_remarks:
        empty_pkt_rows = [r for r in all_rows if not rows_data[r]["packet"] or rows_data[r]["packet"].upper() in ("", "NAN", "NONE")]

        # Index of name → rows currently holding a valid packet, built once and
        # updated incrementally as rows get resolved below. Equivalent to the
        # original full re-scan of all_rows per empty-packet row (a resolved
        # row becomes a candidate match for later rows with the same name,
        # which is what makes ambiguous multi-way name matches get flagged
        # instead of silently resolved) but O(1) per lookup instead of O(n).
        name_to_packet_rows = defaultdict(list)
        for sr in all_rows:
            spkt = rows_data[sr]["packet"]
            if spkt and spkt.upper() not in ("", "NAN", "NONE"):
                name_to_packet_rows[rows_data[sr]["name"].upper().strip()].append(sr)

        for r in empty_pkt_rows:
            rd = rows_data[r]
            status = row_status_upper[r]
            if "CLOSED" in status:
                summary["closed_empty_packet"].append(f"Row {r}: Empty packet (Closed)")
                continue
            name = rd["name"]
            if not name:
                highlight(r, col_packet, ORANGE_FILL)
                summary["unexpected_empty_packet"].append(f"Row {r}: Account {rd['account']} (no name)")
                continue
            name_key = name.upper().strip()
            matching_rows = [sr for sr in name_to_packet_rows.get(name_key, ()) if sr != r]
            if len(matching_rows) == 1:
                sr = matching_rows[0]
                source_acct = rows_data[sr]["account"]
                source_packet = rows_data[sr]["packet"]
                set_val(r, col_packet, source_packet)
                rd["packet"] = source_packet
                name_to_packet_rows[name_key].append(r)
                col_gdr_no = col_map.get("gdr_no")
                if col_gdr_no:
                    set_val(r, col_gdr_no, None)
                col_renewal_date = col_map.get("renewal_date")
                if col_renewal_date:
                    set_val(r, col_renewal_date, None)
                expected_remark = f"This Account No is Topup of Account no {source_acct}, with Same Paket No"
                set_val(r, col_remarks, expected_remark)
                summary["topup_resolved"].append(f"Row {r}: {name} → matched row {sr} (pkt {source_packet})")
                resolved_packet_rows.add(r)
            elif len(matching_rows) > 1:
                highlight(r, col_packet, ORANGE_FILL)
                summary["topup_multiple_matches"].append(f"Row {r}: {name} has {len(matching_rows)} matches, left unresolved")
            else:
                highlight(r, col_packet, ORANGE_FILL)
                summary["topup_no_match"].append(f"Row {r}: {name} has no matching row with packet")

    # ══════════════════════════════════════════════════════════════════
    # PASS 3: Reset Closed/Top-Up rows to default values
    # Verification-result columns are not applicable on these rows: numeric
    # columns get their default (0.000 for weights, 0 for counts/diffs/%),
    # the rest are blanked. Identity/loan columns (packet, account, name,
    # status, TAF/POA, dates, sanction limit, remarks) are kept. Closed rows
    # keep the Renewal/Closed date — that is the closure date itself.
    # ══════════════════════════════════════════════════════════════════
    default_weight_cols = [
        (col_map.get("gdr_gross"), "GDR Gross"),
        (col_map.get("actual_gross"), "Actual Gross"),
        (col_map.get("tare"), "Tare"),
        (col_map.get("gdr_net"), "GDR Net"),
        (col_map.get("actual_net"), "Actual Net"),
        (col_map.get("spur_weight"), "Spurious Weight"),
    ]
    default_count_cols = [
        (col_map.get("gross_diff"), "Gross Diff"),
        (col_map.get("net_diff"), "Net Diff"),
        (col_map.get("ornaments_gdr"), "Ornaments GDR"),
        (col_map.get("ornaments_actual"), "Ornaments Actual"),
        (col_map.get("ornaments_diff"), "Ornaments Diff"),
        (col_map.get("spur_count"), "Spurious Count"),
        (col_map.get("spur_pct"), "Spurious %"),
        (col_map.get("carat_count"), "Carat Mismatch Count"),
        (col_map.get("uncommon_count"), "Uncommon Count"),
    ]
    blank_cols = [
        (col_map.get("gdr_no"), "GDR Number"),
        (col_map.get("magnet"), "Magnet Test"),
        (col_map.get("tampered"), "Packet Tampered"),
    ]
    for r in all_rows:
        is_closed = "CLOSED" in row_status_upper[r]
        is_topup = r in resolved_packet_rows or is_topup_remark(rows_data[r]["remarks"])
        if not (is_closed or is_topup):
            continue
        cleaned_rows.add(r)
        for wc, label in default_weight_cols:
            if wc and num_or_none(val(r, wc)) != 0.0:
                set_val(r, wc, 0.000)
                highlight(r, wc, GREEN_FILL, skip_excluded=False)
                summary["closed_topup_defaulted"].append(f"Row {r}: {label} → 0.000")
        for cc, label in default_count_cols:
            if cc and num_or_none(val(r, cc)) != 0.0:
                set_val(r, cc, 0)
                highlight(r, cc, GREEN_FILL, skip_excluded=False)
                summary["closed_topup_defaulted"].append(f"Row {r}: {label} → 0")
        clear_cols = list(blank_cols)
        if not is_closed:
            clear_cols.append((col_map.get("renewal_date"), "Renewal/Closed Date"))
        for bc, label in clear_cols:
            if bc:
                v = val(r, bc)
                if v is not None and str(v).strip() != "":
                    set_val(r, bc, None)
                    highlight(r, bc, GREEN_FILL, skip_excluded=False)
                    summary["closed_topup_cleared"].append(f"Row {r}: {label} cleared")

    # ══════════════════════════════════════════════════════════════════
    # PASS 4: Date Outlier Detection (Majority Month/Year Rule)
    # ══════════════════════════════════════════════════════════════════
    date_cols = [
        (col_map.get("sanction_date"), "Sanction Date"),
        (col_map.get("verification_date"), "Agency Verification Date"),
        (col_map.get("renewal_date"), "Renewal/Closed Date"),
    ]
    
    def get_year_month(v):
        if v is None:
            return None
        if hasattr(v, "year") and hasattr(v, "month"):
            return v.year, v.month
        s = str(v).strip()
        if not s or s.upper() in ("NAN", "NONE", "NAT", "-"):
            return None
        months_map = {
            "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
            "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
        }
        for sep in ["-", "/", ".", " "]:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) == 3:
                p0, p1, p2 = parts[0], parts[1], parts[2]
                if p2.isdigit() and len(p2) in (2, 4):
                    yr = int(p2)
                    if yr < 100:
                        yr += 2000
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return yr, m
                elif p0.isdigit() and len(p0) == 4:
                    yr = int(p0)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return yr, m
        return None

    for date_col, label in date_cols:
        if date_col:
            row_ym = {}
            for r in all_rows:
                v = val(r, date_col)
                ym = get_year_month(v)
                if ym:
                    row_ym[r] = ym
            
            if row_ym:
                ym_counts = defaultdict(int)
                for ym in row_ym.values():
                    ym_counts[ym] += 1
                
                majority_ym = max(ym_counts, key=ym_counts.get)
                majority_count = ym_counts[majority_ym]
                
                if len(ym_counts) > 1 and majority_count > len(row_ym) / 2:
                    for r, ym in row_ym.items():
                        if ym != majority_ym:
                            highlight(r, date_col, YELLOW_FILL)
                            val_str = str(val(r, date_col)).strip()
                            summary["date_outlier"].append(
                                f"Row {r}: {label} '{val_str}' deviates from majority month/year {majority_ym[1]}/{majority_ym[0]}"
                            )

    # ══════════════════════════════════════════════════════════════════
    # PASS 5: Tampered / Magnet
    # ══════════════════════════════════════════════════════════════════
    col_tampered = col_map.get("tampered")
    if col_tampered:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            if safe_str(val(r, col_tampered)).upper() == "YES":
                highlight(r, col_tampered, ORANGE_FILL)
                summary["tampered"].append(f"Row {r}: Packet Tampered = YES")
    col_magnet = col_map.get("magnet")
    if col_magnet:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            v = safe_str(val(r, col_magnet))
            if v.upper().strip() != "OK":
                highlight(r, col_magnet, ORANGE_FILL)
                summary["magnet_not_ok"].append(f"Row {r}: Magnet = '{v}'")

    # ══════════════════════════════════════════════════════════════════
    # PASS 6: Fresh accounts with Actual Gross = 0
    # ══════════════════════════════════════════════════════════════════
    col_taf = col_map.get("taf")
    col_actual_gross = col_map.get("actual_gross")
    if col_actual_gross:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            status = row_status_upper[r]
            if "FRESH" in status:
                ag_n = num_or_none(val(r, col_actual_gross))
                if ag_n is not None and ag_n == 0:
                    highlight(r, col_actual_gross, ORANGE_FILL)
                    summary["fresh_zero_actual_gross"].append(f"Row {r}: Fresh account but Actual Gross = 0")

    # ══════════════════════════════════════════════════════════════════
    # PASS 7: Gross Diff formula check
    # ══════════════════════════════════════════════════════════════════
    col_gdr_gross = col_map.get("gdr_gross")
    col_gross_diff = col_map.get("gross_diff")
    col_tare = col_map.get("tare")
    if col_gross_diff and col_gdr_gross and col_actual_gross and col_tare:
        gdr_g_let = get_column_letter(col_gdr_gross)
        act_g_let = get_column_letter(col_actual_gross)
        tare_let = get_column_letter(col_tare)
        for r in all_rows:
            status = row_status_upper[r]
            if r in cleaned_rows:
                continue

            gdr_g_n = num_or_none(val(r, col_gdr_gross))
            act_g_n = num_or_none(val(r, col_actual_gross))
            tare_n = num_or_none(val(r, col_tare))
            raw_gd = val(r, col_gross_diff)
            gd_n = num_or_none(val_data(r, col_gross_diff))

            taf_val = safe_str(val(r, col_taf)).upper().strip() if col_taf else ""

            formula = None
            expected = None
            if taf_val == "TAF" and gdr_g_n is not None and tare_n is not None:
                formula = f"={gdr_g_let}{r}-{tare_let}{r}"
                expected = round(gdr_g_n - tare_n, 1)
            elif taf_val == "POA" and gdr_g_n is not None and act_g_n is not None:
                formula = f"={gdr_g_let}{r}-{act_g_let}{r}"
                expected = round(gdr_g_n - act_g_n, 1)
            # Fallback if TAF/POA is empty
            elif not taf_val:
                if "FRESH" in status and gdr_g_n is not None and act_g_n is not None:
                    formula = f"={gdr_g_let}{r}-{act_g_let}{r}"
                    expected = round(gdr_g_n - act_g_n, 1)
                elif ("ALREADY VERIFIED" in status or "VERIFIED" in status) and gdr_g_n is not None and tare_n is not None:
                    formula = f"={gdr_g_let}{r}-{tare_let}{r}"
                    expected = round(gdr_g_n - tare_n, 1)

            # Only touch the cell when its current value is wrong or missing —
            # correct cells keep their original content and formatting. A cell
            # already holding a formula with no cached value can't be verified,
            # so it is left as-is.
            if formula:
                if gd_n is None:
                    if not (isinstance(raw_gd, str) and raw_gd.strip().startswith("=")):
                        set_val(r, col_gross_diff, formula)
                elif abs(expected - round(gd_n, 1)) > 0.05:
                    set_val(r, col_gross_diff, formula)
                    highlight(r, col_gross_diff, YELLOW_FILL)
                    summary["gross_diff_fixed"].append(f"Row {r}: Gross Diff {round(gd_n, 1)}→{expected}")

    # ══════════════════════════════════════════════════════════════════
    # PASS 8: Net Diff check
    # ══════════════════════════════════════════════════════════════════
    col_net_diff = col_map.get("net_diff")
    col_gdr_net = col_map.get("gdr_net")
    col_actual_net = col_map.get("actual_net")
    if col_net_diff and col_gdr_net and col_actual_net:
        gdr_n_let = get_column_letter(col_gdr_net)
        act_n_let = get_column_letter(col_actual_net)
        for r in all_rows:
            status = row_status_upper[r]
            if r in cleaned_rows:
                continue

            gdr_n_n = num_or_none(val(r, col_gdr_net))
            act_n_n = num_or_none(val(r, col_actual_net))
            raw_nd = val(r, col_net_diff)
            nd_n = num_or_none(val_data(r, col_net_diff))

            taf_val = safe_str(val(r, col_taf)).upper().strip() if col_taf else ""

            formula = None
            expected = None
            is_taf = False

            if taf_val == "TAF":
                formula = "=0.000"
                expected = 0.0
                is_taf = True
            elif taf_val == "POA" and gdr_n_n is not None and act_n_n is not None:
                formula = f"={gdr_n_let}{r}-{act_n_let}{r}"
                expected = round(gdr_n_n - act_n_n, 1)
            # Fallback if TAF/POA is empty
            elif not taf_val:
                if "ALREADY VERIFIED" in status:
                    formula = "=0.000"
                    expected = 0.0
                    is_taf = True
                elif "FRESH" in status and gdr_n_n is not None and act_n_n is not None:
                    formula = f"={gdr_n_let}{r}-{act_n_let}{r}"
                    expected = round(gdr_n_n - act_n_n, 1)

            # Only rewrite the cell when its current value is wrong or missing.
            if formula:
                if nd_n is None:
                    if not (isinstance(raw_nd, str) and raw_nd.strip().startswith("=")):
                        set_val(r, col_net_diff, formula)
                elif is_taf:
                    if abs(nd_n) > 0.01:
                        set_val(r, col_net_diff, formula)
                        highlight(r, col_net_diff, YELLOW_FILL)
                        summary["net_diff_av_not_zero"].append(f"Row {r}: Net Diff = {nd_n} (TAF case)")
                elif abs(expected - round(nd_n, 1)) > 0.05:
                    set_val(r, col_net_diff, formula)
                    highlight(r, col_net_diff, YELLOW_FILL)
                    summary["net_diff_fixed"].append(f"Row {r}: Net Diff {round(nd_n, 1)}→{expected} (POA case)")

    # ══════════════════════════════════════════════════════════════════
    # PASS 9: Spurious % check
    # ══════════════════════════════════════════════════════════════════
    col_spur_pct = col_map.get("spur_pct")
    col_spur_weight = col_map.get("spur_weight")
    if col_spur_pct and col_spur_weight and col_gdr_gross:
        spur_w_let = get_column_letter(col_spur_weight)
        gdr_g_let = get_column_letter(col_gdr_gross)
        for r in all_rows:
            if r in cleaned_rows:
                continue
            sw_n = num_or_none(val(r, col_spur_weight))
            gg_n = num_or_none(val(r, col_gdr_gross))
            sp_n = num_or_none(val_data(r, col_spur_pct))
            if sw_n is not None and gg_n is not None and gg_n > 0:
                expected_pct = round((sw_n / gg_n) * 100, 2)
                if abs(expected_pct - (sp_n if sp_n is not None else 0.0)) > 0.5:
                    formula = f"=ROUND(({spur_w_let}{r}/{gdr_g_let}{r})*100,2)"
                    set_val(r, col_spur_pct, formula)
                    highlight(r, col_spur_pct, YELLOW_FILL)
                    summary["spur_pct_mismatch"].append(f"Row {r}: Expected {expected_pct}%, got {sp_n if sp_n is not None else 0.0}%")

    # ══════════════════════════════════════════════════════════════════
    # PASS 10: Ornament count diff
    # ══════════════════════════════════════════════════════════════════
    col_ornaments_diff = col_map.get("ornaments_diff")
    col_ornaments_gdr = col_map.get("ornaments_gdr")
    col_ornaments_actual = col_map.get("ornaments_actual")
    if col_ornaments_diff and col_ornaments_gdr and col_ornaments_actual and col_taf:
        act_orn_let = get_column_letter(col_ornaments_actual)
        gdr_orn_let = get_column_letter(col_ornaments_gdr)
        for r in all_rows:
            if r in cleaned_rows:
                continue

            taf_val = safe_str(val(r, col_taf)).upper()
            if taf_val == "TAF":
                continue
            is_poa = taf_val == "POA"
            go_n = num_or_none(val_data(r, col_ornaments_gdr))
            ao_n = num_or_none(val_data(r, col_ornaments_actual))
            if go_n is not None and ao_n is not None:
                expected_diff = int(round(ao_n)) - int(round(go_n))

                stored = num_or_none(val_data(r, col_ornaments_diff))
                if stored is None or int(round(stored)) != expected_diff:
                    formula = f"={act_orn_let}{r}-{gdr_orn_let}{r}"
                    set_val(r, col_ornaments_diff, formula)

                if expected_diff != 0:
                    highlight(r, col_ornaments_diff, YELLOW_FILL)
                    if not is_poa:
                        highlight(r, col_ornaments_gdr, YELLOW_FILL)
                        highlight(r, col_ornaments_actual, YELLOW_FILL)
                    summary["ornament_diff_fixed"].append(f"Row {r}: Set to {expected_diff}")

    # ── TAF rows: Ornaments Diff = 0 (no highlights) ──
    if col_ornaments_diff and col_taf:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            taf_val = safe_str(val(r, col_taf)).upper()
            if taf_val == "TAF" and num_or_none(val_data(r, col_ornaments_diff)) != 0.0:
                set_val(r, col_ornaments_diff, 0)

    # ══════════════════════════════════════════════════════════════════
    # PASS 11: Highlight zero/non-positive ornaments GDR
    # ══════════════════════════════════════════════════════════════════
    if col_ornaments_gdr:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            n = num_or_none(val(r, col_ornaments_gdr))
            if n is not None and n <= 0:
                highlight(r, col_ornaments_gdr, YELLOW_FILL)
                summary["ornaments_gdr_non_positive"].append(f"Row {r}: Ornaments GDR = {val(r, col_ornaments_gdr)}")

    # ══════════════════════════════════════════════════════════════════
    # PASS 12: Whitespace normalization
    # ══════════════════════════════════════════════════════════════════
    text_cols = [
        col_map.get("account"), col_map.get("applicant"), col_packet,
        col_map.get("gdr_no"), col_remarks, col_taf,
    ]
    for tc in text_cols:
        if tc:
            for r in all_rows:
                v = val(r, tc)
                if isinstance(v, str) and v != v.strip():
                    set_val(r, tc, v.strip())
                    summary["whitespace_fixed"].append(f"Row {r} col {tc}: trimmed")

    # ══════════════════════════════════════════════════════════════════
    # PASS 13: Replace NA with 0.000 in Tare column
    # ══════════════════════════════════════════════════════════════════
    if col_tare:
        for r in all_rows:
            v = val(r, col_tare)
            if isinstance(v, str) and v.strip().upper() in ("NA", "N/A"):
                set_val(r, col_tare, 0.000)

    # ══════════════════════════════════════════════════════════════════
    # PASS 14: Weight values - flag suspiciously large/small
    # ══════════════════════════════════════════════════════════════════
    weight_check_cols = [
        (col_gdr_gross, "GDR Gross"),
        (col_actual_gross, "Actual Gross"),
        (col_tare, "Tare"),
        (col_gdr_net, "GDR Net"),
        (col_actual_net, "Actual Net"),
    ]
    for wc, label in weight_check_cols:
        if wc:
            for r in all_rows:
                v = val(r, wc)
                if v is not None:
                    try:
                        nv = float(v)
                        if nv > 5000:
                            highlight(r, wc, YELLOW_FILL)
                            summary["weight_outlier"].append(f"Row {r}: {label} = {nv} (suspiciously large)")
                        if nv < 0 and wc != col_gross_diff and wc != col_net_diff:
                            highlight(r, wc, YELLOW_FILL)
                            summary["weight_negative"].append(f"Row {r}: {label} = {nv} (negative weight)")
                    except (ValueError, TypeError):
                        pass

    # ══════════════════════════════════════════════════════════════════
    # PASS 15: Gross weight must not be less than Net weight
    # ══════════════════════════════════════════════════════════════════
    gross_net_pairs = [
        (col_gdr_gross, col_gdr_net, "GDR Gross", "GDR Net"),
        (col_actual_gross, col_actual_net, "Actual Gross", "Actual Net"),
    ]
    for gross_col, net_col, gross_label, net_label in gross_net_pairs:
        if gross_col and net_col:
            for r in all_rows:
                g = val(r, gross_col)
                n = val(r, net_col)
                if g is None or n is None:
                    continue
                try:
                    g_n = float(g)
                    n_n = float(n)
                except (ValueError, TypeError):
                    continue
                if g_n < n_n:
                    highlight(r, gross_col, YELLOW_FILL)
                    highlight(r, net_col, YELLOW_FILL)
                    summary["gross_less_than_net"].append(
                        f"Row {r}: {gross_label} ({g_n}) < {net_label} ({n_n})"
                    )

    # ══════════════════════════════════════════════════════════════════
    # PASS 16: Tare weight must not be less than GDR Gross/Net weight
    # The tare (sealed packet incl. packaging) can never weigh less than the
    # declared contents. Skipped when tare is 0/NA (e.g. POA rows where the
    # packet was opened) and on cleaned Closed/Top-Up rows.
    # ══════════════════════════════════════════════════════════════════
    if col_tare:
        tare_weight_pairs = [
            (col_gdr_gross, "GDR Gross"),
            (col_gdr_net, "GDR Net"),
        ]
        for r in all_rows:
            if r in cleaned_rows:
                continue
            t_n = num_or_none(val(r, col_tare))
            if t_n is None or t_n <= 0:
                continue
            for wcol, wlabel in tare_weight_pairs:
                if not wcol:
                    continue
                w_n = num_or_none(val(r, wcol))
                if w_n is not None and t_n < w_n:
                    highlight(r, col_tare, YELLOW_FILL)
                    highlight(r, wcol, YELLOW_FILL)
                    summary["tare_less_than_weight"].append(
                        f"Row {r}: Tare ({t_n}) < {wlabel} ({w_n})"
                    )

    return summary



# ─── Main ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    import glob

    parser = argparse.ArgumentParser(description="Report Validator")
    parser.add_argument("files", nargs="*", help="Excel files to process")
    parser.add_argument("--pdf", help="Optional PDF sequence file")
    args = parser.parse_args()

    files = args.files
    if not files:
        # Auto-detect Excel files in current directory
        files = glob.glob("*_*.xlsx")
        files = [f for f in files if "_VALIDATED" not in f]

    if not files:
        print("Usage: python report_validator.py <file1.xlsx> [file2.xlsx ...] [--pdf sequence.pdf]")
        sys.exit(1)

    all_summaries = {}
    for fp in files:
        s = process_file(fp, pdf_path=args.pdf)
        if s:
            all_summaries[fp] = s

    print(f"\n\n{'='*80}")
    print("SUMMARY ACROSS ALL FILES:")
    print(f"{'='*80}")
    grand_total = 0
    for fp, s in all_summaries.items():
        total = sum(len(v) for v in s.values())
        grand_total += total
        print(f"  {fp}: {total} issues")
    print(f"  ─────────────────────")
    print(f"  TOTAL: {grand_total} issues across {len(all_summaries)} files")
