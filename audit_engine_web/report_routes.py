"""Report Validator API routes (ported from report_automation).

Registers bottle routes on the default app for:
  POST /api/report/upload                  - start a validation job (returns file_id instantly)
  GET  /api/report/status/<file_id>        - poll job progress / completion
  POST /api/report/download                - download the validated/edited workbook

Upload is decoupled from processing: the client uploads once, receives a
`file_id` immediately, then polls `/api/report/status/<file_id>` while the
validation (which can take minutes on large sheets) runs in a background thread.
Session metadata and the generated workbook are persisted under REPORT_STORAGE_DIR
so that a download works even if the in-memory map was cleared by a restart.
"""

import datetime
import os
import re
import shutil
import sqlite3
import threading
import uuid
from copy import copy
from pathlib import Path
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from audit_engine.lib.bottle import route, request, response, static_file, HTTPError

from audit_engine.services import report_validator as rv

# ── Storage ───────────────────────────────────────────────────────────
# REPORT_STORAGE_DIR is meant to live on a persistent disk in production so
# output files survive redeploys. Falls back to a project-local dir.
_STORAGE_ENV = os.environ.get("REPORT_STORAGE_DIR")
if _STORAGE_ENV:
    REPORT_UPLOAD_DIR = Path(_STORAGE_ENV)
    REPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
else:
    REPORT_UPLOAD_DIR = Path(__file__).resolve().parent.parent / "report_uploads"
    REPORT_UPLOAD_DIR.mkdir(exist_ok=True)

# SQLite db holding job metadata (file_id -> filenames / custom name).
DB_PATH = REPORT_UPLOAD_DIR / "report_jobs.db"

# In-memory job status map: file_id -> {status, pct, error, result, ...}
JOBS = {}
JOBS_LOCK = threading.Lock()

FILL_HEX = {
    "DARK_RED": "8B0000",
    "LIGHT_RED": "FFC7CE",
    "ORANGE": "FF8C00",
    "YELLOW": "FFFF00",
    "GREEN": "C6EFCE",
    "LIGHT_BLUE": "BDD7EE",
}

FILL_MAP = {
    "8B0000": "#8B0000",
    "FFC7CE": "#FFC7CE",
    "FF8C00": "#FF8C00",
    "FFFF00": "#FFFF00",
    "C6EFCE": "#C6EFCE",
    "BDD7EE": "#BDD7EE",
}


# ── Persistence helpers ───────────────────────────────────────────────
def _init_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS report_jobs (
            file_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            source_path TEXT,
            source_name TEXT,
            pdf_path TEXT,
            output_path TEXT,
            original_name TEXT,
            custom_filename TEXT,
            stored_source TEXT,
            error TEXT,
            created_at REAL NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


_init_db()


def _upsert_job_meta(file_id, **fields):
    if "created_at" not in fields:
        fields["created_at"] = datetime.datetime.now().timestamp()
    conn = sqlite3.connect(str(DB_PATH))
    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    conn.execute(
        f"INSERT INTO report_jobs (file_id, {cols}) VALUES (?, {placeholders}) "
        f"ON CONFLICT(file_id) DO UPDATE SET "
        + ", ".join(f"{k} = excluded.{k}" for k in fields),
        (file_id, *fields.values()),
    )
    conn.commit()
    conn.close()


def _get_job_meta(file_id):
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM report_jobs WHERE file_id = ?", (file_id,)).fetchone()
    conn.close()
    if not row:
        return None
    return dict(row)


def secure_delete(path):
    """Overwrite file with random data before deleting (unrecoverable)."""
    if not path or not os.path.exists(path):
        return
    try:
        size = os.path.getsize(path)
        with open(path, "wb") as f:
            chunk_size = min(size, 65536)
            for _ in range(3):
                f.seek(0)
                remaining = size
                while remaining > 0:
                    chunk = os.urandom(min(chunk_size, remaining))
                    f.write(chunk)
                    remaining -= len(chunk)
                f.flush()
                os.fsync(f.fileno())
        os.unlink(path)
    except OSError:
        pass


def cell_ref(row, col_idx):
    return f"{get_column_letter(col_idx)}{row}"


def json_safe(val):
    """Convert non-JSON-serializable cell values (dates) to strings."""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if isinstance(val, (datetime.time)):
        return val.isoformat()
    return val


def _set_job(file_id, **fields):
    with JOBS_LOCK:
        job = JOBS.setdefault(file_id, {})
        job.update(fields)


def _job_state(file_id):
    with JOBS_LOCK:
        return dict(JOBS.get(file_id, {}))


def run_validation_and_extract(filepath, file_id, original_name=None, pdf_path=None, on_progress=None):
    """Validate an uploaded workbook and return the payload the web grid renders.

    Thin wrapper over the shared, offline validation core
    (audit_engine.services.report_validator.validate_workbook) so the web
    server and the desktop app run byte-for-byte the same passes.
    """
    result = rv.validate_workbook(
        filepath,
        output_path=REPORT_UPLOAD_DIR / f"{file_id}.xlsx",
        pdf_path=pdf_path,
        on_progress=on_progress,
        build_preview=True,
    )
    result["file_id"] = file_id
    if original_name:
        result["file_name"] = original_name
    return result


def _run_job(file_id, src_path, pdf_path, original_name):
    """Background worker. Mutates JOBS and the jobs DB, then cleans temp files."""
    err = None
    output_path = str(REPORT_UPLOAD_DIR / f"{file_id}.xlsx")
    try:
        _set_job(file_id, status="processing", pct=5, error=None)

        def on_progress(pct, msg=None):
            _set_job(file_id, pct=pct, message=msg)

        result = run_validation_and_extract(
            src_path, file_id, original_name=original_name,
            pdf_path=pdf_path, on_progress=on_progress,
        )
        _set_job(file_id, status="done", pct=100, result=result, message="Complete")
        _upsert_job_meta(file_id, status="done", output_path=output_path,
                         custom_filename=result.get("custom_filename"), error=None)
    except Exception as e:  # noqa: BLE001
        template = str(e)
        if "zipfile" in template.lower() or "bad" in template.lower():
            err = "Invalid file format. Please upload a valid .xlsx file."
        else:
            err = f"Processing error: {template}"
        _set_job(file_id, status="error", pct=100, error=err, message=err)
        _upsert_job_meta(file_id, status="error", error=err)
    finally:
        if pdf_path:
            try:
                secure_delete(pdf_path)
            except Exception:  # noqa: BLE001
                pass


# Routes ─────────────────────────────────────────────────────────────────

@route("/api/report/upload", method=["OPTIONS", "POST"])
def report_upload():
    if request.method == "OPTIONS":
        return {}
    upload = request.files.get("file")
    if not upload:
        response.status = 400
        return {"detail": "No file provided"}
    original_name = upload.filename or "upload.xlsx"
    ext = Path(original_name).suffix.lower()
    if ext != ".xlsx":
        response.status = 400
        return {"detail": "Only .xlsx files are supported (.xls is not supported)"}

    file_id = uuid.uuid4().hex[:12]

    # Persist the uploaded bytes immediately so the client doesn't wait on the
    # heavy validation. Small files are streamed to disk in one shot.
    src_target = REPORT_UPLOAD_DIR / f"src_{file_id}.xlsx"
    with open(str(src_target), "wb") as out:
        shutil.copyfileobj(upload.file, out)

    pdf_target = None
    pdf_file = request.files.get("pdf_file")
    if pdf_file:
        pdf_target = REPORT_UPLOAD_DIR / f"src_{file_id}.pdf"
        with open(str(pdf_target), "wb") as out:
            shutil.copyfileobj(pdf_file.file, out)

    output_path = str(REPORT_UPLOAD_DIR / f"{file_id}.xlsx")
    _upsert_job_meta(file_id, status="queued", original_name=original_name,
                     stored_source=str(src_target), pdf_path=str(pdf_target) if pdf_target else None,
                     output_path=output_path, custom_filename=None, error=None)

    # Fire-and-forget background processing.
    t = threading.Thread(
        target=_run_job,
        args=(file_id, src_target, pdf_target, original_name),
        daemon=True,
    )
    t.start()

    # Instant response: the client now polls status.
    return {
        "file_id": file_id,
        "status": "processing",
    }


@route("/api/report/status/<file_id>", method=["GET"])
def report_status(file_id):
    job = _job_state(file_id)
    if not job or job.get("status") in (None, "queued"):
        meta = _get_job_meta(file_id)
        if meta and meta.get("status") == "done":
            out = REPORT_UPLOAD_DIR / f"{file_id}.xlsx"
            if out.exists():
                return {"status": "done", "file_id": file_id,
                        "custom_filename": meta.get("custom_filename"),
                        "original_name": meta.get("original_name"),
                        "error": None}
        if meta and meta.get("status") == "error":
            return {"status": "error", "file_id": file_id, "error": meta.get("error")}
        response.status = 404
        return {"status": "unknown", "detail": "Job not found"}

    if job.get("status") == "processing":
        return {"status": "processing", "file_id": file_id, "pct": job.get("pct", 5),
                "message": job.get("message")}

    if job.get("status") == "error":
        return {"status": "error", "file_id": file_id, "error": job.get("error")}

    result = job.get("result")
    if result:
        return {"status": "done", "file_id": file_id, "result": result, "pct": 100}
    meta = _get_job_meta(file_id)
    return {"status": "done", "file_id": file_id,
            "result": None, "pct": 100,
            "custom_filename": (meta or {}).get("custom_filename"),
            "original_name": (meta or {}).get("original_name")}


@route("/api/report/download", method=["POST"])
def report_download():
    data = request.json or {}
    file_id = data.get("file_id")
    edits = data.get("edits") or []
    if not file_id:
        response.status = 400
        return {"detail": "Missing file_id"}

    job = _job_state(file_id)
    meta = _get_job_meta(file_id) or {}

    # Resolve the validated output file: prefer on-disk copy, fall back to any
    # in-memory record. This makes downloads work after a restart as long as the
    # file (and metadata) live under REPORT_UPLOAD_DIR.
    output_path = meta.get("output_path") or (job or {}).get("output_path")
    if not output_path:
        output_path = str(REPORT_UPLOAD_DIR / f"{file_id}.xlsx")
    if not os.path.exists(output_path):
        response.status = 404
        return {"detail": "Processed file not found"}

    original_name = meta.get("original_name") or (job or {}).get("original_name") or "validated_report.xlsx"
    custom_name = meta.get("custom_filename") or (job or {}).get("custom_filename")
    download_name = custom_name if custom_name else (Path(original_name).stem + "_VALIDATED.xlsx")

    if not edits:
        return static_file(
            os.path.basename(output_path),
            root=os.path.dirname(output_path),
            download=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    wb = openpyxl.load_workbook(output_path)
    sheet_name = "Purity Verification Format"
    if sheet_name not in wb.sheetnames:
        wb.close()
        response.status = 500
        return {"detail": "Processed file is missing expected sheet"}
    ws = wb[sheet_name]

    wb_orig = None
    ws_orig = None
    src_temp = None
    stored_source = meta.get("stored_source") or (job or {}).get("stored_source")
    if stored_source and os.path.exists(stored_source):
        wb_orig = openpyxl.load_workbook(stored_source)
        ws_orig = wb_orig[sheet_name] if sheet_name in wb_orig.sheetnames else None

    for edit in edits:
        ref = edit.get("ref")
        value = edit.get("value")
        if not ref:
            continue
        match = re.match(r"^([A-Z]+)(\d+)$", str(ref))
        if match:
            col_letter = match.group(1)
            row_num = int(match.group(2))
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            if ws_orig:
                orig_cell = ws_orig.cell(row=row_num, column=col_idx)
                cell.font = copy(orig_cell.font) if orig_cell.font else Font()
                if orig_cell.fill and orig_cell.fill.start_color and orig_cell.fill.start_color.rgb:
                    cell.fill = copy(orig_cell.fill)
                else:
                    cell.fill = PatternFill()
                if orig_cell.border:
                    cell.border = copy(orig_cell.border)
                if orig_cell.number_format and orig_cell.number_format != "General":
                    cell.number_format = orig_cell.number_format
                if orig_cell.alignment:
                    cell.alignment = copy(orig_cell.alignment)
            else:
                cell.fill = PatternFill()

    if wb_orig:
        wb_orig.close()
    if src_temp:
        secure_delete(src_temp.name)

    response_path = REPORT_UPLOAD_DIR / f"final_{file_id}.xlsx"
    wb.save(str(response_path))
    wb.close()

    return static_file(
        response_path.name,
        root=str(REPORT_UPLOAD_DIR),
        download=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )