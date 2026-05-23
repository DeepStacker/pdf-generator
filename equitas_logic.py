"""equitas_logic.py — Core audit logic for Equitas Small Finance Bank.

Two-stage gold loan audit workflow:
  Stage 1: Read master Excel (Normal + JSR sheet pairs), generate per-branch
            audit PDF worksheets and color-coded Excel templates.
  Stage 2: Read manually-filled Stage 1 Excel, consolidate multi-packet
            accounts, produce final account-level audit report.

Single source of truth for all Equitas-specific processing.
"""

import os
import re
import sys
import math
import uuid
from pathlib import Path

import pandas as pd

# --- openpyxl ---
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.page import PageMargins

# --- reportlab ---
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# =========================================================
# CONSTANTS
# =========================================================

ROWS_PER_PAGE = 13

# Column mappings — Normal sheet
NORMAL_COLUMNS = {
    "loan_no": "SVS_LOAN_NO",
    "sole_id": "SOLE_ID",
    "branch": "BRANCH_NAME",
}

# Column mappings — JSR sheet
JSR_COLUMNS = {
    "loan_no": "LOAN NO",
    "loan_date": "DATE OF LOAN",
    "customer": "CUSTOMER NAME",
    "gross_weight": "GROSSWT",
    "net_weight": "NETWT",
    "sanctioned_amount": "SANCTIONED AMOUNT",
    "ornament": "ORNAMENT DETAILS",
    "quantity": "QUANTITY",
    "carat_master": "CARAT MASTER",
    "old_packet_no": "NEW GOLD PACKET NO",
}

# PDF styles
_header_style = ParagraphStyle(
    "eq_header", fontName="Helvetica-Bold", fontSize=7.2,
    alignment=TA_CENTER, leading=8, wordWrap="CJK",
)
_body_style = ParagraphStyle(
    "eq_body", fontName="Helvetica", fontSize=6.2,
    alignment=TA_CENTER, leading=7, wordWrap="CJK",
)
_left_body_style = ParagraphStyle(
    "eq_left_body", fontName="Helvetica", fontSize=6.2,
    alignment=TA_LEFT, leading=7, wordWrap="CJK",
)


# =========================================================
# HELPERS
# =========================================================

def normalize_columns(df):
    """Uppercase and strip all column names."""
    df.columns = [str(col).strip().upper() for col in df.columns]
    return df


def clean_text(value):
    """Clean text by inserting spaces between camelCase and collapsing whitespace."""
    if pd.isna(value):
        return ""
    value = str(value)
    value = re.sub(r'([a-z])([A-Z])', r'\1 \2', value)
    value = re.sub(r'\s+', ' ', value)
    return value.strip()


def safe_value(value):
    """Convert value to stripped string, empty if NaN."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def safe_float(value):
    """Convert to float string with 3 decimals, strip '.0' from integers."""
    if pd.isna(value):
        return ""
    try:
        fval = float(value)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval:.3f}"
    except (ValueError, TypeError):
        return str(value)


def safe_float_numeric(value):
    """Convert to a numeric float, returning 0.0 for invalid/empty values."""
    if pd.isna(value):
        return 0.0
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return 0.0
            if "DIFF:" in value.upper():
                value = value.upper().replace("DIFF:", "").strip()
        fval = float(value)
        if math.isnan(fval) or math.isinf(fval):
            return 0.0
        return round(fval, 3)
    except (ValueError, TypeError):
        return 0.0


def format_date(value):
    """Format a date value to DD-MM-YYYY string."""
    if pd.isna(value):
        return ""
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return str(value)
        return parsed.strftime("%d-%m-%Y")
    except (ValueError, TypeError):
        return str(value)


def base_sheet_name(sheet_name):
    """Strip trailing _JSR suffix to get the base sheet name."""
    name = sheet_name.strip()
    if name.upper().endswith("_JSR"):
        return name[:-4].strip().lower()
    return name.lower()


def generate_verification_id():
    """Generate a 10-digit unique verification ID."""
    return str(uuid.uuid4().int)[:10]


def sanitize_filename(name):
    r"""Remove characters invalid in filenames: \ / * ? : \" < > |"""
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))


def format_3(value):
    """Format a numeric value to 3 decimal places."""
    try:
        return f"{float(value):.3f}"
    except (ValueError, TypeError):
        return "0.000"


# =========================================================
# INPUT VALIDATION
# =========================================================

def validate_equitas_stage1_file(file_path):
    """Validate that the file is a readable Excel with Normal+JSR sheet pairs.

    Returns (True, info_dict) on success, (False, error_message) on failure.
    info_dict contains: sheet_pairs (int), total_rows (int), branches (list).
    """
    if not file_path or not str(file_path).strip():
        return False, "No file path provided."

    file_path = str(file_path).strip()
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False, f"Invalid file type '{ext}'. Expected .xlsx or .xls"

    try:
        excel = pd.ExcelFile(file_path)
        sheet_names = excel.sheet_names

        # Find Normal+JSR pairs
        sheet_map = {}
        for sheet in sheet_names:
            base = base_sheet_name(sheet)
            if base not in sheet_map:
                sheet_map[base] = {"normal": None, "jsr": None}
            if sheet.upper().endswith("_JSR"):
                sheet_map[base]["jsr"] = sheet
            else:
                sheet_map[base]["normal"] = sheet

        pairs = [(b, s) for b, s in sheet_map.items()
                 if s["normal"] and s["jsr"]]

        if not pairs:
            return False, "No valid Normal + JSR sheet pairs found in this file."

        # Quick row count from first pair
        sample_df = pd.read_excel(file_path, sheet_name=pairs[0][1]["normal"])
        row_count = len(sample_df)

        excel.close()

        info = {
            "sheet_pairs": len(pairs),
            "sample_rows": row_count,
            "pair_names": [f"{s['normal']} ↔ {s['jsr']}" for _, s in pairs],
        }
        return True, info

    except PermissionError:
        return False, "File is locked or permission denied. Close it first."
    except Exception as e:
        return False, f"Cannot read file: {e}"


def validate_equitas_stage2_file(file_path):
    """Validate that the file is a Stage 1 audited Excel with required columns.

    Returns (True, info_dict) on success, (False, error_message) on failure.
    """
    if not file_path or not str(file_path).strip():
        return False, "No file path provided."

    file_path = str(file_path).strip()
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False, f"Invalid file type '{ext}'. Expected .xlsx or .xls"

    try:
        df = pd.read_excel(file_path, header=3)
        df.columns = [str(c).strip() for c in df.columns]

        required = ["ACCOUNT NO", "APPLICANT NAME", "SANCTION AMT",
                     "BANK QTY", "ACTUAL QTY", "BANK GROSS", "ACTUAL GROSS",
                     "BANK NET", "ACTUAL NET", "REMARKS"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return False, f"Missing required columns: {', '.join(missing)}"

        df = df.dropna(how="all")
        df = df[df["ACCOUNT NO"].notna()]

        info = {
            "row_count": len(df),
            "account_count": df["ACCOUNT NO"].nunique(),
            "sheet_name": "Audit",
        }
        return True, info

    except PermissionError:
        return False, "File is locked or permission denied. Close it first."
    except Exception as e:
        return False, f"Cannot read file: {e}"


# =========================================================
# STAGE 1 — BUILD MASTER DATAFRAME
# =========================================================

def _validate_required_columns(df, required_columns, sheet_name):
    """Raise if any required columns are missing from the dataframe."""
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in {sheet_name}: {missing}")


def build_master_dataframe(file_path, log_callback=print):
    """Read the master Excel, merge Normal+JSR sheet pairs, return unified DataFrame."""
    excel = pd.ExcelFile(file_path)

    try:
        sheet_map = {}
        for sheet in excel.sheet_names:
            base = base_sheet_name(sheet)
            if base not in sheet_map:
                sheet_map[base] = {"normal": None, "jsr": None}
            if sheet.upper().endswith("_JSR"):
                sheet_map[base]["jsr"] = sheet
            else:
                sheet_map[base]["normal"] = sheet

        final_frames = []

        for base, sheets in sheet_map.items():
            normal_sheet = sheets["normal"]
            jsr_sheet = sheets["jsr"]

            if not normal_sheet or not jsr_sheet:
                log_callback(f"⚠ Skipping incomplete pair: {base}")
                continue

            try:
                log_callback(f"Processing: {normal_sheet} ↔ {jsr_sheet}")

                normal_df = normalize_columns(
                    pd.read_excel(file_path, sheet_name=normal_sheet)
                )
                jsr_df = normalize_columns(
                    pd.read_excel(file_path, sheet_name=jsr_sheet)
                )

                required_normal = [
                    NORMAL_COLUMNS["loan_no"],
                    NORMAL_COLUMNS["sole_id"],
                    NORMAL_COLUMNS["branch"],
                ]
                required_jsr = [v for v in JSR_COLUMNS.values()]

                _validate_required_columns(normal_df, required_normal, normal_sheet)
                _validate_required_columns(jsr_df, required_jsr, jsr_sheet)

                merged = pd.merge(
                    normal_df[required_normal].copy(),
                    jsr_df[required_jsr].copy(),
                    left_on=NORMAL_COLUMNS["loan_no"],
                    right_on=JSR_COLUMNS["loan_no"],
                    how="left",
                )

                # Serial numbering by unique loan number
                merged["_GROUP_KEY"] = (
                    merged[NORMAL_COLUMNS["loan_no"]].astype(str).str.strip()
                )
                merged = merged.sort_values(by=["_GROUP_KEY"]).reset_index(drop=True)
                unique_keys = merged["_GROUP_KEY"].drop_duplicates().tolist()
                serial_map = {key: idx + 1 for idx, key in enumerate(unique_keys)}
                merged["SR_NO"] = merged["_GROUP_KEY"].map(serial_map)
                merged.drop(columns=["_GROUP_KEY"], inplace=True)

                final_frames.append(merged)

            except Exception as e:
                log_callback(f"⚠ Error processing {base}: {e}")

        if not final_frames:
            raise ValueError("No valid data found in any sheet pair.")

        return pd.concat(final_frames, ignore_index=True)

    finally:
        excel.close()


# =========================================================
# STAGE 1 — PDF GENERATION
# =========================================================

def _create_pdf_table_data(rows, branch_name, verification_id,
                           branch_sole_id, page_no, total_pages, total_accounts):
    """Build the table data list for a single PDF page."""
    data = []

    # Row 0: Header metadata
    data.append([
        Paragraph("<b>BRANCH NAME</b>", _header_style), "",
        Paragraph(safe_value(branch_name), _header_style),
        Paragraph("<b>BRANCH SOL ID</b>", _header_style), "",
        Paragraph(safe_value(branch_sole_id), _header_style),
        Paragraph("<b>BRANCH In Time</b>", _header_style), "", "",
        Paragraph(f"<b>PAGE</b><br/>{page_no}/{total_pages}", _header_style),
        Paragraph(f"<b>TOTAL</b><br/>{total_accounts}", _header_style),
        Paragraph("<b>Audit Start Time</b>", _header_style), "",
        "", "",
        Paragraph("<b>Audit Completed Time</b>", _header_style), "",
        "",
    ])

    # Row 1: Section headers
    data.append([
        Paragraph("Sr No", _header_style),
        Paragraph("ACCOUNT NUMBER", _header_style),
        Paragraph("APPLICANT NAME", _header_style),
        Paragraph("<b>Branch Valuation Details</b>", _header_style),
        "", "", "", "", "",
        Paragraph(
            f"<b>Verification Details</b><br/>{verification_id}",
            _header_style,
        ),
        "", "", "", "", "", "", "",
    ])

    # Row 2: Sub-headers
    data.append([
        "", "", "",
        Paragraph("ORNAMENT DETAILS", _header_style),
        Paragraph("SANCTIONED AMOUNT", _header_style),
        Paragraph("OLD PACKET NO", _header_style),
        Paragraph("DATE OF LOAN", _header_style),
        Paragraph("QTY", _header_style),
        Paragraph("GROSS WEIGHT", _header_style),
        Paragraph("TARE WEIGHT", _header_style),
        Paragraph("NEW PACKET", _header_style),
        Paragraph("CARRAT", _header_style),
        Paragraph("ACTUAL QTY", _header_style),
        Paragraph("ACTUAL GROSS", _header_style),
        Paragraph("STONE WT", _header_style),
        Paragraph("IMPURITY", _header_style),
        Paragraph("ACTUAL NET WT", _header_style),
        Paragraph("REMARK", _header_style),
    ])

    # Data rows
    for _, row in rows.iterrows():
        customer_name = clean_text(row.get(JSR_COLUMNS["customer"], ""))
        ornament = clean_text(row.get(JSR_COLUMNS["ornament"], ""))

        data.append([
            Paragraph(safe_value(row["SR_NO"]), _body_style),
            Paragraph(safe_value(row.get(NORMAL_COLUMNS["loan_no"], "")), _body_style),
            Paragraph(customer_name, _left_body_style),
            Paragraph(ornament, _left_body_style),
            Paragraph(safe_float(row.get(JSR_COLUMNS["sanctioned_amount"], "")), _body_style),
            Paragraph(safe_value(row.get(JSR_COLUMNS["old_packet_no"], "")), _body_style),
            Paragraph(format_date(row.get(JSR_COLUMNS["loan_date"], "")), _body_style),
            Paragraph(safe_float(row.get(JSR_COLUMNS["quantity"], "")), _body_style),
            Paragraph(safe_float(row.get(JSR_COLUMNS["gross_weight"], "")), _body_style),
            "", "", "", "", "", "", "", "",
        ])

    return data


def _generate_summary_page(branch_name, total_accounts, branch_sole_id, verification_id):
    """Generate the summary table element for end of branch PDF."""
    data = [
        ["SOL ID", branch_sole_id, "BRANCH NAME", branch_name],
        ["BANK NAME", "Equitas Small Finance Bank", "AUDIT DATE", ""],
        ["BRANCH IN TIME", "", "AUDIT START TIME", ""],
        ["AUDIT COMPLETED TIME", "", "VERIFICATION ID", verification_id],
        ["TOTAL ACCOUNTS RECEIVED", str(total_accounts),
         "TOTAL ACCOUNTS AUDITED", str(total_accounts)],
        ["AUDITOR NAME", "", "AUDITOR SIGNATURE", ""],
        ["BRANCH MANAGER NAME", "", "BRANCH MANAGER SIGNATURE", ""],
        ["REMARKS / OBSERVATIONS", "", "", ""],
    ]

    table = Table(
        data,
        colWidths=[190, 220, 190, 220],
        rowHeights=[40, 40, 55, 55, 45, 75, 75, 140],
    )
    table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1.5, colors.black),
        ("GRID", (0, 0), (-1, -1), 1.2, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("SPAN", (1, 7), (3, 7)),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
    ]))
    return table


def generate_branch_pdf(branch_name, branch_df, output_dir):
    """Generate the per-branch audit PDF worksheet.

    Returns the output file path.
    """
    os.makedirs(output_dir, exist_ok=True)
    safe_branch = sanitize_filename(branch_name)
    output_path = os.path.join(output_dir, f"{safe_branch}_Audit_Report.pdf")

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=4, rightMargin=4,
        topMargin=4, bottomMargin=4,
        title=f"{branch_name} Audit Report",
        author="Audit System",
        subject="Gold Loan Audit Verification",
    )

    elements = []
    branch_df = branch_df.sort_values(by=["SR_NO"])
    usable_width = landscape(A4)[0] - 20

    weights = [2, 7, 11, 9, 5, 6, 5, 3, 4, 5, 7, 3, 3, 5, 5, 4, 5, 5]
    total_weight = sum(weights)
    col_widths = [usable_width * (w / total_weight) for w in weights]

    total_rows = len(branch_df)
    total_pages = (total_rows // ROWS_PER_PAGE) + (
        1 if total_rows % ROWS_PER_PAGE else 0
    )

    verification_id = generate_verification_id()
    branch_sole_id = safe_value(
        branch_df.iloc[0].get(NORMAL_COLUMNS["sole_id"], "")
    )

    page_no = 1
    for start in range(0, total_rows, ROWS_PER_PAGE):
        end = start + ROWS_PER_PAGE
        page_rows = branch_df.iloc[start:end]

        table_data = _create_pdf_table_data(
            page_rows, branch_name, verification_id,
            branch_sole_id, page_no, total_pages, total_rows,
        )

        row_heights = [25, 16, 35] + [36] * len(page_rows)

        table = Table(
            table_data, colWidths=col_widths,
            rowHeights=row_heights, repeatRows=3,
        )

        style = TableStyle([
            ("BOX", (0, 0), (-1, -1), 1.2, colors.black),
            ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
            # Row 0 spans
            ("SPAN", (0, 0), (1, 0)), ("SPAN", (3, 0), (4, 0)),
            ("SPAN", (6, 0), (8, 0)), ("SPAN", (11, 0), (12, 0)),
            ("SPAN", (15, 0), (16, 0)),
            # Row 1-2 spans
            ("SPAN", (0, 1), (0, 2)), ("SPAN", (1, 1), (1, 2)),
            ("SPAN", (2, 1), (2, 2)),
            ("SPAN", (3, 1), (8, 1)), ("SPAN", (9, 1), (16, 1)),
            # Backgrounds
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9d9d9")),
            ("BACKGROUND", (3, 1), (8, 1), colors.HexColor("#ffe699")),
            ("BACKGROUND", (3, 2), (8, 2), colors.HexColor("#fff2cc")),
            ("BACKGROUND", (9, 1), (16, 1), colors.HexColor("#9dc3e6")),
            ("BACKGROUND", (9, 2), (16, 2), colors.HexColor("#ddebf7")),
            ("BACKGROUND", (0, 1), (2, 2), colors.HexColor("#d9d9d9")),
            ("LINEBELOW", (0, 2), (-1, 2), 1.2, colors.black),
            ("FONTNAME", (0, 0), (-1, 2), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ])

        table.setStyle(style)
        elements.append(table)

        if end < total_rows:
            elements.append(PageBreak())
        page_no += 1

    # Summary page
    elements.append(PageBreak())
    elements.append(_generate_summary_page(
        branch_name, total_rows, branch_sole_id, verification_id,
    ))

    doc.build(elements)
    return output_path


# =========================================================
# STAGE 1 — EXCEL GENERATION
# =========================================================

def generate_branch_excel(branch_name, branch_df, output_dir):
    """Generate the per-branch audit Excel template with formulas and styling.

    Returns the output file path.
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True, size=11)

    yellow_fill = PatternFill("solid", fgColor="FFE699")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    grey_fill = PatternFill("solid", fgColor="BFBFBF")
    red_fill = PatternFill("solid", fgColor="F4CCCC")

    # Top header row (branch metadata)
    top_headers = [
        "BRANCH SOL ID", "BRANCH NAME", "STATE", "VERIFIER NAME",
        "VERIFICATION DATE", "BRANCH IN TIME", "AUDIT START TIME",
        "AUDIT END TIME", "TOTAL ACCOUNTS RECEIVED",
        "TOTAL ACCOUNTS AUDITED", "REMAINING ACCOUNTS", "REMARKS",
    ]

    for idx, value in enumerate(top_headers, start=1):
        c = ws.cell(row=1, column=idx, value=value)
        c.font = bold
        c.border = border
        c.alignment = center
        c.fill = grey_fill

    branch_sole_id = safe_value(
        branch_df.iloc[0].get(NORMAL_COLUMNS["sole_id"], "")
    )

    second_row = [
        branch_sole_id, branch_name, "", "", "", "", "", "",
        branch_df["SR_NO"].nunique(), branch_df["SR_NO"].nunique(), "", "",
    ]

    for idx, value in enumerate(second_row, start=1):
        c = ws.cell(row=2, column=idx, value=value)
        c.border = border
        c.alignment = center

    # Main table headers (row 4)
    headers = [
        "SR", "ACCOUNT NO", "APPLICANT NAME",
        "SANCTION AMT", "LOAN DATE",
        "OLD PKT NO", "NEW PKT NO",
        "BANK QTY", "ACTUAL QTY", "QTY DIFF",
        "BANK GROSS", "ACTUAL GROSS", "GROSS DIFF",
        "BANK NET", "ACTUAL NET", "NET DIFF",
        "MASTER CARAT", "ACTUAL CARAT", "CARAT DIFF",
        "REMARKS", "ADDITIONAL DETAILS", "ORNAMENT DETAILS",
    ]

    header_row = 4
    for idx, value in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=value)
        c.font = bold
        c.border = border
        c.alignment = center
        if idx in [1, 2, 3, 4, 5, 6, 8, 11, 14, 17, 22]:
            c.fill = yellow_fill  # Bank data
        elif idx in [7, 9, 12, 15, 18]:
            c.fill = blue_fill  # Manual entry
        else:
            c.fill = green_fill  # System generated

    # Data rows (starting row 5)
    row_num = 5
    LEFT_ALIGN_COLS = [3, 21, 22]

    for _, row in branch_df.iterrows():
        values = [
            row["SR_NO"],
            safe_value(row.get(NORMAL_COLUMNS["loan_no"], "")),
            clean_text(row.get(JSR_COLUMNS["customer"], "")),
            safe_float(row.get(JSR_COLUMNS["sanctioned_amount"], "")),
            format_date(row.get(JSR_COLUMNS["loan_date"], "")),
            safe_value(row.get(JSR_COLUMNS["old_packet_no"], "")),
            "",  # NEW PKT NO (manual)
            safe_float(row.get(JSR_COLUMNS["quantity"], "")),
            "",  # ACTUAL QTY (manual)
            f'=IF(I{row_num}="","",ROUND(I{row_num}-H{row_num},3))',
            safe_float(row.get(JSR_COLUMNS["gross_weight"], "")),
            "",  # ACTUAL GROSS (manual)
            f'=IF(L{row_num}="","",ROUND(L{row_num}-K{row_num},3))',
            safe_float(row.get(JSR_COLUMNS["net_weight"], "")),
            "",  # ACTUAL NET (manual)
            f'=IF(O{row_num}="","",ROUND(O{row_num}-N{row_num},3))',
            safe_value(row.get(JSR_COLUMNS["carat_master"], "")),
            "",  # ACTUAL CARAT (manual)
            f'=IF(R{row_num}="","",IF(R{row_num}=Q{row_num},"DIFF: 0","DIFF: "&ROUND((R{row_num}-Q{row_num}),3)))',
            f'=IF(S{row_num}<>"DIFF: 0","Carat mismatch","OK - no discrepancy")',
            f'=IF(S{row_num}<>"DIFF: 0",'
            f'I{row_num}&" "&V{row_num}&" ("&TEXT(L{row_num},"0.000")&") - ACTUAL "&R{row_num}&" CARAT","")',
            clean_text(row.get(JSR_COLUMNS["ornament"], "")),
        ]

        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=col_idx, value=value)
            c.border = border
            c.alignment = left_align if col_idx in LEFT_ALIGN_COLS else center

        ws.row_dimensions[row_num].height = 32
        row_num += 1

    # Column widths
    widths = {
        1: 6, 2: 16, 3: 24, 4: 14, 5: 14, 6: 14, 7: 14,
        8: 11, 9: 11, 10: 11, 11: 13, 12: 13, 13: 13,
        14: 13, 15: 13, 16: 13, 17: 12, 18: 12, 19: 14,
        20: 24, 21: 45, 22: 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[4].height = 42

    # Conditional formatting — red highlight for non-zero diffs
    for col in ["J", "M", "P"]:
        ws.conditional_formatting.add(
            f"{col}5:{col}{row_num}",
            FormulaRule(formula=[f'{col}5<>0'], fill=red_fill),
        )
    ws.conditional_formatting.add(
        f"S5:S{row_num}",
        FormulaRule(formula=['S5<>"DIFF: 0"'], fill=red_fill),
    )

    ws.auto_filter.ref = f"A4:V{row_num}"
    ws.freeze_panes = "D5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)

    safe_branch = sanitize_filename(branch_name)
    output_path = os.path.join(output_dir, f"{safe_branch}_Audit.xlsx")
    wb.save(output_path)
    wb.close()

    return output_path


# =========================================================
# STAGE 1 — MAIN ENTRY POINT
# =========================================================

def run_equitas_stage1(file_path, output_dir, log_callback=print, cancel_event=None, progress_callback=None, output_format="BOTH", output_mode="FOLDER"):
    """Execute the full Stage 1 pipeline.

    Args:
        file_path: Path to master Excel file.
        output_dir: Base output directory.
        log_callback: Function to receive log messages.
        cancel_event: threading.Event to check for cancellation.
        progress_callback: Function to receive progress updates (0 to 100).
        output_format: "PDF ONLY", "EXCEL ONLY", or "BOTH"
        output_mode: "FOLDER", "ZIP OF PDF", "ZIP OF EXCEL", "ZIP OF BOTH", or "BOTH (FOLDER + ZIP OF ...)"

    Returns:
        (pdf_count, excel_count) tuple.
    """
    pdf_output_dir = os.path.join(output_dir, "output_pdfs")
    excel_output_dir = os.path.join(output_dir, "output_excels")

    log_callback("Building master dataframe from sheet pairs...")
    master_df = build_master_dataframe(file_path, log_callback)

    branch_col = NORMAL_COLUMNS["branch"]
    grouped = master_df.groupby(branch_col)
    total_branches = len(grouped)

    log_callback(f"Found {total_branches} branches. Starting generation...")

    pdf_count = 0
    excel_count = 0

    for branch_name, branch_df in grouped:
        if cancel_event and cancel_event.is_set():
            log_callback(f"⚠ CANCELLED after processing {pdf_count if output_format != 'EXCEL ONLY' else excel_count}/{total_branches} branches.")
            break

        try:
            log_callback(f"Building: {branch_name}")

            if output_format in ("PDF ONLY", "BOTH"):
                pdf_path = generate_branch_pdf(branch_name, branch_df, pdf_output_dir)
                log_callback(f"  ✓ PDF: {os.path.basename(pdf_path)}")
                pdf_count += 1

            if output_format in ("EXCEL ONLY", "BOTH"):
                excel_path = generate_branch_excel(branch_name, branch_df, excel_output_dir)
                log_callback(f"  ✓ Excel: {os.path.basename(excel_path)}")
                excel_count += 1

            if progress_callback:
                # If needs_zip is True, let's keep progress below 100 (e.g. max 90) until compression is done.
                needs_zip = output_mode != "FOLDER"
                pdf_pct_max = 90.0 if needs_zip else 100.0
                gen_count = pdf_count + excel_count
                expected_total = total_branches * (2 if output_format == "BOTH" else 1)
                progress_callback((gen_count / expected_total) * pdf_pct_max)

        except Exception as e:
            log_callback(f"  ✗ Error processing {branch_name}: {e}")

    was_cancelled = cancel_event and cancel_event.is_set()

    if not was_cancelled:
        import zipfile
        import shutil

        def zip_dir(src_dir, zip_filepath):
            if not os.path.exists(src_dir):
                return
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root_dir, _, files in os.walk(src_dir):
                    for file in files:
                        full_path = os.path.join(root_dir, file)
                        rel_path = os.path.relpath(full_path, src_dir)
                        zipf.write(full_path, rel_path)

        zip_pdf = "ZIP OF PDF" in output_mode or "ZIP OF BOTH" in output_mode
        zip_excel = "ZIP OF EXCEL" in output_mode or "ZIP OF BOTH" in output_mode
        delete_raw = output_mode.startswith("ZIP OF")

        if zip_pdf and os.path.exists(pdf_output_dir):
            log_callback("Compressing PDFs...")
            if progress_callback:
                progress_callback(92)
            pdf_zip_path = f"{pdf_output_dir}.zip"
            zip_dir(pdf_output_dir, pdf_zip_path)
            log_callback(f"✓ PDF ZIP created: {os.path.basename(pdf_zip_path)}")
            if delete_raw:
                log_callback("Cleaning up raw PDFs...")
                try:
                    shutil.rmtree(pdf_output_dir)
                except OSError as re:
                    log_callback(f"  ✗ Cleanup Error: {re}")

        if zip_excel and os.path.exists(excel_output_dir):
            log_callback("Compressing Excels...")
            if progress_callback:
                progress_callback(96)
            excel_zip_path = f"{excel_output_dir}.zip"
            zip_dir(excel_output_dir, excel_zip_path)
            log_callback(f"✓ Excel ZIP created: {os.path.basename(excel_zip_path)}")
            if delete_raw:
                log_callback("Cleaning up raw Excels...")
                try:
                    shutil.rmtree(excel_output_dir)
                except OSError as re:
                    log_callback(f"  ✗ Cleanup Error: {re}")

    return pdf_count, excel_count



# =========================================================
# STAGE 2 — CONSOLIDATION HELPERS
# =========================================================

def _safe_str(v):
    if pd.isna(v):
        return ""
    return str(v).strip()


def _avg_int(series):
    vals = []
    for x in series:
        try:
            val = safe_float_numeric(x)
            if val == 0:
                continue
            vals.append(val)
        except (ValueError, TypeError):
            continue
    if not vals:
        return ""
    avg = sum(vals) / len(vals)
    if math.isnan(avg):
        return ""
    return int(round(avg))


def _merge_issue_only(series):
    vals = []
    for x in series:
        x = _safe_str(x)
        if not x:
            continue
        if x.lower() == "ok - no discrepancy":
            continue
        if x not in vals:
            vals.append(x)
    return ", ".join(vals) if vals else "OK - no discrepancy"


def _merge_non_empty(series):
    vals = []
    for x in series:
        x = _safe_str(x)
        if x and x not in vals:
            vals.append(x)
    return ", ".join(vals)


def _first_non_empty(series):
    for x in series:
        x = _safe_str(x)
        if x:
            return x
    return ""


# =========================================================
# STAGE 2 — LOAD & CONSOLIDATE
# =========================================================

def load_stage1_excel(file_path):
    """Load Stage 1 audited Excel (header on row 4, i.e. header=3)."""
    df = pd.read_excel(file_path, header=3)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")
    df = df[df["ACCOUNT NO"].notna()]
    return df


def consolidate_accounts(df, log_callback=print):
    """Consolidate multi-packet records into one row per account."""
    grouped = df.groupby("ACCOUNT NO", dropna=False)
    rows = []
    sr_no = 1

    for acc_no, grp in grouped:
        try:
            sanction_amt = safe_float_numeric(
                _first_non_empty(grp["SANCTION AMT"])
            )

            bank_qty = round(grp["BANK QTY"].apply(safe_float_numeric).sum(), 3)
            actual_qty = round(grp["ACTUAL QTY"].apply(safe_float_numeric).sum(), 3)

            bank_gross = round(grp["BANK GROSS"].apply(safe_float_numeric).sum(), 3)
            actual_gross = round(grp["ACTUAL GROSS"].apply(safe_float_numeric).sum(), 3)

            bank_net = round(grp["BANK NET"].apply(safe_float_numeric).sum(), 3)
            actual_net = round(grp["ACTUAL NET"].apply(safe_float_numeric).sum(), 3)

            remarks = _merge_issue_only(grp["REMARKS"])
            additional_details = _merge_non_empty(grp["ADDITIONAL DETAILS"])
            avg_carat = _avg_int(grp["ACTUAL CARAT"])

            rows.append({
                "S.NO": sr_no,
                "ACCOUNT NUMBER": _safe_str(acc_no),
                "APPLICANT NAME": _first_non_empty(grp["APPLICANT NAME"]),
                "SANCTIONED_AMOUNT": format_3(sanction_amt),
                "DATE_OF_LOAN": _first_non_empty(grp["LOAN DATE"]),
                "OLD PACKET NO.": _first_non_empty(grp["OLD PKT NO"]),
                "NEW PACKET NO.": _first_non_empty(grp["NEW PKT NO"]),
                "Bank Valuation No. Of Ornaments": int(round(bank_qty)),
                "Actual No. of Ornaments": int(round(actual_qty)),
                "Bank Valuation Gross Wt.": format_3(bank_gross),
                "Actual Gross Wt.": format_3(actual_gross),
                "Bank Valuation Net Wt.": format_3(bank_net),
                "Actual Net Wt.": format_3(actual_net),
                "REMARKS": remarks,
                "ADDITIONAL DETAILS": additional_details,
                "AVG CARAT": avg_carat,
            })
            sr_no += 1

        except Exception as e:
            log_callback(f"⚠ Error processing account {acc_no}: {e}")

    return pd.DataFrame(rows)


# =========================================================
# STAGE 2 — EXCEL GENERATION
# =========================================================

def generate_stage2_excel(original_file, consolidated_df, output_dir):
    """Generate the consolidated account-level audit report Excel.

    Returns the output file path.
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Audit"

    # Styles
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True, size=11)

    grey_fill = PatternFill("solid", fgColor="BFBFBF")
    yellow_fill = PatternFill("solid", fgColor="FFE699")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    red_fill = PatternFill("solid", fgColor="F4CCCC")

    # Copy top header from original file
    old_wb = load_workbook(original_file, data_only=True)
    try:
        old_ws = old_wb.active
        for r in range(1, 3):
            for c in range(1, 13):
                val = old_ws.cell(row=r, column=c).value
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                cell.alignment = center
                if r == 1:
                    cell.fill = grey_fill
                    cell.font = bold
    finally:
        old_wb.close()

    # Headers (row 4)
    headers = [
        "S.NO", "ACCOUNT NUMBER", "APPLICANT NAME",
        "SANCTIONED_AMOUNT", "DATE_OF_LOAN",
        "OLD PACKET NO.", "NEW PACKET NO.",
        "Bank Valuation No. Of Ornaments", "Actual No. of Ornaments",
        "Difference in No. Of Ornaments",
        "Bank Valuation Gross Wt.", "Actual Gross Wt.",
        "Gross Weight Difference",
        "Bank Valuation Net Wt.", "Actual Net Wt.",
        "Net Weight Difference",
        "REMARKS", "ADDITIONAL DETAILS",
    ]

    header_row = 4
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = bold
        c.border = border
        c.alignment = center
        c.fill = yellow_fill if idx <= 7 else green_fill

    # Data rows
    row_num = 5
    LEFT_ALIGN_COLS = [3, 17, 18]

    for _, row in consolidated_df.iterrows():
        vals = [
            row["S.NO"],
            row["ACCOUNT NUMBER"],
            row["APPLICANT NAME"],
            row["SANCTIONED_AMOUNT"],
            row["DATE_OF_LOAN"],
            row["OLD PACKET NO."],
            row["NEW PACKET NO."],
            row["Bank Valuation No. Of Ornaments"],
            row["Actual No. of Ornaments"],
            f'=ROUND(I{row_num}-H{row_num},0)',
            row["Bank Valuation Gross Wt."],
            row["Actual Gross Wt."],
            f'=ROUND(L{row_num}-K{row_num},3)',
            row["Bank Valuation Net Wt."],
            row["Actual Net Wt."],
            f'=ROUND(O{row_num}-N{row_num},3)',
            row["REMARKS"],
            row["ADDITIONAL DETAILS"],
        ]

        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=v)
            cell.border = border
            cell.alignment = left_align if col_idx in LEFT_ALIGN_COLS else center

        ws.row_dimensions[row_num].height = 38
        row_num += 1

    # Column widths
    widths = {
        1: 8, 2: 18, 3: 30, 4: 18, 5: 16, 6: 18, 7: 18,
        8: 18, 9: 18, 10: 18, 11: 18, 12: 18, 13: 18,
        14: 18, 15: 18, 16: 18, 17: 26, 18: 65,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[4].height = 42

    # Conditional formatting — red for non-OK remarks
    ws.conditional_formatting.add(
        f"Q5:Q{row_num}",
        FormulaRule(formula=['Q5<>"OK - no discrepancy"'], fill=red_fill),
    )

    ws.auto_filter.ref = f"A4:R{row_num}"
    ws.freeze_panes = "C5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)

    input_name = Path(original_file).stem
    output_path = os.path.join(output_dir, f"{input_name}_Consolidated.xlsx")
    wb.save(output_path)
    wb.close()

    return output_path


# =========================================================
# STAGE 2 — MAIN ENTRY POINT
# =========================================================

def run_equitas_stage2(file_path, output_dir, log_callback=print, cancel_event=None, progress_callback=None):
    """Execute the full Stage 2 consolidation pipeline.

    Args:
        file_path: Path to Stage 1 audited Excel file.
        output_dir: Base output directory.
        log_callback: Function to receive log messages.
        cancel_event: threading.Event to check for cancellation.
        progress_callback: Function to receive progress updates (0 to 100).

    Returns:
        output_path (str) of the consolidated report.
    """
    log_callback("Loading Stage 1 audited Excel...")
    df = load_stage1_excel(file_path)
    log_callback(f"Loaded {len(df)} rows.")

    if cancel_event and cancel_event.is_set():
        log_callback("⚠ CANCELLED before consolidation.")
        return None

    if progress_callback:
        progress_callback(20)

    log_callback("Consolidating accounts...")
    consolidated_df = consolidate_accounts(df, log_callback)
    log_callback(f"Consolidated to {len(consolidated_df)} unique accounts.")

    if progress_callback:
        progress_callback(70)

    if cancel_event and cancel_event.is_set():
        log_callback("⚠ CANCELLED before report generation.")
        return None

    log_callback("Generating consolidated audit report...")
    output_path = generate_stage2_excel(file_path, consolidated_df, output_dir)
    log_callback(f"✓ Report generated: {os.path.basename(output_path)}")

    if progress_callback:
        progress_callback(100)

    return output_path
