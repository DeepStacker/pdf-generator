"""pdf_logic.py — Core PDF generation logic for IDFC Audit Engine Elite.

Single source of truth for:
- Excel reading, validation, and row grouping
- PDF report generation (landscape A4 audit worksheets)
- Font management and resource path resolution
"""

import os
import sys
import openpyxl
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# =========================================================
# UTILITIES
# =========================================================

def get_resource_path(relative_path):
    """Get absolute path to a bundled resource.

    Works both in development (uses script directory) and in
    PyInstaller frozen executables (uses sys._MEIPASS).
    """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


# =========================================================
# CONSTANTS & CONFIGURATION
# =========================================================

# Page layout (landscape A4)
PAGE_SIZE = landscape(A4)
PAGE_MARGIN_LEFT = 50.2
PAGE_MARGIN_RIGHT = 50.2
PAGE_MARGIN_TOP = 48.0
PAGE_MARGIN_BOTTOM = 15.0

# Column widths (points) — matches IDFC audit template spec
COL_WIDTHS = [22.2, 101.1, 118.5, 60.3, 67.2, 125.0, 247.3]

# Row heights (points)
HEADER_ROW_1_HEIGHT = 12.2
HEADER_ROW_2_HEIGHT = 14.2
COLUMN_HEADER_HEIGHT = 22.5
DATA_ROW_HEIGHT = 30.5

# Colors
COLOR_HEADER_BANK = colors.HexColor("#FFFF00")   # Yellow — bank-side columns
COLOR_HEADER_AUDIT = colors.HexColor("#4985E8")   # Blue — auditor-side columns
COLOR_BORDER = colors.black
BORDER_WIDTH = 0.5

# Font configuration
FONT_DIR = get_resource_path("fonts")
FONT_REG_PATH = os.path.join(FONT_DIR, "Carlito-Regular.ttf")
FONT_BOLD_PATH = os.path.join(FONT_DIR, "Arimo-Bold.ttf")

# Font names (set by register_fonts, fallback to Helvetica)
FONT_REG = 'Helvetica'
FONT_BLD = 'Helvetica-Bold'

# Required columns in the input Excel
REQUIRED_COLUMNS = [
    "Prospectno", "CUID", "Tare Weight",
    "State", "CurrentBranch", "CurrentBranchName"
]


def register_fonts():
    """Register custom fonts for PDF generation.

    Falls back to Helvetica/Helvetica-Bold if custom fonts are unavailable.
    """
    global FONT_REG, FONT_BLD
    try:
        if os.path.exists(FONT_REG_PATH) and os.path.exists(FONT_BOLD_PATH):
            pdfmetrics.registerFont(TTFont('Carlito', FONT_REG_PATH))
            pdfmetrics.registerFont(TTFont('ArimoBold', FONT_BOLD_PATH))
            FONT_REG = 'Carlito'
            FONT_BLD = 'ArimoBold'
    except (IOError, OSError, ValueError):
        FONT_REG = 'Helvetica'
        FONT_BLD = 'Helvetica-Bold'


# Register fonts at import time
register_fonts()


# =========================================================
# INPUT VALIDATION
# =========================================================

def validate_excel(excel_path):
    """Validate that the file is a readable Excel file with required columns.

    Returns:
        (True, "") on success.
        (False, error_message) on failure.
    """
    if not excel_path or not excel_path.strip():
        return False, "No file path provided."

    if not os.path.exists(excel_path):
        return False, f"File not found: {excel_path}"

    ext = os.path.splitext(excel_path)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False, f"Invalid file type '{ext}'. Expected .xlsx or .xls"

    try:
        with open(excel_path, 'rb') as f:
            header = f.read(4)
            if len(header) < 4:
                return False, "File is empty or corrupt."
    except PermissionError:
        return False, "File is locked or permission denied. Close it in Excel first."
    except OSError as e:
        return False, f"Cannot read file: {e}"

    return True, ""


def validate_output_dir(output_dir):
    """Validate that the output directory is writable.

    Creates the directory if it doesn't exist.
    Returns (True, "") on success, (False, error_message) on failure.
    """
    if not output_dir or not output_dir.strip():
        return False, "No output directory specified."

    try:
        os.makedirs(output_dir, exist_ok=True)
        test_file = os.path.join(output_dir, ".write_test")
        with open(test_file, 'w') as f:
            f.write("test")
        os.remove(test_file)
    except OSError as e:
        return False, f"Output directory is not writable: {e}"

    return True, ""


# =========================================================
# CORE LOGIC
# =========================================================

def format_tare_weight(val):
    """Format tare weight value for display.

    Strips trailing '.0' from whole numbers, handles NaN/None/empty gracefully.
    """
    if pd.isna(val) or val == "" or val is None:
        return ""
    try:
        fval = float(val)
        return str(int(fval)) if fval == int(fval) else str(fval)
    except (ValueError, TypeError):
        return str(val)


def read_excel(excel_path, log_callback=print):
    """Read and parse an Excel audit file.

    Scans all sheets for one containing the required columns
    (case-insensitive match). Returns (sheet_name, headers, rows)
    where rows is a list of dicts.

    Raises:
        Exception: If no valid sheet is found with all required columns.
    """
    log_callback(f"Reading Excel: {os.path.basename(excel_path)}")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    try:
        required_lower = [c.lower() for c in REQUIRED_COLUMNS]
        target_sheet = None

        for sname in wb.sheetnames:
            ws = wb[sname]
            try:
                header_row = [
                    str(cell.value).strip().replace("\n", "") if cell.value else ""
                    for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                if all(r in [h.lower() for h in header_row] for r in required_lower):
                    target_sheet = sname
                    break
            except StopIteration:
                continue

        if not target_sheet:
            raise Exception(
                f"No valid sheet found. Required columns: {', '.join(REQUIRED_COLUMNS)}"
            )

        log_callback(f"Found valid sheet: {target_sheet}")
        ws = wb[target_sheet]
        headers = [
            str(cell.value).strip().replace("\n", "") if cell.value else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data, all_none = {}, True
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    val = cell.value
                    if val is not None:
                        all_none = False
                    if headers[i] in ("Prospectno", "CUID"):
                        row_data[headers[i]] = str(val).strip() if val is not None else ""
                    else:
                        row_data[headers[i]] = val
            if not all_none:
                rows.append(row_data)

        return target_sheet, headers, rows
    finally:
        wb.close()


def group_by_branch(rows):
    """Group data rows by CurrentBranch code.

    Normalizes empty, 'None', and 'nan' branch codes to 'UNKNOWN'.
    Returns dict of {branch_code: [rows]}.
    """
    groups = {}
    for row in rows:
        branch = str(row.get("CurrentBranch", "UNKNOWN")).strip()
        if branch in ("None", "", "nan"):
            branch = "UNKNOWN"
        groups.setdefault(branch, []).append(row)
    return groups


def generate_pdf(audit_type, branch_code, branch_name, state, rows, output_path):
    """Generate a single branch audit PDF report.

    Creates a landscape A4 page with:
    - Header rows: audit type, branch code, branch name, state
    - Column headers with bank-side (yellow) and audit-side (blue) styling
    - Data rows with Prospectno, CUID, and Tare Weight (bank)
    - Empty columns for auditor: Tare Weight (Audit), Purity Check, Remarks
    """
    if not rows:
        return

    PAGE_WIDTH, PAGE_HEIGHT = PAGE_SIZE
    doc = SimpleDocTemplate(
        output_path,
        pagesize=(PAGE_WIDTH, PAGE_HEIGHT),
        leftMargin=PAGE_MARGIN_LEFT,
        rightMargin=PAGE_MARGIN_RIGHT,
        topMargin=PAGE_MARGIN_TOP,
        bottomMargin=PAGE_MARGIN_BOTTOM,
        title=os.path.basename(output_path)
    )

    style_hdr = ParagraphStyle(
        "ColHdr",
        fontName=FONT_BLD,
        fontSize=9,
        alignment=TA_CENTER,
        leading=10,
        spaceBefore=0,
        spaceAfter=0
    )

    table_data = [
        ["Audit Type :", "", str(audit_type), "", "Branch Name :", "", str(branch_name)],
        ["Branch Code :", "", str(branch_code), "", "State :", "", str(state)],
        [
            Paragraph("Sr<br/>No", style_hdr),
            Paragraph("Prospectno", style_hdr),
            Paragraph("CUID", style_hdr),
            Paragraph("Tare Weight<br/>as per Bank", style_hdr),
            Paragraph("<nobr>Tare Weight as</nobr><br/>per Audit", style_hdr),
            Paragraph("<nobr>Purity Check - 18K and</nobr><br/><nobr>above 18K or Below 18K</nobr>", style_hdr),
            Paragraph("Remarks", style_hdr)
        ]
    ]

    for idx, row in enumerate(rows, 1):
        table_data.append([
            str(idx),
            str(row.get("Prospectno", "")),
            str(row.get("CUID", "")),
            format_tare_weight(row.get("Tare Weight", "")),
            "", "", ""
        ])

    row_heights = (
        [HEADER_ROW_1_HEIGHT, HEADER_ROW_2_HEIGHT, COLUMN_HEADER_HEIGHT]
        + [DATA_ROW_HEIGHT] * (len(table_data) - 3)
    )

    table = Table(table_data, colWidths=COL_WIDTHS, rowHeights=row_heights, repeatRows=3)

    style_cmds = [
        # Span header metadata cells
        ("SPAN", (0, 0), (1, 0)), ("SPAN", (4, 0), (5, 0)),
        ("SPAN", (0, 1), (1, 1)), ("SPAN", (4, 1), (5, 1)),
        # Global alignment
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        # Header font styling (rows 0-1)
        ("FONTNAME", (0, 0), (-1, 1), FONT_BLD),
        ("FONTSIZE", (0, 0), (-1, 1), 9),
        # Data font styling (rows 3+)
        ("FONTNAME", (0, 3), (-1, -1), FONT_REG),
        ("FONTSIZE", (0, 3), (-1, -1), 9),
        # Sr No column bold
        ("FONTNAME", (0, 3), (0, -1), FONT_BLD),
        # Column header backgrounds
        ("BACKGROUND", (0, 2), (3, 2), COLOR_HEADER_BANK),
        ("BACKGROUND", (4, 2), (6, 2), COLOR_HEADER_AUDIT),
        # Tight padding
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        # Borders — precise per-region styling
        ("LINEBEFORE", (0, 0), (0, -1), BORDER_WIDTH, COLOR_BORDER),
        ("LINEAFTER", (6, 0), (6, -1), BORDER_WIDTH, COLOR_BORDER),
        ("LINEABOVE", (0, 0), (-1, 0), BORDER_WIDTH, COLOR_BORDER),
        ("LINEBELOW", (0, -1), (-1, -1), BORDER_WIDTH, COLOR_BORDER),
        ("GRID", (0, 0), (2, 1), BORDER_WIDTH, COLOR_BORDER),
        ("GRID", (4, 0), (6, 1), BORDER_WIDTH, COLOR_BORDER),
        ("GRID", (0, 2), (-1, -1), BORDER_WIDTH, COLOR_BORDER),
    ]

    table.setStyle(TableStyle(style_cmds))
    doc.build([table])
