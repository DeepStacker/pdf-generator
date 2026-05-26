"""Tests for Excel column remapping (preprocess.py)."""

import os

import openpyxl
import pytest

from audit_engine.web.preprocess import preprocess_mapped_excel


@pytest.fixture
def idfc_excel(tmp_path):
    """Create a minimal IDFC-formatted Excel file."""
    p = os.path.join(str(tmp_path), "idfc_input.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = [
        "Prospect No", "Cust ID", "Gross Weight",
        "Region", "Branch Code", "Branch Name"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="P001")
    ws.cell(row=2, column=2, value="C001")
    ws.cell(row=2, column=3, value="50.5")
    ws.cell(row=2, column=4, value="MH")
    ws.cell(row=2, column=5, value="BR001")
    ws.cell(row=2, column=6, value="Mumbai")
    wb.save(str(p))
    wb.close()
    return str(p)


@pytest.fixture
def equitas_excel_single(tmp_path):
    """Create a minimal Equitas single-sheet Excel."""
    p = os.path.join(str(tmp_path), "eq_input.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NORMAL"
    headers = [
        "SVS Loan", "Sole Proprietor ID", "Branch",
        "Customer", "Amount"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="LN001")
    ws.cell(row=2, column=2, value="SP001")
    ws.cell(row=2, column=3, value="Branch A")
    wb.save(str(p))
    wb.close()
    return str(p)


class TestPreprocessIdfc:
    def test_remap_idfc_columns(self, idfc_excel):
        """preprocess_mapped_excel renames IDFC columns."""
        mappings = {
            "prospect": "Prospect No",
            "cuid": "Cust ID",
            "tare": "Gross Weight",
            "branch": "Branch Name",
        }
        result = preprocess_mapped_excel(idfc_excel, mappings, "IDFC First Bank")
        assert "mapped_" in os.path.basename(result)
        assert ".temp_audit_engine" in result
        assert os.path.exists(result)

        wb = openpyxl.load_workbook(result, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Prospectno" in headers
        assert "CUID" in headers
        assert "Tare Weight" in headers
        assert "CurrentBranchName" in headers
        assert "CurrentBranch" in headers
        assert "State" in headers
        wb.close()
        os.unlink(result)

    def test_remap_idfc_partial_mapping(self, idfc_excel):
        """Only the mapped columns get renamed."""
        mappings = {
            "prospect": "Prospect No",
        }
        result = preprocess_mapped_excel(idfc_excel, mappings, "IDFC First Bank")
        wb = openpyxl.load_workbook(result, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Prospectno" in headers
        assert "Cust ID" in headers  # original survives
        wb.close()
        os.unlink(result)

    def test_remap_idfc_missing_branch_header(self, tmp_path):
        """Without a branch mapping, CurrentBranchName is absent so CurrentBranch isn't added, but State is."""
        p = os.path.join(str(tmp_path), "no_branch.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Prospect No")
        ws.cell(row=1, column=2, value="Cust ID")
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value="C001")
        wb.save(str(p))
        wb.close()

        mappings = {"prospect": "Prospect No", "cuid": "Cust ID"}
        result = preprocess_mapped_excel(p, mappings, "IDFC First Bank")
        wb = openpyxl.load_workbook(result, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # CurrentBranch is only added when CurrentBranchName exists; without a branch mapping it's absent
        assert "CurrentBranch" not in headers
        assert "State" in headers
        wb.close()
        os.unlink(result)


class TestPreprocessEquitas:
    def test_remap_equitas_normal_sheet(self, equitas_excel_single):
        mappings = {
            "svs": "SVS Loan",
            "sole": "Sole Proprietor ID",
            "branch": "Branch",
        }
        result = preprocess_mapped_excel(equitas_excel_single, mappings, "Equitas Small Finance Bank")
        assert "mapped_" in os.path.basename(result)

        wb = openpyxl.load_workbook(result, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "SVS_LOAN_NO" in headers
        assert "SOLE_ID" in headers
        assert "BRANCH_NAME" in headers
        wb.close()
        os.unlink(result)

    def test_remap_equitas_no_mapping(self, equitas_excel_single):
        """With empty mapping, the file passes through."""
        result = preprocess_mapped_excel(equitas_excel_single, {}, "Equitas Small Finance Bank")
        assert os.path.exists(result)
        os.unlink(result)

    def test_remap_fallback_on_exception(self, tmp_path):
        """When preprocessing fails, the original filepath is returned."""
        result = preprocess_mapped_excel("/nonexistent/file.xlsx", {"a": "b"}, "IDFC First Bank")
        assert result == "/nonexistent/file.xlsx"
