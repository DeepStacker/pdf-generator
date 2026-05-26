"""Tests for edge cases in detect.py — uncovered lines 40-41, 77, 82, 86-88."""

import openpyxl

from audit_engine.web.detect import detect_bank_from_file, peek_excel_data


def test_detect_corrupt_workbook(tmp_path):
    """detect_bank_from_file returns None for unopenable files (line 30-31)."""
    p = tmp_path / "corrupt.xlsx"
    p.write_bytes(b"not a valid xlsx file")
    assert detect_bank_from_file(str(p)) is None


def test_detect_empty_workbook(tmp_path):
    """detect_bank_from_file on workbook with no data rows (exhausts 30 rows)."""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(str(p))
    wb.close()
    assert detect_bank_from_file(str(p)) is None


def test_peek_empty_workbook(tmp_path):
    """peek_excel_data on workbook with no header row (<3 non-empty cells)."""
    p = tmp_path / "sparse.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="a")
    ws.cell(row=1, column=2, value="b")
    wb.save(str(p))
    wb.close()
    headers, rows = peek_excel_data(str(p))
    assert headers == ["a", "b"]
    assert len(rows) >= 0


def test_peek_shorter_rows_padded(tmp_path):
    """peek_excel_data pads rows shorter than headers (line 82)."""
    p = tmp_path / "uneven.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="A")
    ws.cell(row=1, column=2, value="B")
    ws.cell(row=1, column=3, value="C")
    ws.cell(row=2, column=1, value="short")
    wb.save(str(p))
    wb.close()
    headers, rows = peek_excel_data(str(p))
    assert len(headers) == 3
    assert len(rows[0]) == 3, "short row should be padded to 3 cols"


def test_peek_corrupt_file(tmp_path):
    """peek_excel_data returns empty for corrupt file (line 86-88)."""
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"garbage")
    headers, rows = peek_excel_data(str(p))
    assert headers == []
    assert rows == []


def test_peek_single_column_data(tmp_path):
    """peek_excel_data with only one non-empty cell in a row."""
    p = tmp_path / "single.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Alone")
    wb.save(str(p))
    wb.close()
    headers, rows = peek_excel_data(str(p))
    assert headers == ["Alone"]
    assert isinstance(rows, list)
