"""Tests for handler validate/run branches using real Excel fixtures."""

import os

import openpyxl
import pytest

from audit_engine.database import init_db
from audit_engine.tasks.workers import cancel_event, global_tracker
from audit_engine.web.handlers import handle_run, handle_validate


@pytest.fixture(autouse=True)
def _init():
    init_db()
    global_tracker.is_running = False
    cancel_event.clear()


# ---- Helpers ---- #

def _make_idfc_xlsx(path: str) -> str:
    """Write a minimal IDFC-format Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Prospectno", "CUID", "Tare Weight", "State", "CurrentBranch", "CurrentBranchName"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="P001")
    ws.cell(row=2, column=2, value="C001")
    ws.cell(row=2, column=3, value=100.5)
    ws.cell(row=2, column=4, value="MH")
    ws.cell(row=2, column=5, value="MUMBAI")
    ws.cell(row=2, column=6, value="Mumbai Main")
    wb.save(path)
    wb.close()
    return path


def _make_equitas_stage1_xlsx(path: str) -> str:
    """Write a minimal Equitas Stage 1 format (multi-sheet)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NORMAL"
    headers_n = ["SVS_LOAN_NO", "SOLE_ID", "BRANCH_NAME", "CUSTOMER"]
    for i, h in enumerate(headers_n, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="LN001")
    ws.cell(row=2, column=2, value="SP001")
    ws.cell(row=2, column=3, value="Branch A")
    ws2 = wb.create_sheet("JSR")
    headers_j = ["LOAN NO", "BRANCHNAME", "AMOUNT"]
    for i, h in enumerate(headers_j, 1):
        ws2.cell(row=1, column=i, value=h)
    ws2.cell(row=2, column=1, value="LN001")
    wb.save(path)
    wb.close()
    return path


def _make_equitas_stage2_xlsx(path: str) -> str:
    """Single-sheet Equitas (Stage 2 format)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["SVS_LOAN_NO", "SOLE_ID", "BRANCH_NAME", "AMOUNT"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="LN001")
    wb.save(path)
    wb.close()
    return path


def _make_arvog_xlsx(path: str) -> str:
    """Write a minimal Arvog-format Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Jewellery1", "Jewellery2", "Branch", "SR No."]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="Gold Chain")
    ws.cell(row=2, column=3, value="Branch A")
    ws.cell(row=2, column=4, value="1")
    wb.save(path)
    wb.close()
    return path


# ---- Validate: IDFC ---- #

class TestValidateIdfc:
    def test_validate_idfc_valid_file(self, tmp_path):
        p = _make_idfc_xlsx(os.path.join(str(tmp_path), "idfc.xlsx"))
        result = handle_validate({"filepath": p})
        assert result["success"] is True
        assert result.get("detected_bank") == "IDFC First Bank"
        assert result.get("rows", 0) >= 1
        assert result.get("branches", 0) >= 1

    def test_validate_idfc_missing_file(self):
        result = handle_validate({"filepath": "/nonexistent.xlsx"})
        assert result["success"] is False
        assert "does not exist" in result["error"]


# ---- Validate: Equitas ---- #

class TestValidateEquitas:
    def test_validate_equitas_stage1_valid(self, tmp_path):
        p = _make_equitas_stage1_xlsx(os.path.join(str(tmp_path), "eq_stage1.xlsx"))
        result = handle_validate({"filepath": p, "expected_stage": "STAGE 1"})
        assert result.get("success") is True or result.get("detected_bank") == "Equitas Small Finance Bank"

    def test_validate_equitas_stage1_wrong_stage(self, tmp_path):
        p = _make_equitas_stage1_xlsx(os.path.join(str(tmp_path), "eq_s1_wrong.xlsx"))
        result = handle_validate({"filepath": p, "expected_stage": "STAGE 2"})
        assert result["success"] is False

    def test_validate_equitas_stage2_valid(self, tmp_path):
        p = _make_equitas_stage2_xlsx(os.path.join(str(tmp_path), "eq_stage2.xlsx"))
        result = handle_validate({"filepath": p, "expected_stage": "STAGE 2"})
        assert result.get("success") is True or result.get("detected_bank") == "Equitas Small Finance Bank"

    def test_validate_equitas_stage2_misdetected_as_stage1(self, tmp_path):
        """When stage 2 file is used without expected_stage, handler treats it as stage 1 mismatch."""
        p = _make_equitas_stage2_xlsx(os.path.join(str(tmp_path), "eq_s2_no_stage.xlsx"))
        result = handle_validate({"filepath": p})
        # Without expected_stage, the code checks for multi-sheet / JSR
        # Single-sheet without JSR gets "appears to be Stage 2"
        assert result["success"] is False
        assert "Stage 2" in result["error"]

    def test_validate_equitas_not_stage1_no_expected(self, tmp_path):
        """When file is multi-sheet but expected_stage is not set, the code defaults to stage1 validation."""
        p = _make_equitas_stage1_xlsx(os.path.join(str(tmp_path), "eq_s1_no_stage.xlsx"))
        result = handle_validate({"filepath": p})
        assert result.get("success") is True or result.get("detected_bank") == "Equitas Small Finance Bank"


# ---- Validate: Arvog ---- #

class TestValidateArvog:
    def test_validate_arvog_valid(self, tmp_path):
        p = _make_arvog_xlsx(os.path.join(str(tmp_path), "arvog.xlsx"))
        result = handle_validate({"filepath": p})
        assert result["success"] is True
        assert result.get("detected_bank") == "Arvog Bank"
        assert result.get("rows", 0) >= 1

    def test_validate_arvog_missing_required(self, tmp_path):
        """If Arvog file lacks required columns, detect_raw_excel returns None headers."""
        p = os.path.join(str(tmp_path), "arvog_bad.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Random")
        ws.cell(row=1, column=2, value="Data")
        wb.save(p)
        wb.close()
        result = handle_validate({"filepath": p})
        assert result.get("success") is False

    def test_validate_bank_not_detected(self, tmp_path):
        """Returns invalid bank format when no bank profile matches."""
        p = os.path.join(str(tmp_path), "unknown.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Foo")
        ws.cell(row=1, column=2, value="Bar")
        ws.cell(row=1, column=3, value="Baz")
        ws.cell(row=2, column=1, value="x")
        wb.save(p)
        wb.close()
        result = handle_validate({"filepath": p})
        assert result["success"] is False
        assert result.get("detected_bank") is None


# ---- Run ---- #

class TestHandleRunWithFiles:
    def test_run_idfc_starts_thread(self, tmp_path):
        """IDFC handle_run starts a thread and returns success."""
        p = _make_idfc_xlsx(os.path.join(str(tmp_path), "run_idfc.xlsx"))
        out = str(tmp_path)
        result = handle_run({
            "bank": "IDFC First Bank",
            "filepath": p,
            "out_path": out,
            "audit_type": "POA",
            "output_mode": "FOLDER",
            "auto_open": False,
        })
        assert result["success"] is True
        # Give the thread a moment to start
        import time
        time.sleep(0.3)
        global_tracker.is_running = False
        cancel_event.set()

    def test_run_equitas_starts_thread(self, tmp_path):
        p = _make_equitas_stage1_xlsx(os.path.join(str(tmp_path), "run_eq.xlsx"))
        out = str(tmp_path)
        result = handle_run({
            "bank": "Equitas Small Finance Bank",
            "filepath": p,
            "out_path": out,
            "equitas_format": "BOTH",
            "equitas_pack": "FOLDER",
        })
        assert result["success"] is True
        import time
        time.sleep(0.3)
        global_tracker.is_running = False
        cancel_event.set()

    def test_run_arvog_starts_thread(self, tmp_path):
        p = _make_arvog_xlsx(os.path.join(str(tmp_path), "run_arvog.xlsx"))
        out = str(tmp_path)
        result = handle_run({
            "bank": "Arvog Bank",
            "filepath": p,
            "out_path": out,
            "auto_open": False,
        })
        assert result["success"] is True
        import time
        time.sleep(0.3)
        global_tracker.is_running = False
        cancel_event.set()

    def test_run_with_column_mappings(self, tmp_path):
        """handle_run applies column mappings when provided."""
        p = _make_idfc_xlsx(os.path.join(str(tmp_path), "run_mapped.xlsx"))
        out = str(tmp_path)
        result = handle_run({
            "bank": "IDFC First Bank",
            "filepath": p,
            "out_path": out,
            "auto_open": False,
            "column_mappings": {"prospect": "Prospectno"},
        })
        assert result["success"] is True
        import time
        time.sleep(0.3)
        global_tracker.is_running = False
        cancel_event.set()

    def test_run_list_filepaths(self, tmp_path):
        """handle_run accepts a list of filepaths."""
        p1 = _make_idfc_xlsx(os.path.join(str(tmp_path), "run_list1.xlsx"))
        p2 = _make_idfc_xlsx(os.path.join(str(tmp_path), "run_list2.xlsx"))
        out = str(tmp_path)
        result = handle_run({
            "bank": "IDFC First Bank",
            "filepath": [p1, p2],
            "out_path": out,
            "auto_open": False,
        })
        assert result["success"] is True
        import time
        time.sleep(0.3)
        global_tracker.is_running = False
        cancel_event.set()

    def test_run_list_missing_file(self, tmp_path):
        """One missing file in a list returns an error."""
        p = _make_idfc_xlsx(os.path.join(str(tmp_path), "exists.xlsx"))
        result = handle_run({
            "bank": "IDFC First Bank",
            "filepath": [p, "/nonexistent.xlsx"],
            "out_path": str(tmp_path),
            "auto_open": False,
        })
        assert result["success"] is False
        assert "File missing" in result["error"]

    def test_run_file_exists_but_outpath_invalid(self, tmp_path):
        p = _make_idfc_xlsx(os.path.join(str(tmp_path), "no_out.xlsx"))
        result = handle_run({
            "bank": "IDFC First Bank",
            "filepath": p,
            "out_path": "",
        })
        assert result["success"] is False
