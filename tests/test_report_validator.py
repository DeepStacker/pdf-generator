"""Tests for the Report Validator (audit_engine.services.report_validator).

Covers:
- Validation runs without a PDF (PDF upload is optional).
- Fonts / number formats of untouched cells are preserved exactly.
- Highlight fonts keep the original font family/size.
- Diff cells are only rewritten when their value is actually wrong.
- New tare-weight check (tare must not be less than GDR gross/net).
- Closed/Top-Up rows are reset to default values.
- PDF row reordering moves the full row (including far-right columns).
"""

import importlib
from copy import copy

import openpyxl
import pytest
from openpyxl.styles import Font

from audit_engine.services import report_validator as rv

SHEET = "Purity Verification Format"

HEADERS = [
    ("packet", "PACKET NO."),
    ("account", "ACCOUNT NUMBER"),
    ("applicant", "APPLICANT NAME"),
    ("status", "FRESH/RENEWAL/CLOSED/ALREADY VERIFIED"),
    ("taf", "TAF/POA"),
    ("remarks", "AGENCY REMARKS"),
    ("magnet", "MAGNET TEST RESULT"),
    ("tampered", "PACKET TAMPERED YES/NO"),
    ("gdr_gross", "GROSS WEIGHT AS PER GDR/PACKET"),
    ("actual_gross", "ACTUAL GROSS WEIGHT AS PER FCU AGENCY VERIFICATION"),
    ("tare", "ACTUAL TARE WEIGHT AS PER FCU AGENCY VERIFICATION"),
    ("gross_diff", "DIFFERENCE IN GROSS WEIGHT"),
    ("gdr_net", "NET WEIGHT AS PER GDR/PACKET"),
    ("actual_net", "ACTUAL NET WEIGHT AS PER FCU AGENCY VERIFICATION"),
    ("net_diff", "DIFFERENCE IN NET WEIGHT"),
    ("spur_count", "TOTAL NO.OF SPURIOUS ORNAMENTS"),
    ("spur_weight", "SPURIOUS ORNAMENTS GROSS WEIGHT"),
    ("spur_pct", "% OF SPURIOUS ORNAMENTS GROSS WEIGHT"),
    ("carat_count", "TOTAL NO.OF ORNAMENTS WITH CARAT MISMATCH"),
    ("uncommon_count", "TOTAL NO.OF UNCOMMON ORNAMENTS"),
    ("sanction_date", "SANCTION DATE"),
    ("verification_date", "AGENCY VERIFICATION DATE"),
    ("sanction_limit", "SANCTION LIMIT"),
    ("gdr_no", "GDR NUMBER"),
    ("ornaments_gdr", "TOTAL NO.OF ORNAMENTS AS PER THE GDR/PACKET"),
    ("ornaments_actual", "ACTUAL AVAILABLE ORNAMENTS AT THE TIME OF FCU VERIFICATION"),
    ("ornaments_diff", "DIFFRENCE IN ACTUAL ORNAMENTS"),
    ("renewal_date", "RENEWAL/CLOSED DATE"),
    ("branch", "BRANCH NAME"),
]

KEY_TO_COL = {k: i for i, (k, _h) in enumerate(HEADERS, start=1)}
TEST_FONT = Font(name="Arial", size=9, italic=True)
WEIGHT_KEYS = {"gdr_gross", "actual_gross", "tare", "gdr_net", "actual_net", "spur_weight"}


def make_workbook(path, rows):
    """Write a Purity Verification Format workbook; data starts at row 5.

    Every populated cell gets the Arial-9-italic test font; weight cells get a
    0.000 number format so format-preservation can be asserted after a run.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    for idx, (_k, header) in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=idx, value=header)
    for rn, row in enumerate(rows, start=5):
        for k, v in row.items():
            if v is None:
                continue
            cell = ws.cell(row=rn, column=KEY_TO_COL[k], value=v)
            cell.font = copy(TEST_FONT)
            if k in WEIGHT_KEYS:
                cell.number_format = "0.000"
    wb.save(path)
    wb.close()


def good_row(packet, **over):
    """A consistent POA row that should produce no findings."""
    row = {
        "packet": packet,
        "account": f"9998887776665{packet[-2:]}",
        "applicant": f"NAME {packet}",
        "status": "FRESH",
        "taf": "POA",
        "remarks": "",
        "magnet": "OK",
        "tampered": "NO",
        "gdr_gross": 20.5,
        "actual_gross": 20.0,
        "tare": 21.0,
        "gross_diff": 0.5,
        "gdr_net": 19.0,
        "actual_net": 18.8,
        "net_diff": 0.2,
        "spur_count": 0,
        "spur_weight": 0,
        "spur_pct": 0,
        "carat_count": 0,
        "uncommon_count": 0,
        "gdr_no": f"GDR{packet}",
        "ornaments_gdr": 5,
        "ornaments_actual": 5,
        "ornaments_diff": 0,
        "branch": "TESTBR",
    }
    row.update(over)
    return row


def run_file(tmp_path, rows, pdf_path=None):
    src = tmp_path / "input.xlsx"
    make_workbook(src, rows)
    summary = rv.process_file(str(src), pdf_path=pdf_path)
    out = tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx"
    assert out.exists(), f"expected output workbook at {out}"
    wb = openpyxl.load_workbook(out)
    return summary, wb[SHEET]


def cell_of(ws, row, key):
    return ws.cell(row=row, column=KEY_TO_COL[key])


def fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.start_color is None or fill.start_color.rgb is None:
        return None
    rgb = str(fill.start_color.rgb)
    return rgb[-6:] if rgb not in ("00000000",) else None


class TestPdfOptional:
    def test_runs_without_pdf(self, tmp_path):
        summary, ws = run_file(tmp_path, [good_row("P01"), good_row("P02")])
        assert summary is not None
        # Clean file: no issues at all
        assert sum(len(v) for v in summary.values()) == 0

    def test_pdf_reorder_end_to_end(self, tmp_path):
        # Build a PDF whose account sequence is the reverse of the sheet's
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        rows = [good_row("P01"), good_row("P02"), good_row("P03")]
        accounts = [r["account"] for r in rows]
        pdf = tmp_path / "seq.pdf"
        c = canvas.Canvas(str(pdf), pagesize=A4)
        y = 700
        for acct in reversed(accounts):
            c.drawString(100, y, f"Account: {acct}")
            y -= 40
        c.save()

        _summary, ws = run_file(tmp_path, rows, pdf_path=str(pdf))
        got = [cell_of(ws, r, "account").value for r in (5, 6, 7)]
        assert got == list(reversed(accounts))


class TestFormatPreservation:
    def test_untouched_cells_keep_font_and_number_format(self, tmp_path):
        _summary, ws = run_file(tmp_path, [good_row("P01")])
        for key in ("packet", "applicant", "gdr_gross", "gross_diff", "net_diff"):
            cell = cell_of(ws, 5, key)
            assert cell.font.name == "Arial", key
            assert cell.font.size == 9, key
            assert cell.font.italic is True, key
            assert fill_rgb(cell) is None, key
        assert cell_of(ws, 5, "gdr_gross").number_format == "0.000"

    def test_correct_diff_values_not_rewritten(self, tmp_path):
        _summary, ws = run_file(tmp_path, [good_row("P01")])
        # Values were correct → still static numbers, not formulas
        assert cell_of(ws, 5, "gross_diff").value == 0.5
        assert cell_of(ws, 5, "net_diff").value == 0.2
        assert cell_of(ws, 5, "ornaments_diff").value == 0

    def test_wrong_gross_diff_rewritten_and_flagged(self, tmp_path):
        summary, ws = run_file(tmp_path, [good_row("P01", gross_diff=5.0)])
        v = cell_of(ws, 5, "gross_diff").value
        assert isinstance(v, str) and v.startswith("=")
        assert fill_rgb(cell_of(ws, 5, "gross_diff")) == "FFFF00"
        assert summary["gross_diff_fixed"]

    def test_duplicate_packet_highlight_keeps_font_family(self, tmp_path):
        dup = good_row("P02")
        dup["packet"] = "P01"
        summary, ws = run_file(tmp_path, [good_row("P01"), dup])
        assert summary["duplicate_packet"]
        for r in (5, 6):
            cell = cell_of(ws, r, "packet")
            assert fill_rgb(cell) == "8B0000"
            # emphasis applied…
            assert cell.font.bold is True
            assert str(cell.font.color.rgb)[-6:] == "FFFFFF"
            # …but the original font family/size/italics survive
            assert cell.font.name == "Arial"
            assert cell.font.size == 9
            assert cell.font.italic is True


class TestTareCheck:
    def test_tare_below_gdr_weights_flagged(self, tmp_path):
        summary, ws = run_file(tmp_path, [good_row("P01", tare=15.0)])
        assert summary["tare_less_than_weight"]
        assert fill_rgb(cell_of(ws, 5, "tare")) == "FFFF00"
        assert fill_rgb(cell_of(ws, 5, "gdr_gross")) == "FFFF00"
        assert fill_rgb(cell_of(ws, 5, "gdr_net")) == "FFFF00"

    def test_zero_or_na_tare_not_flagged(self, tmp_path):
        summary, _ws = run_file(tmp_path, [good_row("P01", tare=0), good_row("P02", tare="NA")])
        assert not summary["tare_less_than_weight"]

    def test_closed_row_not_flagged(self, tmp_path):
        summary, _ws = run_file(
            tmp_path, [good_row("P01", status="CLOSED", tare=15.0, renewal_date="01-05-2026")]
        )
        assert not summary["tare_less_than_weight"]

    def test_gross_less_than_net_still_flagged(self, tmp_path):
        summary, ws = run_file(tmp_path, [good_row("P01", gdr_gross=18.0, gross_diff=-2.0)])
        assert summary["gross_less_than_net"]


class TestTopupRemarkMatching:
    @pytest.mark.parametrize(
        "remarks,expected",
        [
            ("Top Up of account 123", True),
            ("Top-up", True),
            ("TopUp", True),
            ("Top  up", True),
            ("This Account No is Topup of Account no 123, with Same Paket No", True),
            ("", False),
            ("Normal remark", False),
            ("Desktop Update pending", False),
            ("Laptop upgrade needed", False),
        ],
    )
    def test_word_boundary_matching(self, remarks, expected):
        assert rv.is_topup_remark(remarks) is expected

    def test_unrelated_remark_not_cleaned(self, tmp_path):
        # A row with an ordinary remark that happens to contain "top"/"up" as
        # parts of other words must NOT be treated as a top-up and wiped.
        rows = [good_row("P01", remarks="Desktop Update pending, ok")]
        summary, ws = run_file(tmp_path, rows)
        assert cell_of(ws, 5, "gdr_gross").value == 20.5
        assert not summary["closed_topup_defaulted"]
        assert not summary["closed_topup_cleared"]


class TestClosedTopupClean:
    def test_topup_row_reset_to_defaults(self, tmp_path):
        rows = [
            good_row("P01"),
            good_row(
                "P02",
                remarks="Top Up of account 123",
                gdr_gross=11.1,
                actual_gross=22.2,
                tare=33.3,
                gross_diff=1.0,
                gdr_net=10.0,
                actual_net=9.0,
                net_diff=1.0,
                spur_count=2,
                spur_weight=3.3,
                spur_pct=15.0,
                carat_count=1,
                uncommon_count=1,
                ornaments_gdr=4,
                ornaments_actual=3,
                ornaments_diff=-1,
                magnet="NOT OK",
                tampered="YES",
                renewal_date="01-05-2026",
            ),
        ]
        summary, ws = run_file(tmp_path, rows)
        r = 6
        for key in ("gdr_gross", "actual_gross", "tare", "gdr_net", "actual_net", "spur_weight"):
            assert cell_of(ws, r, key).value == 0, key
        for key in (
            "gross_diff", "net_diff", "ornaments_gdr", "ornaments_actual",
            "ornaments_diff", "spur_count", "spur_pct", "carat_count", "uncommon_count",
        ):
            assert cell_of(ws, r, key).value == 0, key
        for key in ("gdr_no", "magnet", "tampered"):
            assert cell_of(ws, r, key).value is None, key
        # The Renewal/Closed date is real loan data and must survive on a
        # top-up row, not just on a closed one.
        assert cell_of(ws, r, "renewal_date").value == "01-05-2026"
        # Identity/loan columns untouched
        assert cell_of(ws, r, "packet").value == "P02"
        assert cell_of(ws, r, "status").value == "FRESH"
        assert cell_of(ws, r, "taf").value == "POA"
        # Cleaned row must not be re-flagged by magnet/tampered/ornament checks
        assert not summary["magnet_not_ok"]
        assert not summary["tampered"]
        assert not summary["ornaments_gdr_non_positive"]
        assert summary["closed_topup_defaulted"]
        assert summary["closed_topup_cleared"]

    def test_closed_row_keeps_renewal_date(self, tmp_path):
        rows = [good_row("P01", status="CLOSED", renewal_date="01-05-2026", magnet="NA")]
        _summary, ws = run_file(tmp_path, rows)
        assert cell_of(ws, 5, "renewal_date").value == "01-05-2026"
        assert cell_of(ws, 5, "gdr_gross").value == 0
        assert cell_of(ws, 5, "magnet").value is None

    def test_resolved_topup_row_also_cleaned(self, tmp_path):
        empty_pkt = good_row("P02", applicant="SHARED NAME", gdr_gross=12.5)
        empty_pkt["packet"] = None
        rows = [
            good_row("P01", applicant="SHARED NAME"),
            empty_pkt,
        ]
        summary, ws = run_file(tmp_path, rows)
        assert summary["topup_resolved"]
        # Packet copied from the matching row, remark written
        assert cell_of(ws, 6, "packet").value == "P01"
        assert "Topup" in str(cell_of(ws, 6, "remarks").value)
        # And the row was reset to defaults
        assert cell_of(ws, 6, "gdr_gross").value == 0
        assert cell_of(ws, 6, "gdr_no").value is None


class TestRearrange:
    def _build(self, tmp_path):
        src = tmp_path / "input.xlsx"
        rows = [good_row("P01"), good_row("P02"), good_row("P03")]
        make_workbook(src, rows)
        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        # Far-right extra column, well beyond mapped columns + buffer
        ws.cell(row=2, column=40, value="EXTRA NOTE")
        for i, r in enumerate((5, 6, 7)):
            ws.cell(row=r, column=40, value=f"note-{i + 1}")
        ws.cell(row=5, column=40).font = Font(bold=True)  # style only on row 5
        return wb, ws, [row["account"] for row in rows]

    def test_far_right_columns_move_with_rows(self, tmp_path):
        _wb, ws, accounts = self._build(tmp_path)
        col_map = rv.build_col_map(rv.map_columns(ws))
        rv.rearrange_rows_by_pdf(ws, col_map, [5, 6, 7], list(reversed(accounts)))
        assert [ws.cell(row=r, column=KEY_TO_COL["account"]).value for r in (5, 6, 7)] == list(reversed(accounts))
        assert [ws.cell(row=r, column=40).value for r in (5, 6, 7)] == ["note-3", "note-2", "note-1"]

    def test_no_stale_styles_after_move(self, tmp_path):
        _wb, ws, accounts = self._build(tmp_path)
        col_map = rv.build_col_map(rv.map_columns(ws))
        rv.rearrange_rows_by_pdf(ws, col_map, [5, 6, 7], list(reversed(accounts)))
        # Row 5 now holds old row 7's cells, which had no bold style — the
        # bold style from the previous occupant must not linger.
        assert ws.cell(row=5, column=40).font.bold is not True
        # And old row 5's bold cell moved down to row 7 with its style.
        assert ws.cell(row=7, column=40).font.bold is True


class TestRoutesLayer:
    def test_run_validation_and_extract_without_pdf(self, tmp_path, monkeypatch):
        monkeypatch.setenv("REPORT_STORAGE_DIR", str(tmp_path / "storage"))
        import audit_engine_web.report_routes as rr

        rr = importlib.reload(rr)

        src = tmp_path / "input.xlsx"
        # Top-up row with NO weight cells at all: the clean pass must create
        # them with defaults, and the preview must include those new cells.
        rows = [
            good_row("P01"),
            good_row(
                "P02",
                remarks="Top Up of account 123",
                gdr_gross=None, actual_gross=None, tare=None,
                gdr_net=None, actual_net=None,
            ),
        ]
        make_workbook(src, rows)

        result = rr.run_validation_and_extract(str(src), "testid01", original_name="r.xlsx", pdf_path=None)
        assert (tmp_path / "storage" / "testid01.xlsx").exists()

        row6 = next(r for r in result["rows"] if r["row"] == 6)
        gross_letter = openpyxl.utils.get_column_letter(KEY_TO_COL["gdr_gross"])
        assert row6["cells"].get(gross_letter) == 0
        assert result["highlights"].get(f"{gross_letter}6") == "#C6EFCE"


if __name__ == "__main__":
    import sys

    sys.exit(pytest.main([__file__, "-v"]))
