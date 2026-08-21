"""Tests for the offline/desktop Report Validator path.

The desktop build must validate reports with no HTTP server and no sockets of
any kind: the UI talks to Python through the in-process WebViewBridge, passing
JSON strings only. These tests drive that exact path.
"""

import datetime
import json
import time

import openpyxl
import pytest

from audit_engine.services import report_validator as rv
from audit_engine.web import report_handlers as rh
from tests.test_report_validator import KEY_TO_COL, SHEET, good_row, make_workbook  # noqa: F401


def _wait_idle(timeout=15.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if not rh.report_tracker.is_running:
            return True
        time.sleep(0.05)
    return False


@pytest.fixture(autouse=True)
def _reset_tracker():
    rh.report_tracker.__init__()
    yield
    rh.report_tracker.__init__()


class TestValidateWorkbook:
    """The shared core both the desktop app and web server call."""

    def test_writes_output_beside_source_by_default(self, tmp_path):
        src = tmp_path / "MyBranch.xlsx"
        make_workbook(src, [good_row("P01"), good_row("P02")])

        result = rv.validate_workbook(src)

        assert result["output_path"] == str(tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx")
        assert (tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx").exists()
        assert result["total_issues"] == 0
        assert result["rows_scanned"] == 2
        assert result["pdf_applied"] is False

    def test_explicit_output_path_is_honoured(self, tmp_path):
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])
        out = tmp_path / "nested" / "custom.xlsx"

        result = rv.validate_workbook(src, output_path=out)

        assert result["output_path"] == str(out)
        assert out.exists()

    def test_progress_callback_reaches_completion(self, tmp_path):
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])
        seen = []

        rv.validate_workbook(src, on_progress=lambda pct, msg=None: seen.append(pct))

        assert seen and seen[0] <= 15 and seen[-1] == 100
        assert seen == sorted(seen), "progress must be monotonic"

    def test_no_preview_payload_unless_requested(self, tmp_path):
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])

        assert "rows" not in rv.validate_workbook(src)
        assert "rows" in rv.validate_workbook(src, build_preview=True)

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            rv.validate_workbook(tmp_path / "nope.xlsx")

    def test_wrong_sheet_raises_keyerror_naming_sheets(self, tmp_path):
        src = tmp_path / "wrong.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Some Other Sheet"
        wb.save(src)
        wb.close()

        with pytest.raises(KeyError, match="Some Other Sheet"):
            rv.validate_workbook(src)


class TestDesktopHandlers:
    def test_run_rejects_missing_and_bad_input(self, tmp_path):
        assert rh.handle_report_run({})["success"] is False
        assert rh.handle_report_run({"filepath": str(tmp_path / "ghost.xlsx")})["success"] is False

        xls = tmp_path / "legacy.xls"
        xls.write_text("not really excel")
        assert "xlsx" in rh.handle_report_run({"filepath": str(xls)})["error"].lower()

    def test_full_run_and_progress_cycle(self, tmp_path):
        src = tmp_path / "MyBranch.xlsx"
        rows = [good_row(f"P{i:02d}") for i in range(1, 6)]
        rows[1]["tare"] = 15.0  # tare < GDR gross
        make_workbook(src, rows)

        assert rh.handle_report_run({"filepath": str(src)})["success"] is True
        assert _wait_idle(), "validation did not finish"

        p = rh.handle_report_progress()
        assert p["error"] is None
        assert p["pct"] == 100
        assert p["is_running"] is False
        assert p["summary"]["summary"]["tare_less_than_weight"] == 2
        assert (tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx").exists()
        assert any(log["level"] == "OK" for log in p["logs"])

    def test_missing_sheet_surfaces_as_error_not_crash(self, tmp_path):
        src = tmp_path / "wrong.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Nope"
        wb.save(src)
        wb.close()

        rh.handle_report_run({"filepath": str(src)})
        assert _wait_idle()

        p = rh.handle_report_progress()
        assert p["error"] and "Nope" in p["error"]
        assert p["is_running"] is False, "tracker must not stay stuck running after a failure"

    def test_second_run_refused_while_one_is_active(self, tmp_path):
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])
        rh.report_tracker.is_running = True
        try:
            assert rh.handle_report_run({"filepath": str(src)})["success"] is False
        finally:
            rh.report_tracker.is_running = False

    def test_stale_pdf_path_is_ignored_not_fatal(self, tmp_path):
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])

        assert rh.handle_report_run({
            "filepath": str(src), "pdf_path": str(tmp_path / "gone.pdf"),
        })["success"] is True
        assert _wait_idle()
        p = rh.handle_report_progress()
        assert p["error"] is None
        assert p["summary"]["pdf_applied"] is False

    def test_open_rejects_missing_path(self):
        assert rh.handle_report_open({})["success"] is False
        assert rh.handle_report_open({"path": "/definitely/not/here.xlsx"})["success"] is False


class TestZeroSocketBridge:
    """Drive the endpoints exactly as the desktop JS does — no socket involved."""

    def test_end_to_end_over_bridge(self, tmp_path):
        from audit_engine.app import create_app
        from audit_engine.web.bridge import WebViewBridge

        create_app()
        bridge = WebViewBridge()

        src = tmp_path / "MyBranch.xlsx"
        rows = [good_row(f"P{i:02d}") for i in range(1, 5)]
        rows[2]["status"] = "CLOSED"
        make_workbook(src, rows)

        started = json.loads(bridge.fetch_proxy(
            "POST", "/api/report/run", json.dumps({"filepath": str(src), "pdf_path": ""})
        ))
        assert started["success"] is True

        deadline = time.time() + 15
        payload = None
        while time.time() < deadline:
            payload = json.loads(bridge.fetch_proxy("GET", "/api/report/progress", ""))
            if not payload["is_running"]:
                break
            time.sleep(0.05)

        assert payload is not None and payload["is_running"] is False
        assert payload["error"] is None
        assert payload["summary"]["total_issues"] > 0
        assert (tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx").exists()

    def test_run_response_is_json_serializable(self, tmp_path):
        """Everything crossing the bridge must survive json.dumps (it is a string channel)."""
        src = tmp_path / "in.xlsx"
        rows = [good_row("P01")]
        rows[0]["sanction_date"] = __import__("datetime").date(2026, 5, 1)
        make_workbook(src, rows)

        rh.handle_report_run({"filepath": str(src)})
        assert _wait_idle()
        json.dumps(rh.handle_report_progress())  # must not raise


class TestPdfIsTrulyOptional:
    """A missing/unreadable PDF must never cost the user their validated file.

    The packaged Windows build shipped without pypdf bundled, and because the
    import sat outside the try block a ModuleNotFoundError aborted the whole
    run — the user got no output at all for an optional convenience feature.
    """

    def test_missing_pypdf_degrades_instead_of_failing(self, tmp_path, monkeypatch):
        import builtins

        real_import = builtins.__import__

        def no_pypdf(name, *args, **kwargs):
            if name == "pypdf" or name.startswith("pypdf."):
                raise ImportError("No module named 'pypdf'")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", no_pypdf)

        pdf = tmp_path / "seq.pdf"
        pdf.write_bytes(b"%PDF-1.4 not really a pdf")
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01"), good_row("P02")])

        # Must not raise, and must still produce the validated workbook.
        result = rv.validate_workbook(src, pdf_path=str(pdf))

        assert result["pdf_applied"] is False
        assert result["pdf_warning"]
        assert (tmp_path / "TESTBR_Audit-MIS_UNKNOWN_DATE.xlsx").exists()

    def test_extract_accounts_never_raises(self, tmp_path):
        junk = tmp_path / "junk.pdf"
        junk.write_bytes(b"definitely not a pdf")
        assert rv.extract_accounts_from_pdf(str(junk)) == []
        assert rv.extract_accounts_from_pdf(str(tmp_path / "nonexistent.pdf")) == []

    def test_pdf_applied_reflects_reality_not_just_presence(self, tmp_path):
        """A supplied-but-unusable PDF must report pdf_applied=False."""
        pdf = tmp_path / "empty.pdf"
        pdf.write_bytes(b"not a pdf")
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])

        result = rv.validate_workbook(src, pdf_path=str(pdf))
        assert result["pdf_applied"] is False, "must not claim the PDF was applied"
        assert result["pdf_warning"]

    def test_desktop_worker_surfaces_the_warning(self, tmp_path):
        pdf = tmp_path / "bad.pdf"
        pdf.write_bytes(b"not a pdf")
        src = tmp_path / "in.xlsx"
        make_workbook(src, [good_row("P01")])

        rh.handle_report_run({"filepath": str(src), "pdf_path": str(pdf)})
        assert _wait_idle()

        p = rh.handle_report_progress()
        assert p["error"] is None, "an unusable PDF must not fail the run"
        assert any(log["level"] == "WARN" for log in p["logs"]), "user must be told the PDF was ignored"


def _make_pdf(path, lines):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=A4)
    y = 750
    for line in lines:
        c.drawString(60, y, line)
        y -= 30
    c.save()


class TestPdfResequencingIsRobust:
    """PDF row resequencing must work for real reports, not just one format.

    It previously looked for exactly 15 digits and compared raw cell strings,
    so any other account length — or an account stored as a number by Excel —
    matched nothing and the reorder silently did not happen.
    """

    @pytest.mark.parametrize(
        "accounts,render,label",
        [
            (["999888777666501", "999888777666502", "999888777666503"],
             lambda a: f"Account: {a}", "15-digit"),
            (["100200300401", "100200300402", "100200300403"],
             lambda a: f"A/c {a}", "12-digit"),
            (["1002003004010011", "1002003004010022", "1002003004010033"],
             lambda a: f"AcNo {a}", "16-digit"),
            (["999888777666501", "999888777666502", "999888777666503"],
             lambda a: f"Acct {a[:4]} {a[4:8]} {a[8:]}", "spaced"),
            (["100200300401", "100200300402", "100200300403"],
             lambda a: f"{a[:4]}-{a[4:8]}-{a[8:]}", "hyphenated"),
            (["100200300401", "100200300402", "100200300403"],
             lambda a: f"Ref#77{a}X  Loan", "embedded in text"),
            (["100200300401", "100200300402", "100200300403"],
             lambda a: f"row | {a} | NAME | 12.5", "table row"),
        ],
    )
    def test_reorders_regardless_of_format(self, tmp_path, accounts, render, label):
        rows = []
        for i, acct in enumerate(accounts, 1):
            row = good_row(f"P{i:02d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        pdf = tmp_path / "seq.pdf"
        _make_pdf(pdf, [render(a) for a in reversed(accounts)])

        result = rv.validate_workbook(src, pdf_path=str(pdf))

        assert result["pdf_applied"] is True, f"{label}: reorder did not run"
        assert result["pdf_matched_rows"] == len(accounts), f"{label}: partial match"
        assert result["pdf_warning"] is None

        wb = openpyxl.load_workbook(result["output_path"])
        ws = wb[SHEET]
        got = [str(ws.cell(row=r, column=KEY_TO_COL["account"]).value)
               for r in range(5, 5 + len(accounts))]
        assert got == list(reversed(accounts)), f"{label}: wrong order"

    def test_account_stored_as_number_by_excel(self, tmp_path):
        """openpyxl returns ints/floats for numeric cells; matching must survive that."""
        accounts = [100200300401, 100200300402, 100200300403]
        rows = []
        for i, acct in enumerate(accounts, 1):
            row = good_row(f"P{i:02d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)
        pdf = tmp_path / "seq.pdf"
        _make_pdf(pdf, [f"A/c {a}" for a in reversed(accounts)])

        result = rv.validate_workbook(src, pdf_path=str(pdf))
        assert result["pdf_matched_rows"] == 3

    def test_separate_lines_do_not_fuse_into_one_number(self, tmp_path):
        """A newline between two accounts must not merge them into one blob."""
        text = "100200300401\n100200300402\n100200300403"
        assert rv.normalize_account("1002 0030 0401") == "100200300401"
        pdf = tmp_path / "p.pdf"
        _make_pdf(pdf, text.split("\n"))
        found = rv.extract_accounts_from_pdf(
            str(pdf), known_accounts=["100200300401", "100200300402", "100200300403"]
        )
        assert found == ["100200300401", "100200300402", "100200300403"]

    def test_partial_match_is_reported_not_silent(self, tmp_path):
        accounts = ["100200300401", "100200300402", "100200300403"]
        rows = []
        for i, acct in enumerate(accounts, 1):
            row = good_row(f"P{i:02d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)
        pdf = tmp_path / "seq.pdf"
        _make_pdf(pdf, ["A/c 100200300403"])  # only one of three

        result = rv.validate_workbook(src, pdf_path=str(pdf))
        assert result["pdf_applied"] is True
        assert result["pdf_matched_rows"] == 1
        assert "1 of 3" in result["pdf_warning"]


class TestNormalizeAccount:
    @pytest.mark.parametrize(
        "raw,expected",
        [
            ("123456789012345", "123456789012345"),
            (123456789012345, "123456789012345"),
            ("123456789012345.0", "123456789012345"),
            ("1234 5678 9012 345", "123456789012345"),
            ("1234-5678-9012-345", "123456789012345"),
            ("  123456789012345  ", "123456789012345"),
            ("", ""),
            (None, ""),
        ],
    )
    def test_forms_that_must_compare_equal(self, raw, expected):
        assert rv.normalize_account(raw) == expected


class TestFormulasFollowTheirRow:
    """A moved row's formulas must compute from the row it now occupies.

    Reordering used to copy formula text verbatim, so a row that moved to 46
    still carried "=I26-J26" — the sheet showed another loan's arithmetic, and
    the cells read blank once the referenced row changed.
    """

    @pytest.mark.parametrize(
        "formula,old,new,expected",
        [
            ("=I26-J26", 26, 46, "=I46-J46"),
            ("=ROUND((Q26/I26)*100,2)", 26, 46, "=ROUND((Q46/I46)*100,2)"),
            ("=I$26-J26", 26, 46, "=I$26-J46"),      # absolute row pinned
            ("=SUM(I2:I5)+I26", 26, 46, "=SUM(I2:I5)+I46"),  # other rows untouched
            ("=0.000", 26, 46, "=0.000"),
            ("=I26-J26", 26, 26, "=I26-J26"),        # no move, no change
        ],
    )
    def test_retarget_rules(self, formula, old, new, expected):
        assert rv.retarget_formula(formula, old, new) == expected

    def test_non_formula_values_are_untouched(self):
        for value in (0.5, None, "TEXT", 12, "=notaref"):
            assert rv.retarget_formula(value, 1, 9) == value

    def test_formulas_are_repointed_when_rows_move(self, tmp_path):
        from openpyxl.utils import get_column_letter

        accounts = ["100200300401", "100200300402", "100200300403"]
        rows = []
        for i, acct in enumerate(accounts, 1):
            row = good_row(f"P{i:02d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        gd = KEY_TO_COL["gross_diff"]
        gg = get_column_letter(KEY_TO_COL["gdr_gross"])
        ag = get_column_letter(KEY_TO_COL["actual_gross"])
        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        for r in (5, 6, 7):
            ws.cell(row=r, column=gd).value = f"={gg}{r}-{ag}{r}"
        wb.save(src)
        wb.close()

        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        col_map = rv.build_col_map(rv.map_columns(ws))
        rv.rearrange_rows_by_pdf(ws, col_map, [5, 6, 7], list(reversed(accounts)))

        for r in (5, 6, 7):
            assert ws.cell(row=r, column=gd).value == f"={gg}{r}-{ag}{r}", (
                f"row {r} formula points at the wrong row"
            )

    def test_saved_file_asks_excel_to_recalculate(self, tmp_path):
        """openpyxl drops cached results, so formula cells open blank without this."""
        import re as _re
        import zipfile

        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", gross_diff=9.9)])
        result = rv.validate_workbook(src)

        with zipfile.ZipFile(result["output_path"]) as z:
            workbook_xml = z.read("xl/workbook.xml").decode()
        calc = _re.search(r"<calcPr[^>]*/>", workbook_xml)
        assert calc and 'fullCalcOnLoad="1"' in calc.group(0)


class TestSequenceGapsArePreserved:
    """The PDF is the master order: a missing entry leaves an empty slot."""

    def _run(self, tmp_path, workbook_accounts, pdf_accounts):
        rows = []
        for i, acct in enumerate(workbook_accounts, 1):
            row = good_row(f"P{i:03d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)
        pdf = tmp_path / "seq.pdf"
        _make_pdf(pdf, [f"A/c {a}" for a in pdf_accounts])
        result = rv.validate_workbook(src, pdf_path=str(pdf))
        wb = openpyxl.load_workbook(result["output_path"])
        return result, wb[SHEET]

    def test_missing_entries_leave_blank_rows_in_position(self, tmp_path):
        pdf_accounts = [f"9002003004{i:05d}" for i in range(1, 21)]
        missing = set(pdf_accounts[10:15])                       # entries 11-15
        workbook = [a for a in pdf_accounts if a not in missing]

        result, ws = self._run(tmp_path, workbook, pdf_accounts)

        assert result["pdf_gap_rows"] == 5
        assert result["pdf_matched_rows"] == len(workbook)

        acct_col = KEY_TO_COL["account"]
        laid_out = [ws.cell(row=5 + i, column=acct_col).value for i in range(len(pdf_accounts))]
        for i, expected in enumerate(pdf_accounts):
            if expected in missing:
                assert laid_out[i] is None, f"sequence {i+1} should be blank"
            else:
                assert str(laid_out[i]) == expected, f"sequence {i+1} out of position"

    def test_rows_after_a_gap_keep_their_pdf_position(self, tmp_path):
        """The whole point: later rows must not shift up by the missing count."""
        pdf_accounts = [f"9002003004{i:05d}" for i in range(1, 11)]
        workbook = [a for i, a in enumerate(pdf_accounts) if i != 3]

        _result, ws = self._run(tmp_path, workbook, pdf_accounts)

        acct_col = KEY_TO_COL["account"]
        # sequence 5 (index 4) must still land on its own row, not move up
        assert str(ws.cell(row=5 + 4, column=acct_col).value) == pdf_accounts[4]

    def test_gap_row_is_empty_but_keeps_borders(self, tmp_path):
        from copy import copy as _copy

        pdf_accounts = ["100200300401", "100200300499", "100200300402"]
        workbook = ["100200300401", "100200300402"]

        rows = []
        for i, acct in enumerate(workbook, 1):
            row = good_row(f"P{i:02d}")
            row["account"] = acct
            rows.append(row)
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)
        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        for r in (5, 6):
            for c in range(1, 30):
                ws.cell(row=r, column=c).border = _copy(rv.THIN_BORDER)
        wb.save(src)
        wb.close()

        pdf = tmp_path / "seq.pdf"
        _make_pdf(pdf, [f"A/c {a}" for a in pdf_accounts])
        result = rv.validate_workbook(src, pdf_path=str(pdf))

        wb = openpyxl.load_workbook(result["output_path"])
        ws = wb[SHEET]
        gap = ws.cell(row=6, column=KEY_TO_COL["account"])
        assert gap.value is None
        assert gap.border.left.style, "gap row lost the table border"

    def test_rows_absent_from_the_pdf_go_after_the_sequence(self, tmp_path):
        pdf_accounts = ["100200300401", "100200300402"]
        workbook = ["100200300401", "100200300402", "100200300477"]

        result, ws = self._run(tmp_path, workbook, pdf_accounts)

        acct_col = KEY_TO_COL["account"]
        laid_out = [str(ws.cell(row=5 + i, column=acct_col).value) for i in range(3)]
        assert laid_out == ["100200300401", "100200300402", "100200300477"]
        assert result["pdf_gap_rows"] == 0

    def test_blank_rows_are_not_validated_or_flagged(self, tmp_path):
        """A gap row holds no data and must not raise findings of its own."""
        pdf_accounts = [f"9002003004{i:05d}" for i in range(1, 6)]
        workbook = [a for i, a in enumerate(pdf_accounts) if i != 2]

        result, _ws = self._run(tmp_path, workbook, pdf_accounts)

        assert result["pdf_gap_rows"] == 1
        # a blank row would otherwise trip magnet/tampered/ornament checks
        assert result["summary"].get("magnet_not_ok", 0) == 0
        assert result["summary"].get("ornaments_gdr_non_positive", 0) == 0


class TestTopUpKeepsItsDate:
    def test_renewal_date_survives_on_a_top_up_row(self, tmp_path):
        rows = [
            good_row("P01"),
            good_row("P02", remarks="Top Up of account 123", renewal_date="01-05-2026"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert ws.cell(row=6, column=KEY_TO_COL["renewal_date"]).value == datetime.datetime(2026, 5, 1)

    def test_resolved_top_up_keeps_its_date_too(self, tmp_path):
        """Pass 2 resolves an empty packet by name; it must not wipe the date."""
        first = good_row("P01", applicant="SHARED NAME")
        second = good_row("P02", applicant="SHARED NAME", renewal_date="01-05-2026")
        second["packet"] = None
        src = tmp_path / "B.xlsx"
        make_workbook(src, [first, second])

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert ws.cell(row=6, column=KEY_TO_COL["renewal_date"]).value == datetime.datetime(2026, 5, 1)


def _styled_workbook(path, rows, size=10, horizontal="center"):
    """Write a workbook whose data cells carry the document's own styling."""
    from copy import copy as _copy

    from openpyxl.styles import Alignment, Font

    make_workbook(path, rows)
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    font = Font(name="Calibri", size=size)
    align = Alignment(horizontal=horizontal, vertical="center")
    for r in range(5, 5 + len(rows)):
        for c in range(1, len(KEY_TO_COL) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _copy(font)
            cell.alignment = _copy(align)
    wb.save(path)
    wb.close()


class TestWrittenCellsMatchTheDocument:
    """Values we write must adopt the sheet's own font/alignment.

    A value written into a previously empty cell otherwise lands with Excel's
    defaults (left-aligned, Calibri 11) — which is how a resolved packet
    number came out neither centred nor at the document's font size.
    """

    def test_resolved_packet_adopts_column_font_and_alignment(self, tmp_path):
        first = good_row("P01", applicant="SHARED NAME")
        second = good_row("P02", applicant="SHARED NAME")
        second["packet"] = None
        src = tmp_path / "B.xlsx"
        _styled_workbook(src, [first, second])

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        packet = ws.cell(row=6, column=KEY_TO_COL["packet"])
        assert packet.value == "P01", "packet was not resolved"
        assert packet.alignment.horizontal == "center"
        assert packet.font.size == 10
        assert packet.font.name == "Calibri"

    def test_existing_cells_are_never_restyled(self, tmp_path):
        """We only fill blanks — populated cells keep exactly what they had."""
        src = tmp_path / "B.xlsx"
        _styled_workbook(src, [good_row("P01"), good_row("P02")], size=10, horizontal="right")

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        for key in ("applicant", "account", "status", "gdr_gross"):
            cell = ws.cell(row=5, column=KEY_TO_COL[key])
            assert cell.alignment.horizontal == "right", key
            assert cell.font.size == 10, key
            assert cell.font.name == "Calibri", key


class TestTopUpKeepsExpectedReadings:
    def test_magnet_and_tampered_carry_their_defaults(self, tmp_path):
        rows = [
            good_row("P01"),
            good_row("P02", remarks="Top Up of account 999", magnet="NOT OK", tampered="YES"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        assert ws.cell(row=6, column=KEY_TO_COL["magnet"]).value == "OK"
        assert ws.cell(row=6, column=KEY_TO_COL["tampered"]).value == "NO"
        # a top-up has no GDR of its own
        assert ws.cell(row=6, column=KEY_TO_COL["gdr_no"]).value is None


class TestDateHandling:
    def test_all_dates_become_real_dates_shown_dd_mm_yyyy(self, tmp_path):
        rows = [
            good_row("P01", verification_date="05/05/2026", sanction_date="2026-01-31"),
            good_row("P02", verification_date="6-5-2026", sanction_date="01.02.2026"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        for row, key, expected in [
            (5, "verification_date", datetime.datetime(2026, 5, 5)),
            (6, "verification_date", datetime.datetime(2026, 5, 6)),
        ]:
            cell = ws.cell(row=row, column=KEY_TO_COL[key])
            assert cell.value == expected, f"{key} on row {row}"
            assert cell.number_format == rv.DATE_NUMBER_FORMAT, f"{key} on row {row}"

    def test_sanction_date_is_left_exactly_as_the_source_has_it(self, tmp_path):
        """SANCTION DATE comes from the bank's record and must not be touched."""
        rows = [
            good_row("P01", sanction_date="2026-01-31"),
            good_row("P02", sanction_date="01.02.2026"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        assert ws.cell(row=5, column=KEY_TO_COL["sanction_date"]).value == "2026-01-31"
        assert ws.cell(row=6, column=KEY_TO_COL["sanction_date"]).value == "01.02.2026"

    def test_poa_rows_have_no_renewal_closed_date(self, tmp_path):
        rows = [good_row("P01", taf="POA", renewal_date="21-05-2026")]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert ws.cell(row=5, column=KEY_TO_COL["renewal_date"]).value is None

    def test_taf_rows_keep_their_renewal_date(self, tmp_path):
        rows = [good_row("P01", taf="TAF", renewal_date="21-05-2026")]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert ws.cell(row=5, column=KEY_TO_COL["renewal_date"]).value == datetime.datetime(2026, 5, 21)


class TestOutputFilename:
    def test_uses_branch_and_verification_date_range(self, tmp_path):
        rows = [
            good_row("P01", verification_date="05/05/2026"),
            good_row("P02", verification_date="10/05/2026"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        assert result["custom_filename"] == "TESTBR_Audit-MIS_05-MAY-2026_to_10-MAY-2026.xlsx"

    def test_single_date_is_not_rendered_as_a_range(self, tmp_path):
        rows = [good_row("P01", verification_date="05/05/2026")]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        assert result["custom_filename"] == "TESTBR_Audit-MIS_05-MAY-2026.xlsx"

    def test_falls_back_to_other_date_columns(self, tmp_path):
        """No verification date must not mean UNKNOWN_DATE when other dates exist."""
        row = good_row("P01", taf="TAF", sanction_date="03/03/2026")
        row["verification_date"] = None
        src = tmp_path / "B.xlsx"
        make_workbook(src, [row])

        result = rv.validate_workbook(src)
        assert "UNKNOWN_DATE" not in result["custom_filename"]
        assert "03-MAR-2026" in result["custom_filename"]

    def test_branch_header_variants_are_recognised(self, tmp_path):
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", verification_date="05/05/2026")])
        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        ws.cell(row=2, column=KEY_TO_COL["branch"]).value = "BRANCH"   # not "BRANCH NAME"
        wb.save(src)
        wb.close()

        result = rv.validate_workbook(src)
        assert result["custom_filename"].startswith("TESTBR_")


class TestStateColumn:
    """An unresolved state needs a human to fill it in, so flag it."""

    @pytest.mark.parametrize("value", ["Unknown", "UNKNOWN", "unknown", "Unknown State"])
    def test_unknown_state_is_highlighted(self, tmp_path, value):
        rows = [good_row("P01"), good_row("P02", state=value)]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        flagged = ws.cell(row=6, column=KEY_TO_COL["state"])
        assert str(flagged.fill.start_color.rgb).endswith("FF8C00"), "state cell not highlighted"
        assert result["summary"].get("state_unknown") == 1

    def test_a_real_state_is_left_alone(self, tmp_path):
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", state="KARNATAKA")])

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        cell = ws.cell(row=5, column=KEY_TO_COL["state"])
        assert cell.value == "KARNATAKA"
        assert not str(cell.fill.start_color.rgb).endswith("FF8C00")
        assert "state_unknown" not in result["summary"]

    def test_flagged_even_on_a_top_up_row(self, tmp_path):
        """The branch's state applies regardless of the row's status."""
        rows = [
            good_row("P01"),
            good_row("P02", state="Unknown", remarks="Top Up of account 999"),
        ]
        src = tmp_path / "B.xlsx"
        make_workbook(src, rows)

        result = rv.validate_workbook(src)
        assert result["summary"].get("state_unknown") == 1

    def test_missing_state_column_is_not_an_error(self, tmp_path):
        """Sheets without a STATE column must still validate."""
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01")])
        wb = openpyxl.load_workbook(src)
        ws = wb[SHEET]
        ws.cell(row=2, column=KEY_TO_COL["state"]).value = None
        ws.cell(row=5, column=KEY_TO_COL["state"]).value = None
        wb.save(src)
        wb.close()

        result = rv.validate_workbook(src)
        assert "state_unknown" not in result["summary"]


class TestAutoUpdate:
    """A found update is fetched in the background; the swap waits for a restart.

    Replacing the binary under a running audit would lose work, so only the
    download is automatic.
    """

    @pytest.fixture
    def updater(self, monkeypatch):
        import sys as _sys

        from audit_engine.updater import client as c

        monkeypatch.setattr(_sys, "frozen", True, raising=False)
        st = c.update_state
        st.update_ready = True
        st.binary_url = "https://example.invalid/app.zip"
        st.latest_version = "v9.9.9"
        st.staged_version = ""
        st._is_downloading = False
        st._preflight_pass = True
        from audit_engine.database import legacy
        monkeypatch.setattr(legacy, "get_config", lambda k, d=None: "True")
        return c

    def test_downloads_when_an_update_is_available(self, updater):
        assert updater._should_auto_download() is True

    def test_never_downloads_in_dev_mode(self, updater, monkeypatch):
        import sys as _sys
        monkeypatch.delattr(_sys, "frozen", raising=False)
        assert updater._should_auto_download() is False

    def test_does_not_refetch_a_version_already_staged(self, updater):
        updater.update_state.staged_version = "v9.9.9"
        assert updater._should_auto_download() is False

    def test_fetches_a_newer_version_after_one_was_staged(self, updater):
        updater.update_state.staged_version = "v9.9.8"
        assert updater._should_auto_download() is True

    def test_skips_while_a_download_is_in_flight(self, updater):
        updater.update_state._is_downloading = True
        assert updater._should_auto_download() is False

    def test_skips_when_preflight_fails(self, updater):
        """No disk or no network right now — try again on the next check."""
        updater.update_state._preflight_pass = False
        assert updater._should_auto_download() is False

    def test_skips_when_there_is_no_binary_for_this_os(self, updater):
        updater.update_state.binary_url = ""
        assert updater._should_auto_download() is False

    @pytest.mark.parametrize("value", ["False", "false", "0", "no", "off"])
    def test_config_can_turn_auto_download_off(self, updater, monkeypatch, value):
        from audit_engine.database import legacy
        monkeypatch.setattr(legacy, "get_config", lambda k, d=None: value)
        assert updater._should_auto_download() is False

    def test_defaults_to_on_when_config_is_unreadable(self, updater, monkeypatch):
        from audit_engine.database import legacy

        def boom(*_a, **_k):
            raise RuntimeError("db unavailable")

        monkeypatch.setattr(legacy, "get_config", boom)
        assert updater._auto_download_enabled() is True

    def test_check_reports_whether_the_update_is_on_disk(self, monkeypatch):
        """The banner offers Restart Now, so it must know staged from merely-found."""
        from audit_engine.updater import client as c
        from audit_engine.web import handlers as h

        monkeypatch.setattr(h, "check_latest_release", lambda force=False: {"update_ready": True})
        c.update_state.latest_version = "v9.9.9"

        c.update_state.staged_version = ""
        assert h.handle_update_check()["staged"] is False

        c.update_state.staged_version = "v9.9.9"
        assert h.handle_update_check()["staged"] is True


class TestSmallPoaWeightDifferences:
    """A POA gross/net difference within ±0.200 is weighing noise.

    The FCU figure is aligned to the GDR figure and the row's discrepancy
    remark is reduced to whatever is still outstanding.
    """

    def _run(self, tmp_path, **over):
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", **over)])
        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        return result, ws

    def _cell(self, ws, key):
        return ws.cell(row=5, column=KEY_TO_COL[key])

    def test_gross_within_tolerance_is_aligned_and_remark_cleared(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.5
        assert self._cell(ws, "remarks").value == "Ok - No Discrepancy"

    def test_net_within_tolerance_is_aligned_and_remark_cleared(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.5,
            gdr_net=19.0, actual_net=18.9, remarks="Net Weight Difference",
        )
        assert self._cell(ws, "actual_net").value == 19.0
        assert self._cell(ws, "remarks").value == "Ok - No Discrepancy"

    @pytest.mark.parametrize("delta", [0.2, -0.2])
    def test_the_boundary_is_inclusive(self, tmp_path, delta):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=round(20.5 - delta, 3),
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.5

    def test_just_outside_the_tolerance_is_left_alone(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.299,
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.299
        assert self._cell(ws, "remarks").value == "Gross Weight Difference"

    def test_fixing_one_of_two_leaves_the_other_remark(self, tmp_path):
        """Gross is absorbed, net is a real 0.5 shortfall — say so."""
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=18.5,
            remarks="Gross Weight and Net Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.5
        assert self._cell(ws, "actual_net").value == 18.5, "a real difference must stand"
        assert self._cell(ws, "remarks").value == "Net Weight Difference"

    def test_fixing_both_clears_the_combined_remark(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=18.85,
            remarks="Gross Weight and Net Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.5
        assert self._cell(ws, "actual_net").value == 19.0
        assert self._cell(ws, "remarks").value == "Ok - No Discrepancy"

    def test_does_not_apply_to_taf_rows(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="TAF", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert self._cell(ws, "actual_gross").value == 20.35
        assert self._cell(ws, "remarks").value == "Gross Weight Difference"

    def test_an_unrelated_remark_is_not_rewritten(self, tmp_path):
        _r, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=19.0, remarks="Customer not available",
        )
        assert self._cell(ws, "actual_gross").value == 20.5, "the weight is still corrected"
        assert self._cell(ws, "remarks").value == "Customer not available"

    def test_alignment_is_reported(self, tmp_path):
        result, _ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert result["summary"].get("weight_aligned_to_gdr") == 1
        assert result["summary"].get("discrepancy_remark_amended") == 1


class TestAmendDiscrepancyRemark:
    @pytest.mark.parametrize(
        "remark,fixed_gross,fixed_net,expected",
        [
            ("Gross Weight and Net Weight Difference", True, False, "Net Weight Difference"),
            ("Gross Weight and Net Weight Difference", False, True, "Gross Weight Difference"),
            ("Gross Weight and Net Weight Difference", True, True, "Ok - No Discrepancy"),
            ("Gross Weight Difference", True, False, "Ok - No Discrepancy"),
            ("Net Weight Difference", False, True, "Ok - No Discrepancy"),
            # fixing the other column leaves this one's remark as it was
            ("Gross Weight Difference", False, True, "Gross Weight Difference"),
            ("Net Weight Difference", True, False, "Net Weight Difference"),
            # spacing/case in the source must not matter
            ("gross weight  and net weight difference", True, False, "Net Weight Difference"),
            # anything else is left to the humans
            ("Customer not available", True, True, None),
            ("", True, True, None),
        ],
    )
    def test_rules(self, remark, fixed_gross, fixed_net, expected):
        assert rv.amend_discrepancy_remark(remark, fixed_gross, fixed_net) == expected


class TestTafRowsSkipTheDifferenceColumns:
    """A weight difference does not apply to a TAF row.

    The gross/net difference columns are left exactly as the source has them
    — no formula written, no value corrected, nothing flagged.
    """

    def _run(self, tmp_path, **over):
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", **over)])
        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        return result, ws

    def test_wrong_looking_diffs_are_left_alone(self, tmp_path):
        result, ws = self._run(
            tmp_path, taf="TAF", gdr_gross=20.5, actual_gross=20.0, tare=21.0,
            gdr_net=19.0, actual_net=18.0, gross_diff=99.9, net_diff=88.8,
        )
        assert ws.cell(row=5, column=KEY_TO_COL["gross_diff"]).value == 99.9
        assert ws.cell(row=5, column=KEY_TO_COL["net_diff"]).value == 88.8
        assert "gross_diff_fixed" not in result["summary"]
        assert "net_diff_fixed" not in result["summary"]
        assert "net_diff_av_not_zero" not in result["summary"]

    def test_empty_diffs_stay_empty(self, tmp_path):
        _result, ws = self._run(tmp_path, taf="TAF", gross_diff=None, net_diff=None)
        assert ws.cell(row=5, column=KEY_TO_COL["gross_diff"]).value is None
        assert ws.cell(row=5, column=KEY_TO_COL["net_diff"]).value is None

    def test_poa_rows_are_still_corrected(self, tmp_path):
        """The skip must be TAF-only, not a blanket disable."""
        result, ws = self._run(
            tmp_path, taf="POA", gdr_gross=20.5, actual_gross=19.0,
            gdr_net=19.0, actual_net=19.0, gross_diff=99.9, net_diff=0.0,
        )
        value = ws.cell(row=5, column=KEY_TO_COL["gross_diff"]).value
        assert isinstance(value, str) and value.startswith("=")
        assert result["summary"].get("gross_diff_fixed") == 1

    def test_taf_row_is_not_weight_aligned_either(self, tmp_path):
        """The ±0.200 absorption is POA-only and must stay that way."""
        result, ws = self._run(
            tmp_path, taf="TAF", gdr_gross=20.5, actual_gross=20.35,
            gdr_net=19.0, actual_net=19.0, remarks="Gross Weight Difference",
        )
        assert ws.cell(row=5, column=KEY_TO_COL["actual_gross"]).value == 20.35
        assert "weight_aligned_to_gdr" not in result["summary"]


# A realistic packet column: three series mixed together, roughly sequential
# with gaps, exactly as a branch's sheet looks.
REAL_PACKETS = (
    [f"S824{n}" for n in range(5589, 5645)]
    + [f"M12708{n:02d}" for n in range(21, 57)]
    + [f"XS249{n}" for n in range(4311, 4335)]
)


# Real packet numbers from one branch's report, all correct. Three series
# (S/XS/M), each 7 digits, sequential with gaps.
REAL_PACKETS = [
    "S8245589", "XS2494311", "S8245590", "S8245591", "S8245596", "S8245598",
    "M1270821", "S8245610", "S8245612", "S8245613", "S8245616", "S8245615",
    "S8245617", "S8245621", "S8245622", "S8245625", "S8245629", "S8245630",
    "S8245633", "S8245634", "S8245636", "S8245638", "S8245639", "S8245641",
    "S8245640", "M1270827", "M1270829", "M1270830", "M1270832", "XS2494318",
    "XS2494320", "M1270835", "S8245645", "XS2494321", "XS2494322", "S8245649",
    "S8245650", "M1270849", "M1270838", "S8245652", "S8245654", "S8245657",
    "M1270841", "S8245659", "M1270842", "M1270843", "S8245661", "S8245663",
    "S8245664", "S8245665", "M1270845", "S8245666", "S8245668", "S8245672",
    "S8245674", "M1270848", "S8245675", "S8245676", "S8245677", "S8245678",
    "S8245679", "S8245680", "S8245682", "S8245684", "M1270850", "S8245685",
    "S8245686", "S8245687", "XS2494323", "M1270851", "XS2494324", "M1270852",
    "XS2494325", "S8245689", "M1270854", "S8245691", "S8245694", "S8245693",
    "S8245692", "XS2494326", "S8245695", "S8245696", "S8245697", "S8245698",
    "M1270855", "XS2494328", "XS2494329", "S8245699", "S8245700", "S8245701",
    "S8245702", "S8245703", "S8245705", "S8245706", "XS2494330", "M1270856",
    "XS2494331", "S8245707", "XS2494333", "XS2494332", "XS2494334",
]


class TestPacketSeriesCheck:
    """Only unambiguous typos are reported.

    The check is deliberately narrow: flagging a correct packet costs an
    auditor more than missing a wrong one, because they stop trusting the
    colours entirely.
    """

    def _families(self, values):
        families = rv.learn_packet_families(values)
        return families, rv._established_digit_counts(families)

    def test_series_are_discovered_from_the_column(self):
        families, _ = self._families(REAL_PACKETS)
        assert {(p, n) for p, n in families} == {("S", 7), ("XS", 7), ("M", 7)}

    def test_no_correct_value_is_flagged(self):
        """The whole file is valid — nothing may light up."""
        families, established = self._families(REAL_PACKETS)
        flagged = [
            (v, rv.check_packet(v, families, established))
            for v in REAL_PACKETS
            if rv.check_packet(v, families, established)[0]
        ]
        assert flagged == []

    @pytest.mark.parametrize(
        "typo,why",
        [
            ("S82456211", "one digit too many"),
            ("S824562", "one digit too few"),
            ("S8245O29", "letter O typed for a zero"),
            ("8245629", "prefix missing entirely"),
        ],
    )
    def test_real_typos_are_caught(self, typo, why):
        families, established = self._families([*REAL_PACKETS, typo])
        severity, _explanation = rv.check_packet(typo, families, established)
        assert severity == "error", why

    @pytest.mark.parametrize(
        "value,why",
        [
            ("L1270999", "a prefix this branch simply has not used before"),
            ("S8250001", "series crossed a rounding boundary - still valid"),
            ("S9245589", "a different leading digit is not proof of a typo"),
        ],
    )
    def test_values_that_only_look_unusual_are_left_alone(self, value, why):
        """These were flagged before and were all correct."""
        families, established = self._families([*REAL_PACKETS, value])
        severity, _explanation = rv.check_packet(value, families, established)
        assert severity == "", why

    def test_a_lone_packet_of_a_new_series_is_not_an_error(self):
        families, established = self._families([*REAL_PACKETS, "L1270999"])
        assert rv.check_packet("L1270999", families, established)[0] == ""

    @pytest.mark.parametrize(
        "raw,expect",
        [
            ("s8245589", "S8245589"),      # lower-case prefix
            ("xs2494311", "XS2494311"),
            ("S 8245589", "S8245589"),     # space inside
            ("S-8245589", "S8245589"),     # stray separator
        ],
    )
    def test_case_and_spacing_are_tidied_not_flagged(self, raw, expect):
        cleaned, changed = rv.normalize_packet(raw)
        assert cleaned == expect
        assert changed, "the value differs from the source, so it counts as tidied"

    def test_surrounding_spaces_need_no_tidying(self):
        """They are stripped before the packet ever reaches this check."""
        cleaned, changed = rv.normalize_packet("  S8245589  ")
        assert cleaned == "S8245589"
        assert changed is False


class TestPacketChecksInAFullRun:
    def _rows(self, packets):
        rows = []
        for i, packet in enumerate(packets, 1):
            row = good_row(f"P{i:03d}")
            row["packet"] = packet
            row["account"] = f"90020030040{i:04d}"
            rows.append(row)
        return rows

    def test_a_correct_column_raises_nothing(self, tmp_path):
        src = tmp_path / "B.xlsx"
        make_workbook(src, self._rows(REAL_PACKETS))
        result = rv.validate_workbook(src)
        assert "packet_pattern_break" not in result["summary"]

    def test_a_digit_typo_is_reported(self, tmp_path):
        packets = list(REAL_PACKETS)
        packets[5] = "S82455981"          # one digit too many
        src = tmp_path / "B.xlsx"
        make_workbook(src, self._rows(packets))

        result = rv.validate_workbook(src)
        assert result["summary"].get("packet_pattern_break") == 1
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert str(ws.cell(row=10, column=KEY_TO_COL["packet"]).fill.start_color.rgb).endswith("FF8C00")

    def test_case_and_spaces_are_tidied_in_place(self, tmp_path):
        packets = list(REAL_PACKETS)
        packets[0] = "s8245589"
        packets[1] = "XS 2494311"
        src = tmp_path / "B.xlsx"
        make_workbook(src, self._rows(packets))

        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]
        assert ws.cell(row=5, column=KEY_TO_COL["packet"]).value == "S8245589"
        assert ws.cell(row=6, column=KEY_TO_COL["packet"]).value == "XS2494311"
        assert result["summary"].get("packet_tidied") == 2

    def test_a_short_column_is_left_alone(self, tmp_path):
        """Too few packets to infer anything — stay out of the way."""
        src = tmp_path / "B.xlsx"
        make_workbook(src, self._rows(["S8245589", "S824559", "XS2494311"]))
        result = rv.validate_workbook(src)
        assert "packet_pattern_break" not in result["summary"]


class TestDateSeparatorIsALiteralSlash:
    """dd/mm/yyyy must survive a machine whose locale uses another separator.

    An unescaped "/" in an Excel number format is the *locale* date separator,
    not a slash — so on a machine set to use "-" the cell reads 05-05-2026.
    """

    def test_the_format_escapes_its_slashes(self):
        assert rv.DATE_NUMBER_FORMAT == "DD\\/MM\\/YYYY"
        assert "\\/" in rv.DATE_NUMBER_FORMAT, "an unescaped / follows the machine's locale"

    def test_the_saved_file_carries_the_escaped_format(self, tmp_path):
        import re as _re
        import zipfile

        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", verification_date="05/05/2026")])
        result = rv.validate_workbook(src)

        with zipfile.ZipFile(result["output_path"]) as archive:
            styles = archive.read("xl/styles.xml").decode()
        codes = _re.findall(r'<numFmt numFmtId="\d+" formatCode="([^"]+)"', styles)
        assert "DD\\/MM\\/YYYY" in codes

    def test_normalized_date_cells_use_it(self, tmp_path):
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", taf="TAF", verification_date="05/05/2026",
                                     renewal_date="20-05-2026")])
        result = rv.validate_workbook(src)
        ws = openpyxl.load_workbook(result["output_path"])[SHEET]

        for key in ("verification_date", "renewal_date"):
            assert ws.cell(row=5, column=KEY_TO_COL[key]).number_format == rv.DATE_NUMBER_FORMAT

    def test_the_filename_is_unchanged(self, tmp_path):
        """The filename was never the problem - it must not drift."""
        src = tmp_path / "B.xlsx"
        make_workbook(src, [good_row("P01", verification_date="05/05/2026")])
        result = rv.validate_workbook(src)
        assert result["custom_filename"] == "TESTBR_Audit-MIS_05-MAY-2026.xlsx"
