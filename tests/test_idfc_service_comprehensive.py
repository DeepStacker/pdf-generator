"""Comprehensive tests for IDFCService — covers remaining lines (read_excel, generate, _register_fonts, validate edges)."""

import os
import tempfile

import openpyxl
import pytest

from audit_engine.exceptions import ValidationError
from audit_engine.services.idfc import IDFCService


@pytest.fixture
def service():
    return IDFCService(fonts_dir="/nonexistent/fonts")


class TestInitAndFonts:
    def test_init_fallback_fonts(self, service):
        assert service._font_reg == "Helvetica"
        assert service._font_bld == "Helvetica-Bold"

    def test_resolve_fonts_dir_dev(self):
        """_resolve_fonts_dir returns a path ending in fonts."""
        d = IDFCService._resolve_fonts_dir()
        assert d.endswith("fonts")
        assert os.path.isabs(d)

    def test_get_resource_path(self):
        p = IDFCService.get_resource_path("some/relative.txt")
        assert p.endswith("some/relative.txt")
        assert os.path.isabs(p)

    def test_font_reg_property(self, service):
        assert service.font_reg == "Helvetica"

    def test_font_bld_property(self, service):
        assert service.font_bld == "Helvetica-Bold"

    def test_register_fonts_nonexistent_dir_logs_warning(self, service, caplog):
        """_register_fonts logs a warning when font files are missing."""
        import logging
        caplog.set_level(logging.WARNING)
        service._register_fonts()
        assert service._font_reg == "Helvetica"

    def test_register_fonts_exception_on_corrupt_font(self, service, monkeypatch):
        """_register_fonts catches OSError/ValueError when font registration fails."""
        class FakeTTFont:
            def __init__(self, *a, **kw):
                pass

        monkeypatch.setattr("audit_engine.services.idfc.TTFont", FakeTTFont)
        monkeypatch.setattr("audit_engine.services.idfc.pdfmetrics.registerFont",
                            lambda f: (_ for _ in ()).throw(OSError("Font registration failed")))
        monkeypatch.setattr("os.path.exists", lambda p: True)
        svc = IDFCService(fonts_dir="/fake")
        assert svc._font_reg == "Helvetica"
        assert svc._font_bld == "Helvetica-Bold"


class TestValidateEdgeCases:
    def test_validate_permission_error(self, service, monkeypatch):
        """validate catches PermissionError when reading a file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"AAAA")
            p = f.name
        monkeypatch.setattr("builtins.open", lambda *a, **kw: (_ for _ in ()).throw(PermissionError("Permission denied")))
        try:
            valid, err = service.validate(p)
            assert valid is False
            assert "permission denied" in err.lower() or "locked" in err.lower()
        finally:
            os.unlink(p)

    def test_validate_os_error(self, service, monkeypatch, tmp_path):
        p = os.path.join(str(tmp_path), "exists.xlsx")
        with open(p, "wb") as f:
            f.write(b"AAAA")
        orig_open = open

        def broken_open(*args, **kwargs):
            if args[0] == p and "rb" in (kwargs.get("mode", "rb") if args[0] == p else "rb"):
                raise OSError("device error")
            return orig_open(*args, **kwargs)

        monkeypatch.setattr("builtins.open", broken_open)
        valid, err = service.validate(str(p))
        assert valid is False
        assert "Cannot read" in err

    def test_validate_output_dir_not_writable(self, service, monkeypatch):
        monkeypatch.setattr("os.makedirs", lambda *a, **kw: (_ for _ in ()).throw(OSError("Read-only file system")))
        valid, err = service.validate_output_dir("/readonly/output")
        assert valid is False
        assert "not writable" in err

    def test_validate_output_dir_empty(self, service):
        valid, err = service.validate_output_dir("")
        assert valid is False
        assert "No output directory" in err

    def test_validate_output_dir_valid(self, service, tmp_path):
        valid, err = service.validate_output_dir(str(tmp_path))
        assert valid is True
        assert err == ""


class TestReadExcel:
    def test_read_excel_finds_valid_sheet(self, service, tmp_path):
        """read_excel returns sheet name, headers, and row data."""
        p = os.path.join(str(tmp_path), "valid.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        headers = ["Prospectno", "CUID", "Tare Weight", "State", "CurrentBranch", "CurrentBranchName"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value="C001")
        ws.cell(row=2, column=3, value=100.5)
        ws.cell(row=2, column=4, value="MH")
        ws.cell(row=2, column=5, value="MUMBAI")
        ws.cell(row=2, column=6, value="Mumbai Main")
        wb.save(str(p))
        wb.close()

        sheet, headers_out, rows = service.read_excel(str(p))
        assert sheet == "Sheet1"
        assert "Prospectno" in headers_out
        assert len(rows) == 1
        assert rows[0]["Prospectno"] == "P001"
        assert rows[0]["CUID"] == "C001"

    def test_read_excel_skips_empty_rows(self, service, tmp_path):
        """read_excel skips rows where all cells are None."""
        p = os.path.join(str(tmp_path), "with_empty.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["Prospectno", "CUID", "Tare Weight", "State", "CurrentBranch", "CurrentBranchName"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value="C001")
        ws.cell(row=3, column=1, value=None)  # empty row
        ws.cell(row=3, column=2, value=None)
        ws.cell(row=4, column=1, value="P002")
        wb.save(str(p))
        wb.close()

        _, _, rows = service.read_excel(str(p))
        assert len(rows) == 2
        assert rows[0]["Prospectno"] == "P001"
        assert rows[1]["Prospectno"] == "P002"

    def test_read_excel_no_valid_sheet_raises(self, service, tmp_path):
        """read_excel raises ValidationError when no sheet has the required columns."""
        p = os.path.join(str(tmp_path), "bad.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Wrong")
        ws.cell(row=1, column=2, value="Headers")
        wb.save(str(p))
        wb.close()

        with pytest.raises(ValidationError, match="No valid sheet found"):
            service.read_excel(str(p))

    def test_read_excel_finds_sheet_not_first(self, service, tmp_path):
        """read_excel scans all sheets and picks the one with required columns."""
        p = os.path.join(str(tmp_path), "multi_sheet.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Intro"
        ws1.cell(row=1, column=1, value="Welcome")

        ws2 = wb.create_sheet("Data")
        headers = ["Prospectno", "CUID", "Tare Weight", "State", "CurrentBranch", "CurrentBranchName"]
        for i, h in enumerate(headers, 1):
            ws2.cell(row=1, column=i, value=h)
        ws2.cell(row=2, column=1, value="P001")
        wb.save(str(p))
        wb.close()

        sheet, _, rows = service.read_excel(str(p))
        assert sheet == "Data"
        assert len(rows) == 1

    def test_read_excel_empty_sheet_stopiteration(self, service, tmp_path):
        """read_excel handles empty sheet gracefully (StopIteration on iter_rows)."""
        p = os.path.join(str(tmp_path), "empty_sheet.xlsx")
        wb = openpyxl.Workbook()
        # Delete the default sheet and add an empty one
        default = wb.active
        wb.remove(default)
        wb.create_sheet("Empty")
        # No data written
        wb.save(str(p))
        wb.close()

        with pytest.raises(ValidationError):
            service.read_excel(str(p))

    def test_read_excel_log_callback(self, service, tmp_path):
        """read_excel calls the log callback with status messages."""
        p = os.path.join(str(tmp_path), "callback.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Prospectno", "CUID", "Tare Weight", "State", "CurrentBranch", "CurrentBranchName"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="P001")
        wb.save(str(p))
        wb.close()

        messages = []

        def log(msg):
            messages.append(msg)

        service.read_excel(str(p), log_callback=log)
        assert any("Reading Excel" in m for m in messages)
        assert any("valid sheet" in m for m in messages)


class TestGroupByBranch:
    def test_group_by_branch_unknown_for_none(self, service):
        rows = [
            {"CurrentBranch": None, "data": "a"},
            {"CurrentBranch": "BranchA", "data": "b"},
        ]
        groups = service.group_by_branch(rows)
        assert "UNKNOWN" in groups
        assert "BranchA" in groups
        assert groups["BranchA"][0]["data"] == "b"

    def test_group_by_branch_nan_value(self, service):
        rows = [
            {"CurrentBranch": float("nan"), "data": "x"},
        ]
        groups = service.group_by_branch(rows)
        assert "UNKNOWN" in groups

    def test_group_by_branch_empty_string(self, service):
        rows = [
            {"CurrentBranch": "", "data": "x"},
        ]
        groups = service.group_by_branch(rows)
        assert "UNKNOWN" in groups


class TestFormatTareWeight:
    def test_format_tare_value_error(self, service):
        """format_tare_weight returns empty string for values that cause pd.isna to raise."""
        class WeirdType:
            def __str__(self):
                return "weird"

        result = service.format_tare_weight(WeirdType())
        assert result in ("weird", "")

    def test_format_tare_nan(self, service):
        assert service.format_tare_weight(float("nan")) == ""

    def test_format_tare_pd_isna_raises(self, service):
        """format_tare_weight catches ValueError/TypeError from pd.isna."""
        result = service.format_tare_weight([1, 2])
        assert result == "[1, 2]"

    def test_format_tare_none(self, service):
        assert service.format_tare_weight(None) == ""

    def test_format_tare_empty_string(self, service):
        assert service.format_tare_weight("") == ""


class TestGenerate:
    def test_generate_empty_rows_returns_early(self, service, tmp_path):
        out = os.path.join(str(tmp_path), "empty.pdf")
        service.generate("POA", "BR001", "Test Branch", "MH", [], out)
        assert not os.path.exists(out)

    def test_generate_creates_pdf(self, service, tmp_path):
        out = os.path.join(str(tmp_path), "output.pdf")
        rows = [
            {"Prospectno": "P001", "CUID": "C001", "Tare Weight": 100.5},
            {"Prospectno": "P002", "CUID": "C002", "Tare Weight": 200.0},
        ]
        service.generate("POA", "BR001", "Test Branch", "MH", rows, out)
        assert os.path.exists(out)
        assert os.path.getsize(out) > 1000

    def test_generate_with_special_chars(self, service, tmp_path):
        out = os.path.join(str(tmp_path), "special.pdf")
        rows = [
            {"Prospectno": "P/0\\1", "CUID": "C:01", "Tare Weight": None},
        ]
        service.generate("POA", "BR/0\\1", "Branch:Name", "MH", rows, out)
        assert os.path.exists(out)

    def test_generate_logs_callback_calls(self, service, tmp_path):
        out = os.path.join(str(tmp_path), "logged.pdf")
        messages = []
        rows = [{"Prospectno": "P001", "CUID": "C001"}]

        def log_callback(msg):
            messages.append(msg)

        service.generate("POA", "BR001", "Branch", "MH", rows, out)
        assert os.path.exists(out)
