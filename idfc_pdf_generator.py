import os
import argparse
import openpyxl
import pandas as pd

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import urllib.request

# =========================================================
# DYNAMIC FONT LOADING (Works on ANY machine)
# =========================================================
FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
os.makedirs(FONT_DIR, exist_ok=True)

CALIBRI_REGULAR_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/carlito/Carlito-Regular.ttf"
ARIMO_BOLD_URL = "https://raw.githubusercontent.com/googlefonts/arimo/main/fonts/ttf/Arimo-Bold.ttf"

reg_path = os.path.join(FONT_DIR, "Carlito-Regular.ttf")
bold_path = os.path.join(FONT_DIR, "Arimo-Bold.ttf")


def download_font(url, path):
    if not os.path.exists(path):
        print(f"Downloading required font -> {os.path.basename(path)}...")
        try:
            req = urllib.request.Request(
                url,
                headers={'User-Agent': 'Mozilla/5.0'}
            )

            with urllib.request.urlopen(req) as response, open(path, 'wb') as out_file:
                out_file.write(response.read())

        except Exception as e:
            print(f"Failed to download font: {e}")


download_font(CALIBRI_REGULAR_URL, reg_path)
download_font(ARIMO_BOLD_URL, bold_path)

try:
    pdfmetrics.registerFont(TTFont('Carlito', reg_path))
    pdfmetrics.registerFont(TTFont('Arimo-Bold', bold_path))

    FONT_REGULAR = 'Carlito'
    FONT_BOLD = 'Arimo-Bold'

except Exception as e:
    print(f"Warning: Could not load dynamic fonts ({e}). Falling back to standard fonts.")

    FONT_REGULAR = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'

# =========================================================
# REQUIRED COLUMNS
# =========================================================
REQUIRED_COLUMNS = [
    "Prospectno",
    "CUID",
    "Tare Weight",
    "State",
    "CurrentBranch",
    "CurrentBranchName",
]


# =========================================================
# HELPERS
# =========================================================
def format_tare_weight(val):
    """Format tare weight: remove trailing .0 for whole numbers."""

    if pd.isna(val) or val == "" or val is None:
        return ""

    try:
        fval = float(val)

        if fval == int(fval):
            return str(int(fval))

        return str(fval)

    except (ValueError, TypeError):
        return str(val)


# =========================================================
# READ EXCEL WITH OPENPYXL
# =========================================================
def read_excel_preserving_text(excel_path):

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    required_lower = [c.lower() for c in REQUIRED_COLUMNS]

    target_sheet = None

    for sname in wb.sheetnames:

        ws = wb[sname]

        header_row = []

        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            header_row.append(
                str(cell.value).strip().replace("\n", "")
                if cell.value is not None else ""
            )

        header_lower = [h.lower() for h in header_row]

        if all(r in header_lower for r in required_lower):
            print(f"Valid sheet found: {sname}")
            target_sheet = sname
            break

    if target_sheet is None:
        raise Exception("No valid sheet found.")

    ws = wb[target_sheet]

    print(f"Detected Sheet: {target_sheet}")

    headers = []

    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(
            str(cell.value).strip().replace("\n", "")
            if cell.value is not None else ""
        )

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=False):

        row_data = {}

        all_none = True

        for i, cell in enumerate(row):

            if i < len(headers) and headers[i]:

                val = cell.value

                if val is not None:
                    all_none = False

                if headers[i] in ("Prospectno", "CUID"):
                    row_data[headers[i]] = str(val).strip() if val is not None else ""

                else:
                    row_data[headers[i]] = val

        if not all_none:
            rows.append(row_data)

    return target_sheet, headers, rows


# =========================================================
# PDF GENERATOR
# =========================================================
def generate_pdf(audit_type, branch_code, branch_name, state, rows, output_path):

    PAGE_WIDTH, PAGE_HEIGHT = landscape(A4)

    LEFT_MARGIN = 50.2
    RIGHT_MARGIN = 50.2
    TOP_MARGIN = 48.0

    doc = SimpleDocTemplate(
        output_path,
        pagesize=(PAGE_WIDTH, PAGE_HEIGHT),
        leftMargin=LEFT_MARGIN,
        rightMargin=RIGHT_MARGIN,
        topMargin=TOP_MARGIN,
        bottomMargin=15,
        title=os.path.basename(output_path),
    )

    cw = [22.2, 101.1, 118.5, 60.3, 67.2, 125.0, 247.3]

    style_hdr = ParagraphStyle(
        "ColHdr",
        fontName=FONT_BOLD,
        fontSize=9,
        alignment=TA_CENTER,
        leading=10,
        spaceBefore=0,
        spaceAfter=0,
        leftIndent=0,
        rightIndent=0,
        firstLineIndent=0,
    )

    table_data = []

    # =====================================================
    # HEADER ROW 1
    # =====================================================
    table_data.append([
        "Audit Type :",
        "",
        str(audit_type),
        "",
        "Branch Name :",
        "",
        str(branch_name),
    ])

    # =====================================================
    # HEADER ROW 2
    # =====================================================
    table_data.append([
        "Branch Code :",
        "",
        str(branch_code),
        "",
        "State :",
        "",
        str(state),
    ])

    # =====================================================
    # COLUMN HEADERS
    # =====================================================
    table_data.append([
        Paragraph("Sr<br/>No", style_hdr),
        Paragraph("Prospectno", style_hdr),
        Paragraph("CUID", style_hdr),
        Paragraph("Tare Weight<br/>as per Bank", style_hdr),
        Paragraph("<nobr>Tare Weight as</nobr><br/>per Audit", style_hdr),
        Paragraph(
            "<nobr>Purity Check - 18K and</nobr><br/><nobr>above 18K or Below 18K</nobr>",
            style_hdr
        ),
        Paragraph("Remarks", style_hdr),
    ])

    # =====================================================
    # DATA ROWS
    # =====================================================
    for idx, row in enumerate(rows, 1):

        prospectno = str(row.get("Prospectno", ""))
        cuid = str(row.get("CUID", ""))
        tare_weight = format_tare_weight(row.get("Tare Weight", ""))

        table_data.append([
            str(idx),
            prospectno,
            cuid,
            tare_weight,
            "",
            "",
            "",
        ])

    row_heights = [12.2, 14.2, 22.5] + [30.5] * (len(table_data) - 3)

    yellow = colors.HexColor("#FFFF00")
    blue = colors.HexColor("#4985E8")

    table = Table(
        table_data,
        colWidths=cw,
        rowHeights=row_heights,
        repeatRows=3
    )

    style_cmds = [

        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (4, 0), (5, 0)),
        ("SPAN", (0, 1), (1, 1)),
        ("SPAN", (4, 1), (5, 1)),

        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("FONTNAME", (0, 0), (-1, 1), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 1), 9),

        ("FONTNAME", (0, 3), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 3), (-1, -1), 9),

        ("FONTNAME", (0, 3), (0, -1), FONT_BOLD),

        ("BACKGROUND", (0, 2), (3, 2), yellow),
        ("BACKGROUND", (4, 2), (6, 2), blue),

        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1),

        ("LINEBEFORE", (0, 0), (0, -1), 0.5, colors.black),
        ("LINEAFTER", (6, 0), (6, -1), 0.5, colors.black),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.black),

        ("GRID", (0, 0), (2, 1), 0.5, colors.black),
        ("GRID", (4, 0), (6, 1), 0.5, colors.black),

        ("GRID", (0, 2), (-1, -1), 0.5, colors.black),
    ]

    table.setStyle(TableStyle(style_cmds))

    doc.build([table])


# =========================================================
# PROCESS EXCEL
# =========================================================
def process_excel(input_excel, output_dir, audit_type):

    audit_type = str(audit_type).strip().upper()

    os.makedirs(output_dir, exist_ok=True)

    print(f"Reading Excel: {input_excel}")
    print(f"Audit Type: {audit_type}")

    sheet_name, headers, all_rows = read_excel_preserving_text(input_excel)

    groups = {}

    for row in all_rows:

        branch = str(row.get("CurrentBranch", "UNKNOWN")).strip()

        if branch in ("None", ""):
            branch = "UNKNOWN"

        groups.setdefault(branch, []).append(row)

    total = 0

    for branch_code, branch_rows in sorted(groups.items()):

        try:

            branch_name = str(
                branch_rows[0].get("CurrentBranchName", "")
            ).strip()

            state = str(
                branch_rows[0].get("State", "")
            ).strip()

            safe_branch_name = (
                branch_name
                .replace("/", "_")
                .replace("\\", "_")
            )

            if not safe_branch_name:
                safe_branch_name = (
                    str(branch_code)
                    .replace("/", "_")
                    .replace("\\", "_")
                )

            output_file = os.path.join(
                output_dir,
                f"{safe_branch_name}_{audit_type}.pdf"
            )

            print(f"Generating -> {output_file}")

            generate_pdf(
                audit_type,
                branch_code,
                branch_name,
                state,
                branch_rows,
                output_file
            )

            total += 1

        except Exception as e:

            print(f"Error generating PDF for branch {branch_code}: {e}")

            import traceback
            traceback.print_exc()

    print(f"\nDone. Generated {total} PDFs.")
    print(f"Output Folder: {output_dir}")


# =========================================================
# CLI
# =========================================================
if __name__ == "__main__":

    parser = argparse.ArgumentParser(
        description="Generate branch-wise PDFs from Excel"
    )

    # =====================================================
    # INPUT EXCEL
    # =====================================================
    parser.add_argument(
        "input_file",
        help="Path to Excel file"
    )

    # =====================================================
    # AUDIT TYPE (VERY EASY)
    # =====================================================
    parser.add_argument(
        "-t",
        "--type",
        default="POA",
        help="Pass POA or TAF"
    )

    # =====================================================
    # OUTPUT DIRECTORY
    # =====================================================
    parser.add_argument(
        "-o",
        "--output-dir",
        default="generated_pdfs",
        help="Output folder"
    )

    args = parser.parse_args()

    try:

        process_excel(
            args.input_file,
            args.output_dir,
            args.type
        )

    except Exception as e:
        print(f"Fatal Error: {e}")