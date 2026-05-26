"""Excel column mapping — rename user columns to expected engine columns.

Extracted from routes.py to separate concerns.
"""

import os

import openpyxl

from audit_engine.utils.platform import file_logger


def preprocess_mapped_excel(filepath: str, column_mappings: dict[str, str], bank: str) -> str:
    """Create a temp copy of the Excel with columns renamed per user mappings."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        target_sheet = None
        max_matches = -1
        mapped_headers_set = set(column_mappings.values())

        for sname in wb.sheetnames:
            ws = wb[sname]
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
                matches = sum(1 for h in headers if h in mapped_headers_set)
                if matches > max_matches:
                    max_matches = matches
                    target_sheet = sname
            except StopIteration:
                continue

        if not target_sheet:
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]
        first_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in first_row]
        header_indices = {h: idx for idx, h in enumerate(headers, 1)}

        if bank == "IDFC First Bank":
            _remap_idfc(ws, header_indices, column_mappings)
        else:
            _remap_equitas(wb, column_mappings, mapped_headers_set)

        temp_dir = os.path.join(os.path.expanduser("~"), ".temp_audit_engine")
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, f"mapped_{os.path.basename(filepath)}")
        wb.save(temp_path)
        wb.close()
        return temp_path
    except Exception as e:
        file_logger.warning("Column mapping failed: %s", e)
        return filepath


def _remap_idfc(ws, header_indices: dict, col_map: dict) -> None:
    prospect_col = col_map.get("prospect")
    cuid_col = col_map.get("cuid")
    tare_col = col_map.get("tare")
    branch_col = col_map.get("branch")

    rename: dict[str, str] = {}
    if prospect_col:
        rename["Prospectno"] = prospect_col
    if cuid_col:
        rename["CUID"] = cuid_col
    if tare_col:
        rename["Tare Weight"] = tare_col
    if branch_col:
        rename["CurrentBranchName"] = branch_col

    for target, orig in rename.items():
        if orig in header_indices:
            idx = header_indices[orig]
            ws.cell(row=1, column=idx).value = target

    updated = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "CurrentBranch" not in updated:
        branch_idx = updated.index("CurrentBranchName") + 1 if "CurrentBranchName" in updated else None
        if branch_idx:
            new_idx = len(updated) + 1
            ws.cell(row=1, column=new_idx).value = "CurrentBranch"
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=new_idx).value = ws.cell(row=r, column=branch_idx).value

    updated = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "State" not in updated:
        new_idx = len(updated) + 1
        ws.cell(row=1, column=new_idx).value = "State"
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=new_idx).value = "UNKNOWN"


def _remap_equitas(wb, col_map: dict, mapped_set: set) -> None:
    svs_col = col_map.get("svs")
    sole_col = col_map.get("sole")
    branch_col = col_map.get("branch")
    loan_col = col_map.get("loan")

    normal: dict[str, str] = {}
    if svs_col:
        normal["SVS_LOAN_NO"] = svs_col
    if sole_col:
        normal["SOLE_ID"] = sole_col
    if branch_col:
        normal["BRANCH_NAME"] = branch_col

    jsr: dict[str, str] = {}
    if loan_col:
        jsr["LOAN NO"] = loan_col
    if branch_col:
        jsr["BRANCHNAME"] = branch_col

    for sname in wb.sheetnames:
        ws = wb[sname]
        is_jsr = str(sname).strip().upper().endswith("JSR")
        mapping = jsr if is_jsr else normal
        header_row_idx = 1
        max_matches = -1

        for r_idx in range(1, min(ws.max_row + 1, 31)):
            cells = list(ws.iter_rows(min_row=r_idx, max_row=r_idx))[0]
            hdrs = [str(c.value).strip() if c.value else "" for c in cells]
            matches = sum(1 for h in hdrs if h in mapped_set)
            if matches > max_matches:
                max_matches = matches
                header_row_idx = r_idx

        if max_matches >= 1:
            cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
            hdrs = [str(c.value).strip() if c.value else "" for c in cells]
            indices = {h: idx for idx, h in enumerate(hdrs, 1)}
            for target, orig in mapping.items():
                if orig in indices:
                    ws.cell(row=header_row_idx, column=indices[orig]).value = target
