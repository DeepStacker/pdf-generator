"""Bank auto-detection from Excel header fingerprints.

Defines bank profiles with fingerprint columns and detection thresholds.
The profiles drive detection — adding a new bank means adding a profile entry.
"""

import logging

import openpyxl

from audit_engine.domain.models import BankProfile
from audit_engine.utils.config import fingerprints as fp

_MIN_NON_EMPTY_FOR_HEADER = 3

_logger = logging.getLogger(__name__)

# Bank detection profiles — ordered by specificity (most specific first)
BANK_PROFILES: list[BankProfile] = [
    BankProfile("IDFC First Bank", fp.idfc, 3),
    BankProfile("Equitas Small Finance Bank", fp.equitas, 3),
    BankProfile("Arvog Bank", fp.arvog, 2),
]


def detect_bank_from_file(filepath: str) -> str | None:
    """Peek at Excel headers to determine which bank the file is for."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        _logger.warning("Cannot open %s for bank detection: %s", filepath, e)
        return None

    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r_idx in range(1, min(ws.max_row + 1, 31)):
                try:
                    row_cells = list(ws.iter_rows(min_row=r_idx, max_row=r_idx))[0]
                except Exception:
                    continue
                header_set = {
                    str(cell.value).strip().lower().replace("\n", "")
                    for cell in row_cells if cell.value
                }
                for profile in BANK_PROFILES:
                    if profile.matches(header_set):
                        return profile.name
    finally:
        wb.close()
    return None


def peek_excel_data(filepath: str) -> tuple[list[str], list[list[str]]]:
    """Return headers and first 5 data rows for preview."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        _logger.warning("Failed to peek Excel %s: %s", filepath, e)
        return [], []

    try:
        sheet = wb.active
        headers: list[str] = []
        rows: list[list[str]] = []
        header_row_idx = 0

        all_rows = list(sheet.iter_rows(values_only=True, max_row=30))
        for i, row_cells in enumerate(all_rows):
            non_empty = sum(1 for c in row_cells if c is not None and str(c).strip())
            if non_empty >= _MIN_NON_EMPTY_FOR_HEADER:
                header_row_idx = i
                headers = [str(c).strip() if c is not None else f"Column {idx}" for idx, c in enumerate(row_cells)]
                break

        if not headers and all_rows:
            headers = [str(c).strip() if c is not None else f"Column {idx}" for idx, c in enumerate(all_rows[0])]

        for row_cells in sheet.iter_rows(values_only=True, min_row=header_row_idx + 2, max_row=header_row_idx + 6):
            row_data = [str(c) if c is not None else "" for c in row_cells]
            if len(row_data) < len(headers):
                row_data.extend([""] * (len(headers) - len(row_data)))
            rows.append(row_data[:len(headers)])

        return headers, rows
    except Exception as e:
        _logger.warning("Failed to peek Excel data %s: %s", filepath, e)
        return [], []
    finally:
        wb.close()
