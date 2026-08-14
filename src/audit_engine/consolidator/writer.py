from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, Side

from .config import MD_COLUMNS, PT_COLUMNS

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ConsolidatedWriter:
    """Writes canonical rows into an Excel file matching the existing format."""

    PT_SHEET = "Payment Tracker"
    MD_SHEET = "Master Data"

    def __init__(self, template_path: Path):
        self.template_path = template_path
        self.pt_rows: list[dict[int, Any]] = []
        self.md_rows: list[dict[int, Any]] = []

    def add_pt_rows(self, rows: list[dict[int, Any]]):
        self.pt_rows.extend(rows)

    def add_md_rows(self, rows: list[dict[int, Any]]):
        self.md_rows.extend(rows)

    def write(self, output_path: Path) -> Path:
        wb = openpyxl.load_workbook(self.template_path)

        self._write_sheet(
            wb, self.PT_SHEET, PT_COLUMNS, self.pt_rows,
            total_formulas={13: "M", 19: "S"},
        )
        self._write_sheet(
            wb, self.MD_SHEET, MD_COLUMNS, self.md_rows,
            total_formulas={27: "AA", 37: "AK"},
        )

        wb.save(output_path)
        wb.close()
        return output_path

    def _write_sheet(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        canonical_columns: list[str],
        data_rows: list[dict[int, Any]],
        total_formulas: dict[int, str] | None = None,
    ):
        if sheet_name not in wb:
            raise ValueError(f"Sheet '{sheet_name}' not found in template")

        ws = wb[sheet_name]
        ncols = len(canonical_columns)

        # Ensure enough columns
        while ws.max_column < ncols:
            ws.cell(1, ws.max_column + 1)

        # Capture header formatting from template before clearing
        header_formats = {}
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            header_formats[c] = {
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
            }

        # --- CLEAR ALL EXISTING CONTENT (rows 2+) ---
        max_row = ws.max_row
        needed_rows = 3 + len(data_rows) if data_rows else 2
        for r in range(2, max(max_row, needed_rows) + 5):
            for c in range(1, max(ws.max_column, ncols) + 5):
                cell = ws.cell(r, c)
                cell.value = None
                cell.border = _THIN_BORDER

        # Write header row preserving template formatting
        for c, col_name in enumerate(canonical_columns, 1):
            cell = ws.cell(1, c)
            cell.value = col_name
            fmt = header_formats.get(c, {})
            if fmt.get("font"):
                cell.font = fmt["font"]
            if fmt.get("alignment"):
                cell.alignment = fmt["alignment"]
            if fmt.get("border"):
                cell.border = fmt["border"]
            if fmt.get("fill"):
                cell.fill = fmt["fill"]

        # Write data rows with thin borders
        for r_idx, row_dict in enumerate(data_rows, 2):
            for c_idx in range(ncols):
                cell = ws.cell(r_idx, c_idx + 1)
                cell.border = _THIN_BORDER
                val = row_dict.get(c_idx)
                if val is not None:
                    cell.value = val

        # Write totals row (blank row, then totals)
        if data_rows and total_formulas:
            data_count = len(data_rows)
            formula_end = data_count + 2
            totals_row = data_count + 3

            for col_idx, col_letter in total_formulas.items():
                cell = ws.cell(totals_row, col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{formula_end})"
                cell.border = _THIN_BORDER
                cell.font = Font(bold=True)
