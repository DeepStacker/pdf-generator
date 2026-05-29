from __future__ import annotations

from .config import SourceConfig

# ===========================================================================
# 1) AXIS BANK POA
# Needs overrides for MD: "Location " reads from "State" source header
# (source has no separate "Location" column for MD)
# ===========================================================================

AXIS_POA = SourceConfig(
    file_pattern="Axis Bank POA*",
    client_name="Axis Bank POA",
    expected_pt_rows=(4, 50),
    expected_md_rows=(4, 100),
    column_overrides={
        "Location ": "State",
    },
    needs_manual_enrichment={
        "State": "Derive from Location city and fill in PT col 6",
        "Zone": "Derive from State and fill in PT col 7",
    },
)

# ===========================================================================
# 2) AXIS COLLECTION AGENCY
# Needs overrides for MD: multiple canonicals share the same source column
# ===========================================================================

AXIS_COLLECTION = SourceConfig(
    file_pattern="Axis Collection Agency*",
    client_name="Axis Collection Agency",
    expected_pt_rows=(3, 30),
    expected_md_rows=(3, 50),
    column_overrides={
        "No of days \naudited ": "No of Audits",
        "Audit \ncompletion date": "Audit Date",
        "Total": "Auditor Payment",
        "Assayer fee": "Auditor Base Fee",
        "Additional fee": "Auditor Travel fee",
    },
    needs_manual_enrichment={
        "State": "Derive from Location city and fill in PT col 6",
        "Zone": "Derive from State and fill in PT col 7",
        "MD Zone": "Fill Zone for MD col 3 from Location",
        "MD State": "Fill State for MD col 7 from Location",
    },
)

# ===========================================================================
# 3) SCB
# Needs overrides for MD: "Base Location" reads "Location", dates read
# "Audit Conducted on", "Total" reads "Total Fees in INR"
# ===========================================================================

SCB = SourceConfig(
    file_pattern="SCB*",
    client_name="SCB",  # trailing space matches original
    pt_sheet_hint="Auditor Fees",
    md_sheet_hint="Master data",
    expected_pt_rows=(1, 20),
    expected_md_rows=(1, 50),
    column_overrides={
        "Base Location": "Location",
        "Audit \ncompletion date": "Audit Conducted on",
        "Total": "Total Fees in INR",
        "Assayer fee": "Base Audit Fees in INR",
        "Additional fee": "Additional amount in INR",
    },
    needs_manual_enrichment={
        "State": "Derive from Location city and fill in PT col 6",
        "Zone": "Derive from State and fill in PT col 7",
        "MD Zone": "Fill Zone for MD col 3 from Location",
        "MD State": "Fill State for MD col 7 from Location",
    },
)

# ===========================================================================
# 4) VISTAAR Muthoot
# Needs overrides for MD: "Location " reads from "State" source header
# ===========================================================================

VISTAAR = SourceConfig(
    file_pattern="VISTAAR*",
    client_name="VISTAAR FINANCIAL Muthoot",
    expected_pt_rows=(5, 50),
    expected_md_rows=(5, 50),
    column_overrides={
        "Location ": "State",
        "No of days \naudited For client": "No. of Visit",
        "Additional": "Additonal",
        "Total": "Total",
    },
)

# ===========================================================================
# Registry: all known source templates
# ===========================================================================

ALL_SOURCE_CONFIGS = [
    AXIS_POA,
    AXIS_COLLECTION,
    SCB,
    VISTAAR,
]


def _get_project_root() -> Path:
    """Return project root — works in dev mode and PyInstaller frozen mode."""
    import sys
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    return Path(__file__).parent.parent.parent.parent


def _resolve_settings_path() -> Path:
    """Find settings file, preferring writable user copy for frozen apps."""
    import shutil
    from pathlib import Path
    
    user_copy = Path.home() / ".audit_engine_elite" / "settings" / "mapping_rules.xlsx"
    if user_copy.exists():
        return user_copy
    
    bundled = _get_project_root() / "settings" / "mapping_rules.xlsx"
    if bundled.exists():
        user_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(bundled), str(user_copy))
        return user_copy
    
    return bundled


def load_external_source_configs(default_configs: list[SourceConfig]) -> list[SourceConfig]:
    import openpyxl
    import sys
    from pathlib import Path
    
    rules_path = _resolve_settings_path()
    if not rules_path.exists():
        return default_configs
        
    try:
        wb = openpyxl.load_workbook(rules_path, data_only=True)
        if "Registered_Banks" not in wb.sheetnames or "Column_Mappings" not in wb.sheetnames:
            wb.close()
            return default_configs
            
        ws_reg = wb["Registered_Banks"]
        configs: list[SourceConfig] = []
        config_map: dict[str, SourceConfig] = {}
        
        for r in range(2, ws_reg.max_row + 1):
            client_name = ws_reg.cell(r, 1).value
            pattern = ws_reg.cell(r, 2).value
            if not client_name or not pattern:
                continue
                
            client_name = str(client_name).strip()
            pattern = str(pattern).strip()
            
            pt_hint = ws_reg.cell(r, 3).value or "Payment Tracker"
            md_hint = ws_reg.cell(r, 4).value or "Master Data"
            hdr_row = ws_reg.cell(r, 5).value or 1
            start_row = ws_reg.cell(r, 6).value or 2
            
            pt_min = ws_reg.cell(r, 7).value
            pt_max = ws_reg.cell(r, 8).value
            md_min = ws_reg.cell(r, 9).value
            md_max = ws_reg.cell(r, 10).value
            
            pt_rows = (int(pt_min), int(pt_max)) if pt_min is not None and pt_max is not None else None
            md_rows = (int(md_min), int(md_max)) if md_min is not None and md_max is not None else None
            
            cfg = SourceConfig(
                file_pattern=pattern,
                client_name=client_name,
                pt_sheet_hint=str(pt_hint).strip(),
                md_sheet_hint=str(md_hint).strip(),
                header_row=int(hdr_row),
                data_start_row=int(start_row),
                expected_pt_rows=pt_rows,
                expected_md_rows=md_rows,
            )
            configs.append(cfg)
            config_map[client_name] = cfg
            
        # Preserve code default overrides before applying external ones
        default_overrides: dict[str, dict[str, str]] = {}
        for default in default_configs:
            name_strip = default.client_name.strip()
            if default.column_overrides:
                default_overrides[name_strip] = dict(default.column_overrides)

        ws_map = wb["Column_Mappings"]
        headers = [str(ws_map.cell(1, col).value).strip() for col in range(1, ws_map.max_column + 1)]
        
        client_cols = {}
        for idx, h in enumerate(headers, 1):
            if h in config_map:
                client_cols[h] = idx
                
        for r in range(2, ws_map.max_row + 1):
            canonical = ws_map.cell(r, 1).value
            if not canonical:
                continue
            canonical = str(canonical).strip()
            
            for client_name, col_idx in client_cols.items():
                cell_val = ws_map.cell(r, col_idx).value
                if cell_val is not None:
                    cell_val = str(cell_val).strip()
                    if cell_val and cell_val.lower() != "none":
                        # Skip identity overrides (value == canonical name)
                        # so that duplicate-header pattern matching still works
                        if cell_val.lower() == canonical.lower():
                            continue
                        config_map[client_name].column_overrides[canonical] = cell_val
                        
        # Merge code default overrides (external non-None overrides win)
        for cfg in configs:
            name_strip = cfg.client_name.strip()
            if name_strip in default_overrides:
                for k, v in default_overrides[name_strip].items():
                    if k not in cfg.column_overrides:
                        cfg.column_overrides[k] = v
                        
        for default in default_configs:
            name_strip = default.client_name.strip()
            for cfg in configs:
                if cfg.client_name.strip() == name_strip:
                    cfg.needs_manual_enrichment = default.needs_manual_enrichment
                    cfg.exclude_md_names = default.exclude_md_names
                    
        wb.close()
        return configs if configs else default_configs
    except Exception as e:
        print(f"Warning: Failed to load external source configs: {e}. Using fallback default rules.", file=sys.stderr)
        return default_configs

ALL_SOURCE_CONFIGS = load_external_source_configs(ALL_SOURCE_CONFIGS)

