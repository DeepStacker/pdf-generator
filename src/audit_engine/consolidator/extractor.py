from __future__ import annotations
from pathlib import Path
from typing import Any

import openpyxl

from .config import SourceConfig


class ExtractorError(Exception):
    """Raised when a source file cannot be read as expected."""


class ExtractedSheet:
    """Raw extracted data from one sheet."""

    def __init__(self, headers: list[str], rows: list[dict[str, Any]]):
        self.headers = headers
        self.rows = rows

    def __len__(self) -> int:
        return len(self.rows)

    def __bool__(self) -> bool:
        return len(self.rows) > 0


def _is_summary_row(row: dict[str, Any], headers: list[str]) -> bool:
    """Detect if a row is a totals/summary row (not actual data).

    Strategy:
    1. If first S.no-type column contains "total" text → summary
    2. If first S.no-type column is empty/missing and there are
       numeric financial values → summary
    3. If first column is explicitly non-numeric → summary
    """
    if not headers:
        return False

    # Find S.no-type column index
    sno_idx = 0
    for i, h in enumerate(headers):
        h_clean = h.strip().lower()
        if any(kw in h_clean for kw in ("s.no", "sr no", "sl.no", "sl no", "#", "sno", "srno", "slno")):
            sno_idx = i
            break

    sno_val = row.get(headers[sno_idx])

    # Explicit "Total" keyword
    if sno_val is not None:
        s = str(sno_val).strip().lower()
        if s in ("total", "grand total", "total visit", "totals", "sub total"):
            return True

    if sno_val is None:
        # S.no column empty → check if this is a totals row by looking
        # for numeric values in financial columns while identity
        # columns are empty. Count non-empty identity columns vs total.
        identity_keywords = [
            "name", "code", "phone", "pan", "location", "state",
            "zone", "branch", "contact", "agency", "address",
            "status", "remark", "month", "billing",
        ]
        empty_id_cols = 0
        total_id_cols = 0
        has_financial_val = False

        for h in headers:
            if h == headers[sno_idx]:
                continue
            h_clean = h.strip().lower()
            v = row.get(h)
            is_id = any(kw in h_clean for kw in identity_keywords)

            if is_id:
                total_id_cols += 1
                if v is None or (isinstance(v, str) and not v.strip()):
                    empty_id_cols += 1
            elif not is_id and v is not None:
                if isinstance(v, (int, float)) and v > 0:
                    has_financial_val = True
                elif isinstance(v, str) and v.replace(",", "").replace(".", "").isdigit():
                    has_financial_val = True

        # If most identity columns are empty AND financial values exist
        if total_id_cols > 0:
            ratio = empty_id_cols / total_id_cols
            if ratio > 0.6 and has_financial_val:
                return True

        # Fallback: sno is None and no identity cols found
        if total_id_cols == 0 and has_financial_val:
            return True

        return False

    # Check if S.no value is numeric
    try:
        float(str(sno_val).replace(",", "").strip())
        return False  # Numeric → data row
    except ValueError:
        return True  # Non-numeric → summary row


def _find_sheet_name(
    filepath: Path,
    hint: str,
    must_find: bool = True,
) -> str | None:
    """Find actual sheet name matching the hint (fuzzy)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        all_names = wb.sheetnames
        hint_lower = hint.lower().strip()

        # Exact
        if hint in all_names:
            return hint

        # Case-insensitive
        for name in all_names:
            if name.lower().strip() == hint_lower:
                return name

        # Partial
        for name in all_names:
            if hint_lower in name.lower() or name.lower() in hint_lower:
                return name

        if must_find:
            raise ExtractorError(
                f"Sheet matching '{hint}' not found in {filepath.name}. "
                f"Available: {all_names}"
            )
        return None
    finally:
        wb.close()


class SourceExtractor:
    """Universal extractor that reads any source file."""

    def __init__(self, filepath: Path, config: SourceConfig):
        self.filepath = filepath
        self.config = config

    def get_hidden_sheets(self) -> list[str]:
        hidden = []
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        try:
            for name in wb.sheetnames:
                if wb[name].sheet_state != "visible":
                    hidden.append(name)
        finally:
            wb.close()
        return hidden

    def extract_sheet(
        self,
        sheet_name_hint: str,
        header_row: int = 2,
        data_start_row: int = 3,
    ) -> ExtractedSheet:
        ws_name = _find_sheet_name(self.filepath, sheet_name_hint)
        if not ws_name:
            return ExtractedSheet([], [])

        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        try:
            ws = wb[ws_name]

            # Read header row
            raw_headers: list[str | None] = []
            for c in range(1, ws.max_column + 1):
                raw_headers.append(ws.cell(header_row, c).value)

            headers: list[str] = [
                str(h).strip() if h is not None else "" for h in raw_headers
            ]

            # De-duplicate header names (suffix duplicates with _2, _3, etc.)
            header_count: dict[str, int] = {}
            deduped_headers: list[str] = []
            for h in headers:
                if h:
                    if h in header_count:
                        header_count[h] += 1
                        deduped_headers.append(f"{h}___{header_count[h]}")
                    else:
                        header_count[h] = 1
                        deduped_headers.append(h)
                else:
                    deduped_headers.append(h)
            headers = deduped_headers

            # Trim trailing empty headers
            last_col = 0
            for c, h in enumerate(headers):
                if h:
                    last_col = c + 1
            headers = headers[:last_col]

            # Read data rows, filtering summary rows
            data_rows: list[dict[str, Any]] = []
            for r in range(data_start_row, ws.max_row + 1):
                row: dict[str, Any] = {}
                all_empty = True
                for c in range(1, last_col + 1):
                    val = ws.cell(r, c).value
                    if val is not None:
                        all_empty = False
                    col_name = headers[c - 1] if c - 1 < len(headers) else ""
                    row[col_name] = val

                if all_empty:
                    continue

                if _is_summary_row(row, headers):
                    continue

                data_rows.append(row)

            return ExtractedSheet(headers, data_rows)
        finally:
            wb.close()
