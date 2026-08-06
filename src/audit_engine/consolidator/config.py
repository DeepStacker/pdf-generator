from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Source config: no per-source column rules, just identity
# ---------------------------------------------------------------------------

@dataclass
class SourceConfig:
    file_pattern: str
    client_name: str
    pt_sheet_hint: str = "Payment Tracker"
    md_sheet_hint: str = "Master Data"
    header_row: int = 1
    data_start_row: int = 2
    expected_pt_rows: tuple[int, int] | None = None
    expected_md_rows: tuple[int, int] | None = None
    needs_manual_enrichment: dict[str, str] = field(default_factory=dict)
    exclude_md_names: list[str] = field(default_factory=list)
    column_overrides: dict[str, str] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Canonical schemas
# ---------------------------------------------------------------------------

PT_COLUMNS = [
    "S.no", "client", "Assayer Name", "Assayer Code", "Assayer Phone",
    "Location", "State", "Zone", "Audit Month & Year", "Type of Audit",
    "No. of Visits", "Base Audit Fee", "Total pay (Base)",
    " Travel charges(If any)", "Cancelled visits", "Branch Cancellation Charges",
    " Andaman & Nicobar Branch Expenses", "Error Deduction", "Total pay",
    "Remarks (if any)", "PAN Number", "Bank Name", "A/c Number", "IFSC Code",
]

MD_COLUMNS = [
    "Sr No", "Client", "Month", "Zone", "SOL ID", "BRANCH", "Location ", "State",
    "Total No.of A/cs", "Assayer Name", "Assayer Code", "AssayerPhone",
    "Assayer PAN", "Contact Person", "Schedule date", "Audit \nStatus",
    "Audit \ncompletion date", "No of days \naudited ",
    "No of days \naudited For client", "No of Packets \naudited",
    "Additional Packet", "Assayer Reporting time at Branch", "Audit start time",
    "Audit End Time", "Client fee", "Additional", "Final Client Fees",
    "Assayer fee", "Additional fee", "Distance", "Base Location", "Remarks",
    "Assayer fee", "Additional fee", "Cancelled", "Error Deduciton", "Total",
    "Audit Remarks", "Seeding Status", "Report",
    "Total pouches suggested for audit", "Already Audited", "A/C Closed",
    "A/C Auctioned", "Packet Missing",
    "Actual Audited (except already audited & A/C closed) ",
    "Extra audited pouches", "Total No.of packets actually audited",
    "Type oof Audit", "Aadhar No", "Remarks", "Urban/ Rural",
    "T & F Client fee", "POA Client Fees", "Cancelled\nClient fee",
    "billing  remarks", "Touch and Feel Audit Packet Count",
    "POA \nPacket count", "Additional Packet  T& F", "Additional Packet  POA",
    "oracle_id", "Address",
]


# ---------------------------------------------------------------------------
# Auto-generated canonical columns (not read from source headers)
# ---------------------------------------------------------------------------

AUTO_COLUMNS = {"S.no", "Sr No", "client", "Client", "Month"}


# ---------------------------------------------------------------------------
# Sensible defaults for every canonical column
# ---------------------------------------------------------------------------

DEFAULT_VALUES: dict[str, Any] = {
    # PT defaults
    "S.no": None,          # auto-increment
    "client": "",          # enriched from filename
    "Assayer Name": "",
    "Assayer Code": "",
    "Assayer Phone": "",
    "Location": "",
    "State": "",
    "Zone": "",
    "Audit Month & Year": "",
    "Type of Audit": "",
    "No. of Visits": 0,
    "Base Audit Fee": 0,
    "Total pay (Base)": 0,
    " Travel charges(If any)": 0,
    "Cancelled visits": 0,
    "Branch Cancellation Charges": 0,
    " Andaman & Nicobar Branch Expenses": 0,
    "Error Deduction": 0,
    "Total pay": 0,
    "Remarks (if any)": "",
    "PAN Number": "",
    "Bank Name": "",
    "A/c Number": "",
    "IFSC Code": "",
    # MD defaults
    "Sr No": None,         # auto-increment
    "Client": "",          # enriched from filename
    "Month": "",           # from context
    "Zone": "",
    "SOL ID": "",
    "BRANCH": "",
    "Location ": "",
    "State": "",
    "Total No.of A/cs": 1,
    "Assayer Name": "",
    "Assayer Code": "",
    "AssayerPhone": "",
    "Assayer PAN": "",
    "Contact Person": "",
    "Schedule date": "",
    "Audit \nStatus": "",
    "Audit \ncompletion date": "",
    "No of days \naudited ": 1,
    "No of days \naudited For client": 1,
    "No of Packets \naudited": 0,
    "Additional Packet": "",
    "Assayer Reporting time at Branch": "",
    "Audit start time": "",
    "Audit End Time": "",
    "Client fee": 0,
    "Additional": 0,
    "Final Client Fees": 0,
    "Assayer fee": 0,
    "Additional fee": 0,
    "Distance": "",
    "Base Location": "",
    "Remarks": "",
    "Cancelled": 0,
    "Error Deduciton": "",
    "Total": 0,
    "Audit Remarks": "",
    "Seeding Status": "",
    "Report": "",
    "Total pouches suggested for audit": 0,
    "Already Audited": 0,
    "A/C Closed": 0,
    "A/C Auctioned": 0,
    "Packet Missing": 0,
    "Actual Audited (except already audited & A/C closed) ": 0,
    "Extra audited pouches": 0,
    "Total No.of packets actually audited": 0,
    "Type oof Audit": "",
    "Aadhar No": "",
    "Urban/ Rural": "",
    "T & F Client fee": 0,
    "POA Client Fees": 0,
    "Cancelled\nClient fee": 0,
    "billing  remarks": "",
    "Touch and Feel Audit Packet Count": 0,
    "POA \nPacket count": 0,
    "Additional Packet  T& F": 0,
    "Additional Packet  POA": 0,
    "oracle_id": "",
    "Address": "",
}


# ---------------------------------------------------------------------------
# Global header pattern registry (canonical → list of source header substrings)
# Matching is case-insensitive; any pattern contained in the source header matches.
# ---------------------------------------------------------------------------

HEADER_PATTERNS: dict[str, list[str]] = {
    # Identity
    "Assayer Name": [
        "assayer name", "assayer's name", "name of assayer",
        "name of auditor/assayer", "name of the assayer",
        "name of auditor", "auditor name", "staff name",
        "assayer", "auditor",
    ],
    "Assayer Code": [
        "assayer code", "assayer code", "ass code", "asscode",
        "payee code", "assayer id", "assyrcode", "assy", "assayr",
        "employee code", "emp code", "staff code", "employee id",
    ],
    "Assayer Phone": [
        "assayer phone", "assayerphone", "payee phone", "phone",
        "mobile number", "mobile", "contact number",
    ],
    "Assayer PAN": [
        "assayer pan", "assayerpan", "payee pan", "pan number", "pan",
    ],
    "AssayerPhone": [
        "assayer phone", "assayerphone", "payee phone", "phone",
    ],
    "AssayerPAN": [
        "assayer pan", "assayerpan", "payee pan", "pan number", "pan",
    ],

    # Location hierarchy
    "Location": [
        "location", "branch location", "branch",
    ],
    "Location ": [
        "location", "branch name", "branchname",
    ],
    "State": ["state", "state name"],
    "Zone": ["zone", "zone name"],
    "BRANCH": [
        "branch code", "branch", "agency", "branch name", "agency name",
    ],

    # Dates & time
    "Audit Month & Year": [
        "audit month", "month", "month & year",
    ],
    "Schedule date": [
        "schedule date", "audit schedule date", "audit date",
        "audit conducted on",
    ],
    "Audit \nStatus": [
        "audit status", "status",
    ],
    "Audit \ncompletion date": [
        "completion date", "completion date of audit",
        "end date", "audit completion", "audit completion date",
        "audit conducted on", "audit date",
    ],
    "Assayer Reporting time at Branch": [
        "reporting time", "reporting time at branch",
    ],
    "Audit start time": [
        "start time", "audit start",
    ],
    "Audit End Time": [
        "end time", "audit end",
    ],

    # Contact
    "Contact Person": [
        "contact person", "contact person name",
        "person name", "person to contact",
        "contact", "process manager",
    ],

    # Financial (PT)
    "No. of Visits": [
        "no of visits", "no of visit", "no of audits done", "no of audits",
        "no. of visit", "no. of visits", "number of visits",
    ],
    "No of visits": [
        "no of visits", "no of visit", "no of audits done", "no of audits",
        "no. of visit", "no. of visits", "number of visits",
    ],
    "Base Audit Fee": [
        "base audit fee", "base audit fees", "base fee", "audit fee",
        "base",
    ],
    "Type of Audit": [
        "type of audit", "audit type",
    ],
    "Total pay (Base)": [
        "total base fee", "total pay (base)", "total base", "total pay",
    ],
    " Travel charges(If any)": [
        "travel charge", "travel",
    ],
    "Cancelled visits": [
        "cancelled", "missed",
    ],
    "Branch Cancellation Charges": [
        "cancellation", "branch cancellation", "audit cancellation",
    ],
    " Andaman & Nicobar Branch Expenses": [
        "andaman", "nicobar",
    ],
    "Error Deduction": [
        "error", "deduction", "error (deduct money)",
    ],
    "Total pay": [
        "total pay (final)", "total pay", "total",
    ],
    "Remarks (if any)": [
        "remark", "remark (if any)", "remarks",
    ],

    # Bank info
    "PAN Number": [
        "pan number", "pan", "pan no", "pan no.",
    ],
    "Bank Name": [
        "bank name", "bank",
    ],
    "A/c Number": [
        "account number", "a/c number", "ac number", "account no",
    ],
    "IFSC Code": [
        "ifsc", "ifsc code",
    ],

    # MD specifics
    "SOL ID": [
        "sol id", "sol", "solid",
    ],
    "Total No.of A/cs": [
        "no of audits", "total accounts", "total no of acs",
        "total accounts",
    ],
    "No of Packets \naudited": [
        "packets actually audited", "no of packets audited",
    ],
    "No of days \naudited ": [
        "no of days", "no of visit", "no. of visit",
        "days audited", "naudited", "audit days",
    ],
    "No of days \naudited For client": [
        "days audited for client", "naudited for client",
    ],
    "Additional Packet": [
        "additional packet", "extra packet",
    ],
    "Client fee": [
        "client fee", "client billing", "client fees",
        "client billing per audit",
    ],
    "Additional": [
        "additional", "additonal", "extra",
    ],
    "Final Client Fees": [
        "final client fees", "final client fee",
    ],
    "Assayer fee": [
        "assayer fee", "assayer fees", "auditor base fee", "base audit fees",
        "total fees", "auditor payment",
        "base audit fees in inr",
    ],
    "Additional fee": [
        "additional fee", "additional fees", "auditor travel fee",
        "additional amount", "additional amount in inr",
    ],
    "Distance": [
        "distance",
    ],
    "Base Location": [
        "base location", "assayer base location",
    ],
    "Remarks": [
        "remarks",
    ],
    "Cancelled": [
        "cancelled", "cancellation", "cancel",
    ],
    "Error Deduciton": [
        "error deduction", "error",
    ],
    "Total": [
        "total", "total fee", "total fees", "total fees in inr",
        "auditor payment", "payment",
    ],
    "Audit Remarks": [
        "audit remarks", "audit remark",
    ],
    "Seeding Status": ["seeding", "seeding status"],
    "Report": [],  # no broad pattern to avoid false matches with "reporting" headers
    "Total pouches suggested for audit": [
        "total pouches", "pouches suggested",
    ],
    "Already Audited": [
        "already audited",
    ],
    "A/C Closed": [
        "a/c closed", "account closed",
    ],
    "A/C Auctioned": [
        "a/c auctioned", "account auctioned",
    ],
    "Packet Missing": [
        "packet missing",
    ],
    "Actual Audited (except already audited & A/C closed) ": [
        "actual audited",
    ],
    "Extra audited pouches": [
        "extra audited",
    ],
    "Total No.of packets actually audited": [
        "total no.of packets actually audited",
        "total packets actually audited",
    ],
    "Type oof Audit": [
        "type of audit", "audit type",
    ],
    "Aadhar No": ["aadhar", "aadhar number", "aadhar no"],
    "Urban/ Rural": ["urban", "rural"],
    "T & F Client fee": ["touch and feel", "t & f client"],
    "POA Client Fees": ["poa client"],
    "Cancelled\nClient fee": ["cancelled client"],
    "billing  remarks": ["billing remarks"],
    "Touch and Feel Audit Packet Count": ["touch and feel audit"],
    "POA \nPacket count": ["poa packet"],
    "Additional Packet  T& F": ["additional packet t&f", "additional packet t & f"],
    "Additional Packet  POA": ["additional packet poa"],
    "oracle_id": ["oracle", "oracle id"],
    "Address": ["address"],

    # Generic SrNo patterns (matched last, or auto-increment)
    "S.no": ["s.no", "s no", "sl no", "sl no.", "sr no", "sr no.", "#", "sno"],
    "Sr No": ["sr no", "sr no.", "s no", "s no.", "sl no", "sl no.", "#", "srno", "sno"],
}



# ---------------------------------------------------------------------------
# Processing context
# ---------------------------------------------------------------------------

@dataclass
class ProcessingContext:
    audit_month: str = "Feb'26"

    def get(self, key: str, default: Any = None) -> Any:
        return getattr(self, key, default)


def _get_project_root() -> Path:
    """Return project root — works in dev mode and PyInstaller frozen mode."""
    import sys
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    return Path(__file__).parent.parent.parent.parent


def _resolve_settings_path() -> Path:
    """Find the settings file, preferring a writable user copy."""
    import shutil
    from pathlib import Path
    
    user_copy = Path.home() / ".audit_engine_elite" / "settings" / "mapping_rules.xlsx"
    if user_copy.exists():
        return user_copy
    
    bundled = _get_project_root() / "settings" / "mapping_rules.xlsx"
    if bundled.exists():
        user_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(bundled), str(user_copy))
        return user_copy
    
    return bundled


def load_external_header_patterns(default_patterns: dict[str, list[str]]) -> dict[str, list[str]]:
    import openpyxl
    import sys
    from pathlib import Path
    
    rules_path = _resolve_settings_path()
    if not rules_path.exists():
        return default_patterns
        
    try:
        wb = openpyxl.load_workbook(rules_path, data_only=True)
        if "Column_Mappings" not in wb.sheetnames:
            wb.close()
            return default_patterns
            
        ws = wb["Column_Mappings"]
        patterns: dict[str, list[str]] = {}
        
        # Find backup aliases column by header name (not assuming last column)
        backup_col_idx = None
        for c in range(2, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and "backup aliases" in str(h).strip().lower():
                backup_col_idx = c
                break
        
        for r in range(2, ws.max_row + 1):
            canonical = ws.cell(r, 1).value
            aliases_str = ws.cell(r, backup_col_idx).value if backup_col_idx else None
            if canonical:
                canonical = str(canonical).strip()
                if canonical not in patterns:
                    patterns[canonical] = []
                if aliases_str:
                    raw = str(aliases_str).strip()
                    if raw.lower() != "none":
                        parts = [p.strip() for p in raw.split(",") if p.strip()]
                        for p in parts:
                            if p not in patterns[canonical]:
                                patterns[canonical].append(p)
                            
        wb.close()
        
        for k, v in default_patterns.items():
            if k not in patterns or not patterns[k]:
                patterns[k] = v
                
        return patterns
    except Exception as e:
        print(f"Warning: Failed to load external header patterns: {e}. Using fallback default rules.", file=sys.stderr)
        return default_patterns

HEADER_PATTERNS = load_external_header_patterns(HEADER_PATTERNS)

