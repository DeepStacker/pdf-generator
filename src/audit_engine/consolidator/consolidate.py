import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import json
import io
import argparse
from difflib import SequenceMatcher
from collections import defaultdict

try:
    import rapidfuzz
except ImportError:
    rapidfuzz = None

BASE = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(BASE, "source_files") if os.path.isdir(os.path.join(BASE, "source_files")) else BASE
TEMPLATE_DIR = os.path.join(BASE, "templates") if os.path.isdir(os.path.join(BASE, "templates")) else BASE
OUTPUT_DIR = os.path.join(BASE, "output") if os.path.isdir(os.path.join(BASE, "output")) else BASE
STANDARD_FILE = os.path.join(TEMPLATE_DIR, "Feb'26 consolidated.xlsx")

# ---- CLIENT NAME MAPPING ----
# Build client map from actual filenames on disk
def norm(s):
    """Normalize string: collapse multiple spaces, strip, handle non-breaking space."""
    return ' '.join(s.replace('\xa0', ' ').replace('\u200b', '').split()).lower()


def build_client_map(base_dir):
    """Fuzzy-match filenames to handle non-breaking spaces and typos."""
    known_patterns = [
        ("Axis Bank POA", "Axis Bank POA"),
        ("Axis Bank Muthoot POA", "Axis Bank Muthoot POA"),
        ("Axis Bank Seeding Activity", "Axis Bank Seeding Activity"),
        ("Axis Collection Agency|Axis Bank Collectio Agency", "Axis Collection Agency"),
        ("Cholamandalam POA", "Cholamandalam POA"),
        ("EECPL POA|Enn Enn", "EECPL POA"),
        ("Equitas POA", "Equitas POA"),
        ("FFSL.*Augmont.*POA", "FFSL (Augmont) POA"),
        ("ICICI POA", "ICICI POA"),
        ("IDFC Bank.*Touch.*feel.*POA", "IDFC Bank T&F & POA"),
        ("IDFC IIFL Bank.*Touch.*feel.*POA", "IDFC IIFL T&F & POA"),
        ("IDFC MFL Bank.*Touch.*feel.*POA", "IDFC MFL T&F & POA"),
        ("L & T Finanace|L & T Collection", "L&T Collection Agency"),
        ("RBL.*Ashirwad", "RBL Ashirwad POA"),
        ("RBL.*Indel", "RBL Indel POA"),
        ("^RBL Payment|^RBL -|^RBL\\.", "RBL POA"),
        ("RBL.*IIFL", "RBL IIFL POA"),
        ("RBL.*Muthoot Fincorp|RBL.*Muthoot", "RBL Muthoot POA"),
        ("SCB Mar|^SCB_", "SCB"),
        ("SIB.*Seeding", "SIB Seeding Activity"),
        ("SarvaGram", "SarvaGram POA"),
        ("VISTAAR.*Indel", "VISTAAR Indel"),
        ("VISTAAR.*Manapurram", "VISTAAR Manapurram"),
        ("YES BANK AMC|Yes Bank Forex", "YES BANK AMC"),
        ("Yes Bank Touch and Feel", "Yes Bank T&F & POA"),
    ]

    # Map Type of Audit values to client names (for Payment Tracker Mar26)
    audit_to_client = {
        "AXIS Bank Muthoot POA": "Axis Bank Muthoot POA",
        "AXIS Bank POA": "Axis Bank POA",
        "Axis Bank Seeding Activity": "Axis Bank Seeding Activity",
        "Axis Bank Collectio Agency": "Axis Collection Agency",
        "Cholamandalam POA": "Cholamandalam POA",
        "EECPL POA": "EECPL POA",
        "Equitas POA": "Equitas POA",
        "FFSL(Augmont) POA": "FFSL (Augmont) POA",
        "ICICI POA": "ICICI POA",
        "IDFC Bank  Touch and feel & POA": "IDFC Bank T&F & POA",
        "IDFC IIFL Bank  Touch and feel & POA": "IDFC IIFL T&F & POA",
        "IDFC MFL Bank  Touch and feel & POA": "IDFC MFL T&F & POA",
        "L & T Collection Agency": "L&T Collection Agency",
        "RBL Ashirwad POA": "RBL Ashirwad POA",
        "RBL Indel POA": "RBL Indel POA",
        "RBL POA": "RBL POA",
        "RBL IIFL POA": "RBL IIFL POA",
        "RBL Muthoot Fincorp POA": "RBL Muthoot POA",
        "SarvaGram FinCare Ltd POA": "SarvaGram POA",
        "SCB Coll/Tele/Yard": "SCB",
        "SCB Legal": "SCB",
        "SIB  Seeding Activity": "SIB Seeding Activity",
        "Vistar-Indel-POA": "VISTAAR Indel",
        "Vistar-Manapurram-POA": "VISTAAR Manapurram",
        "Yes Bank Forex Audit": "YES BANK AMC",
        "Yes Bank Touch and Feel & POA Audit": "Yes Bank T&F & POA",
    }

    result = {}
    for fname in os.listdir(base_dir):
        if not fname.lower().endswith(('.xlsx', '.xls')):
            continue
        if fname in ("Feb'26 consolidated.xlsx", "Mar'26 consolidated.xlsx", "consolidate.py"):
            continue
        if fname == "Payment Tracker Mar26.xlsx":
            result[fname] = "__SPLIT_BY_AUDIT__"
            continue
        fname_norm = norm(fname)
        matched = False
        for pattern, client in known_patterns:
            if re.search(pattern, fname_norm, re.IGNORECASE):
                result[fname] = client
                matched = True
                break
        if not matched:
            result[fname] = fname_norm.split(' payment')[0].split(' -')[0][:30].title()
    result['__audit_to_client__'] = audit_to_client
    return result

CLIENT_MAP = build_client_map(SOURCE_DIR)
AUDIT_TO_CLIENT = CLIENT_MAP.pop('__audit_to_client__', {})

# Standard column definitions
PT_COLS = [
    'S.no', 'client', 'Assayer Name', 'Assayer Code', 'Assayer Phone',
    'Location', 'State', 'Zone', 'Audit Month & Year', 'Type of Audit',
    'No. of Visits', 'Base Audit Fee', 'Total pay (Base)',
    ' Travel charges(If any)', 'Cancelled visits', 'Branch Cancellation Charges',
    ' Andaman & Nicobar Branch Expenses', 'Error Deduction', 'Total pay',
    'Remarks (if any)', 'PAN Number', 'Bank Name', 'A/c Number', 'IFSC Code'
]

MD_COLS = [
    'Sr No', 'Client', 'Month', 'Zone', 'SOL ID', 'BRANCH', 'Location ', 'State',
    'Total No.of A/cs', 'Assayer Name', 'Assayer Code', 'AssayerPhone',
    'Assayer PAN', 'Contact Person', 'Schedule date',
    'Audit \nStatus', 'Audit \ncompletion date',
    'No of days \naudited ', 'No of days \naudited For client',
    'No of Packets \naudited', 'Additional Packet',
    'Assayer Reporting time at Branch', 'Audit start time', 'Audit End Time',
    'Client fee', 'Additional', 'Final Client Fees',
    'Assayer fee', 'Additional fee', 'Distance', 'Base Location', 'Remarks',
    'Assayer fee.1', 'Additional fee.1', 'Cancelled', 'Error Deduciton', 'Total',
    'Audit Remarks', 'Seeding Status', 'Report',
    'Total pouches suggested for audit', 'Already Audited', 'A/C Closed',
    'A/C Auctioned', 'Packet Missing',
    'Actual Audited (except already audited & A/C closed) ',
    'Extra audited pouches', 'Total No.of packets actually audited',
    'Type oof Audit', 'Aadhar No', 'Remarks.1', 'Urban/ Rural',
    'T & F Client fee', 'POA Client Fees', 'Cancelled\nClient fee',
    'billing  remarks', 'Touch and Feel Audit Packet Count',
    'POA \nPacket count', 'Additional Packet  T& F', 'Additional Packet  POA',
    'oracle_id', 'Address'
]

IGNORE_SHEETS = {'Summary', 'Auditor Details', 'Sheet6', 'Sheet7'}


from difflib import SequenceMatcher


def clean_col(col):
    return str(col).strip().replace('\n', ' ').replace('\xa0', ' ')


def col_match(a, b):
    return clean_col(a).lower() == clean_col(b).lower()


def find_col(cols, *candidates):
    for c in candidates:
        for col in cols:
            if col_match(col, c):
                return col
    return None


def simplify(s):
    """Remove all non-alphanumeric characters for matching."""
    return re.sub(r'[^a-z0-9]', '', s.lower())


def token_set(s):
    """Word-level token set for fuzzy matching."""
    return set(re.findall(r'[a-z0-9]+', s.lower()))


# ── Multi-strategy column matcher ────────────────────────────────────────────
# Strategies applied in order, each fallback fires only if the previous failed.
# STRATEGIES:
#   1. EXACT      — `clean_col(name).lower() == clean_col(source).lower()`
#   2. SIMPLIFIED — remove non-alphanumeric, compare
#   3. SUBSTRING  — `name` appears inside `source`
#   4. WORD_OVERLAP — ≥50% of tokens in `name` appear in `source`
#   5. FUZZY      — SequenceMatcher ratio ≥ threshold (default 0.75)
#   6. POSITION   — match by position relative to a known-anchored column
#   7. INFER      — infer column type from cell-content patterns (future)

MATCH_CACHE = {}  # (filename, tuple(source_cols), candidate) → match (per-sheet cache cleared per file)


def find_column(candidates, source_cols, name_hint=None,
                anchor_col=None, anchor_pos=None,
                sample_values=None,
                threshold=0.75):
    """
    Multi-strategy column matcher with confidence scoring, RapidFuzz fallback, 
    word-overlap safety rules, and positional validation.
    """
    cleaned = {c: clean_col(c).lower() for c in source_cols}
    simplified = {c: simplify(c) for c in source_cols}
    tokens = {c: token_set(c) for c in source_cols}

    if name_hint:
        cn = clean_col(name_hint).lower()
        if cn not in [clean_col(c).lower() for c in candidates]:
            candidates = [name_hint] + list(candidates)

    # Strategy 1: EXACT match for ALL candidates first (Confidence: 100)
    for candidate in candidates:
        cl = clean_col(candidate).lower()
        for c, cn in cleaned.items():
            if cn == cl:
                return c

    # Strategy 2: SIMPLIFIED match for ALL candidates (Confidence: 95)
    for candidate in candidates:
        sc = simplify(candidate)
        for c, sc_ in simplified.items():
            if sc_ == sc:
                return c

    # Strategy 3: SUBSTRING match for ALL candidates (Confidence: 80)
    for candidate in candidates:
        cl = clean_col(candidate).lower()
        sc = simplify(candidate)
        if not cl and not sc:
            continue
        for c, cn in cleaned.items():
            if (cl and cl in cn) or (sc and sc in cn):
                return c

    # Strategy 4: WORD OVERLAP for ALL candidates (Confidence: 65 - 75)
    # Exclude generic tokens from being sole overlap drivers
    GENERIC_TOKENS = {'date', 'fee', 'fees', 'type', 'no', 'number', 'code', 'name', 'total', 'status', 'charges', 'details'}
    for candidate in candidates:
        ct = token_set(candidate)
        if ct:
            for c, st in tokens.items():
                overlap = ct & st
                non_generic_overlap = overlap - GENERIC_TOKENS
                
                # Rule 1: Multi-word candidate requires at least 2 tokens overlapping OR 1 non-generic token
                if len(ct) >= 2:
                    if len(overlap) >= 2 or (len(overlap) == 1 and len(non_generic_overlap) == 1 and list(non_generic_overlap)[0] not in GENERIC_TOKENS):
                        if len(overlap) >= max(1, len(ct) * 0.5):
                            return c
                # Rule 2: Single-word candidate requires exact token match or non-generic match with single-word source
                elif len(ct) == 1:
                    token = list(ct)[0]
                    if token not in GENERIC_TOKENS and len(st) <= 2 and token in st:
                        return c

    # Strategy 5: FUZZY — SequenceMatcher / RapidFuzz ratio (Confidence: 70 - 90)
    for candidate in candidates:
        cl = clean_col(candidate).lower()
        if not cl:
            continue
        for c, cn in cleaned.items():
            if not cn:
                continue
            if cl[0] != cn[0]:  # first character must match
                continue
            
            if rapidfuzz:
                ratio = rapidfuzz.fuzz.ratio(cl, cn) / 100.0
            else:
                ratio = SequenceMatcher(None, cl, cn).ratio()
                
            if ratio >= threshold:
                return c

    # Strategy 6: POSITIONAL fallback with safety validation
    if anchor_col and anchor_pos is not None:
        known_col = None
        for c, cn in cleaned.items():
            if cn == clean_col(anchor_col).lower():
                known_col = c
                break
        if known_col:
            known_idx = source_cols.index(known_col)
            target_idx = known_idx + (anchor_pos or 0)
            if 0 <= target_idx < len(source_cols):
                guessed_col = source_cols[target_idx]
                # Positional Safety check: Ensure guessed column isn't obviously conflicting
                guessed_cn = cleaned[guessed_col]
                # If target is not completely empty, accept guess
                if guessed_cn:
                    return guessed_col

    # Strategy 7: CONTENT INFERENCE
    if sample_values:
        vals = [str(v) for v in sample_values if pd.notna(v) and str(v).strip() not in ('', 'nan', 'nat', 'none')]
        if len(vals) >= 2:
            pan_ratio = sum(1 for v in vals[:20] if re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', v.strip())) / max(1, len(vals[:20]))
            if pan_ratio >= 0.8:
                for c in source_cols:
                    cn = cleaned[c]
                    if 'pan' in cn or 'tax' in cn or 'assayer' in cn:
                        return c
            digits_only = [re.sub(r'\D', '', v) for v in vals]
            phone_ratio = sum(1 for d in digits_only[:20] if len(d) == 10 and d.isdigit()) / max(1, len(digits_only[:20]))
            if phone_ratio >= 0.8:
                for c in source_cols:
                    cn = cleaned[c]
                    if 'phone' in cn or 'mobile' in cn or 'contact' in cn or 'assayerphone' in cn:
                        return c
            ifsc_ratio = sum(1 for v in vals[:20] if re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', v.strip())) / max(1, len(vals[:20]))
            if ifsc_ratio >= 0.8:
                for c in source_cols:
                    cn = cleaned[c]
                    if 'ifsc' in cn:
                        return c
            states_set = {'andhra pradesh', 'arunachal pradesh', 'assam', 'bihar', 'chhattisgarh', 'goa', 'gujarat', 'haryana', 'himachal pradesh', 'jharkhand', 'karnataka', 'kerala', 'madhya pradesh', 'maharashtra', 'manipur', 'meghalaya', 'mizoram', 'nagaland', 'odisha', 'orissa', 'punjab', 'rajasthan', 'sikkim', 'tamil nadu', 'telangana', 'tripura', 'uttar pradesh', 'uttarakhand', 'west bengal', 'new delhi', 'delhi', 'chandigarh'}
            state_matches = sum(1 for v in vals[:20] if v.strip().lower().strip() in states_set)
            if state_matches >= max(3, len(vals[:20]) * 0.3):
                for c in source_cols:
                    cn = cleaned[c]
                    if 'state' in cn:
                        return c

    return None


class ColumnMappingManager:
    """Persists column mappings per file across runs for schema-drift detection."""

    def __init__(self, path=None):
        self.path = path or os.path.join(BASE, "column_mappings.json")
        self.data = self._load()
        self.changes = []

    def _load(self):
        if os.path.exists(self.path) and os.path.getsize(self.path) > 0:
            with open(self.path) as f:
                return json.load(f)
        return {}

    def _save(self):
        with open(self.path, 'w') as f:
            json.dump(self.data, f, indent=2)

    def get_name_hint(self, filename, template_col):
        mapping = self.data.get(filename, {})
        info = mapping.get(template_col, {})
        return info.get('source_col')

    def get_position(self, filename, template_col):
        mapping = self.data.get(filename, {})
        info = mapping.get(template_col, {})
        return info.get('position')

    def update_mapping(self, filename, template_col, source_col, method, position):
        if filename not in self.data:
            self.data[filename] = {}
        prev = self.data[filename].get(template_col, {})
        prev_source = prev.get('source_col')
        if prev_source and prev_source != source_col:
            self.changes.append(
                f"  {filename}: {template_col} changed "
                f"'{prev_source}' (pos {prev.get('position')}, {prev.get('method')}) → "
                f"'{source_col}' (pos {position}, {method})"
            )
        self.data[filename][template_col] = {
            'source_col': source_col,
            'method': method,
            'position': position,
        }

    def save(self):
        self._save()

    def print_changes(self):
        if self.changes:
            print(f"\n⚠️  Column mapping changes detected ({len(self.changes)}):")
            for c in self.changes:
                print(c)


def build_content_map(df, sample_n=30):
    """
    Analyze column values to infer types, returning a dict:
    {inferred_type: col_name, ...}
    Types: 'pan', 'phone', 'ifsc', 'state', 'zone', 'sol_id', 'date'
    """
    cols = list(df.columns)
    result = {}

    for col in cols:
        vals = [str(v) for v in df[col].dropna().head(sample_n) if str(v).strip() not in ('', 'nan', 'nat', 'none')]
        if len(vals) < 2:
            continue

        pan_matches = sum(1 for v in vals if re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', v.strip()))
        if pan_matches >= max(2, len(vals) * 0.5):
            if 'pan' not in result: result['pan'] = col
            continue

        ifsc_matches = sum(1 for v in vals if re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', v.strip()))
        if ifsc_matches >= max(2, len(vals) * 0.5):
            if 'ifsc' not in result: result['ifsc'] = col
            continue

        digits_only = [re.sub(r'\D', '', v) for v in vals]
        phone_matches = sum(1 for d in digits_only if len(d) == 10 and d.isdigit())
        if phone_matches >= max(2, len(vals) * 0.5):
            if 'phone' not in result: result['phone'] = col
            continue

        states_set = {'andhra pradesh', 'arunachal pradesh', 'assam', 'bihar', 'chhattisgarh',
                      'goa', 'gujarat', 'haryana', 'himachal pradesh', 'jharkhand', 'karnataka',
                      'kerala', 'madhya pradesh', 'maharashtra', 'manipur', 'meghalaya',
                      'mizoram', 'nagaland', 'odisha', 'orissa', 'punjab', 'rajasthan',
                      'sikkim', 'tamil nadu', 'telangana', 'tripura', 'uttar pradesh',
                      'uttarakhand', 'west bengal', 'new delhi', 'delhi', 'chandigarh'}
        state_matches = sum(1 for v in vals if v.strip().lower().strip() in states_set)
        if state_matches >= min(3, len(vals) * 0.3):
            if 'state' not in result: result['state'] = col
            continue

        zones_set = {'north', 'south', 'east', 'west', 'central', 'north east', 'north-east'}
        zone_matches = sum(1 for v in vals if v.strip().lower() in zones_set)
        if zone_matches >= min(2, len(vals) * 0.3):
            if 'zone' not in result: result['zone'] = col
            continue

        sol_matches = sum(1 for v in vals if re.match(r'^\d{4,6}$', v.strip()))
        if sol_matches >= min(3, len(vals) * 0.5):
            if 'sol_id' not in result: result['sol_id'] = col
            continue

        date_matches = 0
        for v in vals[:10]:
            try:
                pd.Timestamp(v)
                date_matches += 1
            except:
                pass
        if date_matches >= max(2, len(vals[:10]) * 0.5):
            if 'date' not in result: result['date'] = col
            continue

    return result


def sum_cols(row, *col_names):
    total = 0
    for c in col_names:
        try:
            v = float(row.get(c, 0) or 0)
            total += v
        except (ValueError, TypeError):
            pass
    return total


def safe_float(val):
    try:
        v = float(val)
        return 0 if pd.isna(v) else v
    except (ValueError, TypeError):
        return 0


def sstr(val):
    """Safely convert value to string, returning empty string for None/NaN."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if isinstance(val, (int, float)):
        # Return integers as-is, floats as rounded to drop .0
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def identify_sheets(xls):
    sheets = xls.sheet_names
    pt_sheet = None
    md_sheet = None
    extras = []

    for s in sheets:
        if s in IGNORE_SHEETS:
            extras.append(s)
            continue

        # Hard rule: "Auditor Details" is always extra
        if 'auditor' in s.lower() and 'detail' in s.lower():
            extras.append(s)
            continue

        df = pd.read_excel(xls, sheet_name=s, nrows=1)
        cols_clean = [clean_col(c).lower() for c in df.columns]

        has_total_pay = any('total pay' in c for c in cols_clean)
        has_pan = any(c in ('pan number', 'assayer pan', 'payee pan') for c in cols_clean)
        has_ifsc = any(c == 'ifsc code' or c == 'ifsc' for c in cols_clean)
        has_bank = any('bank name' in c or 'bank' in c for c in cols_clean)
        has_visits = any('visits' in c or 'audits done' in c for c in cols_clean)
        has_base_fee = any('base audit fee' in c or 'base audit fees' in c for c in cols_clean)
        has_assayer_name = any('assayer name' in c or 'auditor name' in c or 'name of auditor' in c for c in cols_clean)
        has_ac = any('a/c number' in c or 'account no' in c for c in cols_clean)

        has_branch = any('branch' in c for c in cols_clean)
        has_sol = any(c in ('sol id', 'sol_id', 'sole_id', 'sol id') for c in cols_clean)
        has_schedule = any('schedule' in c for c in cols_clean)
        has_audit_status = any('audit status' in c for c in cols_clean)
        has_client_fee = any('client fee' in c for c in cols_clean)
        has_zone = any('zone' in c for c in cols_clean)
        has_distance = any('distance' in c for c in cols_clean)
        has_base_loc = any('base location' in c for c in cols_clean)
        has_vendor = any('vendor' in c for c in cols_clean)
        has_pouches = any('pouch' in c or 'packet' in c for c in cols_clean)

        pt_score = sum([has_total_pay, has_pan, has_ifsc, has_bank, has_visits, has_base_fee, has_assayer_name, has_ac])
        md_score = sum([has_branch, has_sol, has_schedule, has_audit_status, has_client_fee, has_zone, has_distance, has_base_loc, has_vendor, has_pouches])

        # Content fingerprinting fallback when header heuristics are insufficient
        if pt_score == 0 and md_score == 0:
            try:
                df_sample = pd.read_excel(xls, sheet_name=s, nrows=15)
                cmap = build_content_map(df_sample)
                if 'pan' in cmap or 'ifsc' in cmap:
                    pt_score += 4
                if 'sol_id' in cmap:
                    md_score += 4
                if 'phone' in cmap:
                    pt_score += 1
                if 'zone' in cmap or 'state' in cmap:
                    md_score += 1
            except Exception:
                pass

        sn = s.lower()

        # Sheet name boosts
        if sn == 'payments':
            pt_score += 5
        elif 'payment' in sn:
            pt_score += 3
        if 'master' in sn or 'mastr' in sn:
            md_score += 3

        # Tiebreakers for hybrid sheets
        if has_vendor and has_client_fee and not has_ifsc and not has_ac:
            md_score += 2  # YES BANK AMC Master Data

        if pt_score >= md_score and pt_score >= 3:
            if pt_sheet is None or pt_score > 3:
                pt_sheet = s
        elif md_score > pt_score and md_score >= 2:
            if md_sheet is None or md_score > 2:
                md_sheet = s
        else:
            extras.append(s)

    if pt_sheet == md_sheet:
        if pt_sheet and md_sheet:
            extras.append(md_sheet)
            md_sheet = None

    if not pt_sheet and not md_sheet and len(sheets) == 1:
        pt_sheet = sheets[0]

    return pt_sheet, md_sheet, extras


def normalize_pt_sheet(df, client_name, filename, verbose=False, mapping_mgr=None):
    rows = []
    cols_orig = list(df.columns)
    cols_map = {clean_col(c).lower(): c for c in cols_orig}
    content_map = build_content_map(df)

    def g(*names):
        for n in names:
            if n in cols_map:
                return cols_map[n]
        return None

    def gc(*names, intent=None, name_hint=None, prev_position=None, label=None):
        gc._last_method = 'none'
        if name_hint is None and label and mapping_mgr is not None:
            name_hint = mapping_mgr.get_name_hint('PT:' + filename, label)
        if prev_position is None and label and mapping_mgr is not None:
            prev_position = mapping_mgr.get_position('PT:' + filename, label)
        result = find_column(list(names), cols_orig, threshold=0.72, name_hint=name_hint)
        if result:
            gc._last_method = 'multi'
            _log_map(label, result)
            return result
        if intent and intent in content_map:
            gc._last_method = 'content'
            _log_map(label, content_map[intent])
            return content_map[intent]
        if prev_position is not None and 0 <= prev_position < len(cols_orig):
            gc._last_method = 'positional'
            _log_map(label, cols_orig[prev_position])
            return cols_orig[prev_position]
        result = g(*names)
        if result:
            gc._last_method = 'exact'
        _log_map(label, result)
        return result

    def _log_map(label, col):
        if label and mapping_mgr is not None and filename != 'test_obfuscated_headers.xlsx':
            mapping_mgr.update_mapping(
                'PT:' + filename, label, col, gc._last_method or 'none',
                cols_orig.index(col) if col else None
            )

    if verbose:
        all_wanted = [
            ('assayer name', ['assayer name', 'name of auditor/assayer', 'name of auditor', 'auditor name', 'name of assayer the money will be credited to']),
            ('assayer code', ['assayer code', 'payee code', 'auditor code', 'auditor/assayer code']),
            ('assayer phone', ['assayer phone', 'assayerphone', 'payee phone', 'phone']),
            ('location', ['location']),
            ('state', ['state', 'states']),
            ('zone', ['zone', 'region']),
            ('audit month & year', ['audit month & year']),
            ('type of audit', ['type of audit']),
            ('no of visits', ['no of visits', 'no. of visits', 'no of audits done']),
            ('base audit fee', ['base audit fee', 'base audit fees']),
            ('total pay', ['total pay', 'total']),
            ('pan number', ['pan number', 'assayer pan', 'payee pan']),
            ('ifsc code', ['ifsc code', 'ifsc']),
        ]
        for label, candidates in all_wanted:
            match = find_column(candidates, cols_orig, threshold=0.72)
            if not match:
                print(f'    ⚠️  PT col not found: {label}')

    for idx, row in df.iterrows():
        sno_val = sstr(row.get(g('s.no', 'sl. no', 'sl no', 's.n.', '#', 'sr. no.', 'sl no', 'sr no', 's no')))
        if sno_val.lower() in ('total', 'grand total', '', 'nan', 'nat') or ('total' in sstr(row.iloc[0]).lower() and idx > 0):
            continue

        # Find columns (primary: gc = multi-strategy, secondary .1/.2: g = exact-only)
        col_assayer = gc('assayer name', 'name of auditor/assayer', 'name of auditor', 'auditor name', 'name of assayer the money will be credited to', label='Assayer Name')
        col_code = gc('assayer code', 'payee code', 'auditor code', 'auditor/assayer code', label='Assayer Code')
        col_phone = gc('assayer phone', 'assayerphone', 'payee phone', 'phone', intent='phone', label='Assayer Phone')
        col_location = gc('location', label='Location')
        col_state = gc('state', 'states', intent='state', label='State')
        col_zone = gc('zone', 'region', label='Zone')
        col_month = gc('audit month & year', label='Audit Month & Year')
        col_type = gc('type of audit', label='Type of Audit')
        col_visits = gc('no of visits', 'no. of visits', 'no of audits done', label='No. of Visits')
        col_base_fee = gc('base audit fee', 'base audit fees', 'base audit fee', label='Base Audit Fee')
        col_total_base = gc('total pay (base)', 'total base fee', 'total base', 'total pay (base)', label='Total pay (Base)')
        col_travel = gc('travel charges(if any)', 'travel charges  (if any)', 'travel charges(if any)', 'base travel', 'travel fee', label=' Travel charges(If any)')
        col_cancel_visits = gc('cancelled visits', 'audit cancellation fees', 'cancelled branch ', 'cancelled ', label='Cancelled visits')
        col_branch_cancel = gc('branch cancellation charges', label='Branch Cancellation Charges')
        col_andaman = gc(' andaman & nicobar branch expenses', 'andaman & nicobar branch expenses', label=' Andaman & Nicobar Branch Expenses')
        col_error = gc('error deduction', 'deduction', 'error (deduct money)', 'deduction ', 'deduction', label='Error Deduction')
        col_total_pay = gc('total pay', 'total', label='Total pay')
        col_remarks = gc('remarks (if any)', 'remarks', 'additional remarks fyr', label='Remarks (if any)')
        col_pan = gc('pan number', 'assayer pan', 'payee pan', intent='pan', label='PAN Number')
        col_bank = gc('bank name', label='Bank Name')
        col_ac = gc('a/c number', 'account no', 'a/c number', label='A/c Number')
        col_ifsc = gc('ifsc code', 'ifsc', intent='ifsc', label='IFSC Code')

        # Extract values safely
        def ssv(col):
            return sstr(row.get(col)) if col else ''

        assayer = ssv(col_assayer)
        code = ssv(col_code)
        phone = ssv(col_phone)
        location = ssv(col_location)
        state = ssv(col_state)
        zone = ssv(col_zone)
        month = row.get(col_month) if col_month else None
        if pd.notna(month):
            mstr = sstr(month)
            if re.match(r"^[A-Za-z]{3}'\d{2}$", mstr):
                month = mstr
            else:
                try:
                    month = pd.Timestamp(month).strftime("%b'%y")
                except:
                    month = mstr
        else:
            month = ''
        audit_type = ssv(col_type)

        # Visits: handle ICICI multi-visit columns
        visits = safe_float(row.get(col_visits)) if col_visits else 0
        col_v1 = g('no of visits.1')
        col_v2 = g('no of visits.2')
        col_v2b = g('no of visits 2nd branch in day')
        col_vnp = g('no of visits(no payment) 2nd branch')
        if col_v1: visits += safe_float(row.get(col_v1))
        if col_v2: visits += safe_float(row.get(col_v2))
        if col_v2b: visits += safe_float(row.get(col_v2b))
        if col_vnp: visits += safe_float(row.get(col_vnp))
        visits = round(visits, 0)

        base_fee = safe_float(row.get(col_base_fee)) if col_base_fee else 0

        total_base = safe_float(row.get(col_total_base)) if col_total_base else 0
        if not col_total_base:
            total_base = visits * base_fee if base_fee > 0 else 0
        col_tb1 = g('total pay (base).1')
        col_tb2 = g('total pay (base).2')
        if col_tb1: total_base += safe_float(row.get(col_tb1))
        if col_tb2: total_base += safe_float(row.get(col_tb2))

        travel = safe_float(row.get(col_travel)) if col_travel else 0
        col_accom = g('accommodation  charges (if any)', 'accommodation charges (if any)')
        if col_accom:
            travel += safe_float(row.get(col_accom))
        col_courier = g('additional cost (courier charges)', 'additional cost (if any)')
        if col_courier:
            travel += safe_float(row.get(col_courier))

        cancelled_visits = safe_float(row.get(col_cancel_visits)) if col_cancel_visits else 0
        for col_name in cols_orig:
            cn = clean_col(col_name).lower()
            if 'missed' in cn or 'cancell' in cn:
                if col_name not in (col_cancel_visits, col_branch_cancel):
                    cancelled_visits += safe_float(row.get(col_name))

        branch_cancel = safe_float(row.get(col_branch_cancel)) if col_branch_cancel else 0
        andaman = safe_float(row.get(col_andaman)) if col_andaman else 0
        error_ded = safe_float(row.get(col_error)) if col_error else 0
        col_err_rec = g("error recovery sep'2024")
        if col_err_rec:
            error_ded += safe_float(row.get(col_err_rec))

        total_pay = safe_float(row.get(col_total_pay)) if col_total_pay else 0
        if not col_total_pay or total_pay == 0:
            total_pay = total_base + travel + cancelled_visits + branch_cancel + andaman - error_ded

        remarks = ssv(col_remarks)
        pan = ssv(col_pan)
        bank = ssv(col_bank)
        ac = ssv(col_ac)
        ifsc = ssv(col_ifsc)

        # RBL(Muthoot) special: col labeled "Assayer PAN" may contain phone numbers
        if col_pan and clean_col(col_pan).lower() == 'assayer pan':
            val = sstr(row.get(col_pan))
            if val and not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', val.strip()):
                phone = val
                pan = ''
                for c in cols_orig:
                    cc = clean_col(c).lower()
                    if 'pan' in cc and c != col_pan:
                        pv = row.get(c)
                        if pd.notna(pv):
                            pan = sstr(pv)
                            break

        # Clean phone (strip .0)
        phone = phone.replace('.0', '') if phone else ''

        missing_flag = 'MANUAL_ENTRY_NEEDED' if not pan or not bank or not ifsc else ''

        r = {
            'S.no': len(rows) + 1,
            'client': client_name,
            'Assayer Name': assayer,
            'Assayer Code': code,
            'Assayer Phone': phone,
            'Location': location,
            'State': state,
            'Zone': zone,
            'Audit Month & Year': month,
            'Type of Audit': audit_type,
            'No. of Visits': visits,
            'Base Audit Fee': base_fee,
            'Total pay (Base)': round(total_base, 2),
            ' Travel charges(If any)': round(travel, 2),
            'Cancelled visits': round(cancelled_visits, 2),
            'Branch Cancellation Charges': round(branch_cancel, 2),
            ' Andaman & Nicobar Branch Expenses': round(andaman, 2),
            'Error Deduction': round(error_ded, 2),
            'Total pay': round(total_pay, 2),
            'Remarks (if any)': remarks,
            'PAN Number': pan,
            'Bank Name': bank,
            'A/c Number': ac,
            'IFSC Code': ifsc,
            '_missing_flag': missing_flag
        }
        rows.append(r)
    return rows


def normalize_md_sheet(df, client_name, filename, verbose=False, mapping_mgr=None):
    rows = []
    cols_orig = list(df.columns)
    cols_lower = {clean_col(c).lower(): c for c in cols_orig}
    content_map = build_content_map(df)

    def g(*names):
        for n in names:
            if n in cols_lower:
                return cols_lower[n]
        return None

    def v(*names):
        for n in names:
            c = g(n)
            if c:
                return c
        return None

    def sv(row, col_name, default=''):
        if col_name and col_name in row:
            v = row[col_name]
            if pd.notna(v):
                return str(v)
        return default

    def fv(row, col_name, default=0):
        if col_name and col_name in row:
            v = row[col_name]
            try:
                return float(v) if pd.notna(v) else default
            except (ValueError, TypeError):
                return default
        return default

    # Build lookup dict: norm(col) -> original col name
    cols_norm = {}
    for c in cols_orig:
        cn = clean_col(c)
        cols_norm[cn.lower()] = c
        simple = re.sub(r'[^a-z0-9]', '', cn.lower())
        cols_norm[simple] = c

    def v(*names):
        for n in names:
            n_clean = clean_col(n).lower()
            if n_clean in cols_norm:
                return cols_norm[n_clean]
            simple = re.sub(r'[^a-z0-9]', '', n_clean)
            if simple in cols_norm:
                return cols_norm[simple]
            for key, val in cols_norm.items():
                if n_clean in key or simple in key:
                    return val
        return None

    def vc(*names, intent=None, name_hint=None, prev_position=None, label=None):
        vc._last_method = 'none'
        if name_hint is None and label and mapping_mgr is not None:
            name_hint = mapping_mgr.get_name_hint('MD:' + filename, label)
        if prev_position is None and label and mapping_mgr is not None:
            prev_position = mapping_mgr.get_position('MD:' + filename, label)
        candidates = list(names)
        if name_hint and name_hint not in candidates:
            candidates.append(name_hint)
        result = find_column(candidates, cols_orig, threshold=0.80)
        if result:
            vc._last_method = 'multi'
            _log_map(label, result)
            return result
        if intent and intent in content_map:
            vc._last_method = 'content'
            _log_map(label, content_map[intent])
            return content_map[intent]
        if prev_position is not None and 0 <= prev_position < len(cols_orig):
            vc._last_method = 'positional'
            _log_map(label, cols_orig[prev_position])
            return cols_orig[prev_position]
        result = v(*names)
        if result:
            vc._last_method = 'exact'
        else:
            vc._last_method = 'none'
        _log_map(label, result)
        return result

    def _log_map(label, col):
        if label and mapping_mgr is not None and filename != 'test_obfuscated_headers.xlsx':
            mapping_mgr.update_mapping(
                'MD:' + filename, label, col, vc._last_method or 'none',
                cols_orig.index(col) if col else None
            )

    # ── Verbose audit ──────────────────────────────────────────────────
    if verbose:
        md_wanted = [
            ('s.no', ['s.no', 'sr no', 'sr no.', 'sr. no.', 's no', 'sl no', 's.n.', '#']),
            ('zone', ['zone', 'region']),
            ('sol id', ['sol id', 'sol_id', 'sole_id', 'branch code', 'br code']),
            ('branch', ['branch', 'branch name', 'branch_name', 'br. name', 'branchname']),
            ('location', ['location ', 'location', 'city', 'district']),
            ('state', ['state', 'states']),
            ('total no of acs', ['total no.of a/cs', 'total no.of acs', '# account', 'accounts list']),
            ('assayer name', ['assayer name', 'visiting super appraiser name', 'sumeru assayer name', 'auditor name', 'vendor name ']),
            ('assayer code', ['assayer code', 'assayer code.1', 'asasyer code', 'assayer .1']),
            ('assayerphone', ['assayerphone', 'assayer phone', 'payee phone']),
            ('assayer pan', ['assayer pan', 'payee pan']),
            ('contact person', ['contact person', 'conatct person', 'process manager', 'spoc 1']),
            ('schedule date', ['schedule date', 'scheduled date', 'schedule dates', 'audit schedule date', 'audit conducted on', 'audit start date']),
            ('audit status', ['audit \\nstatus', 'audit status', 'status of the activity']),
            ('audit completion date', ['audit \\ncompletion date', 'audit completion date', 'end date', 'activity end date', 'audit end date']),
            ('days audited', ['no of days \\naudited ', 'no of days audited', 'no of days  \\naudited']),
            ('client fee', ['client fee', 'client fees', 'client billing per audit in inr']),
            ('assayer fee', ['assayer fee', 'assayer fees', 'auditor payment ', 'sumeru auditor fee', 'base audit fees in inr', 'audit fee']),
            ('cancelled', ['cancelled', 'cancellation', 'cancelled fee', 'branch cancellation charged']),
            ('total', ['total', 'total assayer fee', 'final total assayer fee', 'assayer total fee', 'total auditor fee']),
        ]
        for label, candidates in md_wanted:
            match = vc(*candidates)
            if not match:
                print(f'    ⚠️  MD col not found: {label} ({candidates[0]})')

    c_sr = vc('s.no', 'sr no', 'sr no.', 'sr. no.', 's no', 'sl no', 's.n.', '#', label='Sr No')
    c_zone = vc('zone', intent='zone', label='Zone')
    c_sol = vc('sol id', 'sol_id', 'sole_id', 'branch code', 'br code', intent='sol_id', label='SOL ID')
    c_branch = vc('branch', 'branch name', 'branch_name', 'br. name', 'branchname', label='BRANCH')
    c_location = vc('location ', 'location', 'city', 'district', label='Location ')
    c_state = vc('state', 'states', intent='state', label='State')
    c_accts = vc('total no.of a/cs', '# account', 'accounts list', 'total no.of acs', label='Total No.of A/cs')
    c_assayer = vc('assayer name', 'visiting super appraiser name', 'sumeru assayer name', 'auditor name', 'vendor name ', label='Assayer Name')
    c_code = vc('assayer code', 'assayer code.1', 'asasyer code', 'assayer .1', label='Assayer Code')
    c_phone = vc('assayerphone', 'assayer phone', 'payee phone', intent='phone', label='AssayerPhone')
    c_pan = vc('assayer pan', 'payee pan', intent='pan', label='Assayer PAN')
    c_contact = vc('contact person', 'conatct person', 'process manager', 'spoc 1', label='Contact Person')
    c_schedule = vc('schedule date', 'scheduled date', 'schedule dates', 'scheduled dates', 'audit schedule date', 'audit conducted on', 'audit start date', label='Schedule date')
    c_status = vc('audit \nstatus', 'audit status', 'status of the activity', label='Audit \nStatus')
    c_end = vc('audit \ncompletion date', 'audit completion date', 'audit completed date', 'end date', 'activity end date', 'audit end date', label='Audit \ncompletion date')
    c_days = vc('no of days \naudited ', 'no of days \naudited', 'no of days audited', 'no of days  \naudited', 'no of days \n audited', label='No of days \naudited ')
    c_days_client = vc('no of days \naudited for client', 'no of days \naudited for client', label='No of days \naudited For client')
    c_packets = vc('no of packets \naudited', 'no of packets \n audited', 'no of packets audited', 'account audited', label='No of Packets \naudited')
    c_addl_pkt = vc('additional packet', 'additonal packet', 'additional', label='Additional Packet')
    c_reporting = vc('assayer reporting time at branch', label='Assayer Reporting time at Branch')
    c_start_time = vc('audit start time', 'audit statrt time', label='Audit start time')
    c_end_time = vc('audit end time', 'audit end \n time', label='Audit End Time')
    c_client_fee = vc('client fee', 'client fees', 'client billing per audit in inr', label='Client fee')
    c_additional = vc('additional', 'additonal', 'additional amount', 'additional amount in inr', 'additonal travelling(approved amount)', label='Additional')
    c_final_client = vc('final client fees', 'total client fee', 'total client fees', 'total payable\n amount', 'total fees in inr', 'client billing', 'final amount', label='Final Client Fees')
    c_assayer_fee = vc('assayer fee', 'assayer fees', 'auditor payment ', 'sumeru auditor fee', 'base audit fees in inr', 'audit fee', label='Assayer fee')
    c_addl_fee = vc('additional fee', 'additional fees', 'additional amount in inr', label='Additional fee')
    c_distance = vc('distance ', 'distance', 'distance to audit location from auditor place', label='Distance')
    c_base_loc = vc('base location', 'assayer base location', 'auditor base location', label='Base Location')
    c_remarks = vc('remarks', 'remarks.1', 'audit remarks', 'other remarks', label='Remarks')
    c_cancelled = vc('cancellation', 'cancelled', 'cancelled fee', 'branch cancellation charged', 'cancelled branches travelling amount', label='Cancelled')
    c_cancelled_client = vc('cancelled fee', 'cancelled', 'cancellation', label='Cancelled\nClient fee')
    c_total = vc('total', 'total assayer fee', 'final total assayer fee', 'assayer total fee', 'total auditor fee', 'final assayer fee', label='Total')

    # Packet analysis
    c_pouches = vc('total pouches suggested for audit', 'total packets suggested for audit', label='Total pouches suggested for audit')
    c_already = vc('already audited', label='Already Audited')
    c_ac_closed = vc('a/c closed', 'loan closing charge', label='A/C Closed')
    c_ac_auctioned = vc('a/c auctioned', label='A/C Auctioned')
    c_pkt_missing = vc('packet missing', label='Packet Missing')
    c_actual_audited = vc('actual audited (except already audited & a/c closed) ', 'actual audited', label='Actual Audited (except already audited & A/C closed) ')
    c_extra = vc('extra audited pouches', 'extra audited', label='Extra audited pouches')
    c_total_packets = vc('total no.of packets actually audited', 'total no.of packets', label='Total No.of packets actually audited')

    c_audit_remarks = vc('audit remarks', 'cancelled remarks', label='Audit Remarks')
    c_seeding = vc('seeding status', label='Seeding Status')
    c_report = vc('report', label='Report')
    c_urban = vc('urban/ rural', label='Urban/ Rural')
    c_tf_client = vc('t & f client fee', label='T & F Client fee')
    c_poa_client = vc('poa client fees', label='POA Client Fees')
    c_tf_packets = vc('touch and feel audit packet count', 't & f count', label='Touch and Feel Audit Packet Count')
    c_poa_packets = vc('poa count', 'poa ', 'total no.of poa audited', 'poa \npacket count', label='POA \nPacket count')
    c_poa = vc('poa', 'poa ', label='POA')  # alt resolution
    c_addl_tf = vc('additional packet  t& f', 'additional packet  t&f', label='Additional Packet  T& F')
    c_addl_poa = vc('additional packet  poa', label='Additional Packet  POA')
    c_address = vc('address', label='Address')

    c_visit_count = vc('no. of visit', 'no of visit', 'no of visits', 'no. of visit', label='No. of Visit')
    # Auxiliary column lookups (move outside row loop — use vc for fallback)
    c_fee1 = vc('assayer fee.1', 'final assayer fees', 'assayer fees.1', 'final assayer fee', 'assayer fee final', label='Assayer fee.1')
    c_addl_fee1 = vc('additional fee.1', 'final additional fees', 'additional fees.1', 'additional fee final', label='Additional fee.1')
    c_error = vc('error deduction', 'error deduciton', label='Error Deduciton')
    # If no explicit Cancelled column, try inferring from Audit Status
    infer_cancel = c_cancelled is None and c_status is not None

    for idx, row in df.iterrows():
        sr_val = sv(row, c_sr, '')
        if str(sr_val).lower().strip() in ('total', 'grand total') or (idx > 0 and 'total' in str(row.iloc[0]).lower()):
            continue
        if c_sr is not None and str(sr_val).lower().strip() == '':
            continue

        # Core identity
        assayer_name = sv(row, c_assayer)
        zone = sv(row, c_zone)
        branch = sv(row, c_branch)
        state = sv(row, c_state)

        schedule = row.get(c_schedule) if c_schedule else pd.NaT
        if pd.notna(schedule):
            schedule = str(schedule)
        else:
            schedule = ''

        status_val = sv(row, c_status)

        end_date = row.get(c_end) if c_end else pd.NaT
        if pd.notna(end_date):
            end_date = str(end_date)
        else:
            end_date = ''

        days = fv(row, c_days)
        days_client = fv(row, c_days_client)
        packets = fv(row, c_packets)
        addl_pkt = fv(row, c_addl_pkt)
        client_fee = fv(row, c_client_fee)
        additional = fv(row, c_additional)
        final_client = fv(row, c_final_client)
        assayer_fee = fv(row, c_assayer_fee)
        addl_fee = fv(row, c_addl_fee)
        distance = sv(row, c_distance)
        base_loc = sv(row, c_base_loc)
        remarks = sv(row, c_remarks)

        # Compute fee.1 = assayer_fee * visit_count or use source if available
        visit_count = fv(row, c_visit_count)

        if c_fee1:
            fee1 = fv(row, c_fee1)
        else:
            fee1 = assayer_fee * visit_count if visit_count > 0 else assayer_fee

        if c_addl_fee1:
            addl_fee1 = fv(row, c_addl_fee1)
        else:
            addl_fee1 = addl_fee * visit_count if visit_count > 0 else addl_fee

        if c_cancelled:
            cancelled = fv(row, c_cancelled)
        elif infer_cancel:
            st = status_val.lower()
            if 'cancel' in st:
                cancelled = assayer_fee if assayer_fee > 0 else (visit_count * fv(row, c_client_fee) if c_client_fee else 0)
            else:
                cancelled = 0
        else:
            cancelled = 0

        # Client-level cancellation (separate column if available)
        if c_cancelled_client and c_cancelled_client != c_cancelled:
            cancelled_client = fv(row, c_cancelled_client)
        else:
            cancelled_client = cancelled

        total_row = fv(row, c_total)
        error_ded = fv(row, c_error)

        # Packet analysis
        pouches = fv(row, c_pouches)
        already = fv(row, c_already)
        ac_closed = fv(row, c_ac_closed)
        ac_auctioned = fv(row, c_ac_auctioned)
        pkt_missing = fv(row, c_pkt_missing)
        actual_audited = fv(row, c_actual_audited)
        extra_pouches = fv(row, c_extra)
        total_audited = fv(row, c_total_packets)

        # Build month from filename or source
        month = client_name

        # Get month from Month column if exists
        c_month = vc('month', 'audit month & year', 'month')
        if c_month:
            mval = row.get(c_month)
            if pd.notna(mval):
                mstr = sv(row, c_month, '')
                if re.match(r"^[A-Za-z]{3}'\d{2}$", mstr):
                    month = mstr
                else:
                    try:
                        month = pd.Timestamp(mval).strftime("%b'%y")
                    except:
                        month = mstr[:10]
            else:
                month = "Mar'26"
        else:
            month = "Mar'26"

        sr_num = fv(row, c_sr) if c_sr else 0

        r = {
            'Sr No': len(rows) + 1,
            'Client': client_name,
            'Month': month,
            'Zone': zone,
            'SOL ID': sv(row, c_sol),
            'BRANCH': branch,
            'Location ': sv(row, c_location),
            'State': state,
            'Total No.of A/cs': packets if not c_accts else fv(row, c_accts),
            'Assayer Name': assayer_name,
            'Assayer Code': sv(row, c_code),
            'AssayerPhone': sv(row, c_phone),
            'Assayer PAN': sv(row, c_pan),
            'Contact Person': sv(row, c_contact),
            'Schedule date': schedule,
            'Audit \nStatus': status_val,
            'Audit \ncompletion date': end_date,
            'No of days \naudited ': days,
            'No of days \naudited For client': days_client,
            'No of Packets \naudited': packets,
            'Additional Packet': addl_pkt,
            'Assayer Reporting time at Branch': sv(row, c_reporting),
            'Audit start time': sv(row, c_start_time),
            'Audit End Time': sv(row, c_end_time),
            'Client fee': client_fee,
            'Additional': additional,
            'Final Client Fees': final_client,
            'Assayer fee': assayer_fee,
            'Additional fee': addl_fee,
            'Distance': distance,
            'Base Location': base_loc,
            'Remarks': remarks,
            'Assayer fee.1': round(fee1, 2),
            'Additional fee.1': round(addl_fee1, 2),
            'Cancelled': cancelled,
            'Error Deduciton': error_ded,
            'Total': round(total_row, 2),
            'Audit Remarks': sv(row, c_audit_remarks),
            'Seeding Status': sv(row, c_seeding),
            'Report': sv(row, c_report),
            'Total pouches suggested for audit': pouches,
            'Already Audited': already,
            'A/C Closed': ac_closed,
            'A/C Auctioned': ac_auctioned,
            'Packet Missing': pkt_missing,
            'Actual Audited (except already audited & A/C closed) ': actual_audited,
            'Extra audited pouches': extra_pouches,
            'Total No.of packets actually audited': total_audited,
            'Type oof Audit': '',
            'Aadhar No': '',
            'Remarks.1': '',
            'Urban/ Rural': sv(row, c_urban),
            'T & F Client fee': fv(row, c_tf_client),
            'POA Client Fees': fv(row, c_poa_client),
            'Cancelled\nClient fee': cancelled_client,
            'billing  remarks': '',
            'Touch and Feel Audit Packet Count': fv(row, c_tf_packets),
            'POA \nPacket count': fv(row, c_poa_packets) if c_poa_packets else fv(row, c_poa),
            'Additional Packet  T& F': fv(row, c_addl_tf),
            'Additional Packet  POA': fv(row, c_addl_poa),
            'oracle_id': '',
            'Address': sv(row, c_address),
        }
        rows.append(r)
    return rows


def process_payment_tracker_mar26(fname, fpath):
    """Special handling: Payment Tracker Mar26.xlsx has all clients in one file, no PAN/Bank."""
    rows = []
    df = pd.read_excel(fpath, sheet_name="Sheet2")
    # Drop all-NaN rows and Grand Total rows
    df = df.dropna(how='all').reset_index(drop=True)
    df = df[df['Type of Audit'].notna() & (df['Type of Audit'].astype(str).str.strip() != '')]
    for audit_type, group in df.groupby('Type of Audit'):
        client_name = AUDIT_TO_CLIENT.get(str(audit_type).strip(), str(audit_type).strip())
        r = normalize_pt_sheet(group, client_name, f"{fname}[{audit_type}]")
        for row in r:
            row['_missing_flag'] = 'MANUAL_ENTRY_NEEDED'
        rows.extend(r)
    return rows


def consolidate_in_memory(files_dict, save_mapping=False):
    """
    Process multiple Excel workbooks purely in-memory.
    
    Parameters
    ----------
    files_dict : dict
        Dict mapping filename string to bytes stream (io.BytesIO) or file path.
    save_mapping : bool
        Whether to persist column mapping JSON updates.
        
    Returns
    -------
    tuple: (excel_bytes: bytes, summary: dict)
    """
    all_pt_rows = []
    all_md_rows = []
    flags = []
    file_summaries = []
    mapping_mgr = ColumnMappingManager()

    known_patterns = [
        ("Axis Bank POA", "Axis Bank POA"),
        ("Axis Bank Muthoot POA", "Axis Bank Muthoot POA"),
        ("Axis Bank Seeding Activity", "Axis Bank Seeding Activity"),
        ("Axis Collection Agency|Axis Bank Collectio Agency", "Axis Collection Agency"),
        ("Cholamandalam POA", "Cholamandalam POA"),
        ("EECPL POA|Enn Enn", "EECPL POA"),
        ("Equitas POA", "Equitas POA"),
        ("FFSL.*Augmont.*POA", "FFSL (Augmont) POA"),
        ("ICICI POA", "ICICI POA"),
        ("IDFC Bank.*Touch.*feel.*POA", "IDFC Bank T&F & POA"),
        ("IDFC IIFL Bank.*Touch.*feel.*POA", "IDFC IIFL T&F & POA"),
        ("IDFC MFL Bank.*Touch.*feel.*POA", "IDFC MFL T&F & POA"),
        ("L & T Finanace|L & T Collection", "L&T Collection Agency"),
        ("RBL.*Ashirwad", "RBL Ashirwad POA"),
        ("RBL.*Indel", "RBL Indel POA"),
        ("^RBL Payment|^RBL -|^RBL\\.", "RBL POA"),
        ("RBL.*IIFL", "RBL IIFL POA"),
        ("RBL.*Muthoot Fincorp|RBL.*Muthoot", "RBL Muthoot POA"),
        ("SCB Mar|^SCB_", "SCB"),
        ("SIB.*Seeding", "SIB Seeding Activity"),
        ("SarvaGram", "SarvaGram POA"),
        ("VISTAAR.*Indel", "VISTAAR Indel"),
        ("VISTAAR.*Manapurram", "VISTAAR Manapurram"),
        ("YES BANK AMC|Yes Bank Forex", "YES BANK AMC"),
        ("Yes Bank Touch and Feel", "Yes Bank T&F & POA"),
    ]

    for fname, source in files_dict.items():
        if fname == "Payment Tracker Mar26.xlsx":
            file_summaries.append({"filename": fname, "client": "META_FILE", "pt_rows": 0, "md_rows": 0, "status": "SKIPPED"})
            continue

        fname_norm = norm(fname)
        client_name = None
        for pattern, client in known_patterns:
            if re.search(pattern, fname_norm, re.IGNORECASE):
                client_name = client
                break
        if not client_name:
            client_name = fname_norm.split(' payment')[0].split(' -')[0][:30].title()

        try:
            xls = pd.ExcelFile(source)
            pt_sheet, md_sheet, extras = identify_sheets(xls)

            pt_count = 0
            md_count = 0
            if pt_sheet:
                df_pt = pd.read_excel(xls, sheet_name=pt_sheet)
                pt_rows = normalize_pt_sheet(df_pt, client_name, fname, mapping_mgr=mapping_mgr)
                all_pt_rows.extend(pt_rows)
                pt_count = len(pt_rows)
            else:
                flags.append(f"{fname}: No PT sheet identified")

            if md_sheet:
                df_md = pd.read_excel(xls, sheet_name=md_sheet)
                md_rows = normalize_md_sheet(df_md, client_name, fname, mapping_mgr=mapping_mgr)
                all_md_rows.extend(md_rows)
                md_count = len(md_rows)
            else:
                flags.append(f"{fname}: No MD sheet identified")

            file_summaries.append({
                "filename": fname,
                "client": client_name,
                "pt_rows": pt_count,
                "md_rows": md_count,
                "status": "SUCCESS"
            })

        except Exception as e:
            flags.append(f"{fname}: {str(e)[:80]}")
            file_summaries.append({
                "filename": fname,
                "client": client_name,
                "pt_rows": 0,
                "md_rows": 0,
                "status": f"ERROR: {str(e)[:50]}"
            })

    pt_df = pd.DataFrame(all_pt_rows)
    for c in PT_COLS:
        if c not in pt_df.columns:
            pt_df[c] = ''
    pt_df = pt_df[PT_COLS]

    md_df = pd.DataFrame(all_md_rows)
    for c in MD_COLS:
        if c not in md_df.columns:
            md_df[c] = ''
    md_df = md_df[MD_COLS]

    for df_ in [pt_df, md_df]:
        for c in df_.columns:
            if df_[c].dtype == object:
                df_[c] = df_[c].fillna('').apply(
                    lambda v: '' if (isinstance(v, float) and pd.isna(v)) or v is None or str(v).lower() in ('nan', 'nat', 'none') else v
                )

    def norm_zone(z):
        if z and isinstance(z, str) and z.strip():
            return z.strip().title()
        return ''

    if 'Zone' in md_df.columns:
        md_df['Zone'] = md_df['Zone'].apply(norm_zone)

    pt_no_zone = pt_df[pt_df['Zone'].isin(['', None])]['client'].unique()
    md_zone_map = md_df[md_df['Zone'].notna() & (md_df['Zone'] != '')].groupby('Client')['Zone'].first().to_dict()
    for cl in pt_no_zone:
        if cl in md_zone_map:
            pt_df.loc[pt_df['client'] == cl, 'Zone'] = md_zone_map[cl]

    if 'Zone' in pt_df.columns:
        pt_df['Zone'] = pt_df['Zone'].apply(norm_zone)

    pt_df = pt_df.sort_values(['client', 'Assayer Name']).reset_index(drop=True)
    md_df = md_df.sort_values(['Client', 'Assayer Name']).reset_index(drop=True)

    pt_df['S.no'] = range(1, len(pt_df) + 1)
    md_df['Sr No'] = range(1, len(md_df) + 1)

    output_buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def write_formatted(sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        headers = list(df.columns)

        hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin_border

        for r_idx, row in df.iterrows():
            for c_idx, h in enumerate(headers):
                cell = ws.cell(r_idx + 2, c_idx + 1, row.get(h))
                cell.alignment = data_align
                cell.border = thin_border
                cell.font = Font(size=10)
                if r_idx % 2 == 1:
                    cell.fill = alt_fill

        col_widths = {
            'S.no': 6, 'client': 22, 'Sr No': 6, 'Client': 22,
            'Assayer Name': 24, 'Assayer Code': 12, 'Assayer Phone': 15,
            'Location': 16, 'State': 14, 'Zone': 10, 'Audit Month & Year': 14,
            'Type of Audit': 22, 'BRANCH': 20, 'SOL ID': 8, 'Total pay': 12,
            'Total pay (Base)': 14, ' Travel charges(If any)': 12,
            'Cancelled visits': 12, 'Branch Cancellation Charges': 12,
            ' Andaman & Nicobar Branch Expenses': 12, 'Error Deduction': 12,
            'PAN Number': 16, 'Bank Name': 20, 'A/c Number': 18, 'IFSC Code': 16,
        }
        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 14)

        ws.freeze_panes = 'A2'

    write_formatted('Payment Tracker', pt_df)
    write_formatted('Master Data', md_df)
    wb.save(output_buffer)
    output_buffer.seek(0)

    if save_mapping:
        mapping_mgr.save()

    total_pay = float(pt_df['Total pay'].sum())
    summary = {
        "total_pay": total_pay,
        "pt_rows": len(pt_df),
        "md_rows": len(md_df),
        "pt_clients": sorted([str(c) for c in pt_df['client'].dropna().unique()]),
        "md_clients": sorted([str(c) for c in md_df['Client'].dropna().unique()]),
        "file_summaries": file_summaries,
        "flags": flags,
    }

    return output_buffer.getvalue(), summary


def main(dry_run=False, exclude_tests=False):
    all_pt_rows = []
    all_md_rows = []
    flags = []

    files = sorted([f for f in os.listdir(SOURCE_DIR) if f.endswith(('.xlsx', '.XLSX', '.xls')) 
                    and not f.startswith('.')
                    and f not in ("Feb'26 consolidated.xlsx", "Mar'26 consolidated.xlsx", "consolidate.py")])

    if exclude_tests:
        files = [f for f in files if not f.startswith('test_')]

    print(f"Processing {len(files)} files...\n")

    mapping_mgr = ColumnMappingManager()

    for fname in files:
        fpath = os.path.join(SOURCE_DIR, fname)
        client_name = CLIENT_MAP.get(fname, fname.split(' Payment')[0].split(' -')[0])
        print(f"[{files.index(fname)+1}/{len(files)}] {fname[:55]:55s} → ", end="")

        if fname == "Payment Tracker Mar26.xlsx":
            print("SKIPPED (meta-file, all clients have source files)")
            continue

        print(f"{client_name}")

        try:
            xls = pd.ExcelFile(fpath)
            pt_sheet, md_sheet, extras = identify_sheets(xls)

            if pt_sheet:
                df_pt = pd.read_excel(xls, sheet_name=pt_sheet)
                pt_rows = normalize_pt_sheet(df_pt, client_name, fname, mapping_mgr=mapping_mgr)
                all_pt_rows.extend(pt_rows)
                print(f"          PT: {pt_sheet} ({len(pt_rows)} rows)")
            else:
                print(f"          PT: NOT FOUND ❌")
                flags.append(f"{fname}: No PT sheet identified")

            if md_sheet:
                df_md = pd.read_excel(xls, sheet_name=md_sheet)
                md_rows = normalize_md_sheet(df_md, client_name, fname, mapping_mgr=mapping_mgr)
                all_md_rows.extend(md_rows)
                print(f"          MD: {md_sheet} ({len(md_rows)} rows)")
            else:
                print(f"          MD: NOT FOUND ❌")
                flags.append(f"{fname}: No MD sheet identified")

            if extras:
                print(f"          Skipped: {extras}")

        except Exception as e:
            print(f"          ERROR: {e}")
            flags.append(f"{fname}: {str(e)[:80]}")

    pt_df = pd.DataFrame(all_pt_rows)

    for c in PT_COLS:
        if c not in pt_df.columns:
            pt_df[c] = ''
    pt_df = pt_df[PT_COLS]

    md_df = pd.DataFrame(all_md_rows)
    for c in MD_COLS:
        if c not in md_df.columns:
            md_df[c] = ''
    md_df = md_df[MD_COLS]

    for df_ in [pt_df, md_df]:
        for c in df_.columns:
            if df_[c].dtype == object:
                df_[c] = df_[c].fillna('').apply(
                    lambda v: '' if (isinstance(v, float) and pd.isna(v)) or v is None or str(v).lower() in ('nan', 'nat', 'none') else v
                )

    def norm_zone(z):
        if z and isinstance(z, str) and z.strip():
            return z.strip().title()
        return ''

    if 'Zone' in md_df.columns:
        md_df['Zone'] = md_df['Zone'].apply(norm_zone)

    pt_no_zone = pt_df[pt_df['Zone'].isin(['', None])]['client'].unique()
    md_zone_map = md_df[md_df['Zone'].notna() & (md_df['Zone'] != '')].groupby('Client')['Zone'].first().to_dict()
    for cl in pt_no_zone:
        if cl in md_zone_map:
            pt_df.loc[pt_df['client'] == cl, 'Zone'] = md_zone_map[cl]

    if 'Zone' in pt_df.columns:
        pt_df['Zone'] = pt_df['Zone'].apply(norm_zone)

    pt_df = pt_df.sort_values(['client', 'Assayer Name']).reset_index(drop=True)
    md_df = md_df.sort_values(['Client', 'Assayer Name']).reset_index(drop=True)

    pt_df['S.no'] = range(1, len(pt_df) + 1)
    md_df['Sr No'] = range(1, len(md_df) + 1)

    output_path = os.path.join(OUTPUT_DIR, "Mar'26 consolidated.xlsx")

    def write_formatted(sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        headers = list(df.columns)

        hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin_border

        for r_idx, row in df.iterrows():
            for c_idx, h in enumerate(headers):
                cell = ws.cell(r_idx + 2, c_idx + 1, row.get(h))
                cell.alignment = data_align
                cell.border = thin_border
                cell.font = Font(size=10)
                if r_idx % 2 == 1:
                    cell.fill = alt_fill

        col_widths = {
            'S.no': 6, 'client': 22, 'Sr No': 6, 'Client': 22,
            'Assayer Name': 24, 'Assayer Code': 12, 'Assayer Phone': 15,
            'Location': 16, 'State': 14, 'Zone': 10, 'Audit Month & Year': 14,
            'Type of Audit': 22, 'BRANCH': 20, 'SOL ID': 8, 'Total pay': 12,
            'Total pay (Base)': 14, ' Travel charges(If any)': 12,
            'Cancelled visits': 12, 'Branch Cancellation Charges': 12,
            ' Andaman & Nicobar Branch Expenses': 12, 'Error Deduction': 12,
            'PAN Number': 16, 'Bank Name': 20, 'A/c Number': 18, 'IFSC Code': 16,
        }
        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = col_widths.get(h, 14)

        ws.freeze_panes = 'A2'

    if not dry_run:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        write_formatted('Payment Tracker', pt_df)
        write_formatted('Master Data', md_df)
        wb.save(output_path)
        print(f"\n{'='*60}")
        print(f"OUTPUT SAVED: {output_path}")
        print(f"{'='*60}")
    else:
        print(f"\n{'='*60}")
        print(f"DRY RUN COMPLETE — NO FILE SAVED")
        print(f"{'='*60}")

    print(f"Payment Tracker: {len(pt_df)} rows, {pt_df['client'].nunique()} clients")
    print(f"Master Data:     {len(md_df)} rows, {md_df['Client'].nunique()} clients")
    print(f"\nClients in PT: {sorted(pt_df['client'].dropna().unique())}")
    print(f"\nTotal pay sum: {pt_df['Total pay'].sum():,.2f}")

    if flags:
        print(f"\n⚠️  FLAGS ({len(flags)}):")
        for f in flags:
            print(f"  - {f}")

    if not dry_run:
        mapping_mgr.save()
    mapping_mgr.print_changes()

    return pt_df, md_df


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Consolidate Mar'26 Payment Tracker and Master Data workbooks.")
    parser.add_argument('--dry-run', action='store_true', help="Run consolidation without saving Excel file or mapping changes.")
    parser.add_argument('--exclude-tests', action='store_true', help="Exclude test_*.xlsx files from processing.")
    args = parser.parse_args()

    pt_df, md_df = main(dry_run=args.dry_run, exclude_tests=args.exclude_tests)
