"""Arvog Bank PDF generation service.

The ArvogService class provides a clean API. Module-level legacy functions
delegate to a default singleton so existing callers work unchanged.
"""

import argparse
import html
import logging
import os
from collections.abc import Callable

import openpyxl
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

_MAX_JEWELLERY_COLUMNS = 20
logger = logging.getLogger(__name__)


# =========================================================
# CONSTANTS
# =========================================================

REQUIRED_COLUMNS = [
    "SR No.",
    "Branch",
    "Customer Name",
    "Loan No.",
    "Jewellery No.",
    "Jewellery1",
    "Gross Wt.",
    "Stone Wt.",
    "Net Wt.",
    "Karat",
    "Packets Number",
]


# =========================================================
# ArvogService class
# =========================================================

class ArvogService:
    """Service for Arvog Bank PDF generation from Excel data."""

    name = "Arvog Bank"
    REQUIRED_COLUMNS = list(REQUIRED_COLUMNS)

    # --- Static helpers -----------------------------------------------

    @staticmethod
    def normalize_columns(cols):
        return [str(c).strip().lower() for c in cols]

    @staticmethod
    def format_value(value):
        if pd.isna(value):
            return ""
        if isinstance(value, (int, float)):
            if float(value).is_integer():
                return str(int(value))
            return str(value)
        val_str = str(value).strip()
        try:
            if val_str.endswith(".0"):
                fval = float(val_str)
                return str(int(fval))
        except Exception:
            pass
        return val_str

    @staticmethod
    def make_cell(value, style):
        text = ArvogService.format_value(value)
        if not text:
            return ""
        escaped = html.escape(text)
        formatted = escaped.replace("\n", "<br/>")
        return Paragraph(formatted, style)

    # --- Detection ----------------------------------------------------

    def detect_raw_excel(self, excel_path: str) -> tuple:
        try:
            xls = pd.ExcelFile(excel_path)
            for sheet_name in xls.sheet_names:
                df_temp = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=3, header=None)
                for r_idx in range(len(df_temp)):
                    row_vals = [str(x).strip().lower() for x in df_temp.iloc[r_idx] if pd.notna(x)]
                    if "jewellery1" in row_vals and "jewellery2" in row_vals:
                        return sheet_name, r_idx
        except Exception:
            pass
        return None, None

    def detect_valid_sheet(self, excel_path: str) -> str:
        xls = pd.ExcelFile(excel_path)
        required_normalized = [c.lower() for c in self.REQUIRED_COLUMNS]

        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                cols = self.normalize_columns(df.columns)
                if all(col in cols for col in required_normalized):
                    return sheet_name
            except Exception:
                continue

        raise Exception("No valid sheet found.")

    # --- Conversion ---------------------------------------------------

    def convert_raw_to_tall(
        self,
        excel_path: str,
        sheet_name: str,
        header_row: int,
        cluster_manager: str | None = None,
        log_func: Callable = logger.info,
    ) -> pd.DataFrame:
        df_raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=header_row)
        df_raw.columns = [str(c).strip() for c in df_raw.columns]

        manual_ref_dir = "manual_converted"
        base_name = os.path.basename(excel_path)
        manual_path = os.path.join(manual_ref_dir, base_name)

        manual_headers = None
        if os.path.exists(manual_path):
            try:
                df_man = pd.read_excel(manual_path, nrows=1)
                manual_headers = [str(c).strip() for c in df_man.columns]
                log_func(f"Aligning schema with manual reference sheet: {manual_path}")
            except Exception:
                pass

        if manual_headers is not None:
            tall_cols = manual_headers
        else:
            raw_cols_set = {str(c).strip() for c in df_raw.columns}

            beg_group = ["SR No.", "Branch", "Region", "Date", "Customer Name", "Loan No.", "Loan type", "Partner Name", "Customer ID"]
            beg_present = [c for c in beg_group if c in raw_cols_set]

            end_group = ["Loan Amount", "Auditor Name", "Appraiser Name", "Packets Location", "Cluster Manager", "Packets Number", "After audit packet number"]
            end_present = [c for c in end_group if c in raw_cols_set]

            jewellery_cols = [
                "Jewellery No.", "Jewellery1", "Gross Wt.", "Stone Wt.", "Net Wt.",
                "Karat", "Purity % %", "Net weight after Purity %"
            ]
            tall_cols = beg_present + jewellery_cols + end_present

        converted_rows = []

        for idx, row in df_raw.iterrows():
            if pd.isna(row.get("Loan No.")) and pd.isna(row.get("SR No.")):
                continue

            first_ornament_for_loan = True

            for i in range(1, 21):
                j_col = f"Jewellery{i}"
                if j_col not in df_raw.columns:
                    continue

                ornament_name = row[j_col]
                if pd.isna(ornament_name):
                    continue
                ornament_name_str = str(ornament_name).strip()
                if ornament_name_str in ("-", ""):
                    continue

                if i == 1:
                    gw_col = "Gross Wt."
                    sw_col = "Stone Wt."
                    nw_col = "Net Wt."
                    k_col = "Karat"
                    pur_col = "Purity % %"
                    nwp_col = "Net weight after Purity %"
                else:
                    gw_col = f"Gross Wt..{i-1}"
                    sw_col = f"Stone Wt..{i-1}"
                    nw_col = f"Net Wt..{i-1}"
                    k_col = f"Karat.{i-1}"
                    pur_col = f"Purity % %.{i-1}"
                    nwp_col = "Final net weight after purity %" if i == _MAX_JEWELLERY_COLUMNS else f"Net weight after Purity %.{i - 1}"

                gross_wt = row.get(gw_col) if gw_col in df_raw.columns else None
                stone_wt = row.get(sw_col) if sw_col in df_raw.columns else None
                net_wt = row.get(nw_col) if nw_col in df_raw.columns else None
                karat = row.get(k_col) if k_col in df_raw.columns else None
                purity = row.get(pur_col) if pur_col in df_raw.columns else None
                net_wt_purity = row.get(nwp_col) if nwp_col in df_raw.columns else None

                if isinstance(karat, str):
                    karat = karat.strip()

                new_row = {}

                if first_ornament_for_loan:
                    for col in tall_cols:
                        if col not in ["Jewellery No.", "Jewellery1", "Gross Wt.", "Stone Wt.", "Net Wt.", "Karat", "Purity % %", "Net weight after Purity %"]:
                            new_row[col] = row.get(col)

                    if "Cluster Manager" in tall_cols and pd.isna(new_row.get("Cluster Manager")):
                        region_val = str(row.get("Region", "")).strip().upper()
                        if cluster_manager:
                            cm_val = cluster_manager
                        elif region_val == "WARANGAL":
                            cm_val = "D.Praveen Kumar"
                        else:
                            cm_val = None
                        new_row["Cluster Manager"] = cm_val

                    first_ornament_for_loan = False
                else:
                    for col in tall_cols:
                        if col not in ["Jewellery No.", "Jewellery1", "Gross Wt.", "Stone Wt.", "Net Wt.", "Karat", "Purity % %", "Net weight after Purity %"]:
                            new_row[col] = None

                new_row["Jewellery No."] = f"Jewellery{i}"
                new_row["Jewellery1"] = ornament_name_str
                new_row["Gross Wt."] = gross_wt
                new_row["Stone Wt."] = stone_wt
                new_row["Net Wt."] = net_wt
                new_row["Karat"] = karat
                new_row["Purity % %"] = purity
                new_row["Net weight after Purity %"] = net_wt_purity

                ordered_row = {col: new_row.get(col) for col in tall_cols}
                converted_rows.append(ordered_row)

        return pd.DataFrame(converted_rows)

    # --- Formatting ---------------------------------------------------

    def apply_excel_formatting(self, excel_path: str) -> None:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for col_idx, h in enumerate(headers):
            if h == "Date":
                fmt = "d-mmm-yy"
            elif h in ("Gross Wt.", "Stone Wt.", "Net Wt.", "Net weight after Purity %"):
                fmt = "0.00"
            elif h == "Purity % %":
                fmt = "0%"
            elif h in ("Jewellery No.", "Packets Number"):
                fmt = "General"
            elif h == "After audit packet number":
                fmt = "d-mmm-yy"
            elif h in (
                "SR No.", "Branch", "Region", "Customer Name", "Loan No.",
                "Partner Name", "Customer ID", "Jewellery1", "Karat",
                "Auditor Name", "Appraiser Name", "Packets Location",
                "Cluster Manager", "Loan Amount",
            ):
                fmt = "0"
            else:
                fmt = "General"

            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx + 1)
                cell.number_format = fmt

        wb.save(excel_path)

    # --- Data cleaning ------------------------------------------------

    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")

        missing = [c for c in self.REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            raise Exception(f"Missing columns: {missing}")

        return df

    # --- PDF generation -----------------------------------------------

    def generate_pdf(self, branch_name: str, df: pd.DataFrame, output_path: str) -> None:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(A3),
            leftMargin=85.95,
            rightMargin=52.5,
            topMargin=53.78,
            bottomMargin=100.0,
            title=os.path.basename(output_path),
            author="Audit Automation",
        )

        body_style = ParagraphStyle(
            'BodyStyle',
            fontName='Times-Bold',
            fontSize=8.19,
            leading=9.5,
            alignment=TA_CENTER,
            textColor=colors.black,
        )

        table_data = []

        table_data.append([
            f"Branch Name - {branch_name}",
            "", "", "", "", "", "", "", "", "",
            "Sumeru",
            "", "", "", "", ""
        ])

        table_data.append([
            "SR\nNo.",
            "Customer Name",
            "Loan No.",
            "PACKET\nNUMBER",
            "Ornament\nNo.",
            "Ornament\nName",
            "Gross\nWt.",
            "Stone\nWt.",
            "Karat",
            "Net Wt.",
            "Count of\nOrnamen",
            "Gross Wt.",
            "Stone Wt",
            "Net Weight",
            "Karat",
            "Remarks",
        ])

        for _, row in df.iterrows():
            table_data.append([
                self.make_cell(row.get("SR No."), body_style),
                self.make_cell(row.get("Customer Name"), body_style),
                self.make_cell(row.get("Loan No."), body_style),
                self.make_cell(row.get("Packets Number"), body_style),
                self.make_cell(row.get("Jewellery No."), body_style),
                self.make_cell(row.get("Jewellery1"), body_style),
                self.make_cell(row.get("Gross Wt."), body_style),
                self.make_cell(row.get("Stone Wt."), body_style),
                self.make_cell(row.get("Karat"), body_style),
                self.make_cell(row.get("Net Wt."), body_style),
                "", "", "", "", "", ""
            ])

        row_heights = [24.75, 24.75] + [38.7] * len(df)

        col_widths = [
            24.75, 116.55, 57.60, 59.40, 49.95, 57.60,
            35.10, 36.00, 45.90, 33.30, 40.50,
            77.40, 77.40, 77.40, 77.40, 185.85,
        ]

        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights, repeatRows=2)

        style = TableStyle([
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 8.78),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("BACKGROUND", (10, 0), (15, 0), colors.yellow),
            ("BACKGROUND", (10, 1), (15, 1), colors.yellow),
            ("SPAN", (0, 0), (9, 0)),
            ("SPAN", (10, 0), (15, 0)),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])

        table.setStyle(style)
        doc.build([table])

    # --- Main processing pipeline -------------------------------------

    def process_excel(
        self,
        input_excel: str,
        output_dir: str,
        cluster_manager: str | None = None,
        convert_only: bool = False,
        log_func: Callable = logger.info,
    ) -> None:
        os.makedirs(output_dir, exist_ok=True)

        log_func(f"Reading Excel: {input_excel}")

        sheet_name, header_row = self.detect_raw_excel(input_excel)
        if sheet_name is not None:
            log_func(f"Detected Raw Wide-Format Bank Excel in sheet '{sheet_name}' (row {header_row}). Converting...")
            df = self.convert_raw_to_tall(
                input_excel, sheet_name, header_row,
                cluster_manager=cluster_manager, log_func=log_func,
            )
            base_name = os.path.basename(input_excel)
            converted_excel_path = os.path.join(output_dir, base_name)
            log_func(f"Saving converted Excel to: {converted_excel_path}")
            df.to_excel(converted_excel_path, index=False)
            self.apply_excel_formatting(converted_excel_path)

            if convert_only:
                log_func("Conversion completed successfully. Exiting.")
                return
        else:
            sheet_name = self.detect_valid_sheet(input_excel)
            log_func(f"Detected Converted Tall-Format Excel Sheet: {sheet_name}")
            df = pd.read_excel(input_excel, sheet_name=sheet_name)

        df = self.clean_dataframe(df)
        df["Branch"] = df["Branch"].ffill()
        grouped = df.groupby("Branch", sort=False)

        total = 0
        for branch, grp_df in grouped:
            branch_df = grp_df.reset_index(drop=True)
            branch_name = self.format_value(branch)
            safe_branch = branch_name.replace("/", "_")
            output_file = os.path.join(output_dir, f"{safe_branch}.pdf")
            log_func(f"Generating -> {output_file}")
            self.generate_pdf(branch_name=branch_name, df=branch_df, output_path=output_file)
            total += 1

        log_func(f"Done. Generated {total} PDFs.")
        log_func(f"Output Folder: {output_dir}")


# =========================================================
# Backward-compatible module-level wrappers
# =========================================================

_default_service = ArvogService()


def normalize_columns(cols):
    return ArvogService.normalize_columns(cols)


def format_value(value):
    return ArvogService.format_value(value)


def make_cell(value, style):
    return ArvogService.make_cell(value, style)


def detect_raw_excel(excel_path):
    return _default_service.detect_raw_excel(excel_path)


def detect_valid_sheet(excel_path):
    return _default_service.detect_valid_sheet(excel_path)


def convert_raw_to_tall(excel_path, sheet_name, header_row, cluster_manager=None, log_func=print):
    return _default_service.convert_raw_to_tall(
        excel_path, sheet_name, header_row,
        cluster_manager=cluster_manager, log_func=log_func,
    )


def apply_excel_formatting(excel_path):
    return _default_service.apply_excel_formatting(excel_path)


def clean_dataframe(df):
    return _default_service.clean_dataframe(df)


def generate_pdf(branch_name, df, output_path):
    return _default_service.generate_pdf(branch_name, df, output_path)


def process_excel(input_excel, output_dir, cluster_manager=None, convert_only=False, log_func=print):
    return _default_service.process_excel(
        input_excel, output_dir,
        cluster_manager=cluster_manager, convert_only=convert_only, log_func=log_func,
    )


# =========================================================
# CLI
# =========================================================

def parse_args():
    parser = argparse.ArgumentParser(description="Generate branch-wise PDFs from Excel")
    parser.add_argument("input_file", help="Path to Excel file")
    parser.add_argument("--output-dir", default="generated_pdfs", help="Output folder")
    parser.add_argument("--cluster-manager", default=None, help="Cluster Manager name to populate")
    parser.add_argument("--convert-only", action="store_true", help="Only convert raw wide-format Excel to tall-format and do not generate PDFs")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    process_excel(
        input_excel=args.input_file,
        output_dir=args.output_dir,
        cluster_manager=args.cluster_manager,
        convert_only=args.convert_only,
    )
