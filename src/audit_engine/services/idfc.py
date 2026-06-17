"""IDFC PDF generation service.

Encapsulates Excel validation, data extraction, branch grouping, and
PDF report generation for IDFC audit worksheets. No module-level state.
"""

import logging
import math
import os
import sys
from typing import Any

import openpyxl
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from audit_engine.exceptions import ValidationError

_HEADER_BYTES = 4

logger = logging.getLogger("audit_engine.idfc")


class IDFCService:
    """Generates IDFC bank audit PDF reports from Excel master data."""

    # ------------------------------------------------------------------
    # Page layout
    # ------------------------------------------------------------------

    PAGE_SIZE = landscape(A4)

    PAGE_MARGIN_LEFT = 50.2
    PAGE_MARGIN_RIGHT = 50.2
    PAGE_MARGIN_TOP = 48.0
    PAGE_MARGIN_BOTTOM = 15.0

    # Column widths
    COL_WIDTHS = [
        22.2,
        101.1,
        118.5,
        60.3,
        67.2,
        125.0,
        247.3,
    ]

    # Row heights
    HEADER_ROW_1_HEIGHT = 12.2
    HEADER_ROW_2_HEIGHT = 14.2
    COLUMN_HEADER_HEIGHT = 22.5
    DATA_ROW_HEIGHT = 30.5

    # Colors
    COLOR_HEADER_BANK = colors.HexColor("#FFFF00")
    COLOR_HEADER_AUDIT = colors.HexColor("#4985E8")
    COLOR_BORDER = colors.black
    BORDER_WIDTH = 0.5

    REQUIRED_COLUMNS = [
        "Prospectno",
        "CUID",
        "Tare Weight",
        "State",
        "CurrentBranch",
        "CurrentBranchName",
    ]

    def __init__(self, fonts_dir: str | None = None) -> None:
        self._fonts_dir = fonts_dir or self._resolve_fonts_dir()

        self._font_reg = "Helvetica"
        self._font_bld = "Helvetica-Bold"

        self._register_fonts()

    # ------------------------------------------------------------------
    # Resource resolution
    # ------------------------------------------------------------------

    @staticmethod
    def _resolve_fonts_dir() -> str:
        base = getattr(sys, "_MEIPASS", None) or os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        return os.path.join(base, "fonts")

    @staticmethod
    def get_resource_path(relative_path: str) -> str:
        """Resolve a resource path for dev and PyInstaller."""

        base = getattr(sys, "_MEIPASS", None) or os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

        return os.path.join(base, relative_path)

    # ------------------------------------------------------------------
    # Font registration
    # ------------------------------------------------------------------

    def _register_fonts(self) -> None:
        reg_path = os.path.join(self._fonts_dir, "Carlito-Regular.ttf")
        bld_path = os.path.join(self._fonts_dir, "Arimo-Bold.ttf")

        try:
            if os.path.exists(reg_path) and os.path.exists(bld_path):
                pdfmetrics.registerFont(TTFont("Carlito", reg_path))
                pdfmetrics.registerFont(TTFont("ArimoBold", bld_path))

                self._font_reg = "Carlito"
                self._font_bld = "ArimoBold"

        except (OSError, ValueError) as exc:
            logger.warning(
                "Font registration failed, using Helvetica fallback: %s",
                exc,
            )

    @property
    def font_reg(self) -> str:
        return self._font_reg

    @property
    def font_bld(self) -> str:
        return self._font_bld

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    @staticmethod
    def validate(excel_path: str | None) -> tuple[bool, str]:

        if not excel_path or not excel_path.strip():
            return False, "No file path provided."

        if not os.path.exists(excel_path):
            return False, f"File not found: {excel_path}"

        if os.path.getsize(excel_path) == 0:
            return False, "File is empty."

        ext = os.path.splitext(excel_path)[1].lower()

        if ext not in (".xlsx", ".xls"):
            return (
                False,
                f"Invalid file type '{ext}'. Expected .xlsx or .xls",
            )

        try:
            with open(excel_path, "rb") as f:
                if len(f.read(_HEADER_BYTES)) < _HEADER_BYTES:
                    return False, "File is empty or corrupt."

        except PermissionError:
            return (
                False,
                "File is locked or permission denied. Close it in Excel first.",
            )

        except OSError as exc:
            return False, f"Cannot read file: {exc}"

        return True, ""

    @staticmethod
    def validate_output_dir(output_dir: str | None) -> tuple[bool, str]:

        if not output_dir or not output_dir.strip():
            return False, "No output directory specified."

        try:
            os.makedirs(output_dir, exist_ok=True)

            test_file = os.path.join(
                output_dir,
                ".write_test",
            )

            with open(test_file, "w") as f:
                f.write("test")

            os.remove(test_file)

        except OSError as exc:
            return (
                False,
                f"Output directory is not writable: {exc}",
            )

        return True, ""

    # ------------------------------------------------------------------
    # Excel Reading
    # ------------------------------------------------------------------

    @staticmethod
    def read_excel(
        excel_path: str,
        log_callback: Any = logger.info,
    ) -> tuple[str, list[str], list[dict[str, Any]]]:

        log_callback(f"Reading Excel: {os.path.basename(excel_path)}")

        try:
            wb = openpyxl.load_workbook(
                excel_path,
                data_only=True,
                read_only=True,
            )
        except Exception as exc:
            raise ValidationError(f"Unable to open workbook: {exc}") from exc

        try:
            required_lower = {col.lower() for col in IDFCService.REQUIRED_COLUMNS}

            target_sheet = None

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                try:
                    header_row = [
                        (str(cell.value).replace("\n", " ").strip() if cell.value is not None else "")
                        for cell in next(
                            ws.iter_rows(
                                min_row=1,
                                max_row=1,
                            )
                        )
                    ]

                    header_lower = {h.lower() for h in header_row if h}

                    if required_lower.issubset(header_lower):
                        target_sheet = sheet_name
                        break

                except StopIteration:
                    continue

            if target_sheet is None:
                raise ValidationError("No valid sheet found. Required columns: " + ", ".join(IDFCService.REQUIRED_COLUMNS))

            log_callback(f"Found valid sheet: {target_sheet}")

            ws = wb[target_sheet]

            headers = [
                (str(cell.value).replace("\n", " ").strip() if cell.value is not None else "")
                for cell in next(
                    ws.iter_rows(
                        min_row=1,
                        max_row=1,
                    )
                )
            ]

            # Required columns are matched case-insensitively above, but
            # row_data keys must match REQUIRED_COLUMNS exactly (e.g.
            # "Tare Weight") or downstream .get() lookups silently miss
            # the data. Normalize any required header to its canonical
            # casing while leaving non-required headers untouched.
            canonical_lookup = {col.lower(): col for col in IDFCService.REQUIRED_COLUMNS}

            headers = [canonical_lookup.get(h.lower(), h) for h in headers]

            header_lower = {h.lower() for h in headers if h}

            missing = [col for col in IDFCService.REQUIRED_COLUMNS if col.lower() not in header_lower]

            if missing:
                raise ValidationError("Missing required columns: " + ", ".join(missing))

            rows: list[dict[str, Any]] = []

            for row in ws.iter_rows(
                min_row=2,
                values_only=False,
            ):
                row_data: dict[str, Any] = {}
                all_blank = True

                for i, cell in enumerate(row):
                    if i >= len(headers):
                        continue

                    header = headers[i].strip()

                    if not header:
                        continue

                    value = cell.value

                    if isinstance(value, str):
                        value = value.strip()

                    if value not in (None, ""):
                        all_blank = False

                    if header in ("Prospectno", "CUID"):
                        row_data[header] = str(value).strip() if value is not None else ""
                    else:
                        row_data[header] = value

                if not all_blank:
                    rows.append(row_data)

            return (
                target_sheet,
                headers,
                rows,
            )

        finally:
            wb.close()

    # ------------------------------------------------------------------
    # Data processing
    # ------------------------------------------------------------------

    @staticmethod
    def _clean_text(value: Any) -> str:
        """Convert a value to display text without float artifacts.

        Excel numeric cells (e.g. a branch code) often come back from
        openpyxl as floats, so a plain str() turns "1023" into
        "1023.0". Collapse whole-number floats to their int form first.
        """

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    @staticmethod
    def group_by_branch(
        rows: list[dict[str, Any]],
    ) -> dict[str, list[dict[str, Any]]]:

        groups: dict[str, list[dict[str, Any]]] = {}

        for row in rows:
            branch = IDFCService._clean_text(row.get("CurrentBranch", ""))

            if branch.lower() in {
                "",
                "none",
                "nan",
            }:
                branch = "UNKNOWN"

            groups.setdefault(branch, []).append(row)

        return groups

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    @staticmethod
    def format_tare_weight(val: Any) -> str:
        """
        Format tare weight.

        - Empty -> ""
        - 10.0 -> "10"
        - 10.5000 -> "10.5"
        - 10.0110000001 -> "10.01"
        - Maximum 2 decimal places
        - Removes floating point artifacts
        """

        try:
            if val is None or pd.isna(val):
                return ""
        except (TypeError, ValueError):
            pass

        try:
            num = float(val)

            if not math.isfinite(num):
                return ""

            # Round first so artifacts (e.g. 9.999999999999) collapse
            # to a clean value before the integer/decimal check below.
            num = round(num, 2)

            if num.is_integer():
                return str(int(num))

            return f"{num:.2f}".rstrip("0").rstrip(".")

        except (TypeError, ValueError):
            return str(val).strip()

    # ------------------------------------------------------------------
    # PDF Generation
    # ------------------------------------------------------------------

    def generate(
        self,
        audit_type: str,
        branch_code: str,
        branch_name: str,
        state: str,
        rows: list[dict[str, Any]],
        output_path: str,
    ) -> None:
        """Generate a single branch audit PDF report."""

        if not isinstance(rows, list):
            raise ValidationError("Rows must be a list.")

        if not rows:
            return

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        page_w, page_h = self.PAGE_SIZE

        doc = SimpleDocTemplate(
            output_path,
            pagesize=(page_w, page_h),
            leftMargin=self.PAGE_MARGIN_LEFT,
            rightMargin=self.PAGE_MARGIN_RIGHT,
            topMargin=self.PAGE_MARGIN_TOP,
            bottomMargin=self.PAGE_MARGIN_BOTTOM,
            title=os.path.basename(output_path),
        )

        style_hdr = ParagraphStyle(
            "ColHdr",
            fontName=self._font_bld,
            fontSize=9,
            alignment=TA_CENTER,
            leading=10,
            spaceBefore=0,
            spaceAfter=0,
        )

        table_data: list[list[Any]] = [
            [
                "Audit Type :",
                "",
                self._clean_text(audit_type),
                "",
                "Branch Name :",
                "",
                self._clean_text(branch_name),
            ],
            [
                "Branch Code :",
                "",
                self._clean_text(branch_code),
                "",
                "State :",
                "",
                self._clean_text(state),
            ],
            [
                Paragraph("Sr<br/>No", style_hdr),
                Paragraph("Prospectno", style_hdr),
                Paragraph("CUID", style_hdr),
                Paragraph("Tare Weight<br/>as per Bank", style_hdr),
                Paragraph("<nobr>Tare Weight as</nobr><br/>per Audit", style_hdr),
                Paragraph(
                    "<nobr>Purity Check - 18K and</nobr><br/>above 18K or Below 18K",
                    style_hdr,
                ),
                Paragraph("Remarks", style_hdr),
            ],
        ]

        for idx, row in enumerate(rows, start=1):
            table_data.append(
                [
                    str(idx),
                    str(row.get("Prospectno", "")).strip(),
                    str(row.get("CUID", "")).strip(),
                    self.format_tare_weight(row.get("Tare Weight", "")),
                    "",
                    "",
                    "",
                ]
            )

        row_heights = [
            self.HEADER_ROW_1_HEIGHT,
            self.HEADER_ROW_2_HEIGHT,
            self.COLUMN_HEADER_HEIGHT,
        ] + [self.DATA_ROW_HEIGHT] * (len(table_data) - 3)

        table = Table(
            table_data,
            colWidths=self.COL_WIDTHS,
            rowHeights=row_heights,
            repeatRows=3,
        )

        style_cmds = [
            ("SPAN", (0, 0), (1, 0)),
            ("SPAN", (4, 0), (5, 0)),
            ("SPAN", (0, 1), (1, 1)),
            ("SPAN", (4, 1), (5, 1)),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 1), self._font_bld),
            ("FONTSIZE", (0, 0), (-1, 1), 9),
            ("FONTNAME", (0, 3), (-1, -1), self._font_reg),
            ("FONTSIZE", (0, 3), (-1, -1), 9),
            ("FONTNAME", (0, 3), (0, -1), self._font_bld),
            ("BACKGROUND", (0, 2), (3, 2), self.COLOR_HEADER_BANK),
            ("BACKGROUND", (4, 2), (6, 2), self.COLOR_HEADER_AUDIT),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("LINEBEFORE", (0, 0), (0, -1), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("LINEAFTER", (6, 0), (6, -1), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("LINEABOVE", (0, 0), (-1, 0), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("LINEBELOW", (0, -1), (-1, -1), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("GRID", (0, 0), (2, 1), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("GRID", (4, 0), (6, 1), self.BORDER_WIDTH, self.COLOR_BORDER),
            ("GRID", (0, 2), (-1, -1), self.BORDER_WIDTH, self.COLOR_BORDER),
        ]

        table.setStyle(TableStyle(style_cmds))

        doc.build([table])


# ----------------------------------------------------------------------
# Backward-compatible singleton
# ----------------------------------------------------------------------

_service = IDFCService()

validate_excel = _service.validate
validate_output_dir = _service.validate_output_dir
read_excel = _service.read_excel
group_by_branch = _service.group_by_branch
generate_pdf = _service.generate
format_tare_weight = _service.format_tare_weight
register_fonts = _service._register_fonts  # noqa: SLF001
get_resource_path = IDFCService.get_resource_path

FONT_REG = _service.font_reg
FONT_BLD = _service.font_bld
REQUIRED_COLUMNS = _service.REQUIRED_COLUMNS
COL_WIDTHS = _service.COL_WIDTHS
