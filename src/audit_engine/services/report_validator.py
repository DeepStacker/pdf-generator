import contextlib
import logging
import re
import sys
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from datetime import time as time_cls
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Every date in the report is displayed this way.
DATE_NUMBER_FORMAT = "DD/MM/YYYY"

# Highlight fill (ARGB, as openpyxl stores it) → CSS color for the web grid.
PREVIEW_FILL_MAP = {
    "8B0000": "#8B0000",
    "FFC7CE": "#FFC7CE",
    "FF8C00": "#FF8C00",
    "FFFF00": "#FFFF00",
    "C6EFCE": "#C6EFCE",
    "BDD7EE": "#BDD7EE",
}

# ─── Highlight Colors ─────────────────────────────────────────────────
DARK_RED_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def map_columns(ws):
    """Auto-detect column mapping from header row (row 2) and section row (row 1).
    Uses ws._cells to avoid iterating empty columns (performance).
    First occurrence wins (main display columns take priority over extra columns)."""
    mapping = {}
    row2 = {}
    row1 = {}
    for (_r, _c), _cell in ws._cells.items():
        if _r == 1 and _cell.value is not None:
            row1[_c] = str(_cell.value).strip()
        elif _r == 2 and _cell.value is not None:
            row2[_c] = str(_cell.value).strip()
    for c in sorted(set(row2.keys()) | set(row1.keys())):
        h = row2.get(c)
        s = row1.get(c)
        if h:
            key = h.upper()
            if key not in mapping:
                mapping[key] = c
        if s and not h:
            key = f"__SECTION__{s.upper()}"
            if key not in mapping:
                mapping[key] = c
    return mapping


def get_col(mapping, *names):
    """Get column number for any of the given names."""
    for name in names:
        key = name.upper()
        if key in mapping:
            return mapping[key]
    return None


def build_col_map(col):
    """Build the canonical column map from a map_columns() header mapping."""
    return {
        "packet": get_col(col, "PACKET NO."),
        "account": get_col(col, "ACCOUNT NUMBER"),
        "applicant": get_col(col, "APPLICANT NAME"),
        "status": get_col(col, "FRESH/RENEWAL/CLOSED/ALREADY VERIFIED"),
        "taf": get_col(col, "TAF/POA"),
        "remarks": get_col(col, "AGENCY REMARKS"),
        "magnet": get_col(col, "MAGNET TEST RESULT"),
        "tampered": get_col(col, "PACKET TAMPERED YES/NO"),
        "gdr_gross": get_col(col, "GROSS WEIGHT AS PER GDR/PACKET"),
        "actual_gross": get_col(col, "ACTUAL GROSS WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "tare": get_col(col, "ACTUAL TARE WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "gross_diff": get_col(col, "DIFFERENCE IN GROSS WEIGHT"),
        "gdr_net": get_col(col, "NET WEIGHT AS PER GDR/PACKET"),
        "actual_net": get_col(col, "ACTUAL NET WEIGHT AS PER FCU AGENCY VERIFICATION"),
        "net_diff": get_col(col, "DIFFERENCE IN NET WEIGHT"),
        "spur_count": get_col(col, "TOTAL NO.OF SPURIOUS ORNAMENTS"),
        "spur_weight": get_col(col, "SPURIOUS ORNAMENTS GROSS WEIGHT"),
        "spur_pct": get_col(col, "% OF SPURIOUS ORNAMENTS GROSS WEIGHT"),
        "carat_count": get_col(col, "TOTAL NO.OF ORNAMENTS WITH CARAT MISMATCH"),
        "uncommon_count": get_col(col, "TOTAL NO.OF UNCOMMON ORNAMENTS"),
        "sanction_date": get_col(col, "SANCTION DATE"),
        "verification_date": get_col(col, "AGENCY VERIFICATION DATE"),
        "sanction_limit": get_col(col, "SANCTION LIMIT"),
        "gdr_no": get_col(col, "GDR NUMBER"),
        "ornaments_gdr": get_col(col, "TOTAL NO.OF ORNAMENTS AS PER THE GDR/PACKET"),
        "ornaments_actual": get_col(col, "ACTUAL AVAILABLE ORNAMENTS AT THE TIME OF FCU VERIFICATION"),
        "ornaments_diff": get_col(col, "DIFFRENCE IN ACTUAL ORNAMENTS"),
        "renewal_date": get_col(col, "RENEWAL/CLOSED DATE"),
        "state": get_col(col, "STATE", "STATE NAME", "BRANCH STATE"),
    }


# The weight-discrepancy remarks the report uses, and the text that replaces
# them once nothing is outstanding.
GROSS_DIFF_REMARK = "Gross Weight Difference"
NET_DIFF_REMARK = "Net Weight Difference"
BOTH_DIFF_REMARK = "Gross Weight and Net Weight Difference"
NO_DISCREPANCY_REMARK = "Ok - No Discrepancy"

# A gross/net difference no larger than this is measurement noise rather than
# a real discrepancy.
WEIGHT_MATCH_TOLERANCE = 0.200


def amend_discrepancy_remark(remark, fixed_gross, fixed_net):
    """Rewrite a weight-discrepancy remark to name only what is still outstanding.

    Returns the replacement text, or None when the remark is not one of the
    discrepancy remarks and should be left exactly as it is.
    """
    normalized = re.sub(r"\s+", " ", str(remark or "")).strip().upper()
    if not normalized:
        return None

    # Check the combined wording first: it contains "NET WEIGHT DIFFERENCE"
    # as a substring, so testing the single-column wording first would
    # misread it.
    both = "GROSS WEIGHT AND NET WEIGHT DIFFERENCE" in normalized
    has_gross = both or "GROSS WEIGHT DIFFERENCE" in normalized
    has_net = both or "NET WEIGHT DIFFERENCE" in normalized
    if not (has_gross or has_net):
        return None

    still_gross = has_gross and not fixed_gross
    still_net = has_net and not fixed_net
    if still_gross and still_net:
        return BOTH_DIFF_REMARK
    if still_gross:
        return GROSS_DIFF_REMARK
    if still_net:
        return NET_DIFF_REMARK
    return NO_DISCREPANCY_REMARK


def is_topup_remark(remarks):
    """True if the remarks text marks the row as a top-up ("Top Up", "Top-up",
    "Topup of account ..." and similar spellings).

    Matches on a word boundary so unrelated remarks that merely contain "top"
    and "up" as parts of other words (e.g. "Desktop Update pending") are not
    mistaken for a top-up row.
    """
    return bool(re.search(r"\bTOP[\s\-]*UP\b", str(remarks or "").upper()))


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def find_data_end(ws, data_start=5):
    """Return the last row index that contains actual cell *values*.

    Real-world sheets often carry stray formatting / empty styled cells far
    below the real data (ws.max_row can be 200k+ while only a few hundred rows
    hold values). Iterating all_rows up to ws.max_row then becomes quadratic
    and can take minutes. We bound the scan to the highest row that holds a
    non-empty cell value, ignoring style-only cells."""
    max_row = ws.max_row
    if max_row <= data_start:
        return max_row
    end = data_start
    for (r, c), cell in ws._cells.items():
        if r < data_start or r > max_row:
            continue
        v = cell.value
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        end = max(end, r)
    return end


def is_empty(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.upper() == "NAN" or s == "None"


def parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.date()
        return v
    s = str(v).strip()
    if not s or s.upper() in ("NAN", "NONE", "NAT", "-"):
        return None

    months_map = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
    }

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    for sep in ["-", "/", ".", " "]:
        parts = [p.strip() for p in s.split(sep) if p.strip()]
        if len(parts) == 3:
            p0, p1, p2 = parts[0], parts[1], parts[2]
            try:
                if p2.isdigit() and len(p2) in (2, 4):
                    yr = int(p2)
                    if yr < 100:
                        yr += 2000
                    day = int(p0)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return date(yr, m, day)
                elif p0.isdigit() and len(p0) == 4:
                    yr = int(p0)
                    day = int(p2)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return date(yr, m, day)
            except Exception:
                pass
    return None


def get_custom_output_filename(ws, col, col_map, all_rows, _cell_cache, default_filename):
    col_branch = get_col(col, "BRANCH NAME", "BRANCH", "BRANCH_NAME", "NAME OF BRANCH", "BRANCH NAME ")
    branch_name = "UNKNOWN"
    if col_branch:
        for r in all_rows:
            cell_obj = _cell_cache.get((r, col_branch))
            v = cell_obj.value if cell_obj else None
            if v and str(v).strip():
                branch_name = str(v).strip().upper()
                break

    import re
    branch_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', branch_name).strip()
    branch_name = re.sub(r'_+', '_', branch_name)

    # Prefer the agency verification date, but fall back to the other date
    # columns rather than giving up and naming the file UNKNOWN_DATE.
    valid_dates = []
    for date_key in ("verification_date", "sanction_date", "renewal_date"):
        date_col = col_map.get(date_key)
        if not date_col:
            continue
        for r in all_rows:
            cell_obj = _cell_cache.get((r, date_col))
            v = cell_obj.value if cell_obj else None
            if v:
                parsed_d = parse_any_date(v)
                if parsed_d:
                    valid_dates.append(parsed_d)
        if valid_dates:
            break

    if valid_dates:
        unique_dates = sorted(list(set(valid_dates)))
        min_date = unique_dates[0]
        max_date = unique_dates[-1]

        months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        def fmt_d(d):
            return f"{d.day:02d}-{months[d.month-1]}-{d.year}"

        date_str = fmt_d(min_date) if min_date == max_date else f"{fmt_d(min_date)}_to_{fmt_d(max_date)}"
    else:
        date_str = "UNKNOWN_DATE"

    return f"{branch_name}_Audit-MIS_{date_str}.xlsx"


def normalize_account(value):
    """Digits-only comparison form for an account number.

    The same account reaches us in several shapes: as text in the PDF
    ("1234 5678 9012 345"), and from openpyxl as a float or in scientific
    notation when the cell is numeric ("123456789012345.0", "1.23457E+14").
    Comparing raw strings therefore misses matches that are plainly the same
    account, which is what made resequencing silently do nothing.
    """
    s = safe_str(value)
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):                      # 123....0 from a float cell
        s = s.split(".", 1)[0]
    elif re.fullmatch(r"\d+(?:\.\d+)?[eE][+-]?\d+", s):  # 1.23457E+14
        with contextlib.suppress(ValueError, OverflowError):
            s = format(int(float(s)), "d")
    return re.sub(r"\D", "", s)


def extract_accounts_from_pdf(pdf_path, known_accounts=None):
    """Return account numbers in the order they appear in the PDF.

    ``known_accounts`` should be the accounts actually present in the
    workbook. Given those, the PDF is searched for *them* rather than for a
    guessed digit count — account numbers are not universally 15 digits, and
    the previous fixed ``\\d{15}`` pattern matched nothing at all on any other
    format, so resequencing quietly did nothing.

    Never raises: the PDF only controls row order, so an unreadable file must
    not cost the user their validated workbook.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning(
            "pypdf is unavailable, so PDF row resequencing was skipped. "
            "The workbook is still validated; row order is left unchanged."
        )
        return []

    text = ""
    try:
        reader = PdfReader(pdf_path)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:  # noqa: BLE001 - any malformed PDF must not be fatal
        logger.warning("Could not read PDF %s: %s", pdf_path, e)
        return []

    if not text.strip():
        logger.warning("No extractable text in %s (a scanned/image PDF needs OCR).", pdf_path)
        return []

    known = {a for a in (normalize_account(k) for k in (known_accounts or ())) if a}

    # The workbook's own accounts define the account *shape* (how many digits);
    # every entry of that shape is then found in the PDF, including ones the
    # workbook does not contain — those are exactly the sequence positions that
    # must be left blank. With no workbook context (direct/CLI use), accept the
    # common account lengths instead.
    lengths = sorted({len(a) for a in known}, reverse=True) if known else list(range(18, 7, -1))

    # Exactly L digits, tolerating a single separator between them, and not
    # butted against further digits — so "1002-0030-0401 100200300402" reads
    # as two entries and a longer number is not mistaken for an account.
    pattern = "|".join(rf"(?<!\d)(?:\d[\s\-]?){{{length - 1}}}\d(?!\d)" for length in lengths)

    found = []
    for m in re.finditer(pattern, text):
        digits = re.sub(r"\D", "", m.group(0))
        if digits:
            found.append((m.start(), digits))

    if not known:
        return [d for _pos, d in found]

    # The strict boundary above is what makes discovering *unknown* entries
    # safe, but it also skips a known account printed inside a longer digit
    # run ("Ref#77<account>X"). Those are known to be real accounts, so look
    # for any that the scan missed and slot them in at their position.
    seen_digits = {d for _pos, d in found}
    for acct in known - seen_digits:
        m = re.search(r"[\s\-]?".join(acct), text)
        if m:
            found.append((m.start(), acct))

    found.sort(key=lambda pair: pair[0])

    # Keep PDF order; drop repeats so a summary line cannot duplicate a slot.
    ordered, seen = [], set()
    for _pos, acct in found:
        if acct not in seen:
            ordered.append(acct)
            seen.add(acct)
    return ordered


# A1-style reference: optional $ before the column and/or the row.
_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def retarget_formula(formula, old_row, new_row):
    """Repoint a moved row's own-row references at the row it now occupies.

    These sheets use same-row arithmetic (``=I26-J26`` on row 26). Copying that
    text to row 46 as-is leaves it computing from row 26 — the row shows a
    formula for a different loan, or blanks once the referenced row changes.
    Only relative references to the row's *own* row are shifted; absolute rows
    ($26) and references to other rows are left exactly as written.
    """
    if old_row == new_row or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _sub(m):
        dollar_col, col, dollar_row, row = m.groups()
        if not dollar_row and int(row) == old_row:
            return f"{dollar_col}{col}{dollar_row}{new_row}"
        return m.group(0)

    return _CELL_REF_RE.sub(_sub, formula)


def sheet_accounts(ws, col_map, all_rows):
    """Every account number present in the data rows, normalized."""
    col_acct = col_map.get("account")
    if not col_acct:
        return []
    out = []
    for r in all_rows:
        acct = normalize_account(ws.cell(row=r, column=col_acct).value)
        if acct:
            out.append(acct)
    return out


def rearrange_rows_by_pdf(ws, col_map, all_rows, pdf_accounts):
    """Lay the data rows out in the PDF's account order.

    The PDF is the master sequence. Every entry in it gets a slot: if the
    workbook has that account the row moves there, and if it does not the slot
    is left as an empty (but still bordered) row, so every later row keeps
    lining up with the PDF's position instead of shifting up by the number of
    missing entries. Rows whose account never appears in the PDF are placed
    after the sequence rather than being interleaved.

    Returns {"matched": rows placed from the PDF, "blank_rows": [row numbers
    left empty], "last_row": last row written}.
    """
    col_acct = col_map.get("account")
    if not col_acct or not all_rows:
        return {"matched": 0, "blank_rows": [], "last_row": all_rows[-1] if all_rows else 0}

    # Normalized on both sides so a numeric Excel cell still matches the text
    # printed in the PDF.
    acct_to_rows = defaultdict(list)
    for r in all_rows:
        acct = normalize_account(ws.cell(row=r, column=col_acct).value)
        if acct:
            acct_to_rows[acct].append(r)

    # One slot per PDF entry; None means "no such row in the workbook".
    slots = []
    used_rows = set()
    for acct in pdf_accounts:
        rows_for_acct = [r for r in acct_to_rows.get(acct, ()) if r not in used_rows]
        if rows_for_acct:
            for r in rows_for_acct:
                slots.append(r)
                used_rows.add(r)
        else:
            slots.append(None)

    matched_rows = len(used_rows)
    # Rows the PDF never mentions follow the sequence.
    slots.extend(r for r in all_rows if r not in used_rows)

    # Bound the column range to the true data width: the rightmost column that
    # holds an actual value in the header rows or the data rows. This keeps
    # extra columns beyond the mapped ones moving together with their rows
    # (a fixed cap would silently leave far-right data behind, misaligning
    # rows), while still ignoring style-only stray cells that can push
    # ws.max_column into the thousands.
    row_set = set(all_rows)
    max_val_col = 0
    for (r, c), cl in ws._cells.items():
        if c <= max_val_col or (r not in row_set and r not in (1, 2)):
            continue
        v = cl.value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        max_val_col = c
    known_cols = [c for c in col_map.values() if c]
    if known_cols:
        max_val_col = max(max_val_col, *known_cols)
    max_col = min(ws.max_column, max_val_col + 2) if max_val_col else min(ws.max_column, 150)

    def _snapshot(r):
        cells = {}
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            styled = cell.has_style
            cells[c] = {
                "value": cell.value,
                "font": copy(cell.font) if styled and cell.font else None,
                "fill": copy(cell.fill) if styled and cell.fill else None,
                "border": copy(cell.border) if styled and cell.border else None,
                "alignment": copy(cell.alignment) if styled and cell.alignment else None,
                "number_format": cell.number_format,
            }
        return cells

    row_data_cache = {r: _snapshot(r) for r in all_rows}
    row_heights = {r: ws.row_dimensions[r].height for r in all_rows}

    # Borders for a gap row are taken from the first data row so the table
    # stays visually continuous where an entry is missing.
    template = row_data_cache[all_rows[0]]

    first_row = all_rows[0]
    blank_rows = []
    for offset, src in enumerate(slots):
        new_r = first_row + offset
        if src is None:
            blank_rows.append(new_r)
            ws.row_dimensions[new_r].height = row_heights[all_rows[0]]
            for c in range(1, max_col + 1):
                cell = ws.cell(row=new_r, column=c)
                cell.value = None
                cell.font = Font()
                cell.fill = PatternFill()
                cell.border = copy(template[c]["border"]) if template[c]["border"] else Border()
                cell.alignment = copy(template[c]["alignment"]) if template[c]["alignment"] else Alignment()
                cell.number_format = template[c]["number_format"]
            continue

        ws.row_dimensions[new_r].height = row_heights[src]
        cells_dict = row_data_cache[src]
        for c in range(1, max_col + 1):
            cell = ws.cell(row=new_r, column=c)
            data = cells_dict[c]
            # If the source cell carried no explicit style, the target must be
            # reset to defaults — otherwise the moved row keeps the previous
            # occupant's formatting.
            target_had_style = cell.has_style
            # A formula must follow the row it belongs to, not keep pointing
            # at the row number it was written for.
            cell.value = retarget_formula(data["value"], src, new_r)
            if data["font"]:
                cell.font = data["font"]
            elif target_had_style:
                cell.font = Font()
            if data["fill"]:
                cell.fill = data["fill"]
            elif target_had_style:
                cell.fill = PatternFill()
            if data["border"]:
                cell.border = data["border"]
            elif target_had_style:
                cell.border = Border()
            if data["alignment"]:
                cell.alignment = data["alignment"]
            elif target_had_style:
                cell.alignment = Alignment()
            cell.number_format = data["number_format"]

    last_row = first_row + len(slots) - 1
    # The layout can be shorter than the original block (duplicate accounts
    # collapse); wipe anything left over so stale rows are not orphaned below.
    for r in range(last_row + 1, all_rows[-1] + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    return {"matched": matched_rows, "blank_rows": blank_rows, "last_row": last_row}


def _json_safe(val):
    """Convert cell values that json can't serialize (dates/times) to strings."""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, time_cls):
        return val.isoformat()
    return val


def validate_workbook(
    src_path,
    output_path=None,
    pdf_path=None,
    on_progress=None,
    build_preview=False,
):
    """Validate one "Purity Verification Format" workbook, file in → file out.

    This is the single entry point shared by the desktop app (called in-process
    over the zero-socket JS bridge) and the web server (called from its upload
    worker). It is entirely offline: it opens a path, writes a path, and never
    touches the network — which is what lets the desktop build run in fully
    restricted environments.

    Args:
        src_path: workbook to validate.
        output_path: where to write the validated copy. Defaults to the
            generated "<BRANCH>_Audit-MIS_<dates>.xlsx" beside src_path.
        pdf_path: optional PDF whose account sequence reorders the rows. When
            omitted, row order is left untouched and no PDF is parsed.
        on_progress: optional callback(pct: int, message: str) for live UI.
        build_preview: also return per-cell values/highlights for an in-app
            grid. The desktop UI does not need this; the web UI does.

    Returns:
        A JSON-serializable dict (safe to hand straight to the bridge).

    Raises:
        FileNotFoundError: src_path does not exist.
        KeyError: the workbook has no "Purity Verification Format" sheet.
    """
    def report(pct, msg=None):
        if on_progress:
            on_progress(pct, msg)

    src_path = Path(src_path)
    if not src_path.exists():
        raise FileNotFoundError(f"File not found: {src_path}")

    report(5, "Reading workbook…")
    wb = openpyxl.load_workbook(src_path, keep_vba=False, keep_links=False)
    wb_data = openpyxl.load_workbook(src_path, data_only=True, keep_vba=False, keep_links=False)
    sheet_name = "Purity Verification Format"
    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        wb_data.close()
        raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]

    try:
        report(12, "Mapping columns…")
        col = map_columns(ws)
        data_start = 5
        data_end = find_data_end(ws, data_start)
        col_map = build_col_map(col)
        all_rows = list(range(data_start, data_end + 1))

        # PDF is optional: without one, row order is left exactly as-is.
        pdf_applied = False
        pdf_warning = None
        pdf_matched_rows = 0
        pdf_gap_rows = 0
        if pdf_path:
            report(15, "Parsing PDF sequence…")
            # Search the PDF for the accounts this workbook actually contains,
            # rather than for a guessed digit pattern.
            wanted = sheet_accounts(ws, col_map, all_rows)
            pdf_accounts = extract_accounts_from_pdf(pdf_path, known_accounts=wanted)
            if pdf_accounts:
                report(20, "Rearranging rows by PDF…")
                original_count = len(all_rows)
                layout = rearrange_rows_by_pdf(ws, col_map, all_rows, pdf_accounts)
                rearrange_rows_by_pdf(ws_data, col_map, all_rows, pdf_accounts)
                pdf_matched_rows = layout["matched"]
                pdf_applied = pdf_matched_rows > 0
                pdf_gap_rows = len(layout["blank_rows"])

                # The layout may be longer than the source block (a gap row is
                # added wherever the PDF lists an account the workbook lacks),
                # so the row range has to be re-derived. Gap rows are excluded:
                # they hold no data and must not be validated or flagged.
                blanks = set(layout["blank_rows"])
                all_rows = [r for r in range(all_rows[0], layout["last_row"] + 1) if r not in blanks]

                notes = []
                if pdf_matched_rows < original_count:
                    notes.append(
                        f"{pdf_matched_rows} of {original_count} rows were matched to the PDF; "
                        "unmatched rows were placed after the sequence"
                    )
                if pdf_gap_rows:
                    notes.append(
                        f"{pdf_gap_rows} row(s) left blank where the PDF lists an account "
                        "this workbook does not contain"
                    )
                if notes:
                    pdf_warning = ". ".join(notes) + "."
                    logger.warning(pdf_warning)
            if not pdf_applied:
                pdf_warning = (
                    "None of this workbook's account numbers were found in the PDF, so "
                    "row order was left unchanged. The workbook was still validated."
                )
                logger.warning(pdf_warning)

        _cell_cache = dict(ws._cells)
        rows_data = _build_rows_data(all_rows, col_map, _cell_cache)
        _data_cache = dict(ws_data._cells)

        report(30, "Running validations…")
        summary = defaultdict(list)
        run_validation(
            ws, ws_data, col_map, all_rows, rows_data, summary,
            _cell_cache=_cell_cache, _data_cache=_data_cache,
        )

        # Validation can write into cells that did not exist before (defaults
        # on Closed/Top-Up rows), so refresh the cache before reading back.
        _cell_cache = dict(ws._cells)

        custom_filename = get_custom_output_filename(ws, col, col_map, all_rows, _cell_cache, src_path.name)
        out_path = Path(output_path) if output_path else src_path.parent / custom_filename

        result = {
            "file_name": src_path.name,
            "custom_filename": custom_filename,
            "summary": {k: len(v) for k, v in sorted(summary.items())},
            "summary_details": {k: list(v) for k, v in sorted(summary.items())},
            "total_issues": sum(len(v) for v in summary.values()),
            "rows_scanned": len(all_rows),
            "pdf_applied": pdf_applied,
            "pdf_matched_rows": pdf_matched_rows,
            "pdf_gap_rows": pdf_gap_rows,
            "pdf_warning": pdf_warning,
        }

        if build_preview:
            report(60, "Building preview…")
            result.update(_build_preview(all_rows, _cell_cache, _data_cache))

        report(85, "Saving validated workbook…")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        # openpyxl does not carry cached formula results into the saved file, so
        # every formula cell would open blank until something recalculates it.
        # Ask Excel to do a full recalculation on load.
        with contextlib.suppress(AttributeError):
            wb.calculation.fullCalcOnLoad = True
        wb.save(str(out_path))
        result["output_path"] = str(out_path)
    finally:
        wb.close()
        wb_data.close()

    report(100, "Complete")
    return result


def _build_rows_data(all_rows, col_map, cell_cache):
    """Extract the text fields the validation passes key off, per row."""
    fields = (
        ("packet", "packet"), ("account", "account"), ("name", "applicant"),
        ("status", "status"), ("taf", "taf"), ("remarks", "remarks"),
        ("magnet", "magnet"), ("tampered", "tampered"),
    )
    rows_data = {}
    for r in all_rows:
        row = {}
        for key, col_key in fields:
            c = col_map.get(col_key)
            cell = cell_cache.get((r, c)) if c else None
            row[key] = safe_str(cell.value if cell is not None else None)
        rows_data[r] = row
    return rows_data


def _build_preview(all_rows, cell_cache, data_cache):
    """Build the cell/highlight payload the web grid renders."""
    column_indices, column_letters, column_headers = [], [], []
    for c in range(1, 101):
        cell = cell_cache.get((2, c))
        header_val = cell.value if cell else None
        if header_val and str(header_val).strip():
            column_indices.append(c)
            column_letters.append(get_column_letter(c))
            column_headers.append(str(header_val).strip())

    col_letter_map = {c: get_column_letter(c) for c in column_indices}
    rows_json, highlights, issues = [], {}, []

    for r in all_rows:
        row_data = {"row": r, "cells": {}}
        for c in column_indices:
            cell_obj = cell_cache.get((r, c))
            if cell_obj is None:
                continue
            col_letter = col_letter_map[c]
            raw = cell_obj.value
            if raw is None or (isinstance(raw, str) and raw.startswith("=")):
                dc = data_cache.get((r, c))
                if dc is not None:
                    raw = dc.value
            row_data["cells"][col_letter] = _json_safe(raw)
            fill = cell_obj.fill
            if fill and fill.start_color and fill.start_color.rgb:
                rgb = str(fill.start_color.rgb)
                if rgb.startswith("00"):
                    rgb = rgb[2:]
                if rgb in PREVIEW_FILL_MAP and rgb != "00000000":
                    key = f"{col_letter}{r}"
                    highlights[key] = PREVIEW_FILL_MAP[rgb]
                    issues.append({"ref": key, "row": r, "col": col_letter, "color": PREVIEW_FILL_MAP[rgb]})
        rows_json.append(row_data)

    return {
        "columns": column_headers,
        "column_letters": column_letters,
        "rows": rows_json,
        "highlights": highlights,
        "issues": issues,
    }


def process_file(filepath: str, output_suffix: str = "_VALIDATED", pdf_path: str = None):
    """Validate one file and print a CLI report. Thin wrapper over validate_workbook()."""
    print(f"\n{'='*80}")
    print(f"PROCESSING: {Path(filepath).name}")
    if pdf_path:
        print(f"USING PDF SEQUENCE: {Path(pdf_path).name}")
    print(f"{'='*80}")

    try:
        result = validate_workbook(filepath, pdf_path=pdf_path)
    except FileNotFoundError:
        print(f"ERROR: File not found: {filepath}")
        return
    except KeyError as e:
        print(f"ERROR: {e}")
        return

    details = result["summary_details"]
    print(f"\n{'-'*60}")
    print(f"PROCESSING COMPLETE: {Path(result['output_path']).name}")
    print(f"{'-'*60}")

    for key, items in sorted(details.items()):
        if not items:
            continue
        color = ""
        reset = ""
        if any(kw in key for kw in ["duplicate", "outlier", "mismatch", "error", "negative"]):
            color = "\033[91m"  # Red
        elif any(kw in key for kw in ["found", "updated", "fixed", "copied", "defaulted", "cleared"]):
            color = "\033[92m"  # Green
        elif any(kw in key for kw in ["warning", "multiple", "unexpected", "no_match"]):
            color = "\033[93m"  # Yellow

        label = key.replace("_", " ").title()
        print(f"\n  {color}{label}: {len(items)}{reset}")
        for item in items[:10]:
            print(f"    - {item}")
        if len(items) > 10:
            print(f"    ... and {len(items)-10} more")

    print(f"\n  TOTAL ISSUES FOUND: {result['total_issues']}")
    print(f"  Saved to: {result['output_path']}")
    # defaultdict so callers can probe categories that produced no findings.
    return defaultdict(list, details)


def run_validation(ws, ws_data, col_map, all_rows, rows_data, summary, _cell_cache=None, _data_cache=None):
    """Run all validation passes against a worksheet.

    Args:
        ws: openpyxl Worksheet (for writing)
        ws_data: data_only Worksheet (for formula values)
        col_map: dict of column name → column index
        all_rows: list of row numbers to process
        rows_data: dict of row → {packet, account, name, status, taf, remarks, magnet, tampered}
        summary: defaultdict(list) to collect issues
        _cell_cache: optional dict {(row, col): cell} for fast access (app.py), or None

    Returns:
        summary (same dict, mutated in place)
    """
    resolved_packet_rows = set()

    # Computed once (not per-highlight-call): columns excluded from highlighting
    # on rows already resolved by the Pass 2 top-up packet match.
    _diff_cols_excluded = frozenset(
        c for c in (
            col_map.get("gdr_gross"), col_map.get("actual_gross"), col_map.get("tare"), col_map.get("gross_diff"),
            col_map.get("gdr_net"), col_map.get("actual_net"), col_map.get("net_diff"),
            col_map.get("ornaments_gdr"), col_map.get("ornaments_actual"), col_map.get("ornaments_diff"),
        )
        if c is not None
    )

    # Computed once (not re-derived in every pass): uppercased status keyed by row.
    row_status_upper = {}
    for r in all_rows:
        rd = rows_data[r]
        row_status_upper[r] = rd["status"].upper() if rd["status"] else ""

    # Rows reset to default values by the Closed/Top-Up cleaning pass; later
    # passes skip these so cleaned defaults aren't re-flagged as issues.
    cleaned_rows = set()

    # ── Cell helpers (use cache if available, else direct cell access) ──
    def cell(row, c):
        if c is None:
            return None
        if _cell_cache is not None:
            cl = _cell_cache.get((row, c))
            if cl is not None:
                return cl
        return ws.cell(row=row, column=c)

    def val(row, c):
        cl = cell(row, c)
        return cl.value if cl is not None else None

    _column_style = {}

    def column_style(c):
        """Font/alignment/number format this column's populated cells use.

        A value written into a cell that was empty would otherwise land with
        Excel's defaults (left-aligned, Calibri 11) instead of the document's
        own styling — which is how resolved packet numbers ended up neither
        centred nor at the sheet's font size.
        """
        if c in _column_style:
            return _column_style[c]
        style = None
        for r in all_rows:
            cl = cell(r, c)
            if cl is None or not cl.has_style:
                continue
            v = cl.value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            style = (copy(cl.font), copy(cl.alignment), cl.number_format)
            break
        _column_style[c] = style
        return style

    def set_val(row, c, new_val, inherit_style=True):
        cl = cell(row, c)
        if cl is None:
            return
        prev = cl.value
        was_empty = prev is None or (isinstance(prev, str) and not prev.strip())
        cl.value = new_val
        if inherit_style and was_empty and new_val is not None:
            style = column_style(c)
            if style:
                font, alignment, number_format = style
                cl.font = copy(font)
                cl.alignment = copy(alignment)
                if cl.number_format == "General":
                    cl.number_format = number_format

    def highlight(row, c, fill, font=None, force=False, skip_excluded=True):
        if skip_excluded and row in resolved_packet_rows and c in _diff_cols_excluded:
            return
        cl = cell(row, c)
        if cl is not None:
            if not force and cl.fill and cl.fill.start_color and cl.fill.start_color.rgb:
                rgb = str(cl.fill.start_color.rgb)
                if not rgb.endswith("000000") and rgb != "00000000":
                    return
            cl.fill = copy(fill)
            if font:
                # Preserve the cell's original font (family, size, italics…)
                # and only apply the requested emphasis attributes on top.
                new_font = copy(cl.font) if cl.font else Font()
                if font.color is not None:
                    new_font.color = copy(font.color)
                if font.bold:
                    new_font.bold = True
                cl.font = new_font

    def num_or_none(v):
        try:
            s = str(v).strip()
            if not s or s.upper() in ("NA", "N/A", "N A", "NAN", "NONE", "-"):
                return None
            return float(s)
        except (ValueError, TypeError):
            return None

    def val_data(row, c):
        v = val(row, c)
        if v is not None and isinstance(v, str) and v.startswith("="):
            if _data_cache is not None:
                cv_cell = _data_cache.get((row, c))
                cv = cv_cell.value if cv_cell is not None else None
            else:
                cv = ws_data.cell(row=row, column=c).value if ws_data else None
            return cv if cv is not None else v
        return v

    # ══════════════════════════════════════════════════════════════════
    # PASS 1: Duplicate Packet Numbers
    # ══════════════════════════════════════════════════════════════════
    col_packet = col_map.get("packet")
    if col_packet:
        packet_map = defaultdict(list)
        for r in all_rows:
            pkt = rows_data[r]["packet"]
            if pkt and pkt.upper() not in ("", "NAN", "NONE"):
                packet_map[pkt].append(r)
        for pkt, dup_rows in packet_map.items():
            if len(dup_rows) > 1:
                for r in dup_rows:
                    is_topup = is_topup_remark(rows_data[r]["remarks"])
                    if is_topup:
                        highlight(r, col_packet, LIGHT_BLUE_FILL)
                        summary["duplicate_packet_topup"].append(f"Row {r}: Packet '{pkt}' (Top-Up)")
                    else:
                        highlight(r, col_packet, DARK_RED_FILL, WHITE_FONT)
                        summary["duplicate_packet"].append(f"Row {r}: Packet '{pkt}' (appears {len(dup_rows)}x)")

    # ══════════════════════════════════════════════════════════════════
    # PASS 2: Empty Packet → Resolve by Applicant Name
    # ══════════════════════════════════════════════════════════════════
    col_remarks = col_map.get("remarks")
    if col_packet and col_remarks:
        empty_pkt_rows = [r for r in all_rows if not rows_data[r]["packet"] or rows_data[r]["packet"].upper() in ("", "NAN", "NONE")]

        # Index of name → rows currently holding a valid packet, built once and
        # updated incrementally as rows get resolved below. Equivalent to the
        # original full re-scan of all_rows per empty-packet row (a resolved
        # row becomes a candidate match for later rows with the same name,
        # which is what makes ambiguous multi-way name matches get flagged
        # instead of silently resolved) but O(1) per lookup instead of O(n).
        name_to_packet_rows = defaultdict(list)
        for sr in all_rows:
            spkt = rows_data[sr]["packet"]
            if spkt and spkt.upper() not in ("", "NAN", "NONE"):
                name_to_packet_rows[rows_data[sr]["name"].upper().strip()].append(sr)

        for r in empty_pkt_rows:
            rd = rows_data[r]
            status = row_status_upper[r]
            if "CLOSED" in status:
                summary["closed_empty_packet"].append(f"Row {r}: Empty packet (Closed)")
                continue
            name = rd["name"]
            if not name:
                highlight(r, col_packet, ORANGE_FILL)
                summary["unexpected_empty_packet"].append(f"Row {r}: Account {rd['account']} (no name)")
                continue
            name_key = name.upper().strip()
            matching_rows = [sr for sr in name_to_packet_rows.get(name_key, ()) if sr != r]
            if len(matching_rows) == 1:
                sr = matching_rows[0]
                source_acct = rows_data[sr]["account"]
                source_packet = rows_data[sr]["packet"]
                set_val(r, col_packet, source_packet)
                rd["packet"] = source_packet
                name_to_packet_rows[name_key].append(r)
                col_gdr_no = col_map.get("gdr_no")
                if col_gdr_no:
                    set_val(r, col_gdr_no, None)
                expected_remark = f"This Account is Top Up of Account No.{source_acct} , With Same Packet No"
                set_val(r, col_remarks, expected_remark)
                summary["topup_resolved"].append(f"Row {r}: {name} → matched row {sr} (pkt {source_packet})")
                resolved_packet_rows.add(r)
            elif len(matching_rows) > 1:
                highlight(r, col_packet, ORANGE_FILL)
                summary["topup_multiple_matches"].append(f"Row {r}: {name} has {len(matching_rows)} matches, left unresolved")
            else:
                highlight(r, col_packet, ORANGE_FILL)
                summary["topup_no_match"].append(f"Row {r}: {name} has no matching row with packet")

    # ══════════════════════════════════════════════════════════════════
    # PASS 3: Reset Closed/Top-Up rows to default values
    # Verification-result columns are not applicable on these rows: numeric
    # columns get their default (0.000 for weights, 0 for counts/diffs/%),
    # the rest are blanked. Identity/loan columns (packet, account, name,
    # status, TAF/POA, dates, sanction limit, remarks) are kept. Closed rows
    # keep the Renewal/Closed date — that is the closure date itself.
    # ══════════════════════════════════════════════════════════════════
    default_weight_cols = [
        (col_map.get("gdr_gross"), "GDR Gross"),
        (col_map.get("actual_gross"), "Actual Gross"),
        (col_map.get("tare"), "Tare"),
        (col_map.get("gdr_net"), "GDR Net"),
        (col_map.get("actual_net"), "Actual Net"),
        (col_map.get("spur_weight"), "Spurious Weight"),
    ]
    default_count_cols = [
        (col_map.get("gross_diff"), "Gross Diff"),
        (col_map.get("net_diff"), "Net Diff"),
        (col_map.get("ornaments_gdr"), "Ornaments GDR"),
        (col_map.get("ornaments_actual"), "Ornaments Actual"),
        (col_map.get("ornaments_diff"), "Ornaments Diff"),
        (col_map.get("spur_count"), "Spurious Count"),
        (col_map.get("spur_pct"), "Spurious %"),
        (col_map.get("carat_count"), "Carat Mismatch Count"),
        (col_map.get("uncommon_count"), "Uncommon Count"),
    ]
    # These are not "not applicable" on a closed/top-up row — the expected
    # reading is simply the clean one, so they carry their default value
    # rather than being emptied.
    default_text_cols = [
        (col_map.get("magnet"), "Magnet Test", "OK"),
        (col_map.get("tampered"), "Packet Tampered", "NO"),
    ]
    # A top-up has no GDR of its own, so this one genuinely blanks.
    blank_cols = [
        (col_map.get("gdr_no"), "GDR Number"),
    ]
    for r in all_rows:
        is_closed = "CLOSED" in row_status_upper[r]
        is_topup = r in resolved_packet_rows or is_topup_remark(rows_data[r]["remarks"])
        if not (is_closed or is_topup):
            continue
        cleaned_rows.add(r)
        for wc, label in default_weight_cols:
            if wc and num_or_none(val(r, wc)) != 0.0:
                set_val(r, wc, 0.000)
                highlight(r, wc, GREEN_FILL, skip_excluded=False)
                summary["closed_topup_defaulted"].append(f"Row {r}: {label} → 0.000")
        for cc, label in default_count_cols:
            if cc and num_or_none(val(r, cc)) != 0.0:
                set_val(r, cc, 0)
                highlight(r, cc, GREEN_FILL, skip_excluded=False)
                summary["closed_topup_defaulted"].append(f"Row {r}: {label} → 0")
        # The Renewal/Closed date is real loan data on both closed and top-up
        # rows, so it is preserved rather than blanked.
        for tc, label, default in default_text_cols:
            if tc and safe_str(val(r, tc)).upper() != default:
                set_val(r, tc, default)
                highlight(r, tc, GREEN_FILL, skip_excluded=False)
                summary["closed_topup_defaulted"].append(f"Row {r}: {label} → {default}")
        for bc, label in blank_cols:
            if bc:
                v = val(r, bc)
                if v is not None and str(v).strip() != "":
                    set_val(r, bc, None)
                    highlight(r, bc, GREEN_FILL, skip_excluded=False)
                    summary["closed_topup_cleared"].append(f"Row {r}: {label} cleared")

    # ══════════════════════════════════════════════════════════════════
    # PASS 3b: Normalize dates to dd/mm/yyyy
    # Dates arrive as text in half a dozen spellings. Store them as real
    # dates with one display format so the sheet is consistent and anything
    # reading them back (the output filename, the outlier check) can parse
    # them without guessing.
    # SANCTION DATE is deliberately excluded — it comes from the bank's own
    # record and is left exactly as the source has it.
    # ══════════════════════════════════════════════════════════════════
    for date_key in ("verification_date", "renewal_date"):
        dc = col_map.get(date_key)
        if not dc:
            continue
        for r in all_rows:
            cl = cell(r, dc)
            if cl is None:
                continue
            v = cl.value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            if isinstance(v, str) and v.startswith("="):
                continue
            parsed = parse_any_date(v)
            if parsed is None:
                continue
            if not isinstance(v, (datetime, date)):
                cl.value = parsed
                summary["date_reformatted"].append(f"Row {r}: {date_key} '{v}' → {parsed:%d/%m/%Y}")
            cl.number_format = DATE_NUMBER_FORMAT

    # ══════════════════════════════════════════════════════════════════
    # PASS 3c: POA rows carry no Renewal/Closed date
    # An ordinary POA row is an active loan being physically verified, so it
    # has no renewal or closure date. Closed and top-up rows are excluded:
    # there the date is real loan data and must survive (see PASS 3).
    # ══════════════════════════════════════════════════════════════════
    col_renewal = col_map.get("renewal_date")
    col_taf_early = col_map.get("taf")
    if col_renewal and col_taf_early:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            if safe_str(val(r, col_taf_early)).upper() != "POA":
                continue
            v = val(r, col_renewal)
            if v is not None and str(v).strip() != "":
                set_val(r, col_renewal, None)
                summary["poa_renewal_date_cleared"].append(f"Row {r}: Renewal/Closed Date cleared (POA)")

    # ══════════════════════════════════════════════════════════════════
    # PASS 4: Date Outlier Detection (Majority Month/Year Rule)
    # ══════════════════════════════════════════════════════════════════
    date_cols = [
        (col_map.get("sanction_date"), "Sanction Date"),
        (col_map.get("verification_date"), "Agency Verification Date"),
        (col_map.get("renewal_date"), "Renewal/Closed Date"),
    ]

    def get_year_month(v):
        if v is None:
            return None
        if hasattr(v, "year") and hasattr(v, "month"):
            return v.year, v.month
        s = str(v).strip()
        if not s or s.upper() in ("NAN", "NONE", "NAT", "-"):
            return None
        months_map = {
            "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
            "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
        }
        for sep in ["-", "/", ".", " "]:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) == 3:
                p0, p1, p2 = parts[0], parts[1], parts[2]
                if p2.isdigit() and len(p2) in (2, 4):
                    yr = int(p2)
                    if yr < 100:
                        yr += 2000
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return yr, m
                elif p0.isdigit() and len(p0) == 4:
                    yr = int(p0)
                    m = int(p1) if p1.isdigit() else months_map.get(p1.upper()[:3], 1)
                    return yr, m
        return None

    for date_col, label in date_cols:
        if date_col:
            row_ym = {}
            for r in all_rows:
                v = val(r, date_col)
                ym = get_year_month(v)
                if ym:
                    row_ym[r] = ym

            if row_ym:
                ym_counts = defaultdict(int)
                for ym in row_ym.values():
                    ym_counts[ym] += 1

                majority_ym = max(ym_counts, key=ym_counts.get)
                majority_count = ym_counts[majority_ym]

                if len(ym_counts) > 1 and majority_count > len(row_ym) / 2:
                    for r, ym in row_ym.items():
                        if ym != majority_ym:
                            highlight(r, date_col, YELLOW_FILL)
                            val_str = str(val(r, date_col)).strip()
                            summary["date_outlier"].append(
                                f"Row {r}: {label} '{val_str}' deviates from majority month/year {majority_ym[1]}/{majority_ym[0]}"
                            )

    # ══════════════════════════════════════════════════════════════════
    # PASS 4b: State not resolved
    # An "Unknown" state means the branch could not be placed, so the row
    # needs a human to fill it in — flag it wherever it appears, including on
    # closed/top-up rows, since the branch's state applies regardless.
    # ══════════════════════════════════════════════════════════════════
    col_state = col_map.get("state")
    if col_state:
        for r in all_rows:
            v = safe_str(val(r, col_state))
            if v and "UNKNOWN" in v.upper():
                highlight(r, col_state, ORANGE_FILL, skip_excluded=False)
                summary["state_unknown"].append(f"Row {r}: State = '{v}'")

    # ══════════════════════════════════════════════════════════════════
    # PASS 5: Tampered / Magnet
    # ══════════════════════════════════════════════════════════════════
    col_tampered = col_map.get("tampered")
    if col_tampered:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            if safe_str(val(r, col_tampered)).upper() == "YES":
                highlight(r, col_tampered, ORANGE_FILL)
                summary["tampered"].append(f"Row {r}: Packet Tampered = YES")
    col_magnet = col_map.get("magnet")
    if col_magnet:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            v = safe_str(val(r, col_magnet))
            if v.upper().strip() != "OK":
                highlight(r, col_magnet, ORANGE_FILL)
                summary["magnet_not_ok"].append(f"Row {r}: Magnet = '{v}'")

    # ══════════════════════════════════════════════════════════════════
    # PASS 6: Fresh accounts with Actual Gross = 0
    # ══════════════════════════════════════════════════════════════════
    col_taf = col_map.get("taf")
    col_actual_gross = col_map.get("actual_gross")
    if col_actual_gross:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            status = row_status_upper[r]
            if "FRESH" in status:
                ag_n = num_or_none(val(r, col_actual_gross))
                if ag_n is not None and ag_n == 0:
                    highlight(r, col_actual_gross, ORANGE_FILL)
                    summary["fresh_zero_actual_gross"].append(f"Row {r}: Fresh account but Actual Gross = 0")

    # ══════════════════════════════════════════════════════════════════
    # PASS 6b: Absorb small POA weight differences
    # On a POA row a gross/net difference within ±0.200 is weighing noise, not
    # a real shortfall, so the FCU figure is aligned to the GDR figure and the
    # row's discrepancy remark is reduced to whatever is still outstanding.
    # Runs before the diff columns are checked so those see the corrected
    # weights.
    # ══════════════════════════════════════════════════════════════════
    col_gdr_gross_a = col_map.get("gdr_gross")
    col_actual_gross_a = col_map.get("actual_gross")
    col_gdr_net_a = col_map.get("gdr_net")
    col_actual_net_a = col_map.get("actual_net")
    col_taf_a = col_map.get("taf")
    col_remarks_a = col_map.get("remarks")
    if col_taf_a and col_gdr_gross_a and col_actual_gross_a and col_gdr_net_a and col_actual_net_a:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            if safe_str(val(r, col_taf_a)).upper() != "POA":
                continue

            fixed_gross = False
            fixed_net = False

            for gdr_col, actual_col, label, flag in (
                (col_gdr_gross_a, col_actual_gross_a, "Actual Gross", "gross"),
                (col_gdr_net_a, col_actual_net_a, "Actual Net", "net"),
            ):
                gdr_v = num_or_none(val(r, gdr_col))
                actual_v = num_or_none(val(r, actual_col))
                if gdr_v is None or actual_v is None:
                    continue
                difference = gdr_v - actual_v
                if difference == 0 or abs(difference) > WEIGHT_MATCH_TOLERANCE + 1e-9:
                    continue
                set_val(r, actual_col, gdr_v)
                highlight(r, actual_col, GREEN_FILL)
                summary["weight_aligned_to_gdr"].append(
                    f"Row {r}: {label} {actual_v} → {gdr_v} (difference {difference:+.3f})"
                )
                if flag == "gross":
                    fixed_gross = True
                else:
                    fixed_net = True

            if (fixed_gross or fixed_net) and col_remarks_a:
                current = safe_str(val(r, col_remarks_a))
                amended = amend_discrepancy_remark(current, fixed_gross, fixed_net)
                if amended is not None and amended != current:
                    set_val(r, col_remarks_a, amended)
                    summary["discrepancy_remark_amended"].append(
                        f"Row {r}: '{current}' → '{amended}'"
                    )

    # ══════════════════════════════════════════════════════════════════
    # PASS 7: Gross Diff formula check
    # ══════════════════════════════════════════════════════════════════
    col_gdr_gross = col_map.get("gdr_gross")
    col_gross_diff = col_map.get("gross_diff")
    col_tare = col_map.get("tare")
    if col_gross_diff and col_gdr_gross and col_actual_gross and col_tare:
        gdr_g_let = get_column_letter(col_gdr_gross)
        act_g_let = get_column_letter(col_actual_gross)
        tare_let = get_column_letter(col_tare)
        for r in all_rows:
            status = row_status_upper[r]
            if r in cleaned_rows:
                continue

            gdr_g_n = num_or_none(val(r, col_gdr_gross))
            act_g_n = num_or_none(val(r, col_actual_gross))
            tare_n = num_or_none(val(r, col_tare))
            raw_gd = val(r, col_gross_diff)
            gd_n = num_or_none(val_data(r, col_gross_diff))

            taf_val = safe_str(val(r, col_taf)).upper().strip() if col_taf else ""

            formula = None
            expected = None
            if taf_val == "TAF" and gdr_g_n is not None and tare_n is not None:
                formula = f"={gdr_g_let}{r}-{tare_let}{r}"
                expected = round(gdr_g_n - tare_n, 1)
            elif taf_val == "POA" and gdr_g_n is not None and act_g_n is not None:
                formula = f"={gdr_g_let}{r}-{act_g_let}{r}"
                expected = round(gdr_g_n - act_g_n, 1)
            # Fallback if TAF/POA is empty
            elif not taf_val:
                if "FRESH" in status and gdr_g_n is not None and act_g_n is not None:
                    formula = f"={gdr_g_let}{r}-{act_g_let}{r}"
                    expected = round(gdr_g_n - act_g_n, 1)
                elif ("ALREADY VERIFIED" in status or "VERIFIED" in status) and gdr_g_n is not None and tare_n is not None:
                    formula = f"={gdr_g_let}{r}-{tare_let}{r}"
                    expected = round(gdr_g_n - tare_n, 1)

            # Only touch the cell when its current value is wrong or missing —
            # correct cells keep their original content and formatting. A cell
            # already holding a formula with no cached value can't be verified,
            # so it is left as-is.
            if formula:
                if gd_n is None:
                    if not (isinstance(raw_gd, str) and raw_gd.strip().startswith("=")):
                        set_val(r, col_gross_diff, formula)
                elif abs(expected - round(gd_n, 1)) > 0.05:
                    set_val(r, col_gross_diff, formula)
                    highlight(r, col_gross_diff, YELLOW_FILL)
                    summary["gross_diff_fixed"].append(f"Row {r}: Gross Diff {round(gd_n, 1)}→{expected}")

    # ══════════════════════════════════════════════════════════════════
    # PASS 8: Net Diff check
    # ══════════════════════════════════════════════════════════════════
    col_net_diff = col_map.get("net_diff")
    col_gdr_net = col_map.get("gdr_net")
    col_actual_net = col_map.get("actual_net")
    if col_net_diff and col_gdr_net and col_actual_net:
        gdr_n_let = get_column_letter(col_gdr_net)
        act_n_let = get_column_letter(col_actual_net)
        for r in all_rows:
            status = row_status_upper[r]
            if r in cleaned_rows:
                continue

            gdr_n_n = num_or_none(val(r, col_gdr_net))
            act_n_n = num_or_none(val(r, col_actual_net))
            raw_nd = val(r, col_net_diff)
            nd_n = num_or_none(val_data(r, col_net_diff))

            taf_val = safe_str(val(r, col_taf)).upper().strip() if col_taf else ""

            formula = None
            expected = None
            is_taf = False

            if taf_val == "TAF":
                formula = "=0.000"
                expected = 0.0
                is_taf = True
            elif taf_val == "POA" and gdr_n_n is not None and act_n_n is not None:
                formula = f"={gdr_n_let}{r}-{act_n_let}{r}"
                expected = round(gdr_n_n - act_n_n, 1)
            # Fallback if TAF/POA is empty
            elif not taf_val:
                if "ALREADY VERIFIED" in status:
                    formula = "=0.000"
                    expected = 0.0
                    is_taf = True
                elif "FRESH" in status and gdr_n_n is not None and act_n_n is not None:
                    formula = f"={gdr_n_let}{r}-{act_n_let}{r}"
                    expected = round(gdr_n_n - act_n_n, 1)

            # Only rewrite the cell when its current value is wrong or missing.
            if formula:
                if nd_n is None:
                    if not (isinstance(raw_nd, str) and raw_nd.strip().startswith("=")):
                        set_val(r, col_net_diff, formula)
                elif is_taf:
                    if abs(nd_n) > 0.01:
                        set_val(r, col_net_diff, formula)
                        highlight(r, col_net_diff, YELLOW_FILL)
                        summary["net_diff_av_not_zero"].append(f"Row {r}: Net Diff = {nd_n} (TAF case)")
                elif abs(expected - round(nd_n, 1)) > 0.05:
                    set_val(r, col_net_diff, formula)
                    highlight(r, col_net_diff, YELLOW_FILL)
                    summary["net_diff_fixed"].append(f"Row {r}: Net Diff {round(nd_n, 1)}→{expected} (POA case)")

    # ══════════════════════════════════════════════════════════════════
    # PASS 9: Spurious % check
    # ══════════════════════════════════════════════════════════════════
    col_spur_pct = col_map.get("spur_pct")
    col_spur_weight = col_map.get("spur_weight")
    if col_spur_pct and col_spur_weight and col_gdr_gross:
        spur_w_let = get_column_letter(col_spur_weight)
        gdr_g_let = get_column_letter(col_gdr_gross)
        for r in all_rows:
            if r in cleaned_rows:
                continue
            sw_n = num_or_none(val(r, col_spur_weight))
            gg_n = num_or_none(val(r, col_gdr_gross))
            sp_n = num_or_none(val_data(r, col_spur_pct))
            if sw_n is not None and gg_n is not None and gg_n > 0:
                expected_pct = round((sw_n / gg_n) * 100, 2)
                if abs(expected_pct - (sp_n if sp_n is not None else 0.0)) > 0.5:
                    formula = f"=ROUND(({spur_w_let}{r}/{gdr_g_let}{r})*100,2)"
                    set_val(r, col_spur_pct, formula)
                    highlight(r, col_spur_pct, YELLOW_FILL)
                    summary["spur_pct_mismatch"].append(f"Row {r}: Expected {expected_pct}%, got {sp_n if sp_n is not None else 0.0}%")

    # ══════════════════════════════════════════════════════════════════
    # PASS 10: Ornament count diff
    # ══════════════════════════════════════════════════════════════════
    col_ornaments_diff = col_map.get("ornaments_diff")
    col_ornaments_gdr = col_map.get("ornaments_gdr")
    col_ornaments_actual = col_map.get("ornaments_actual")
    if col_ornaments_diff and col_ornaments_gdr and col_ornaments_actual and col_taf:
        act_orn_let = get_column_letter(col_ornaments_actual)
        gdr_orn_let = get_column_letter(col_ornaments_gdr)
        for r in all_rows:
            if r in cleaned_rows:
                continue

            taf_val = safe_str(val(r, col_taf)).upper()
            if taf_val == "TAF":
                continue
            is_poa = taf_val == "POA"
            go_n = num_or_none(val_data(r, col_ornaments_gdr))
            ao_n = num_or_none(val_data(r, col_ornaments_actual))
            if go_n is not None and ao_n is not None:
                expected_diff = int(round(ao_n)) - int(round(go_n))

                stored = num_or_none(val_data(r, col_ornaments_diff))
                if stored is None or int(round(stored)) != expected_diff:
                    formula = f"={act_orn_let}{r}-{gdr_orn_let}{r}"
                    set_val(r, col_ornaments_diff, formula)

                if expected_diff != 0:
                    highlight(r, col_ornaments_diff, YELLOW_FILL)
                    if not is_poa:
                        highlight(r, col_ornaments_gdr, YELLOW_FILL)
                        highlight(r, col_ornaments_actual, YELLOW_FILL)
                    summary["ornament_diff_fixed"].append(f"Row {r}: Set to {expected_diff}")

    # ── TAF rows: Ornaments Diff = 0 (no highlights) ──
    if col_ornaments_diff and col_taf:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            taf_val = safe_str(val(r, col_taf)).upper()
            if taf_val == "TAF" and num_or_none(val_data(r, col_ornaments_diff)) != 0.0:
                set_val(r, col_ornaments_diff, 0)

    # ══════════════════════════════════════════════════════════════════
    # PASS 11: Highlight zero/non-positive ornaments GDR
    # ══════════════════════════════════════════════════════════════════
    if col_ornaments_gdr:
        for r in all_rows:
            if r in cleaned_rows:
                continue
            n = num_or_none(val(r, col_ornaments_gdr))
            if n is not None and n <= 0:
                highlight(r, col_ornaments_gdr, YELLOW_FILL)
                summary["ornaments_gdr_non_positive"].append(f"Row {r}: Ornaments GDR = {val(r, col_ornaments_gdr)}")

    # ══════════════════════════════════════════════════════════════════
    # PASS 12: Whitespace normalization
    # ══════════════════════════════════════════════════════════════════
    text_cols = [
        col_map.get("account"), col_map.get("applicant"), col_packet,
        col_map.get("gdr_no"), col_remarks, col_taf,
    ]
    for tc in text_cols:
        if tc:
            for r in all_rows:
                v = val(r, tc)
                if isinstance(v, str) and v != v.strip():
                    set_val(r, tc, v.strip())
                    summary["whitespace_fixed"].append(f"Row {r} col {tc}: trimmed")

    # ══════════════════════════════════════════════════════════════════
    # PASS 13: Replace NA with 0.000 in Tare column
    # ══════════════════════════════════════════════════════════════════
    if col_tare:
        for r in all_rows:
            v = val(r, col_tare)
            if isinstance(v, str) and v.strip().upper() in ("NA", "N/A"):
                set_val(r, col_tare, 0.000)

    # ══════════════════════════════════════════════════════════════════
    # PASS 14: Weight values - flag suspiciously large/small
    # ══════════════════════════════════════════════════════════════════
    weight_check_cols = [
        (col_gdr_gross, "GDR Gross"),
        (col_actual_gross, "Actual Gross"),
        (col_tare, "Tare"),
        (col_gdr_net, "GDR Net"),
        (col_actual_net, "Actual Net"),
    ]
    for wc, label in weight_check_cols:
        if wc:
            for r in all_rows:
                v = val(r, wc)
                if v is not None:
                    try:
                        nv = float(v)
                        if nv > 5000:
                            highlight(r, wc, YELLOW_FILL)
                            summary["weight_outlier"].append(f"Row {r}: {label} = {nv} (suspiciously large)")
                        if nv < 0 and wc not in (col_gross_diff, col_net_diff):
                            highlight(r, wc, YELLOW_FILL)
                            summary["weight_negative"].append(f"Row {r}: {label} = {nv} (negative weight)")
                    except (ValueError, TypeError):
                        pass

    # ══════════════════════════════════════════════════════════════════
    # PASS 15: Gross weight must not be less than Net weight
    # ══════════════════════════════════════════════════════════════════
    gross_net_pairs = [
        (col_gdr_gross, col_gdr_net, "GDR Gross", "GDR Net"),
        (col_actual_gross, col_actual_net, "Actual Gross", "Actual Net"),
    ]
    for gross_col, net_col, gross_label, net_label in gross_net_pairs:
        if gross_col and net_col:
            for r in all_rows:
                g = val(r, gross_col)
                n = val(r, net_col)
                if g is None or n is None:
                    continue
                try:
                    g_n = float(g)
                    n_n = float(n)
                except (ValueError, TypeError):
                    continue
                if g_n < n_n:
                    highlight(r, gross_col, YELLOW_FILL)
                    highlight(r, net_col, YELLOW_FILL)
                    summary["gross_less_than_net"].append(
                        f"Row {r}: {gross_label} ({g_n}) < {net_label} ({n_n})"
                    )

    # ══════════════════════════════════════════════════════════════════
    # PASS 16: Tare weight must not be less than GDR Gross/Net weight
    # The tare (sealed packet incl. packaging) can never weigh less than the
    # declared contents. Skipped when tare is 0/NA (e.g. POA rows where the
    # packet was opened) and on cleaned Closed/Top-Up rows.
    # ══════════════════════════════════════════════════════════════════
    if col_tare:
        tare_weight_pairs = [
            (col_gdr_gross, "GDR Gross"),
            (col_gdr_net, "GDR Net"),
        ]
        for r in all_rows:
            if r in cleaned_rows:
                continue
            t_n = num_or_none(val(r, col_tare))
            if t_n is None or t_n <= 0:
                continue
            for wcol, wlabel in tare_weight_pairs:
                if not wcol:
                    continue
                w_n = num_or_none(val(r, wcol))
                if w_n is not None and t_n < w_n:
                    highlight(r, col_tare, YELLOW_FILL)
                    highlight(r, wcol, YELLOW_FILL)
                    summary["tare_less_than_weight"].append(
                        f"Row {r}: Tare ({t_n}) < {wlabel} ({w_n})"
                    )

    return summary



# ─── Main ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    import glob

    parser = argparse.ArgumentParser(description="Report Validator")
    parser.add_argument("files", nargs="*", help="Excel files to process")
    parser.add_argument("--pdf", help="Optional PDF sequence file")
    args = parser.parse_args()

    files = args.files
    if not files:
        # Auto-detect Excel files in current directory
        files = glob.glob("*_*.xlsx")
        files = [f for f in files if "_VALIDATED" not in f]

    if not files:
        print("Usage: python report_validator.py <file1.xlsx> [file2.xlsx ...] [--pdf sequence.pdf]")
        sys.exit(1)

    all_summaries = {}
    for fp in files:
        s = process_file(fp, pdf_path=args.pdf)
        if s:
            all_summaries[fp] = s

    print(f"\n\n{'='*80}")
    print("SUMMARY ACROSS ALL FILES:")
    print(f"{'='*80}")
    grand_total = 0
    for fp, s in all_summaries.items():
        total = sum(len(v) for v in s.values())
        grand_total += total
        print(f"  {fp}: {total} issues")
    print("  ─────────────────────")
    print(f"  TOTAL: {grand_total} issues across {len(all_summaries)} files")
